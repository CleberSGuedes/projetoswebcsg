from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
import pandas as pd
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from models import db
from openpyxl.styles import Font

BATCH_SIZE = 200
UPLOAD_DIR = Path("upload") / "fip_613"
OUTPUT_DIR = Path("outputs") / "fip_613"


def ensure_dirs():
    for base in (UPLOAD_DIR, OUTPUT_DIR, UPLOAD_DIR / "tmp", OUTPUT_DIR / "tmp"):
        base.mkdir(parents=True, exist_ok=True)


def move_existing_to_tmp(base_dir: Path):
    tmp = base_dir / "tmp"
    tmp.mkdir(parents=True, exist_ok=True)
    for f in base_dir.glob("*.xlsx"):
        dest = tmp / f"{f.stem}_{datetime.now().strftime('%Y%m%d%H%M%S')}{f.suffix}"
        try:
            f.rename(dest)
        except OSError:
            pass


def get_year_from_file(file_path, sheet_name="FIPLAN"):
    try:
        raw_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        for _, row in raw_data.iterrows():
            for cell in row:
                if isinstance(cell, str) and "Exercício igual a" in cell:
                    year = int(cell.split()[-1])
                    return year
    except Exception:
        return None
    return None


def load_clean_data(file_path, sheet_name="FIPLAN"):
    raw_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    header_row_index = None
    for i, row in raw_data.iterrows():
        if "UO" in row.values and "UG" in row.values:
            header_row_index = i
            break
    if header_row_index is None:
        return None

    data = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)
    data = data.dropna(how="all").reset_index(drop=True)
    data = data.dropna(subset=["UO", "UG", "Função", "Subfunção", "Programa", "Projeto/Atividade"])

    total_row_index = data[
        data.apply(lambda row: row.astype(str).str.contains("Total UO 14101").any(), axis=1)
    ].index
    if not total_row_index.empty:
        data = data.iloc[: total_row_index[0]]

    data.rename(
        columns={
            "UO": "uo",
            "UG": "ug",
            "Função": "funcao",
            "Subfunção": "subfuncao",
            "Programa": "programa",
            "Projeto/Atividade": "projeto_atividade",
            "Regional": "regional",
            "Natureza de Despesa": "natureza_despesa",
            "Fonte de Recurso": "fonte_recurso",
            "Iduso": "iduso",
            "Tipo de Recurso": "tipo_recurso",
            "Dotação Inicial": "dotacao_inicial",
            "Créd. Suplementar": "cred_suplementar",
            "Créd. Especial": "cred_especial",
            "Créd. Extraordinário": "cred_extraordinario",
            "Redução": "reducao",
            "Créd. Autorizado": "cred_autorizado",
            "Bloqueado/Conting.": "bloqueado_conting",
            "Reserva Empenho": "reserva_empenho",
            "Saldo de Destaque": "saldo_destaque",
            "Saldo Dotação": "saldo_dotacao",
            "Empenhado": "empenhado",
            "Liquidado": "liquidado",
            "A liquidar": "a_liquidar",
            "Valor Pago": "valor_pago",
            "Valor a Pagar": "valor_a_pagar",
        },
        inplace=True,
    )

    numeric_columns = [
        "dotacao_inicial",
        "cred_suplementar",
        "cred_especial",
        "cred_extraordinario",
        "reducao",
        "cred_autorizado",
        "bloqueado_conting",
        "reserva_empenho",
        "saldo_destaque",
        "saldo_dotacao",
        "empenhado",
        "liquidado",
        "a_liquidar",
        "valor_pago",
        "valor_a_pagar",
    ]

    for col in numeric_columns:
        data[col] = data[col].astype(str).str.replace(".", "", regex=False)
        data[col] = data[col].str.replace(",", ".", regex=False)
        data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0.0)

    data["iduso"] = pd.to_numeric(data["iduso"], errors="coerce").fillna(0).astype(int)
    # manter natureza/fonte como texto (evita notação científica)
    for col in ["natureza_despesa", "fonte_recurso"]:
        data[col] = data[col].apply(lambda v: str(v).split(".")[0] if pd.notna(v) else "")
    return data


def save_clean_data(data, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    move_existing_to_tmp(output_dir)
    filename = f"fip613_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = output_dir / filename
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        data.to_excel(writer, index=False, sheet_name="FIP613")
        ws = writer.sheets["FIP613"]
        # definir largura básica das colunas
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 40)
        # formato numérico
        num_fmt = "#,##0.00"
        for col in range(12, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = num_fmt
        # fonte e autofiltro
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Helvetica", size=8)
        ws.auto_filter.ref = ws.dimensions
    return output_path


def update_database(data, ano, data_arquivo, user_email, upload_id):
    insert_sql = text(
        """
        INSERT INTO fip613 (
            upload_id, uo, ug, funcao, subfuncao, programa, projeto_atividade, regional, natureza_despesa,
            fonte_recurso, iduso, tipo_recurso, dotacao_inicial, cred_suplementar, cred_especial,
            cred_extraordinario, reducao, cred_autorizado, bloqueado_conting, reserva_empenho,
            saldo_destaque, saldo_dotacao, empenhado, liquidado, a_liquidar, valor_pago,
            valor_a_pagar, data_atualizacao, ano, data_arquivo, user_email, ativo
        )
        VALUES (
            :upload_id, :uo, :ug, :funcao, :subfuncao, :programa, :projeto_atividade, :regional, :natureza_despesa,
            :fonte_recurso, :iduso, :tipo_recurso, :dotacao_inicial, :cred_suplementar, :cred_especial,
            :cred_extraordinario, :reducao, :cred_autorizado, :bloqueado_conting, :reserva_empenho,
            :saldo_destaque, :saldo_dotacao, :empenhado, :liquidado, :a_liquidar, :valor_pago,
            :valor_a_pagar, :data_atualizacao, :ano, :data_arquivo, :user_email, :ativo
        )
        """
    )

    # desativa versões anteriores
    try:
        db.session.execute(text("UPDATE fip613 SET ativo = 0 WHERE ativo = 1"))
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        raise

    rows = data.to_dict(orient="records")
    total = 0
    for start in range(0, len(rows), BATCH_SIZE):
        chunk = rows[start : start + BATCH_SIZE]
        for r in chunk:
            r["data_atualizacao"] = datetime.utcnow()
            r["ano"] = ano
            r["data_arquivo"] = data_arquivo
            r["user_email"] = user_email
            r["upload_id"] = upload_id
            r["ativo"] = True
        try:
            db.session.execute(insert_sql, chunk)
            db.session.commit()
            total += len(chunk)
        except SQLAlchemyError as exc:
            db.session.rollback()
            raise exc
    return total


def run_fip613(file_path: Path, data_arquivo: datetime, user_email: str, upload_id: int) -> tuple[int, Path]:
    ensure_dirs()
    ano = get_year_from_file(file_path)
    data = load_clean_data(file_path)
    if data is None or ano is None:
        raise RuntimeError("Não foi possível ler o arquivo FIP 613 (cabeçalho ou ano ausente).")

    output_path = save_clean_data(data, OUTPUT_DIR)
    total = update_database(data, ano, data_arquivo, user_email, upload_id)
    return total, output_path
