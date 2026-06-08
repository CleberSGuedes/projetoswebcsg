from flask import Blueprint, jsonify, render_template, request, abort, g, session, send_file, current_app
from functools import wraps, lru_cache
import re
from datetime import datetime, timedelta
from decimal import Decimal
import os
from io import BytesIO
import json
import unicodedata
import subprocess
import sys
import threading
import secrets
import hashlib
import pytz
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
from models import (
    Usuario,
    Perfil,
    PerfilPermissao,
    NivelPermissao,
    Fip613Upload,
    Fip613Registro,
    Plan20Upload,
    PedUpload,
    PedRegistro,
    EmpUpload,
    EmpRegistro,
    EstEmpUpload,
    EstEmpRegistro,
    NobUpload,
    NobRegistro,
    Plan21Nger,
    Adj,
    Dotacao,
    CadastrarSubacao,
    CadastrarEtapa,
    AlterarMeta,
    AlterarMetaItem,
    ChavePlanejamentoRegra,
    ApiClient,
    ApiClientScope,
    ApiRefreshToken,
    ApiAccessLog,
    ApiKey,
    ActiveSession,
    db,
)
from sqlalchemy.exc import (
    ProgrammingError,
    IntegrityError,
    NoSuchColumnError,
    OperationalError,
    ResourceClosedError,
    SQLAlchemyError,
)
from services.auth import login_required, role_required, current_user
from services.emp_record import get_emp_record_snapshot
from services.features import FEATURES, flatten_features, build_parent_map
from services.fip613_runner import run_fip613, UPLOAD_DIR
from services.plan20_runner import run_plan20
from services.ped_runner import (
    run_ped,
    move_existing_to_tmp,
    INPUT_DIR as PED_UPLOAD_DIR,
    OUTPUT_DIR as PED_OUTPUT_DIR,
)
from services.est_emp_runner import (
    run_est_emp,
    INPUT_DIR as EST_EMP_UPLOAD_DIR,
    OUTPUT_DIR as EST_EMP_OUTPUT_DIR,
    move_existing_to_tmp as move_est_emp_existing_to_tmp,
)
from services.job_status import read_status, set_cancel_flag, update_status_fields, write_status
from pathlib import Path
from sqlalchemy import text, func, or_, select, inspect
from urllib.parse import unquote

home_bp = Blueprint("home", __name__)

EMP_UPLOAD_DIR = Path("upload/emp")
EMP_OUTPUT_DIR = Path("outputs/td_emp")
NOB_UPLOAD_DIR = Path("upload/nob")
NOB_OUTPUT_DIR = Path("outputs/td_nob")
NODE_RUNNER = Path(__file__).resolve().parents[1] / "node_runners" / "run.js"
NODE_EXE = os.getenv("NODE_EXE", "node")
META_FISICA_OPTIONS_CACHE_TTL_S = max(
    0, int(os.getenv("META_FISICA_OPTIONS_CACHE_TTL_S", "60") or "60")
)
_meta_fisica_options_cache_lock = threading.Lock()
_meta_fisica_options_cache: dict[tuple, tuple[float, dict]] = {}
_etapa_options_cache_lock = threading.Lock()
_etapa_options_cache: dict[tuple, tuple[float, dict]] = {}


def _cache_get_meta_fisica_options(cache_key: tuple):
    if META_FISICA_OPTIONS_CACHE_TTL_S <= 0:
        return None
    now_ts = datetime.utcnow().timestamp()
    with _meta_fisica_options_cache_lock:
        item = _meta_fisica_options_cache.get(cache_key)
        if not item:
            return None
        expires_at, payload = item
        if expires_at <= now_ts:
            _meta_fisica_options_cache.pop(cache_key, None)
            return None
        # cópia rasa para evitar mutação acidental fora do cache
        return dict(payload)


def _cache_set_meta_fisica_options(cache_key: tuple, payload: dict) -> None:
    if META_FISICA_OPTIONS_CACHE_TTL_S <= 0:
        return
    expires_at = datetime.utcnow().timestamp() + META_FISICA_OPTIONS_CACHE_TTL_S
    with _meta_fisica_options_cache_lock:
        _meta_fisica_options_cache[cache_key] = (expires_at, dict(payload))


def _cache_invalidate_meta_fisica_options() -> None:
    with _meta_fisica_options_cache_lock:
        _meta_fisica_options_cache.clear()


def _cache_get_etapa_options(cache_key: tuple):
    if META_FISICA_OPTIONS_CACHE_TTL_S <= 0:
        return None
    now_ts = datetime.utcnow().timestamp()
    with _etapa_options_cache_lock:
        item = _etapa_options_cache.get(cache_key)
        if not item:
            return None
        expires_at, payload = item
        if expires_at <= now_ts:
            _etapa_options_cache.pop(cache_key, None)
            return None
        return dict(payload)


def _cache_set_etapa_options(cache_key: tuple, payload: dict) -> None:
    if META_FISICA_OPTIONS_CACHE_TTL_S <= 0:
        return
    expires_at = datetime.utcnow().timestamp() + META_FISICA_OPTIONS_CACHE_TTL_S
    with _etapa_options_cache_lock:
        _etapa_options_cache[cache_key] = (expires_at, dict(payload))


def _cache_invalidate_etapa_options() -> None:
    with _etapa_options_cache_lock:
        _etapa_options_cache.clear()


def _find_upload_path(base_dir: Path, stored_filename: str) -> Path | None:
    if not stored_filename:
        return None
    candidate = base_dir / stored_filename
    if candidate.exists():
        return candidate
    tmp_dir = base_dir / "tmp"
    if not tmp_dir.exists():
        return None
    stem = Path(stored_filename).stem
    matches = sorted(tmp_dir.glob(f"{stem}_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0] if matches else None


def _normalize_header_token(value: object) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    text = text.replace("º", "O").replace("°", "O")
    text = re.sub(r"[^A-Z0-9]+", " ", text).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _detect_upload_layouts(file_path: Path) -> set[str]:
    try:
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None, dtype=str, nrows=260)
    except Exception:
        return set()
    if df_raw is None or df_raw.empty:
        return set()

    rows_tokens: list[set[str]] = []
    all_tokens: list[str] = []
    for _, row in df_raw.iterrows():
        vals = [str(v).strip() if pd.notna(v) else "" for v in row.tolist()]
        vals = [v for v in vals if v]
        if not vals:
            continue
        tokens = {_normalize_header_token(v) for v in vals if v}
        if not tokens:
            continue
        rows_tokens.append(tokens)
        all_tokens.extend(tokens)

    if not rows_tokens:
        return set()

    blob = " ".join(all_tokens)
    detected: set[str] = set()

    def _has(tokens: set[str], token: str) -> bool:
        return token in tokens

    def _has_prefix(tokens: set[str], prefix: str) -> bool:
        return any(t.startswith(prefix) for t in tokens)

    for tokens in rows_tokens:
        # NOB: aceita tanto layout clássico (NOB) quanto layout GCV/NOB.
        nob_core = _has(tokens, "N NOB") and (
            _has(tokens, "N NOB ESTORNO ESTORNADO")
            or _has(tokens, "VALOR NOB")
            or _has(tokens, "N GCV")
            or _has(tokens, "VALOR DA GCV")
        )
        nob_aux = (
            _has(tokens, "N LIQ")
            or _has(tokens, "DATA NOB")
            or _has(tokens, "DATA CADASTRO NOB")
            or _has(tokens, "DATA DA GCV")
            or _has(tokens, "HISTORICO DA GCV")
            or _has(tokens, "N NOBLIST")
            or _has(tokens, "SITUACAO GCV")
        )
        if nob_core and nob_aux:
            detected.add("nob")

        has_exerc = any(t.startswith("EXERC") for t in tokens)
        if not has_exerc:
            # FIP613 costuma não ter "Exercício" no mesmo padrão.
            fip_markers = {"UO", "UG", "FUNCAO", "SUBFUNCAO", "PROGRAMA", "PROJETO ATIVIDADE"}
            if len(fip_markers.intersection(tokens)) >= 5:
                detected.add("fip613")
            continue

        # PED
        ped_core = _has(tokens, "N PED ESTORNO ESTORNADO") or _has(tokens, "VALOR PED")
        ped_aux = {"N PED", "N EMP", "N CAD", "N NOBLIST", "N OS", "VALOR DO ESTORNO"}
        if ped_core and len(ped_aux.intersection(tokens)) >= 3:
            detected.add("ped")

        # EMP
        emp_core = _has(tokens, "N EMP") and _has(tokens, "N PED") and _has(tokens, "VALOR EMP")
        emp_forbidden = _has(tokens, "N PED ESTORNO ESTORNADO") or _has(tokens, "VALOR PED")
        emp_aux = {"HISTORICO", "DOTACAO ORCAMENTARIA", "TIPO EMPENHO", "SITUACAO", "DEVOLUCAO GCV"}
        if emp_core and (not emp_forbidden) and len(emp_aux.intersection(tokens)) >= 2:
            detected.add("emp")

        # EST_EMP
        est_emp_core = _has(tokens, "N EST") and _has(tokens, "N EMP") and _has(tokens, "N PED")
        est_emp_aux = _has_prefix(tokens, "VALOR EST EMP")
        if est_emp_core and est_emp_aux:
            detected.add("est_emp")

    if (
        "EXERCICIO IGUAL A" in blob
        and "PROGRAMA" in blob
        and ("ACAO P A OE" in blob or "ACAO PAOE" in blob)
        and ("PUBLICO TRANSVERSAL" in blob or "PLANO DE ACAO POR PRODUTO" in blob)
    ):
        detected.add("plan20")

    return detected


def _validate_upload_layout(file_path: Path, expected_kind: str, expected_label: str) -> str | None:
    detected = _detect_upload_layouts(file_path)
    # Alguns layouts compartilham colunas por natureza (ex.: EST_EMP herda vários
    # campos de EMP). Permitimos somente as sobreposições explicitamente seguras.
    allowed_sets: dict[str, set[str]] = {
        "fip613": {"fip613"},
        "ped": {"ped"},
        "emp": {"emp"},
        "est_emp": {"est_emp", "emp"},
        "nob": {"nob"},
        "plan20": {"plan20"},
    }
    allowed = allowed_sets.get(expected_kind, {expected_kind})

    if expected_kind in detected and detected.issubset(allowed):
        return None

    if not detected:
        detected_text = "desconhecido"
    else:
        display_map = {
            "ped": "PED",
            "emp": "EMP",
            "est_emp": "EST_EMP",
            "nob": "NOB",
            "fip613": "FIP613",
            "plan20": "PLAN20",
        }
        detected_text = ", ".join(display_map.get(k, k.upper()) for k in sorted(detected))
    if len(detected) > 1:
        return (
            f"Arquivo invalido para {expected_label}: layout ambiguo detectado ({detected_text}). "
            f"Envie um arquivo exclusivo de {expected_label}."
        )
    return (
        f"Arquivo invalido para {expected_label}: layout detectado ({detected_text}). "
        f"Envie um arquivo de {expected_label}."
    )


def _run_node(kind: str, file_path: Path, user_email: str, data_arquivo, upload_id: int) -> dict:
    node_env = os.environ.copy()
    node_max_old_space_mb = os.getenv("NODE_MAX_OLD_SPACE_MB", "4096").strip()
    if node_max_old_space_mb.isdigit() and int(node_max_old_space_mb) > 0:
        extra_opt = f"--max-old-space-size={int(node_max_old_space_mb)}"
        existing_opts = str(node_env.get("NODE_OPTIONS") or "").strip()
        if extra_opt not in existing_opts:
            node_env["NODE_OPTIONS"] = f"{existing_opts} {extra_opt}".strip()

    args = [
        NODE_EXE,
        str(NODE_RUNNER),
        "--kind",
        kind,
        "--file",
        str(file_path),
        "--upload-id",
        str(upload_id),
        "--user-email",
        user_email or "desconhecido",
    ]
    if data_arquivo:
        try:
            args.extend(["--data-arquivo", data_arquivo.isoformat()])
        except Exception:
            args.extend(["--data-arquivo", str(data_arquivo)])
    proc = subprocess.run(
        args,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        cwd=str(NODE_RUNNER.parent),
        env=node_env,
    )
    if proc.returncode != 0:
        err = (proc.stderr or proc.stdout or "").strip()
        raise RuntimeError(f"Node runner falhou: {err or 'erro desconhecido'}")
    raw = (proc.stdout or "").strip()
    try:
        payload = json.loads(raw) if raw else {}
    except json.JSONDecodeError as exc:
        parsed = None
        for line in reversed(raw.splitlines()):
            text = line.strip()
            if not text:
                continue
            try:
                parsed = json.loads(text)
                break
            except Exception:
                continue
        if parsed is None:
            raise RuntimeError(f"Resposta invalida do Node: {exc}") from exc
        payload = parsed
    if not payload.get("ok"):
        raise RuntimeError(f"Node runner falhou: {payload.get('error')}")
    return payload


def _move_existing_to_tmp(base_dir: Path) -> None:
    tmp = base_dir / "tmp"
    tmp.mkdir(parents=True, exist_ok=True)
    for f in base_dir.glob("*.xlsx"):
        dest = tmp / f"{f.stem}_{datetime.now().strftime('%Y%m%d%H%M%S')}{f.suffix}"
        try:
            f.rename(dest)
        except OSError:
            pass


def _send_excel_bytes(buffer: BytesIO, filename: str):
    buffer.seek(0)
    resp = current_app.response_class(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


def _next_pk(model) -> int:
    max_id = db.session.query(func.max(model.id)).scalar() or 0
    return int(max_id) + 1


def _row_value(row, key: str, index: int | None = None):
    if row is None:
        return None
    mapping = getattr(row, "_mapping", None)
    if mapping is not None and key in mapping:
        return mapping[key]
    try:
        return getattr(row, key)
    except Exception:
        pass
    if index is not None:
        try:
            return row[index]
        except Exception:
            return None
    return None


def _safe_session_rollback() -> None:
    try:
        db.session.rollback()
    except Exception:
        pass
    try:
        db.session.remove()
    except Exception:
        pass
    try:
        db.engine.dispose()
    except Exception:
        pass


def _execute_with_retry(stmt, attempts: int | None = None, backoff_s: float | None = None):
    attempts = attempts or int(os.getenv("DB_RETRY_ATTEMPTS", "2"))
    backoff_s = backoff_s if backoff_s is not None else float(os.getenv("DB_RETRY_BACKOFF", "0.2"))
    for idx in range(max(1, attempts)):
        try:
            return db.session.execute(stmt)
        except (OperationalError, ResourceClosedError) as exc:
            _safe_session_rollback()
            try:
                db.engine.dispose()
            except Exception:
                pass
            try:
                with db.engine.connect() as conn:
                    return conn.execute(stmt)
            except Exception:
                if idx < attempts - 1:
                    try:
                        current_app.logger.warning(
                            "DB retry %s/%s failed: %s",
                            idx + 1,
                            attempts,
                            exc,
                        )
                    except Exception:
                        pass
                    try:
                        import time

                        time.sleep(backoff_s * (idx + 1))
                    except Exception:
                        pass
                continue
    return None


def _fetch_all_with_retry(stmt, attempts: int | None = None, backoff_s: float | None = None):
    attempts = attempts or int(os.getenv("DB_RETRY_ATTEMPTS", "2"))
    backoff_s = backoff_s if backoff_s is not None else float(os.getenv("DB_RETRY_BACKOFF", "0.2"))
    for idx in range(max(1, attempts)):
        try:
            return db.session.execute(stmt).all()
        except (OperationalError, ResourceClosedError, SQLAlchemyError, IndexError) as exc:
            _safe_session_rollback()
            try:
                db.engine.dispose()
            except Exception:
                pass
            if idx < attempts - 1:
                try:
                    current_app.logger.warning(
                        "DB fetch retry %s/%s failed: %s",
                        idx + 1,
                        attempts,
                        exc,
                    )
                except Exception:
                    pass
                try:
                    import time

                    time.sleep(backoff_s * (idx + 1))
                except Exception:
                    pass
    return []


def _fetch_scalars_all_with_retry(stmt, attempts: int | None = None, backoff_s: float | None = None):
    attempts = attempts or int(os.getenv("DB_RETRY_ATTEMPTS", "2"))
    backoff_s = backoff_s if backoff_s is not None else float(os.getenv("DB_RETRY_BACKOFF", "0.2"))
    for idx in range(max(1, attempts)):
        try:
            return db.session.execute(stmt).scalars().all()
        except (OperationalError, ResourceClosedError, SQLAlchemyError, IndexError) as exc:
            _safe_session_rollback()
            try:
                db.engine.dispose()
            except Exception:
                pass
            if idx < attempts - 1:
                try:
                    current_app.logger.warning(
                        "DB scalar fetch retry %s/%s failed: %s",
                        idx + 1,
                        attempts,
                        exc,
                    )
                except Exception:
                    pass
                try:
                    import time

                    time.sleep(backoff_s * (idx + 1))
                except Exception:
                    pass
    return []


def _debug_probe(event: str, **fields) -> None:
    if os.getenv("AUTH_DEBUG_PRINTS", "false").strip().lower() != "true":
        return
    try:
        ts = datetime.utcnow().isoformat()
        payload = " ".join(f"{k}={fields[k]}" for k in sorted(fields))
        print(f"[AUTH_DEBUG] ts={ts} event={event} {payload}".strip(), flush=True)
    except Exception:
        pass


def _process_emp_upload(upload_id: int) -> None:
    registro = db.session.get(EmpUpload, upload_id)
    if not registro:
        raise RuntimeError(f"Upload EMP nao encontrado: {upload_id}")
    file_path = _find_upload_path(EMP_UPLOAD_DIR, registro.stored_filename)
    if not file_path:
        raise RuntimeError(f"Arquivo EMP nao encontrado: {EMP_UPLOAD_DIR / registro.stored_filename}")
    payload = _run_node("emp", file_path, registro.user_email, registro.data_arquivo, registro.id)
    registro.output_filename = str(payload.get("output_filename") or "")
    db.session.commit()
    write_status(
        "emp",
        upload_id,
        "processamento finalizado",
        f"Processado com sucesso. Registros: {payload.get('total')}.",
        payload.get("output_filename"),
        progress=100,
    )


def _process_nob_upload(upload_id: int) -> None:
    registro = db.session.get(NobUpload, upload_id)
    if not registro:
        raise RuntimeError(f"Upload NOB nao encontrado: {upload_id}")
    file_path = _find_upload_path(NOB_UPLOAD_DIR, registro.stored_filename)
    if not file_path:
        raise RuntimeError(f"Arquivo NOB nao encontrado: {NOB_UPLOAD_DIR / registro.stored_filename}")
    payload = _run_node("nob", file_path, registro.user_email, registro.data_arquivo, registro.id)
    registro.output_filename = str(payload.get("output_filename") or "")
    db.session.commit()
    write_status(
        "nob",
        upload_id,
        "processamento finalizado",
        f"Processado com sucesso. Registros: {payload.get('total')}.",
        payload.get("output_filename"),
        progress=100,
    )


def _start_thread(kind: str, upload_id: int) -> None:
    app = current_app._get_current_object()

    def _runner() -> None:
        with app.app_context():
            try:
                write_status(
                    kind,
                    upload_id,
                    "em processamento",
                    "Processamento iniciado (thread).",
                    progress=0,
                )
                if kind == "emp":
                    _process_emp_upload(upload_id)
                else:
                    _process_nob_upload(upload_id)
            except Exception as exc:
                msg = f"{type(exc).__name__}: {exc}"
                if "PROCESSAMENTO_CANCELADO" in msg:
                    write_status(kind, upload_id, "processamento cancelado", "Cancelado pelo usuario.")
                else:
                    write_status(kind, upload_id, "falha no processamento", msg)
            finally:
                db.session.remove()

    thread = threading.Thread(target=_runner, daemon=True)
    thread.start()


def _start_worker(kind: str, upload_id: int) -> None:
    worker_path = Path(__file__).resolve().parents[1] / "worker.py"
    creationflags = 0
    if sys.platform.startswith("win"):
        creationflags = getattr(subprocess, "DETACHED_PROCESS", 0) | getattr(
            subprocess, "CREATE_NEW_PROCESS_GROUP", 0
        )
    try:
        log_dir = Path("outputs") / "status"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / f"worker_{kind}_{upload_id}.log"
        log_handle = open(log_path, "a", encoding="utf-8")
        proc = subprocess.Popen(
            [sys.executable, str(worker_path), "--kind", kind, "--upload-id", str(upload_id)],
            cwd=str(worker_path.parent),
            stdout=log_handle,
            stderr=log_handle,
            creationflags=creationflags,
        )
        update_status_fields(
            kind,
            upload_id,
            message="Worker externo iniciado.",
            pid=proc.pid,
        )
    except Exception as exc:
        update_status_fields(
            kind,
            upload_id,
            message=f"Falha ao iniciar worker externo ({type(exc).__name__}: {exc}). Usando thread.",
        )
        _start_thread(kind, upload_id)


@home_bp.route("/")
@login_required
def index():
    # initial_content tells JS which partial to load first
    allowed = _permissoes_with_parents(
        getattr(g, "user_perfil_id", None),
        getattr(g, "user_nivel", None),
    )
    return render_template("base.html", initial_content="dashboard", initial_features=allowed)


@home_bp.route("/partial/dashboard")
@login_required
def partial_dashboard():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    nivel_raw = getattr(g, "user_nivel", 99)
    try:
        nivel_int = int(nivel_raw)
    except (TypeError, ValueError):
        nivel_int = 99
    warn_ms = int(os.getenv("DB_SLOW_QUERY_MS", "2000"))

    def _log_timing(label: str, started: datetime, count: int):
        try:
            elapsed_ms = int((datetime.utcnow() - started).total_seconds() * 1000)
            if elapsed_ms >= warn_ms:
                current_app.logger.warning(
                    "dashboard %s rows=%s ms=%s (slow)",
                    label,
                    count,
                    elapsed_ms,
                )
            else:
                current_app.logger.info(
                    "dashboard %s rows=%s ms=%s",
                    label,
                    count,
                    elapsed_ms,
                )
        except Exception:
            pass
    can_view_sessions = nivel_int in (1, 2)
    active_sessions = []
    if can_view_sessions:
        cutoff = datetime.utcnow() - timedelta(hours=2)
        try:
            t0 = datetime.utcnow()
            rows = _fetch_all_with_retry(
                select(ActiveSession.email, ActiveSession.last_activity)
                .where(ActiveSession.last_activity >= cutoff)
                .order_by(ActiveSession.last_activity.desc())
            )
            _log_timing("active_sessions", t0, len(rows))
            emails = [(_row_value(r, "email", 0) or "") for r in rows if _row_value(r, "email", 0)]
            usuarios = {}
            if emails:
                try:
                    t1 = datetime.utcnow()
                    users_rows = _fetch_all_with_retry(
                        select(Usuario.email, Usuario.nome).where(Usuario.email.in_(emails))
                    )
                    usuarios = dict(users_rows)
                    _log_timing("active_sessions_users", t1, len(usuarios))
                except NotImplementedError:
                    _safe_session_rollback()
                    with db.engine.connect() as conn:
                        usuarios = dict(
                            conn.execute(
                                select(Usuario.email, Usuario.nome).where(Usuario.email.in_(emails))
                            ).all()
                        )
            for r in rows:
                email = _row_value(r, "email", 0)
                active_sessions.append(
                    {
                        "email": email,
                        "nome": usuarios.get(email, email),
                        "last_activity": _row_value(r, "last_activity", 1),
                    }
                )
        except (ProgrammingError, NoSuchColumnError, OperationalError, ResourceClosedError):
            _safe_session_rollback()
            active_sessions = []
    ped_dotacao_missing = session.get("ped_dotacao_missing", [])
    ped_planejamento_missing_lines = session.get("ped_planejamento_missing_lines", [])
    if not ped_dotacao_missing:
        try:
            t2 = datetime.utcnow()
            ped_keys = _fetch_scalars_all_with_retry(
                select(PedRegistro.chave).where(PedRegistro.ativo == True)  # noqa: E712
            )
            ped_keys = [
                _normalize_dotacao_key(k)
                for k in ped_keys
                if k and str(k).strip().upper().startswith("DOT.")
            ]
            ped_keys = {k for k in ped_keys if k}
            if ped_keys:
                dotacao_keys = _fetch_all_with_retry(
                    select(Dotacao.chave_dotacao).where(Dotacao.chave_dotacao.isnot(None))
                )
                dotacao_keys = {_normalize_dotacao_key(k[0]) for k in dotacao_keys if k and k[0]}
                missing = sorted([k for k in ped_keys if k not in dotacao_keys])
                ped_dotacao_missing = missing
            _log_timing("ped_dotacao_missing", t2, len(ped_dotacao_missing))
        except Exception:
            _safe_session_rollback()
            ped_dotacao_missing = []
    emp_planejamento_missing_lines: list[int] = []
    emp_dotacao_missing: list[str] = []
    try:
        last_emp = (
            _fetch_scalars_all_with_retry(
                select(EmpUpload).order_by(EmpUpload.uploaded_at.desc()).limit(1)
            )
            or [None]
        )[0]
        if last_emp:
            status_data = read_status("emp", last_emp.id) or {}
            raw_lines = status_data.get("planejamento_missing_lines") or []
            if isinstance(raw_lines, list):
                emp_planejamento_missing_lines = [int(v) for v in raw_lines if str(v).isdigit()]
            raw_dot = status_data.get("dotacao_missing_keys") or []
            if isinstance(raw_dot, list):
                emp_dotacao_missing = [str(v) for v in raw_dot if str(v).strip()]
    except Exception:
        emp_planejamento_missing_lines = []
        emp_dotacao_missing = []

    aprovar_por_nivel2 = ""
    try:
        perfil_n2 = (
            _fetch_scalars_all_with_retry(
                select(Perfil)
                .where(Perfil.nivel == 2)
                .where(or_(Perfil.ativo.is_(None), Perfil.ativo == 1))
                .order_by(Perfil.id.asc())
                .limit(1)
            )
            or [None]
        )[0]
        aprovar_por_nivel2 = (getattr(perfil_n2, "nome", "") or "").strip()
    except Exception:
        aprovar_por_nivel2 = ""
    t3 = datetime.utcnow()
    try:
        pendentes_raw = _fetch_all_with_retry(
            select(
                Dotacao.chave_dotacao,
                Dotacao.valor_atual,
                Dotacao.valor_dotacao,
                Dotacao.status_aprovacao,
                Dotacao.adj_concedente,
            )
            .where(Dotacao.ativo == True)  # noqa: E712
            .where(func.lower(Dotacao.status_aprovacao) == "aguardando")
            .order_by(Dotacao.id.desc())
        )
        _log_timing("pendentes_raw", t3, len(pendentes_raw))
    except (ProgrammingError, NoSuchColumnError, OperationalError, ResourceClosedError, SQLAlchemyError) as exc:
        _safe_session_rollback()
        current_app.logger.warning("dashboard pendentes_raw unavailable: %s", exc)
        pendentes_raw = []
    pendentes = []
    for chave, valor_atual, valor_dot, status, adj_concedente in pendentes_raw:
        valor_base = valor_atual if valor_atual is not None else valor_dot
        valor_fmt = _dec_or_zero(valor_base)
        valor_str = f"{valor_fmt:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        pendentes.append(
            {
                "chave_dotacao": chave or "",
                "valor_atual": valor_str,
                "status_aprovacao": status or "",
                "adj_concedente": adj_concedente or "",
            }
        )
    estornos_aguardando = []
    try:
        t4 = datetime.utcnow()
        est_raw = _fetch_all_with_retry(
            text(
                """
                SELECT id, chave_dotacao, valor_a_ser_est, valor_dotacao, status_aprovacao, perfil_id
                FROM est_dotacao
                WHERE ativo = 1 AND lower(status_aprovacao) = 'aguardando'
                ORDER BY id DESC
                """
            )
        )
        _log_timing("estornos_raw", t4, len(est_raw))
    except Exception:
        est_raw = []
    est_perfil_ids = [r[5] for r in est_raw if len(r) > 5 and r[5]]
    est_adj_map = {}
    if est_perfil_ids:
        perfis = _fetch_scalars_all_with_retry(select(Perfil).where(Perfil.id.in_(est_perfil_ids)))
        est_adj_map = {p.id: p.nome for p in perfis if p and p.nome}
    for row in est_raw:
        chave = row[1] if len(row) > 1 else ""
        valor_est = row[2] if len(row) > 2 else None
        valor_dot = row[3] if len(row) > 3 else None
        status = row[4] if len(row) > 4 else ""
        perfil_id = row[5] if len(row) > 5 else None
        valor_base = valor_est if valor_est is not None else valor_dot
        valor_fmt = _dec_or_zero(valor_base)
        valor_str = f"{valor_fmt:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        estornos_aguardando.append(
            {
                "chave_dotacao": chave or "",
                "valor_estorno": valor_str,
                "status_aprovacao": status or "",
                "adj_solicitante": est_adj_map.get(perfil_id, ""),
            }
        )
    subacoes_aguardando = []
    try:
        t5 = datetime.utcnow()
        subacoes_raw = _fetch_scalars_all_with_retry(
            select(CadastrarSubacao)
            .where(CadastrarSubacao.ativo == True)  # noqa: E712
            .where(func.lower(CadastrarSubacao.status_aprovacao) == "aguardando")
            .order_by(CadastrarSubacao.id.desc())
        )
        _log_timing("subacoes_aguardando", t5, len(subacoes_raw))
    except Exception:
        subacoes_raw = []
    if subacoes_raw:
        usuario_ids = {s.usuario_id for s in subacoes_raw if s.usuario_id}
        usuarios_map = {}
        perfil_map = {}
        perfil_ids = set()
        if usuario_ids:
            usuarios = _fetch_scalars_all_with_retry(select(Usuario).where(Usuario.id.in_(usuario_ids)))
            usuarios_map = {u.id: u for u in usuarios}
            perfil_ids = {u.perfil_id for u in usuarios if u.perfil_id}
        if perfil_ids:
            perfis = _fetch_scalars_all_with_retry(select(Perfil).where(Perfil.id.in_(perfil_ids)))
            perfil_map = {p.id: p for p in perfis}
        for s in subacoes_raw:
            usuario = usuarios_map.get(s.usuario_id)
            perfil_id = getattr(usuario, "perfil_id", None)
            perfil_nome = ""
            if perfil_id:
                perfil_row = perfil_map.get(perfil_id)
                perfil_nome = (getattr(perfil_row, "nome", "") or "").strip()
            exercicio = getattr(s, "exercicio", "") or ""
            if perfil_nome:
                controle = f"SUB.{exercicio}.{perfil_nome}.{s.id}"
            else:
                controle = f"SUB.{exercicio}.{s.id}"
            subacoes_aguardando.append(
                {
                    "controle_subacao": controle,
                    "status_aprovacao": getattr(s, "status_aprovacao", "") or "",
                    "tipo_solicitacao": getattr(s, "tipo_solicitacao", "") or "",
                    "perfil": aprovar_por_nivel2 or perfil_nome,
                }
            )
    metas_aguardando = []
    try:
        t6 = datetime.utcnow()
        metas_raw = _fetch_scalars_all_with_retry(
            select(AlterarMeta)
            .where(AlterarMeta.ativo == True)  # noqa: E712
            .where(func.lower(AlterarMeta.status_aprovacao) == "aguardando")
            .order_by(AlterarMeta.id.desc())
        )
        _log_timing("metas_aguardando", t6, len(metas_raw))
    except Exception:
        metas_raw = []
    if metas_raw:
        meta_user_ids = {m.usuario_id for m in metas_raw if m.usuario_id}
        meta_usuarios_map = {}
        meta_perfil_map = {}
        if meta_user_ids:
            meta_usuarios = _fetch_scalars_all_with_retry(select(Usuario).where(Usuario.id.in_(list(meta_user_ids))))
            meta_usuarios_map = {u.id: u for u in meta_usuarios}
            meta_perfil_ids = {u.perfil_id for u in meta_usuarios if u.perfil_id}
            if meta_perfil_ids:
                meta_perfis = _fetch_scalars_all_with_retry(select(Perfil).where(Perfil.id.in_(list(meta_perfil_ids))))
                meta_perfil_map = {p.id: p for p in meta_perfis}

        for m in metas_raw:
            usuario = meta_usuarios_map.get(getattr(m, "usuario_id", None))
            exercicio_ctrl = str(getattr(m, "exercicio", "") or _now_local().year)
            adj_nome = (getattr(m, "responsavel_acao", "") or "").strip()
            if not adj_nome and usuario:
                perfil_id = getattr(usuario, "perfil_id", None)
                if perfil_id:
                    perfil_row = meta_perfil_map.get(perfil_id)
                    adj_nome = (getattr(perfil_row, "nome", "") or "").strip()
                if not adj_nome:
                    adj_nome = (getattr(usuario, "perfil", "") or "").strip()
            adj_token = _normalize_controle_token(adj_nome) or "SEMADJ"
            controle_meta = f"META.{exercicio_ctrl}.{adj_token}.{m.id}"
            metas_aguardando.append(
                {
                    "controle_meta": controle_meta,
                    "status_aprovacao": getattr(m, "status_aprovacao", "") or "",
                    "produto_acao": (getattr(m, "produto_acao", "") or "").strip(),
                    "perfil": aprovar_por_nivel2 or "",
                }
            )
    etapas_aguardando = []
    try:
        t7 = datetime.utcnow()
        etapas_raw = _fetch_scalars_all_with_retry(
            select(CadastrarEtapa)
            .where(CadastrarEtapa.ativo == True)  # noqa: E712
            .where(func.lower(CadastrarEtapa.status_aprovacao) == "aguardando")
            .order_by(CadastrarEtapa.id.desc())
        )
        _log_timing("etapas_aguardando", t7, len(etapas_raw))
    except Exception:
        etapas_raw = []
    if etapas_raw:
        etapa_user_ids = {e.usuario_id for e in etapas_raw if e.usuario_id}
        etapa_usuarios_map = {}
        etapa_perfil_map = {}
        if etapa_user_ids:
            etapa_usuarios = _fetch_scalars_all_with_retry(select(Usuario).where(Usuario.id.in_(list(etapa_user_ids))))
            etapa_usuarios_map = {u.id: u for u in etapa_usuarios}
            etapa_perfil_ids = {u.perfil_id for u in etapa_usuarios if u.perfil_id}
            if etapa_perfil_ids:
                etapa_perfis = _fetch_scalars_all_with_retry(select(Perfil).where(Perfil.id.in_(list(etapa_perfil_ids))))
                etapa_perfil_map = {p.id: p for p in etapa_perfis}
        for e in etapas_raw:
            usuario = etapa_usuarios_map.get(getattr(e, "usuario_id", None))
            perfil_nome = ""
            perfil_id = getattr(usuario, "perfil_id", None) if usuario else None
            if perfil_id:
                perfil_nome = (getattr(etapa_perfil_map.get(perfil_id), "nome", "") or "").strip()
            exercicio_ctrl = str(getattr(e, "exercicio", "") or _now_local().year)
            controle_etapa = (
                f"ETAPA.{exercicio_ctrl}.{perfil_nome}.{e.id}"
                if perfil_nome
                else f"ETAPA.{exercicio_ctrl}.{e.id}"
            )
            etapas_aguardando.append(
                {
                    "controle_etapa": controle_etapa,
                    "status_aprovacao": getattr(e, "status_aprovacao", "") or "",
                    "tipo_solicitacao": getattr(e, "tipo_solicitacao", "") or "",
                    "etapa": (getattr(e, "etapa_nova", "") or getattr(e, "etapa_origem", "") or "").strip(),
                    "perfil": aprovar_por_nivel2 or "",
                }
            )
    try:
        emp_record = get_emp_record_snapshot()
    except Exception:
        emp_record = {}
    emp_record_payload = {
        "dias_sem_erro": int(emp_record.get("dias_sem_erro") or 0),
        "recorde": int(emp_record.get("recorde") or 0),
        "ult_upload_at": _as_iso(emp_record.get("ult_upload_at")),
        "penult_upload_at": _as_iso(emp_record.get("penult_upload_at")),
    }
    return render_template(
        "partials/dashboard.html",
        can_view_sessions=can_view_sessions,
        active_sessions=active_sessions,
        ped_dotacao_missing=ped_dotacao_missing,
        ped_planejamento_missing_lines=ped_planejamento_missing_lines,
        emp_planejamento_missing_lines=emp_planejamento_missing_lines,
        emp_dotacao_missing=emp_dotacao_missing,
        dotacoes_aguardando=pendentes,
        estornos_aguardando=estornos_aguardando,
        subacoes_aguardando=subacoes_aguardando,
        metas_aguardando=metas_aguardando,
        etapas_aguardando=etapas_aguardando,
        emp_record=emp_record_payload,
    )


def ensure_admin_nivel1():
    nivel = getattr(g, "user_nivel", None)
    if nivel != 1:
        abort(403)

def has_permission(feature: str) -> bool:
    if getattr(g, "user_nivel", None) == 1:
        return True
    perfil_id = getattr(g, "user_perfil_id", None)
    nivel = getattr(g, "user_nivel", None)
    if not perfil_id and nivel is None:
        return False
    try:
        if perfil_id:
            exists = (
                db.session.query(PerfilPermissao.id)
                .filter(PerfilPermissao.perfil_id == perfil_id, PerfilPermissao.feature == feature)
                .first()
            )
            if exists:
                return True
    except ProgrammingError:
        db.session.rollback()
    try:
        if nivel is not None:
            exists = (
                db.session.query(NivelPermissao.id)
                .filter(NivelPermissao.nivel == nivel, NivelPermissao.feature == feature)
                .first()
            )
            return bool(exists)
    except ProgrammingError:
        db.session.rollback()
    return False


def _load_permissoes_perfil(perfil_id: int | None):
    if perfil_id is None:
        return []
    try:
        rows = _fetch_scalars_all_with_retry(
            select(PerfilPermissao.feature).where(
                PerfilPermissao.perfil_id == perfil_id,
                PerfilPermissao.ativo == True,  # noqa: E712
                PerfilPermissao.feature.isnot(None),
            )
        )
        return [f for f in rows if f]
    except (ProgrammingError, NoSuchColumnError, ResourceClosedError, OperationalError, SQLAlchemyError, IndexError):
        _safe_session_rollback()
        return []


def _load_permissoes_nivel(nivel: int | None):
    if nivel is None:
        return []
    try:
        rows = _fetch_scalars_all_with_retry(
            select(NivelPermissao.feature).where(
                NivelPermissao.nivel == nivel,
                NivelPermissao.ativo == True,  # noqa: E712
                NivelPermissao.feature.isnot(None),
            )
        )
        return [f for f in rows if f]
    except (ProgrammingError, NoSuchColumnError, ResourceClosedError, OperationalError, SQLAlchemyError, IndexError):
        _safe_session_rollback()
        return []


def _add_parent_features(features: list[str]) -> list[str]:
    parent_map = build_parent_map()
    feats = list(features)
    for feat in list(feats):
        parent = parent_map.get(feat)
        if parent and parent not in feats:
            feats.append(parent)
    return feats


def _load_permissoes_por_nivel_perfis(nivel: int):
    try:
        perfis = _fetch_scalars_all_with_retry(select(Perfil).where(Perfil.nivel == nivel))
        perfil_ids = [p.id for p in perfis]
        if not perfil_ids:
            return []
        feats = _fetch_scalars_all_with_retry(
            select(PerfilPermissao.feature).where(
                PerfilPermissao.perfil_id.in_(perfil_ids),
                PerfilPermissao.ativo == True,  # noqa: E712
                PerfilPermissao.feature.isnot(None),
            )
        )
        feats = [f for f in feats if f]
        return _add_parent_features(list(set(feats)))
    except (ProgrammingError, NoSuchColumnError, ResourceClosedError, OperationalError, SQLAlchemyError, IndexError):
        _safe_session_rollback()
        return []


def _permissoes_with_parents(perfil_id: int | None, nivel: int | None = None):
    locked = [f["id"] for f in FEATURES if f.get("locked")]
    parent_map = build_parent_map()
    feats = _load_permissoes_perfil(perfil_id) + _load_permissoes_nivel(nivel)
    # include parents of children
    feats = _add_parent_features(feats)
    # add locked always
    for l in locked:
        if l not in feats:
            feats.append(l)
    return feats


def require_feature(feature_id):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if getattr(g, "user_nivel", None) == 1:
                return view(*args, **kwargs)
            if has_permission(feature_id):
                return view(*args, **kwargs)
            abort(403)

        return wrapped

    return decorator


def _perfil_by_nome(nome: str | None):
    if not nome:
        return None
    nome_strip = nome.strip()
    return (
        Perfil.query.filter(
            func.lower(func.ltrim(func.rtrim(Perfil.nome))) == nome_strip.lower()
        ).first()
        or Perfil.query.filter(Perfil.nome.ilike(nome_strip)).first()
    )


def _perfil_id_by_nome(nome: str | None) -> int | None:
    perfil = _perfil_by_nome(nome)
    return perfil.id if perfil else None


def _perfil_by_id(perfil_id) -> Perfil | None:
    if perfil_id is None:
        return None
    try:
        return db.session.get(Perfil, int(perfil_id))
    except Exception:
        return None


def _is_nivel1(perfil_id=None) -> bool:
    perfil_row = _perfil_by_id(perfil_id)
    return bool(perfil_row and perfil_row.nivel == 1)


@home_bp.route("/partial/usuarios")
@login_required
@require_feature("usuarios")
def partial_usuarios():
    return partial_usuarios_cadastrar()


@home_bp.route("/partial/usuarios/cadastrar")
@login_required
@require_feature("usuarios/cadastrar")
def partial_usuarios_cadastrar():
    usuarios = Usuario.query.order_by(Usuario.nome).all()
    perfis_query = Perfil.query.filter_by(ativo=True)
    caller_nivel = getattr(g, "user_nivel", None)
    # Somente nivel 1 enxerga perfis de nivel 1 (admin)
    if caller_nivel != 1:
        perfis_query = perfis_query.filter(Perfil.nivel != 1)
    perfis = perfis_query.order_by(Perfil.nivel, Perfil.nome).all()
    return render_template("partials/usuarios_cadastrar.html", usuarios=usuarios, perfis=perfis)


@home_bp.route("/partial/usuarios/editar")
@login_required
@require_feature("usuarios/editar")
def partial_usuarios_editar():
    usuarios = Usuario.query.order_by(Usuario.nome).all()
    perfis_query = Perfil.query.filter_by(ativo=True)
    caller_nivel = getattr(g, "user_nivel", None)
    if caller_nivel != 1:
        perfis_query = perfis_query.filter(Perfil.nivel != 1)
    perfis = perfis_query.order_by(Perfil.nivel, Perfil.nome).all()
    return render_template("partials/usuarios_editar.html", usuarios=usuarios, perfis=perfis)


@home_bp.route("/partial/usuarios/perfil")
@login_required
@role_required("admin")
def partial_usuarios_perfil():
    perfis = Perfil.query.order_by(Perfil.nivel, Perfil.nome).all()
    return render_template("partials/usuarios_perfil.html", perfis=perfis)


@home_bp.route("/partial/usuarios/senha")
@login_required
@require_feature("usuarios/senha")
def partial_usuarios_senha():
    return render_template("partials/usuarios_senha.html")


API_SCOPE_OPTIONS = [
    {"scope": "dados.emp.read", "label": "EMP (somente leitura)"},
    {"scope": "dados.nob.read", "label": "NOB (somente leitura)"},
    {"scope": "dados.ped.read", "label": "PED (somente leitura)"},
    {"scope": "dados.est_emp.read", "label": "EST_EMP (somente leitura)"},
    {"scope": "dados.fip613.read", "label": "FIP613 (somente leitura)"},
]
API_SCOPE_SET = {item["scope"] for item in API_SCOPE_OPTIONS}
API_SCOPE_LABEL = {item["scope"]: item["label"] for item in API_SCOPE_OPTIONS}
API_SCOPE_META_SERVIDOR_CAD = "__meta.servidor_cadastrado__"
API_ACCESS_TOKEN_MINUTES = int(os.getenv("API_ACCESS_TOKEN_MINUTES", "60") or "60")
API_ACCESS_TOKEN_MINUTES = max(5, API_ACCESS_TOKEN_MINUTES)


def _sha256_hex(text: str) -> str:
    return hashlib.sha256(str(text or "").encode("utf-8")).hexdigest()


def _as_json_value(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        try:
            return float(value)
        except Exception:
            return str(value)
    if isinstance(value, (datetime,)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value


def _log_api_access(
    client_id: int | None,
    endpoint: str,
    metodo: str,
    status_code: int,
    duration_ms: int | None = None,
):
    try:
        qp = None
        try:
            qp = json.dumps(dict(request.args), ensure_ascii=False)
        except Exception:
            qp = None
        db.session.add(
            ApiAccessLog(
                client_id=client_id,
                request_id=str(secrets.token_hex(8)),
                endpoint=str(endpoint or "")[:255],
                metodo=str(metodo or "")[:10],
                query_params=qp,
                status_code=int(status_code or 0),
                ip=(request.headers.get("X-Forwarded-For") or request.remote_addr or "")[:45],
                user_agent=(request.headers.get("User-Agent") or "")[:255],
                duration_ms=int(duration_ms) if duration_ms is not None else None,
            )
        )
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def _api_scope_display(scope: str) -> str:
    return API_SCOPE_LABEL.get(scope or "", scope or "")


def _parse_dt_local_iso(value) -> datetime | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1]
    text = text.replace("T", " ")
    try:
        return datetime.fromisoformat(text)
    except Exception:
        return None


def _refresh_api_client_expirations() -> int:
    now = _now_local()
    rows = (
        ApiClient.query.filter(ApiClient.status != "revogado")
        .filter(ApiClient.acesso_fim_em.isnot(None))
        .filter(ApiClient.acesso_fim_em < now)
        .all()
    )
    if not rows:
        return 0
    for row in rows:
        row.status = "revogado"
        row.revoked_at = now
    db.session.commit()
    return len(rows)


def _get_api_client_scopes(client_id: int) -> list[str]:
    rows = (
        ApiClientScope.query.filter_by(client_id=client_id)
        .order_by(ApiClientScope.scope.asc())
        .all()
    )
    return [str(r.scope or "").strip() for r in rows if str(r.scope or "").strip()]


def _extract_bearer_token() -> str:
    auth = str(request.headers.get("Authorization") or "").strip()
    if not auth.lower().startswith("bearer "):
        return ""
    return auth[7:].strip()


def require_api_scope(required_scope: str):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            _refresh_api_client_expirations()
            token = _extract_bearer_token()
            if not token:
                _log_api_access(None, request.path, request.method, 401)
                return jsonify({"error": "Token Bearer obrigatorio."}), 401

            now = _now_local()
            token_hash = _sha256_hex(token)
            api_key = ApiKey.query.filter_by(key_hash=token_hash).first()
            if not api_key or (api_key.status or "").lower() != "ativo":
                _log_api_access(None, request.path, request.method, 401)
                return jsonify({"error": "Token invalido."}), 401

            if api_key.expires_at and api_key.expires_at < now:
                try:
                    api_key.status = "expirado"
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                _log_api_access(api_key.client_id, request.path, request.method, 401)
                return jsonify({"error": "Token expirado."}), 401

            client = db.session.get(ApiClient, api_key.client_id)
            if not client:
                _log_api_access(None, request.path, request.method, 401)
                return jsonify({"error": "Cliente da API nao encontrado."}), 401

            if (client.status or "").lower() != "ativo":
                _log_api_access(client.id, request.path, request.method, 403)
                return jsonify({"error": "Cliente da API inativo ou revogado."}), 403
            if client.acesso_inicio_em and now < client.acesso_inicio_em:
                _log_api_access(client.id, request.path, request.method, 403)
                return jsonify({"error": "Acesso ainda nao iniciado."}), 403
            if client.acesso_fim_em and now > client.acesso_fim_em:
                try:
                    client.status = "revogado"
                    client.revoked_at = now
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                _log_api_access(client.id, request.path, request.method, 403)
                return jsonify({"error": "Acesso expirado."}), 403

            scopes = _get_api_client_scopes(client.id)
            allowed_scopes = {s for s in scopes if s in API_SCOPE_SET}
            if required_scope not in allowed_scopes:
                _log_api_access(client.id, request.path, request.method, 403)
                return jsonify({"error": "Escopo insuficiente para este recurso."}), 403

            started = datetime.utcnow()
            g.api_client = client
            g.api_scopes = list(allowed_scopes)
            try:
                response = current_app.make_response(view(*args, **kwargs))
            except Exception:
                duration = int((datetime.utcnow() - started).total_seconds() * 1000)
                _log_api_access(client.id, request.path, request.method, 500, duration)
                raise

            try:
                api_key.last_used_at = now
                client.last_used_at = now
                db.session.commit()
            except Exception:
                db.session.rollback()
            duration = int((datetime.utcnow() - started).total_seconds() * 1000)
            _log_api_access(client.id, request.path, request.method, int(response.status_code or 200), duration)
            return response

        return wrapped

    return decorator


def _serialize_api_client(row: ApiClient, scopes_map: dict[int, list[str]] | None = None) -> dict:
    raw_scopes = list((scopes_map or {}).get(row.id, []))
    servidor_cadastrado = API_SCOPE_META_SERVIDOR_CAD in raw_scopes
    scopes = [s for s in raw_scopes if s in API_SCOPE_SET]
    last_used = None
    try:
        last_used = row.last_used_at.isoformat() if row.last_used_at else None
    except Exception:
        last_used = str(row.last_used_at or "") or None
    try:
        inicio = row.acesso_inicio_em.isoformat() if row.acesso_inicio_em else None
    except Exception:
        inicio = str(row.acesso_inicio_em or "") or None
    try:
        fim = row.acesso_fim_em.isoformat() if row.acesso_fim_em else None
    except Exception:
        fim = str(row.acesso_fim_em or "") or None
    return {
        "id": row.id,
        "nome": (row.nome or "").strip(),
        "client_id": (row.client_id or "").strip(),
        "status": (row.status or "").strip(),
        "acesso_inicio_em": inicio,
        "acesso_fim_em": fim,
        "last_used_at": last_used,
        "servidor_cadastrado": bool(servidor_cadastrado),
        "scopes": scopes,
        "scopes_label": [_api_scope_display(s) for s in scopes],
        "created_at": row.created_at.isoformat() if getattr(row, "created_at", None) else None,
    }


def _generate_client_id() -> str:
    for _ in range(8):
        candidate = f"cli_{secrets.token_hex(8)}"
        exists = ApiClient.query.filter_by(client_id=candidate).first()
        if not exists:
            return candidate
    return f"cli_{secrets.token_hex(12)}"


def _generate_client_secret() -> str:
    return f"sec_{secrets.token_urlsafe(24)}"


def _replace_client_scopes(client_id: int, scopes: list[str]) -> None:
    ApiClientScope.query.filter_by(client_id=client_id).delete(synchronize_session=False)
    now = _now_local()
    for sc in scopes:
        db.session.add(ApiClientScope(client_id=client_id, scope=sc, created_at=now))


@home_bp.route("/partial/usuarios/api-acessos")
@login_required
@require_feature("usuarios/api-acessos")
def partial_usuarios_api_acessos():
    return render_template("partials/usuarios_api_acessos.html", scope_options=API_SCOPE_OPTIONS)


@home_bp.route("/api/api-clients", methods=["GET"])
@login_required
@require_feature("usuarios/api-acessos")
def api_clients_list():
    _refresh_api_client_expirations()
    rows = ApiClient.query.order_by(ApiClient.created_at.desc(), ApiClient.id.desc()).all()
    ids = [r.id for r in rows]
    scopes_map = {i: [] for i in ids}
    if ids:
        scopes = (
            ApiClientScope.query.filter(ApiClientScope.client_id.in_(ids))
            .order_by(ApiClientScope.client_id.asc(), ApiClientScope.scope.asc())
            .all()
        )
        for sc in scopes:
            scopes_map.setdefault(sc.client_id, []).append((sc.scope or "").strip())
    return jsonify({"ok": True, "rows": [_serialize_api_client(r, scopes_map) for r in rows]})


@home_bp.route("/api/api-clients", methods=["POST"])
@login_required
@require_feature("usuarios/api-acessos")
def api_clients_create():
    payload = request.get_json(silent=True) or {}
    nome = str(payload.get("nome") or "").strip()
    status = str(payload.get("status") or "ativo").strip().lower()
    scopes_raw = payload.get("scopes") or []
    servidor_cadastrado = bool(payload.get("servidor_cadastrado", False))
    acesso_inicio_em = _parse_dt_local_iso(payload.get("acesso_inicio_em"))
    acesso_fim_em = _parse_dt_local_iso(payload.get("acesso_fim_em"))
    scopes = []
    for item in scopes_raw:
        scope = str(item or "").strip()
        if scope and scope in API_SCOPE_SET and scope not in scopes:
            scopes.append(scope)

    if not nome:
        return jsonify({"error": "Nome do servidor e obrigatorio."}), 400
    if status not in {"ativo", "inativo", "revogado"}:
        return jsonify({"error": "Status invalido."}), 400
    if not scopes:
        return jsonify({"error": "Selecione ao menos um escopo."}), 400
    if not acesso_inicio_em:
        return jsonify({"error": "Data/hora de inicio e obrigatoria."}), 400
    if acesso_fim_em and acesso_fim_em <= acesso_inicio_em:
        return jsonify({"error": "Data/hora final deve ser maior que a inicial."}), 400
    scopes_to_save = list(scopes)
    if servidor_cadastrado:
        scopes_to_save.append(API_SCOPE_META_SERVIDOR_CAD)

    plain_secret = _generate_client_secret()
    row = ApiClient(
        nome=nome,
        client_id=_generate_client_id(),
        client_secret_hash=generate_password_hash(plain_secret),
        status=status,
        created_by_user_id=_resolve_usuario_id(),
        acesso_inicio_em=acesso_inicio_em,
        acesso_fim_em=acesso_fim_em,
        revoked_at=_now_local() if status == "revogado" else None,
    )
    db.session.add(row)
    db.session.flush()
    _replace_client_scopes(row.id, scopes_to_save)
    db.session.commit()

    return jsonify(
        {
            "ok": True,
            "message": "Servidor de API criado.",
            "row": _serialize_api_client(row, {row.id: scopes_to_save}),
            "credentials": {"client_id": row.client_id, "client_secret": plain_secret},
        }
    )


@home_bp.route("/api/api-clients/<int:client_id>", methods=["PUT"])
@login_required
@require_feature("usuarios/api-acessos")
def api_clients_update(client_id: int):
    _refresh_api_client_expirations()
    row = db.session.get(ApiClient, client_id)
    if not row:
        return jsonify({"error": "Servidor nao encontrado."}), 404

    payload = request.get_json(silent=True) or {}
    nome = str(payload.get("nome") or row.nome or "").strip()
    status = str(payload.get("status") or row.status or "ativo").strip().lower()
    scopes_raw = payload.get("scopes") or []
    servidor_cadastrado = bool(payload.get("servidor_cadastrado", False))
    acesso_inicio_em = _parse_dt_local_iso(payload.get("acesso_inicio_em"))
    acesso_fim_em = _parse_dt_local_iso(payload.get("acesso_fim_em"))
    scopes = []
    for item in scopes_raw:
        scope = str(item or "").strip()
        if scope and scope in API_SCOPE_SET and scope not in scopes:
            scopes.append(scope)

    if not nome:
        return jsonify({"error": "Nome do servidor e obrigatorio."}), 400
    if status not in {"ativo", "inativo", "revogado"}:
        return jsonify({"error": "Status invalido."}), 400
    if not scopes:
        return jsonify({"error": "Selecione ao menos um escopo."}), 400
    if not acesso_inicio_em:
        return jsonify({"error": "Data/hora de inicio e obrigatoria."}), 400
    if acesso_fim_em and acesso_fim_em <= acesso_inicio_em:
        return jsonify({"error": "Data/hora final deve ser maior que a inicial."}), 400
    scopes_to_save = list(scopes)
    if servidor_cadastrado:
        scopes_to_save.append(API_SCOPE_META_SERVIDOR_CAD)

    row.nome = nome
    row.status = status
    row.acesso_inicio_em = acesso_inicio_em
    row.acesso_fim_em = acesso_fim_em
    if status == "revogado":
        row.revoked_at = _now_local()
    elif status in {"ativo", "inativo"}:
        row.revoked_at = None

    _replace_client_scopes(row.id, scopes_to_save)
    db.session.commit()
    return jsonify({"ok": True, "message": "Servidor atualizado.", "row": _serialize_api_client(row, {row.id: scopes_to_save})})


@home_bp.route("/api/api-clients/<int:client_id>/rotate-secret", methods=["POST"])
@login_required
@require_feature("usuarios/api-acessos")
def api_clients_rotate_secret(client_id: int):
    _refresh_api_client_expirations()
    row = db.session.get(ApiClient, client_id)
    if not row:
        return jsonify({"error": "Servidor nao encontrado."}), 404

    plain_secret = _generate_client_secret()
    row.client_secret_hash = generate_password_hash(plain_secret)
    if row.status == "revogado":
        row.status = "ativo"
        row.revoked_at = None
    db.session.commit()

    return jsonify(
        {
            "ok": True,
            "message": "Segredo rotacionado.",
            "credentials": {"client_id": row.client_id, "client_secret": plain_secret},
        }
    )


@home_bp.route("/api/api-clients/<int:client_id>/revoke", methods=["POST"])
@login_required
@require_feature("usuarios/api-acessos")
def api_clients_revoke(client_id: int):
    row = db.session.get(ApiClient, client_id)
    if not row:
        return jsonify({"error": "Servidor nao encontrado."}), 404
    row.status = "revogado"
    row.revoked_at = _now_local()
    db.session.commit()
    return jsonify({"ok": True, "message": "Servidor revogado."})


@home_bp.route("/api/api-clients/<int:client_id>", methods=["DELETE"])
@login_required
@require_feature("usuarios/api-acessos")
def api_clients_delete(client_id: int):
    row = db.session.get(ApiClient, client_id)
    if not row:
        return jsonify({"error": "Servidor nao encontrado."}), 404
    try:
        ApiClientScope.query.filter_by(client_id=row.id).delete(synchronize_session=False)
        ApiKey.query.filter_by(client_id=row.id).delete(synchronize_session=False)
        ApiRefreshToken.query.filter_by(client_id=row.id).delete(synchronize_session=False)
        db.session.delete(row)
        db.session.commit()
        return jsonify({"ok": True, "message": "Servidor excluido."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir servidor: {exc}"}), 500


@home_bp.route("/api/auth/token", methods=["POST"])
def api_auth_token():
    started = datetime.utcnow()
    _refresh_api_client_expirations()
    payload = request.get_json(silent=True) or {}
    client_id = str(payload.get("client_id") or "").strip()
    client_secret = str(payload.get("client_secret") or "")
    client = ApiClient.query.filter_by(client_id=client_id).first() if client_id else None

    def _deny(message: str, code: int = 401):
        duration = int((datetime.utcnow() - started).total_seconds() * 1000)
        _log_api_access(client.id if client else None, request.path, request.method, code, duration)
        return jsonify({"error": message}), code

    if not client_id or not client_secret:
        return _deny("client_id e client_secret sao obrigatorios.", 400)
    if not client or not check_password_hash(client.client_secret_hash or "", client_secret):
        return _deny("Credenciais invalidas.", 401)
    if (client.status or "").lower() != "ativo":
        return _deny("Cliente da API inativo ou revogado.", 403)

    now = _now_local()
    if client.acesso_inicio_em and now < client.acesso_inicio_em:
        return _deny("Acesso ainda nao iniciado.", 403)
    if client.acesso_fim_em and now > client.acesso_fim_em:
        try:
            client.status = "revogado"
            client.revoked_at = now
            db.session.commit()
        except Exception:
            db.session.rollback()
        return _deny("Acesso expirado.", 403)

    all_scopes = _get_api_client_scopes(client.id)
    scopes = [s for s in all_scopes if s in API_SCOPE_SET]
    if not scopes:
        return _deny("Cliente sem escopos habilitados.", 403)

    expires_at = now + timedelta(minutes=API_ACCESS_TOKEN_MINUTES)
    if client.acesso_fim_em and client.acesso_fim_em < expires_at:
        expires_at = client.acesso_fim_em
    if expires_at <= now:
        return _deny("Acesso expirado.", 403)

    token = f"atk_{secrets.token_urlsafe(32)}"
    token_hash = _sha256_hex(token)
    api_key = ApiKey(
        client_id=client.id,
        key_prefix=token[:16],
        key_hash=token_hash,
        status="ativo",
        expires_at=expires_at,
    )
    try:
        db.session.add(api_key)
        client.last_used_at = now
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return _deny(f"Falha ao emitir token: {exc}", 500)

    duration = int((datetime.utcnow() - started).total_seconds() * 1000)
    _log_api_access(client.id, request.path, request.method, 200, duration)
    return jsonify(
        {
            "ok": True,
            "token_type": "Bearer",
            "access_token": token,
            "expires_at": expires_at.isoformat(),
            "expires_in": int((expires_at - now).total_seconds()),
            "scopes": scopes,
        }
    )


def _api_query_model(model, table_name: str):
    reserved = {"limit", "offset", "order_by", "order_dir", "fields"}
    mapper = inspect(model)
    col_keys = [c.key for c in mapper.columns]
    if "ativo" not in col_keys:
        return jsonify({"error": f"Tabela {table_name} sem coluna ativo."}), 500

    try:
        limit = int(str(request.args.get("limit", "500") or "500"))
    except Exception:
        limit = 500
    try:
        offset = int(str(request.args.get("offset", "0") or "0"))
    except Exception:
        offset = 0
    limit = max(1, min(limit, 5000))
    offset = max(0, offset)

    query = db.session.query(model).filter(getattr(model, "ativo") == True)  # noqa: E712
    for key, val in request.args.items():
        if key in reserved:
            continue
        if key not in col_keys:
            continue
        text_val = str(val or "").strip()
        if not text_val:
            continue
        if "," in text_val:
            vals = [v.strip() for v in text_val.split(",") if v.strip()]
            if vals:
                query = query.filter(getattr(model, key).in_(vals))
        else:
            query = query.filter(getattr(model, key) == text_val)

    order_by = str(request.args.get("order_by", "id") or "id")
    order_dir = str(request.args.get("order_dir", "desc") or "desc").lower()
    if order_by not in col_keys:
        order_by = "id" if "id" in col_keys else col_keys[0]
    order_col = getattr(model, order_by)
    query = query.order_by(order_col.asc() if order_dir == "asc" else order_col.desc())

    fields_param = str(request.args.get("fields") or "").strip()
    if fields_param:
        asked = [f.strip() for f in fields_param.split(",") if f.strip()]
        fields = [f for f in asked if f in col_keys]
    else:
        fields = [f for f in col_keys if f != "raw_payload"]
    if not fields:
        fields = [f for f in col_keys if f != "raw_payload"] or col_keys

    total = query.count()
    rows = query.offset(offset).limit(limit).all()
    data = []
    for row in rows:
        item = {}
        for field in fields:
            item[field] = _as_json_value(getattr(row, field, None))
        data.append(item)

    return jsonify(
        {
            "ok": True,
            "table": table_name,
            "total": int(total),
            "limit": int(limit),
            "offset": int(offset),
            "count": len(data),
            "data": data,
        }
    )


@home_bp.route("/api/v1/dados/emp", methods=["GET"])
@require_api_scope("dados.emp.read")
def api_dados_emp():
    return _api_query_model(EmpRegistro, "emp")


@home_bp.route("/api/v1/dados/nob", methods=["GET"])
@require_api_scope("dados.nob.read")
def api_dados_nob():
    return _api_query_model(NobRegistro, "nob")


@home_bp.route("/api/v1/dados/ped", methods=["GET"])
@require_api_scope("dados.ped.read")
def api_dados_ped():
    return _api_query_model(PedRegistro, "ped")


@home_bp.route("/api/v1/dados/est-emp", methods=["GET"])
@require_api_scope("dados.est_emp.read")
def api_dados_est_emp():
    return _api_query_model(EstEmpRegistro, "est_emp")


@home_bp.route("/api/v1/dados/fip613", methods=["GET"])
@require_api_scope("dados.fip613.read")
def api_dados_fip613():
    return _api_query_model(Fip613Registro, "fip613")


@home_bp.route("/partial/painel")
@login_required
def partial_painel():
    if not has_permission("painel"):
        abort(403)
    perfis = Perfil.query.order_by(Perfil.nivel, Perfil.nome).all()
    features = FEATURES
    allowed_perfil = {}
    for perfil in perfis:
        allowed_perfil[perfil.id] = _load_permissoes_perfil(perfil.id)
    niveis = [1, 2, 3, 4, 5]
    allowed_nivel = {}
    for nivel in niveis:
        allowed_nivel[nivel] = _load_permissoes_nivel(nivel)
    return render_template(
        "partials/painel.html",
        perfis=perfis,
        features=features,
        allowed_perfil=allowed_perfil,
        allowed_nivel=allowed_nivel,
        niveis=niveis,
    )


@home_bp.route("/partial/atualizar/fip613")
@login_required
@require_feature("atualizar/fip613")
def partial_atualizar_fip613():
    return render_template("partials/atualizar_fip613.html")


@home_bp.route("/partial/atualizar/ped")
@login_required
@require_feature("atualizar/ped")
def partial_atualizar_ped():
    return render_template("partials/atualizar_ped.html")


@home_bp.route("/partial/atualizar/emp")
@login_required
@require_feature("atualizar/emp")
def partial_atualizar_emp():
    return render_template("partials/atualizar_emp.html")


@home_bp.route("/partial/atualizar/chave_planejamento_regra")
@login_required
@require_feature("atualizar/chave_planejamento_regra")
def partial_atualizar_chave_planejamento_regra():
    regras = (
        ChavePlanejamentoRegra.query.order_by(
            ChavePlanejamentoRegra.tipo_regra.asc(),
            ChavePlanejamentoRegra.chave_origem.asc(),
            ChavePlanejamentoRegra.id.desc(),
        ).all()
    )
    return render_template("partials/atualizar_chave_planejamento_regra.html", regras=regras)


def _serialize_chave_regra(r: ChavePlanejamentoRegra) -> dict:
    return {
        "id": r.id,
        "tipo_regra": (r.tipo_regra or "").strip(),
        "chave_origem": (r.chave_origem or "").strip(),
        "chave_destino": (r.chave_destino or "").strip(),
        "observacao": (r.observacao or "").strip(),
        "usuario_id": r.usuario_id,
        "ativo": bool(r.ativo),
        "criado_em": r.criado_em.isoformat() if getattr(r, "criado_em", None) else "",
        "alterado_em": r.alterado_em.isoformat() if getattr(r, "alterado_em", None) else "",
    }


@home_bp.route("/api/chave-planejamento-regra", methods=["GET"])
@login_required
@require_feature("atualizar/chave_planejamento_regra")
def api_chave_planejamento_regra_list():
    regras = (
        ChavePlanejamentoRegra.query.order_by(
            ChavePlanejamentoRegra.tipo_regra.asc(),
            ChavePlanejamentoRegra.chave_origem.asc(),
            ChavePlanejamentoRegra.id.desc(),
        ).all()
    )
    return jsonify({"ok": True, "rows": [_serialize_chave_regra(r) for r in regras]})


@home_bp.route("/api/chave-planejamento-regra", methods=["POST"])
@login_required
@require_feature("atualizar/chave_planejamento_regra")
def api_chave_planejamento_regra_create():
    payload = request.get_json(silent=True) or {}
    tipo = str(payload.get("tipo_regra") or "").strip()
    origem = str(payload.get("chave_origem") or "").strip()
    destino = str(payload.get("chave_destino") or "").strip()
    observacao = str(payload.get("observacao") or "").strip()
    ativo = bool(payload.get("ativo", True))

    tipos_validos = {"chaves_planejamento", "chave_arrumar", "forcar_chave"}
    if tipo not in tipos_validos:
        return jsonify({"error": "Tipo de regra invalido."}), 400
    can_upload_planejamento = has_permission("atualizar/chaves_planejamento_upload") or has_permission("painel/chaves_planejamento_upload")
    if tipo == "chaves_planejamento" and not can_upload_planejamento:
        return jsonify({"error": "Usuario sem permissao para cadastrar chaves_planejamento."}), 403
    if not origem:
        return jsonify({"error": "Chave de origem obrigatoria."}), 400
    if tipo in {"chave_arrumar", "forcar_chave"} and not destino:
        return jsonify({"error": "Chave de destino obrigatoria para este tipo."}), 400

    now = _now_local()
    user_id = _resolve_usuario_id()
    try:
        existente = (
            ChavePlanejamentoRegra.query.filter_by(tipo_regra=tipo, chave_origem=origem).first()
        )
        if existente:
            existente.chave_destino = destino or None
            existente.observacao = observacao or None
            existente.ativo = ativo
            existente.usuario_id = user_id or existente.usuario_id
            existente.excluido_em = None if ativo else now
            existente.alterado_em = now
            db.session.flush()
            row = existente
            msg = "Regra atualizada."
        else:
            row = ChavePlanejamentoRegra(
                tipo_regra=tipo,
                chave_origem=origem,
                chave_destino=destino or None,
                observacao=observacao or None,
                usuario_id=user_id,
                ativo=ativo,
                criado_em=now,
                alterado_em=None,
                excluido_em=None if ativo else now,
            )
            db.session.add(row)
            db.session.flush()
            msg = "Regra cadastrada."
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "Ja existe uma regra com este tipo e chave de origem."}), 409
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar regra: {exc}"}), 500

    return jsonify({"ok": True, "message": msg, "row": _serialize_chave_regra(row)})


@home_bp.route("/api/chave-planejamento-regra/import", methods=["POST"])
@login_required
@require_feature("atualizar/chave_planejamento_regra")
def api_chave_planejamento_regra_import():
    can_upload_planejamento = has_permission("atualizar/chaves_planejamento_upload") or has_permission("painel/chaves_planejamento_upload")
    if not can_upload_planejamento:
        return jsonify({"error": "Usuario sem permissao para importar chaves_planejamento."}), 403
    arquivo = request.files.get("arquivo")
    if not arquivo or not getattr(arquivo, "filename", ""):
        return jsonify({"error": "Arquivo .xlsx Ã© obrigatÃ³rio."}), 400
    if not str(arquivo.filename).lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx vÃ¡lido."}), 400

    observacao_padrao = str(request.form.get("observacao") or "").strip()
    ativo_flag = str(request.form.get("ativo") or "1").strip() not in {"0", "false", "False"}
    user_id = _resolve_usuario_id()
    now = _now_local()

    try:
        df = pd.read_excel(arquivo, dtype=str)
    except Exception as exc:
        return jsonify({"error": f"Falha ao ler o arquivo .xlsx: {exc}"}), 400

    if df is None or df.empty:
        return jsonify({"error": "Arquivo sem dados para importaÃ§Ã£o."}), 400

    cols_norm = {str(c).strip().lower(): c for c in df.columns}
    col_origem = cols_norm.get("chave_origem") or cols_norm.get("origem")
    col_obs = cols_norm.get("observacao") or cols_norm.get("observaÃ§Ã£o")
    if not col_origem:
        return jsonify({"error": "Coluna obrigatÃ³ria nÃ£o encontrada: chave_origem."}), 400

    inseridas = 0
    atualizadas = 0
    ignoradas_duplicadas_arquivo = 0
    ignoradas_vazias = 0
    vistos = set()

    try:
        for _, raw in df.iterrows():
            origem = str(raw.get(col_origem) or "").strip()
            if not origem or origem.lower() in {"nan", "none"}:
                ignoradas_vazias += 1
                continue
            if origem in vistos:
                ignoradas_duplicadas_arquivo += 1
                continue
            vistos.add(origem)

            obs_linha = observacao_padrao
            if col_obs:
                obs_raw = str(raw.get(col_obs) or "").strip()
                if obs_raw and obs_raw.lower() not in {"nan", "none"}:
                    obs_linha = obs_raw

            existente = (
                ChavePlanejamentoRegra.query.filter_by(
                    tipo_regra="chaves_planejamento",
                    chave_origem=origem,
                ).first()
            )
            if existente:
                existente.chave_destino = None
                existente.observacao = obs_linha or None
                existente.ativo = ativo_flag
                existente.usuario_id = user_id or existente.usuario_id
                existente.excluido_em = None if ativo_flag else now
                existente.alterado_em = now
                atualizadas += 1
            else:
                db.session.add(
                    ChavePlanejamentoRegra(
                        tipo_regra="chaves_planejamento",
                        chave_origem=origem,
                        chave_destino=None,
                        observacao=obs_linha or None,
                        usuario_id=user_id,
                        ativo=ativo_flag,
                        criado_em=now,
                        alterado_em=None,
                        excluido_em=None if ativo_flag else now,
                    )
                )
                inseridas += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao importar regras: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "message": "ImportaÃ§Ã£o de chaves concluÃ­da.",
            "inseridas": inseridas,
            "atualizadas": atualizadas,
            "ignoradas_duplicadas_arquivo": ignoradas_duplicadas_arquivo,
            "ignoradas_vazias": ignoradas_vazias,
            "total_processadas": inseridas + atualizadas,
        }
    )


@home_bp.route("/api/chave-planejamento-regra/<int:regra_id>", methods=["PUT"])
@login_required
@require_feature("atualizar/chave_planejamento_regra")
def api_chave_planejamento_regra_update(regra_id: int):
    row = ChavePlanejamentoRegra.query.filter_by(id=regra_id).first()
    if not row:
        return jsonify({"error": "Regra nao encontrada."}), 404

    payload = request.get_json(silent=True) or {}
    tipo = str(payload.get("tipo_regra") or row.tipo_regra or "").strip()
    origem = str(payload.get("chave_origem") or row.chave_origem or "").strip()
    destino = str(payload.get("chave_destino") or "").strip()
    observacao = str(payload.get("observacao") or "").strip()
    ativo = bool(payload.get("ativo", row.ativo))

    tipos_validos = {"chaves_planejamento", "chave_arrumar", "forcar_chave"}
    if tipo not in tipos_validos:
        return jsonify({"error": "Tipo de regra invalido."}), 400
    can_upload_planejamento = has_permission("atualizar/chaves_planejamento_upload") or has_permission("painel/chaves_planejamento_upload")
    if tipo == "chaves_planejamento" and not can_upload_planejamento:
        return jsonify({"error": "Usuario sem permissao para alterar chaves_planejamento."}), 403
    if not origem:
        return jsonify({"error": "Chave de origem obrigatoria."}), 400
    if tipo in {"chave_arrumar", "forcar_chave"} and not destino:
        return jsonify({"error": "Chave de destino obrigatoria para este tipo."}), 400

    now = _now_local()
    user_id = _resolve_usuario_id()
    try:
        row.tipo_regra = tipo
        row.chave_origem = origem
        row.chave_destino = destino or None
        row.observacao = observacao or None
        row.ativo = ativo
        row.usuario_id = user_id or row.usuario_id
        row.excluido_em = None if ativo else now
        row.alterado_em = now
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "Ja existe uma regra com este tipo e chave de origem."}), 409
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao atualizar regra: {exc}"}), 500

    return jsonify({"ok": True, "message": "Regra atualizada.", "row": _serialize_chave_regra(row)})


@home_bp.route("/api/chave-planejamento-regra/<int:regra_id>", methods=["DELETE"])
@login_required
@require_feature("atualizar/chave_planejamento_regra")
def api_chave_planejamento_regra_delete(regra_id: int):
    row = ChavePlanejamentoRegra.query.filter_by(id=regra_id).first()
    if not row:
        return jsonify({"error": "Regra nao encontrada."}), 404
    try:
        now = _now_local()
        row.ativo = False
        row.excluido_em = now
        row.alterado_em = now
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir regra: {exc}"}), 500
    return jsonify({"ok": True, "message": "Regra desativada com sucesso."})


@home_bp.route("/partial/atualizar/est-emp")
@login_required
@require_feature("atualizar/est-emp")
def partial_atualizar_est_emp():
    return render_template("partials/atualizar_est_emp.html")


@home_bp.route("/partial/atualizar/nob")
@login_required
@require_feature("atualizar/nob")
def partial_atualizar_nob():
    return render_template("partials/atualizar_nob.html")


@home_bp.route("/partial/atualizar/plan20-seduc")
@login_required
@require_feature("atualizar/plan20-seduc")
def partial_atualizar_plan20():
    return render_template("partials/atualizar_plan20.html")


@home_bp.route("/partial/cadastrar/dotacao")
@home_bp.route("/partial/cadastrar/dotacao/formulario")
@home_bp.route("/partial/cadastrar/dotacao/consultar")
@login_required
@require_feature("cadastrar/dotacao")
def partial_cadastrar_dotacao():
    view_mode = "consultar" if request.path.rstrip("/").endswith("/consultar") else "formulario"
    user_session = session.get("user") or {}
    user_email = (user_session.get("email") or "").strip()
    user_nome = ""
    user_id = ""
    if user_email:
        usuario_row = Usuario.query.filter_by(email=user_email).first()
        if usuario_row:
            user_nome = (usuario_row.nome or "").strip()
            user_id = str(usuario_row.id or "")
    if view_mode != "consultar":
        return render_template(
            "partials/cadastrar_dotacao.html",
            dotacoes=[],
            user_id=user_id,
            user_nome=user_nome,
            view_mode=view_mode,
        )
    rows = (
        db.session.query(Dotacao)
        .filter(
            or_(
                Dotacao.ativo == True,  # noqa: E712
                func.lower(Dotacao.status_aprovacao) == "rejeitado",
            )
        )
        .order_by(Dotacao.id.desc())
        .all()
    )
    perfil_ids = [dot.perfil_id for dot in rows if getattr(dot, "perfil_id", None)]
    adj_map = {}
    if perfil_ids:
        perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
        adj_map = {p.id: p.nome for p in perfis if p and p.nome}
    usuarios_ids = [dot.usuarios_id for dot in rows if getattr(dot, "usuarios_id", None)]
    usuarios_map = {}
    usuarios_perfil_map = {}
    usuarios_perfil_id_map = {}
    if usuarios_ids:
        usuarios = Usuario.query.filter(Usuario.id.in_(usuarios_ids)).all()
        usuarios_map = {u.id: u.nome for u in usuarios}
        usuarios_perfil_map = {u.id: u.perfil for u in usuarios if getattr(u, "perfil", None)}
        usuarios_perfil_id_map = {
            u.id: str(getattr(u, "perfil_id", "") or "").strip()
            for u in usuarios
            if getattr(u, "id", None)
        }

    dotacoes = []
    atualizar_ids: list[int] = []
    atualizar_ped_emp: list[Decimal] = []
    atualizar_estorno: list[Decimal] = []
    atualizar_situacao: list[str] = []
    atualizar_atual: list[Decimal] = []
    est_map, situacao_map = _build_estorno_maps()
    aprovado_ids: set[int] = set()
    for dot in rows:
        if getattr(dot, "aprovado_por", None):
            try:
                aprovado_ids.add(int(dot.aprovado_por))
            except Exception:
                pass
    aprovado_map = {}
    aprovado_perfil_map = {}
    if aprovado_ids:
        usuarios_aprov = Usuario.query.filter(Usuario.id.in_(aprovado_ids)).all()
        aprovado_map = {u.id: u.nome for u in usuarios_aprov}
        aprovado_perfil_map = {u.id: u.perfil for u in usuarios_aprov if getattr(u, "perfil", None)}

    for dot in rows:
        adj_nome = (adj_map.get(dot.perfil_id) or "").strip()
        ped_sum = _calc_ped_sum_for_dotacao(dot.chave_dotacao)
        emp_sum = _calc_emp_sum_for_dotacao(dot.chave_dotacao)
        key_norm = _normalize_dotacao_key(dot.chave_dotacao)
        est_sum = est_map.get(key_norm, Decimal("0"))
        est_situacao = situacao_map.get(key_norm, "")
        ped_emp_sum = _dec_or_zero(ped_sum) + _dec_or_zero(emp_sum)
        valor_dot = _dec_or_zero(dot.valor_dotacao)
        valor_atual = valor_dot - _dec_or_zero(est_sum) - ped_emp_sum
        if (
            _dec_or_zero(dot.valor_ped_emp) != ped_emp_sum
            or _dec_or_zero(dot.valor_estorno) != _dec_or_zero(est_sum)
            or _dec_or_zero(dot.valor_atual) != valor_atual
            or (est_situacao and (dot.situacao or "").strip() != est_situacao)
        ):
            atualizar_ids.append(dot.id)
            atualizar_ped_emp.append(ped_emp_sum)
            atualizar_estorno.append(_dec_or_zero(est_sum))
            atualizar_situacao.append(est_situacao)
            atualizar_atual.append(valor_atual)
        dotacoes.append(
            {
                "id": dot.id,
                "exercicio": dot.exercicio,
                "perfil_id": dot.perfil_id,
                "adj_abreviacao": adj_nome,
                "chave_planejamento": dot.chave_planejamento,
                "chave_dotacao": dot.chave_dotacao,
                "adj_concedente": getattr(dot, "adj_concedente", "") or "",
                "adj_concedente_id": _perfil_id_by_nome(getattr(dot, "adj_concedente", "") or ""),
                "status_aprovacao": getattr(dot, "status_aprovacao", "") or "",
                "aprovado_por": getattr(dot, "aprovado_por", "") or "",
                "aprovado_por_nome": aprovado_map.get(getattr(dot, "aprovado_por", None), ""),
                "aprovado_por_perfil": aprovado_perfil_map.get(getattr(dot, "aprovado_por", None), ""),
                "data_aprovacao": dot.data_aprovacao.isoformat() if getattr(dot, "data_aprovacao", None) else "",
                "motivo_rejeicao": getattr(dot, "motivo_rejeicao", "") or "",
                "uo": dot.uo,
                "programa": dot.programa,
                "acao_paoe": dot.acao_paoe,
                "produto": dot.produto,
                "ug": dot.ug,
                "regiao": dot.regiao,
                "subacao_entrega": dot.subacao_entrega,
                "etapa": dot.etapa,
                "natureza_despesa": dot.natureza_despesa,
                "elemento": dot.elemento,
                "subelemento": dot.subelemento,
                "fonte": dot.fonte,
                "iduso": dot.iduso,
                "valor_dotacao": dot.valor_dotacao,
                "valor_atual": valor_atual,
                "justificativa_historico": dot.justificativa_historico,
                "usuario_nome": usuarios_map.get(dot.usuarios_id, ""),
                "usuario_perfil": usuarios_perfil_map.get(dot.usuarios_id, ""),
                "criador_perfil_id": usuarios_perfil_id_map.get(dot.usuarios_id, ""),
                "criado_em": dot.criado_em.isoformat() if dot.criado_em else "",
                "alterado_em": dot.alterado_em.isoformat() if dot.alterado_em else "",
            }
        )
    if atualizar_ids:
        try:
            for idx, dot_id in enumerate(atualizar_ids):
                db.session.execute(
                    Dotacao.__table__.update()
                    .where(Dotacao.id == dot_id)
                    .values(
                        valor_ped_emp=atualizar_ped_emp[idx],
                        valor_estorno=atualizar_estorno[idx],
                        situacao=atualizar_situacao[idx],
                        valor_atual=atualizar_atual[idx],
                        alterado_em=_now_local(),
                    )
                )
            db.session.commit()
        except Exception:
            db.session.rollback()
    return render_template(
        "partials/cadastrar_dotacao.html",
        dotacoes=dotacoes,
        user_id=user_id,
        user_nome=user_nome,
        view_mode=view_mode,
    )


@home_bp.route("/partial/cadastrar/plan_21-nger/subacao")
@home_bp.route("/partial/cadastrar/plan_21-nger/subacao/formulario")
@home_bp.route("/partial/cadastrar/plan_21-nger/subacao/consultar")
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def partial_cadastrar_plan21_subacao():
    view_mode = "consultar" if request.path.rstrip("/").endswith("/consultar") else "formulario"
    user_id = _resolve_usuario_id() or ""
    user_perfil_id = str(_resolve_usuario_perfil_id() or "").strip()
    user_session = session.get("user") or {}
    user_nome = (user_session.get("nome") or "").strip()
    subacoes = (
        CadastrarSubacao.query.filter(
            (CadastrarSubacao.ativo == True)  # noqa: E712
            | (func.lower(CadastrarSubacao.status_aprovacao) == "rejeitado")
        )
        .order_by(CadastrarSubacao.id.desc())
        .all()
    )
    usuario_ids = {s.usuario_id for s in subacoes if s.usuario_id}
    aprovado_ids = set()
    for s in subacoes:
        aprovado_id = getattr(s, "aprovado_por", None)
        if not aprovado_id:
            continue
        try:
            aprovado_ids.add(int(aprovado_id))
        except Exception:
            continue
    all_user_ids = set(usuario_ids)
    all_user_ids.update(aprovado_ids)
    usuarios_map = {}
    perfil_ids = set()
    if all_user_ids:
        usuarios = Usuario.query.filter(Usuario.id.in_(all_user_ids)).all()
        usuarios_map = {u.id: u for u in usuarios}
        perfil_ids = {u.perfil_id for u in usuarios if u.perfil_id}
    perfil_map = {}
    if perfil_ids:
        perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
        perfil_map = {p.id: p for p in perfis}
    plan21_ids_all: set[int] = set()
    subacao_plan_ids: dict[int, list[int]] = {}
    for s in subacoes:
        ids = _extract_plan21_ids(s)
        if ids:
            subacao_plan_ids[s.id] = ids
            plan21_ids_all.update(ids)
    plan21_subacao_map: dict[int, str] = {}
    if plan21_ids_all:
        plan_rows = Plan21Nger.query.filter(Plan21Nger.id.in_(list(plan21_ids_all))).all()
        for r in plan_rows:
            try:
                key = int(getattr(r, "id", 0) or 0)
            except Exception:
                key = 0
            if key > 0:
                plan21_subacao_map[key] = (getattr(r, "subacao_entrega", "") or "").strip()
    for s in subacoes:
        usuario = usuarios_map.get(s.usuario_id)
        perfil_id = getattr(usuario, "perfil_id", None)
        perfil_nome = ""
        if perfil_id:
            perfil_row = perfil_map.get(perfil_id)
            perfil_nome = (getattr(perfil_row, "nome", "") or "").strip()
        exercicio = getattr(s, "exercicio", "") or ""
        if perfil_nome:
            s.controle_subacao = f"SUB.{exercicio}.{perfil_nome}.{s.id}"
        else:
            s.controle_subacao = f"SUB.{exercicio}.{s.id}"
        s.usuario_nome = (
            (getattr(usuario, "nome", "") or getattr(usuario, "email", "") or "").strip()
            if usuario
            else ""
        )
        s.usuario_perfil = perfil_nome
        s.criador_perfil_id = str(perfil_id or "").strip() if perfil_id else ""
        aprovado_nome = ""
        aprovado_perfil = ""
        aprovado_id = getattr(s, "aprovado_por", None)
        if aprovado_id:
            try:
                aprovado_id = int(aprovado_id)
            except Exception:
                aprovado_id = None
        if aprovado_id:
            aprovado_usuario = usuarios_map.get(aprovado_id)
            aprovado_nome = (
                (getattr(aprovado_usuario, "nome", "") or getattr(aprovado_usuario, "email", "") or "").strip()
                if aprovado_usuario
                else ""
            )
            aprovado_perfil_id = getattr(aprovado_usuario, "perfil_id", None) if aprovado_usuario else None
            if aprovado_perfil_id:
                aprovado_perfil_row = perfil_map.get(aprovado_perfil_id)
                aprovado_perfil = (getattr(aprovado_perfil_row, "nome", "") or "").strip()
        s.aprovado_por_nome = aprovado_nome
        s.aprovado_por_perfil = aprovado_perfil
        s.criado_em_iso = s.criado_em.isoformat() if getattr(s, "criado_em", None) else ""
        s.data_aprovacao_iso = (
            s.data_aprovacao.isoformat() if getattr(s, "data_aprovacao", None) else ""
        )
        s.subacao_origem = ""
        if (getattr(s, "tipo_edicao", "") or "").strip().lower() == "subacao_name":
            ids = subacao_plan_ids.get(s.id) or []
            if ids:
                s.subacao_origem = plan21_subacao_map.get(ids[0], "")
    return render_template(
        "partials/cadastrar_plan21_nger_subacao.html",
        subacoes=subacoes,
        user_id=user_id,
        user_perfil_id=user_perfil_id,
        user_nome=user_nome,
        current_year=str(_now_local().year),
        view_mode=view_mode,
    )


@home_bp.route("/partial/cadastrar/plan_21-nger/etapa")
@home_bp.route("/partial/cadastrar/plan_21-nger/etapa/formulario")
@home_bp.route("/partial/cadastrar/plan_21-nger/etapa/consultar")
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def partial_cadastrar_plan21_etapa():
    view_mode = "consultar" if request.path.rstrip("/").endswith("/consultar") else "formulario"
    user_id = _resolve_usuario_id() or ""
    user_perfil_id = str(_resolve_usuario_perfil_id() or "").strip()
    user_session = session.get("user") or {}
    user_nome = (user_session.get("nome") or "").strip()
    etapas = (
        CadastrarEtapa.query.filter(
            (CadastrarEtapa.ativo == True)  # noqa: E712
            | (func.lower(CadastrarEtapa.status_aprovacao) == "rejeitado")
        )
        .order_by(CadastrarEtapa.id.desc())
        .all()
    )
    usuario_ids = {e.usuario_id for e in etapas if e.usuario_id}
    aprovado_ids = set()
    for e in etapas:
        aprovado_id = getattr(e, "aprovado_por", None)
        if not aprovado_id:
            continue
        try:
            aprovado_ids.add(int(aprovado_id))
        except Exception:
            continue
    all_user_ids = set(usuario_ids)
    all_user_ids.update(aprovado_ids)
    usuarios_map = {}
    perfil_map = {}
    if all_user_ids:
        usuarios = Usuario.query.filter(Usuario.id.in_(all_user_ids)).all()
        usuarios_map = {u.id: u for u in usuarios}
        perfil_ids = {u.perfil_id for u in usuarios if u.perfil_id}
        if perfil_ids:
            perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
            perfil_map = {p.id: p for p in perfis}
    for e in etapas:
        usuario = usuarios_map.get(e.usuario_id)
        perfil_id = getattr(usuario, "perfil_id", None)
        perfil_nome = ""
        if perfil_id:
            perfil_nome = (getattr(perfil_map.get(perfil_id), "nome", "") or "").strip()
        exercicio = getattr(e, "exercicio", "") or ""
        e.controle_etapa = f"ETAPA.{exercicio}.{perfil_nome}.{e.id}" if perfil_nome else f"ETAPA.{exercicio}.{e.id}"
        e.usuario_nome = (
            (getattr(usuario, "nome", "") or getattr(usuario, "email", "") or "").strip()
            if usuario
            else ""
        )
        e.usuario_perfil = perfil_nome
        e.criador_perfil_id = str(perfil_id or "").strip() if perfil_id else ""
        aprovado_nome = ""
        aprovado_perfil = ""
        aprovado_id = getattr(e, "aprovado_por", None)
        if aprovado_id:
            try:
                aprovado_id = int(aprovado_id)
            except Exception:
                aprovado_id = None
        if aprovado_id:
            aprovado_usuario = usuarios_map.get(aprovado_id)
            aprovado_nome = (
                (getattr(aprovado_usuario, "nome", "") or getattr(aprovado_usuario, "email", "") or "").strip()
                if aprovado_usuario
                else ""
            )
            aprovado_perfil_id = getattr(aprovado_usuario, "perfil_id", None) if aprovado_usuario else None
            if aprovado_perfil_id:
                aprovado_perfil = (getattr(perfil_map.get(aprovado_perfil_id), "nome", "") or "").strip()
        e.aprovado_por_nome = aprovado_nome
        e.aprovado_por_perfil = aprovado_perfil
        e.criado_em_iso = e.criado_em.isoformat() if getattr(e, "criado_em", None) else ""
        e.data_aprovacao_iso = e.data_aprovacao.isoformat() if getattr(e, "data_aprovacao", None) else ""
    return render_template(
        "partials/cadastrar_plan21_nger_etapa.html",
        etapas=etapas,
        user_id=user_id,
        user_perfil_id=user_perfil_id,
        user_nome=user_nome,
        user_nivel=str(getattr(g, "user_nivel", "") or ""),
        current_year=str(_now_local().year),
        view_mode=view_mode,
    )


@home_bp.route("/partial/cadastrar/plan_21-nger/meta_fisica")
@home_bp.route("/partial/cadastrar/plan_21-nger/meta_fisica/formulario")
@home_bp.route("/partial/cadastrar/plan_21-nger/meta_fisica/consultar")
@login_required
@require_feature("cadastrar/plan_21-nger/meta_fisica")
def partial_cadastrar_plan21_meta_fisica():
    view_mode = "consultar" if request.path.rstrip("/").endswith("/consultar") else "formulario"
    user_id = _resolve_usuario_id() or ""
    user_session = session.get("user") or {}
    user_nome = (user_session.get("nome") or "").strip()
    user_perfil_id = ""
    user_perfil_nome = ""
    try:
        user_id_int = int(user_id) if str(user_id).strip() else 0
    except Exception:
        user_id_int = 0
    if user_id_int > 0:
        user_row = Usuario.query.filter_by(id=user_id_int).first()
        if user_row:
            if getattr(user_row, "perfil_id", None):
                user_perfil_id = str(user_row.perfil_id)
                perfil_row_user = Perfil.query.filter_by(id=user_row.perfil_id).first()
                user_perfil_nome = (getattr(perfil_row_user, "nome", "") or "").strip() if perfil_row_user else ""
            if not user_perfil_nome:
                user_perfil_nome = (getattr(user_row, "perfil", "") or "").strip()
    metas = (
        AlterarMeta.query.filter(
            or_(
                AlterarMeta.ativo == True,  # noqa: E712
                func.lower(func.coalesce(AlterarMeta.status_aprovacao, "")) == "rejeitado",
            )
        )
        .order_by(AlterarMeta.id.desc())
        .all()
    )
    usuario_ids = {m.usuario_id for m in metas if m.usuario_id}
    aprovado_ids = set()
    for m in metas:
        aprovado_id = getattr(m, "aprovado_por", None)
        if not aprovado_id:
            continue
        try:
            aprovado_ids.add(int(aprovado_id))
        except Exception:
            continue
    usuarios_map = {}
    perfil_map = {}
    all_user_ids = set(usuario_ids)
    all_user_ids.update(aprovado_ids)
    if all_user_ids:
        usuarios = Usuario.query.filter(Usuario.id.in_(list(all_user_ids))).all()
        usuarios_map = {u.id: u for u in usuarios}
        perfil_ids = {u.perfil_id for u in usuarios if u.perfil_id}
        if perfil_ids:
            perfis = Perfil.query.filter(Perfil.id.in_(list(perfil_ids))).all()
            perfil_map = {p.id: p for p in perfis}
    for m in metas:
        usuario = usuarios_map.get(getattr(m, "usuario_id", None))
        m.criador_nome = (getattr(usuario, "nome", "") or getattr(usuario, "email", "") or "").strip() if usuario else ""
        m.criador_perfil_nome = ""
        if usuario:
            perfil_id = getattr(usuario, "perfil_id", None)
            if perfil_id:
                perfil_row = perfil_map.get(perfil_id)
                m.criador_perfil_nome = (getattr(perfil_row, "nome", "") or "").strip() if perfil_row else ""
            if not m.criador_perfil_nome:
                m.criador_perfil_nome = (getattr(usuario, "perfil", "") or "").strip()
        m.criador_perfil_id = (
            str(getattr(usuario, "perfil_id", "") or "").strip() if usuario else ""
        )
        exercicio_ctrl = str(getattr(m, "exercicio", "") or _now_local().year)
        adj_nome = (getattr(m, "responsavel_acao", "") or "").strip()
        if not adj_nome:
            if usuario:
                perfil_id = getattr(usuario, "perfil_id", None)
                if perfil_id:
                    perfil_row = perfil_map.get(perfil_id)
                    adj_nome = (getattr(perfil_row, "nome", "") or "").strip()
                if not adj_nome:
                    adj_nome = (getattr(usuario, "perfil", "") or "").strip()
        adj_token = _normalize_controle_token(adj_nome) or "SEMADJ"
        m.controle_meta = f"META.{exercicio_ctrl}.{adj_token}.{m.id}"
        aprovado_nome = ""
        aprovado_perfil_nome = ""
        aprovado_id = getattr(m, "aprovado_por", None)
        if aprovado_id:
            try:
                aprovado_id = int(aprovado_id)
            except Exception:
                aprovado_id = None
        if aprovado_id:
            aprovado_usuario = usuarios_map.get(aprovado_id)
            aprovado_nome = (
                (getattr(aprovado_usuario, "nome", "") or getattr(aprovado_usuario, "email", "") or "").strip()
                if aprovado_usuario
                else ""
            )
            aprovado_perfil_id = getattr(aprovado_usuario, "perfil_id", None) if aprovado_usuario else None
            if aprovado_perfil_id:
                aprovado_perfil_row = perfil_map.get(aprovado_perfil_id)
                aprovado_perfil_nome = (getattr(aprovado_perfil_row, "nome", "") or "").strip()
            if not aprovado_perfil_nome and aprovado_usuario:
                aprovado_perfil_nome = (getattr(aprovado_usuario, "perfil", "") or "").strip()
        m.aprovado_por_nome = aprovado_nome
        m.aprovado_por_perfil = aprovado_perfil_nome
        m.data_aprovacao_iso = (
            getattr(m, "data_aprovacao", None).isoformat() if getattr(m, "data_aprovacao", None) else ""
        )
        m.regioes_preview = ""
        m.linhas_count = 0
        m.regioes_payload = []
        m.criado_em_iso = getattr(m, "criado_em", None).isoformat() if getattr(m, "criado_em", None) else ""
        reg_raw = (getattr(m, "regiao_produto", "") or "").strip()
        if reg_raw.startswith("[") and reg_raw.endswith("]"):
            try:
                parsed = json.loads(reg_raw)
            except Exception:
                parsed = []
            if isinstance(parsed, list):
                m.regioes_payload = parsed
                regioes = []
                for item in parsed:
                    reg = str((item or {}).get("regiao_produto") or "").strip()
                    if reg:
                        regioes.append(reg)
                if regioes:
                    m.regioes_preview = ", ".join(regioes[:5])
                    if len(regioes) > 5:
                        m.regioes_preview += "..."
                    m.linhas_count = len(regioes)
    return render_template(
        "partials/cadastrar_plan21_nger_meta_fisica.html",
        metas=metas,
        user_id=user_id,
        user_nome=user_nome,
        user_perfil_id=user_perfil_id,
        user_perfil_nome=user_perfil_nome,
        user_nivel=getattr(g, "user_nivel", "") or "",
        current_year=str(_now_local().year),
        view_mode=view_mode,
    )


@home_bp.route("/partial/cadastrar/est-dotacao")
@home_bp.route("/partial/cadastrar/est-dotacao/formulario")
@home_bp.route("/partial/cadastrar/est-dotacao/consultar")
@login_required
@require_feature("cadastrar/est-dotacao")
def partial_cadastrar_est_dotacao():
    view_mode = "consultar" if request.path.rstrip("/").endswith("/consultar") else "formulario"
    user_session = session.get("user") or {}
    user_email = (user_session.get("email") or "").strip()
    user_nome = ""
    user_id = ""
    if user_email:
        usuario_row = Usuario.query.filter_by(email=user_email).first()
        if usuario_row:
            user_nome = (usuario_row.nome or "").strip()
            user_id = str(usuario_row.id or "")
    rows = (
        db.session.query(Dotacao)
        .filter(Dotacao.ativo == True)  # noqa: E712
        .filter(func.lower(Dotacao.status_aprovacao) == "aprovado")
        .order_by(Dotacao.id.desc())
        .all()
    )
    perfil_ids = [dot.perfil_id for dot in rows if getattr(dot, "perfil_id", None)]
    adj_map = {}
    if perfil_ids:
        perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
        adj_map = {p.id: p.nome for p in perfis if p and p.nome}

    dotacoes = []
    for dot in rows:
        adj_nome = (adj_map.get(dot.perfil_id) or "").strip()
        dotacoes.append(
            {
                "id": dot.id,
                "exercicio": dot.exercicio,
                "perfil_id": dot.perfil_id,
                "adj_abreviacao": adj_nome,
                "chave_planejamento": dot.chave_planejamento,
                "chave_dotacao": dot.chave_dotacao,
                "status_aprovacao": getattr(dot, "status_aprovacao", "") or "",
                "uo": dot.uo,
                "programa": dot.programa,
                "acao_paoe": dot.acao_paoe,
                "produto": dot.produto,
                "ug": dot.ug,
                "regiao": dot.regiao,
                "subacao_entrega": dot.subacao_entrega,
                "etapa": dot.etapa,
                "natureza_despesa": dot.natureza_despesa,
                "elemento": dot.elemento,
                "subelemento": dot.subelemento,
                "fonte": dot.fonte,
                "iduso": dot.iduso,
                "valor_dotacao": dot.valor_dotacao,
                "valor_atual": dot.valor_atual,
                "justificativa_historico": dot.justificativa_historico,
            }
        )

    estorno_rows = []
    try:
        raw = db.session.execute(text("SELECT * FROM est_dotacao WHERE ativo = 1")).mappings().all()
    except Exception:
        raw = []
    est_perfil_ids = {r.get("perfil_id") for r in raw if r.get("perfil_id")}
    est_adj_map = {}
    if est_perfil_ids:
        est_perfis = Perfil.query.filter(Perfil.id.in_(est_perfil_ids)).all()
        est_adj_map = {p.id: p.nome for p in est_perfis if p and p.nome}
    est_usuario_ids = {r.get("usuarios_id") for r in raw if r.get("usuarios_id")}
    est_usuario_perfil_id_map = {}
    if est_usuario_ids:
        est_usuarios = Usuario.query.filter(Usuario.id.in_(list(est_usuario_ids))).all()
        est_usuario_perfil_id_map = {
            u.id: str(getattr(u, "perfil_id", "") or "").strip()
            for u in est_usuarios
            if getattr(u, "id", None)
        }
    for r in raw:
        def pick(*keys):
            for k in keys:
                if k in r and r[k] is not None:
                    return r[k]
            return ""

        est_id = pick("id")
        chave = pick("chave_dotacao", "chave", "controle_dotacao")
        status = pick("status_aprovacao", "status")
        justificativa = pick("justificativa_estorno", "justificativa", "historico", "motivo", "observacao")
        valor_dot = pick("valor_dotacao", "valor")
        valor_est = pick("valor_a_ser_est", "valor_estorno", "estorno")
        saldo_est = pick("saldo_dotacao_apos", "saldo_estorno", "saldo")
        adj_solic = pick("adjunta_solicitante", "adj_solicitante", "adjunta", "adj")
        exercicio = pick("exercicio", "ano")
        programa = pick("programa", "programa_governo")
        paoe = pick("paoe", "acao_paoe")
        chave_planejamento = pick("chave_planejamento")
        uo = pick("uo")
        ug = pick("ug")
        regiao = pick("regiao")
        subacao_entrega = pick("subacao_entrega")
        etapa = pick("etapa")
        natureza_despesa = pick("natureza_despesa")
        elemento = pick("elemento")
        subelemento = pick("subelemento")
        fonte = pick("fonte")
        iduso = pick("iduso")
        produto = pick("produto")
        situacao = pick("situacao")
        if not adj_solic:
            perfil_id = r.get("perfil_id")
            adj_solic = est_adj_map.get(perfil_id, "")

        def fmt_num(val):
            v = _parse_decimal_value(val)
            return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        estorno_rows.append(
            {
                "id": str(est_id or ""),
                "perfil_id": str(r.get("perfil_id") or ""),
                "status": str(status or ""),
                "chave_dotacao": str(chave or ""),
                "justificativa_estorno": str(justificativa or ""),
                "valor_dotacao": fmt_num(valor_dot),
                "valor_estorno": fmt_num(valor_est),
                "saldo_estorno": fmt_num(saldo_est),
                "adjunta_solicitante": str(adj_solic or ""),
                "exercicio": str(exercicio or ""),
                "programa": str(programa or ""),
                "paoe": str(paoe or ""),
                "chave_planejamento": str(chave_planejamento or ""),
                "uo": str(uo or ""),
                "ug": str(ug or ""),
                "regiao": str(regiao or ""),
                "subacao_entrega": str(subacao_entrega or ""),
                "etapa": str(etapa or ""),
                "natureza_despesa": str(natureza_despesa or ""),
                "elemento": str(elemento or ""),
                "subelemento": str(subelemento or ""),
                "fonte": str(fonte or ""),
                "iduso": str(iduso or ""),
                "produto": str(produto or ""),
                "situacao": str(situacao or ""),
                "criador_perfil_id": est_usuario_perfil_id_map.get(r.get("usuarios_id"), ""),
            }
        )

    return render_template(
        "partials/cadastrar_est_dotacao.html",
        dotacoes=dotacoes,
        estornos=estorno_rows,
        user_id=user_id,
        user_nome=user_nome,
        view_mode=view_mode,
    )


@home_bp.route("/partial/institucional/diretrizes")
@login_required
def partial_institucional_diretrizes():
    return render_template("partials/institucional_diretrizes.html")


@home_bp.route("/partial/institucional/repositorio")
@login_required
def partial_institucional_repositorio():
    return render_template("partials/institucional_repositorio.html")


@home_bp.route("/partial/institucional/legislacao")
@login_required
def partial_institucional_legislacao():
    return render_template("partials/institucional_legislacao.html")


@home_bp.route("/partial/institucional/parceiros")
@login_required
def partial_institucional_parceiros():
    return render_template("partials/institucional_parceiros.html")


@home_bp.route("/partial/relatorios/fip613")
@login_required
@require_feature("relatorios/fip613")
def partial_relatorios_fip613():
    return render_template("partials/relatorios_fip613.html")


@home_bp.route("/partial/relatorios/ped")
@login_required
@require_feature("relatorios/ped")
def partial_relatorios_ped():
    return render_template("partials/relatorios_ped.html")


@home_bp.route("/partial/relatorios/emp")
@login_required
@require_feature("relatorios/emp")
def partial_relatorios_emp():
    return render_template("partials/relatorios_emp.html")


@home_bp.route("/partial/relatorios/dotacao")
@login_required
@require_feature("relatorios/dotacao")
def partial_relatorios_dotacao():
    return render_template("partials/relatorios_dotacao.html")


@home_bp.route("/partial/relatorios/est-dotacao")
@login_required
@require_feature("relatorios/est-dotacao")
def partial_relatorios_est_dotacao():
    return render_template("partials/relatorios_est_dotacao.html")


@home_bp.route("/partial/relatorios/est-emp")
@login_required
@require_feature("relatorios/est-emp")
def partial_relatorios_est_emp():
    return render_template("partials/relatorios_est_emp.html")


@home_bp.route("/partial/relatorios/nob")
@login_required
@require_feature("relatorios/nob")
def partial_relatorios_nob():
    return render_template("partials/relatorios_nob.html")


@home_bp.route("/partial/relatorios/plan20-seduc")
@login_required
@require_feature("relatorios/plan20-seduc")
def partial_relatorios_plan20():
    return render_template("partials/relatorios_plan20.html")


@home_bp.route("/partial/relatorios/plan21-nger")
@login_required
@require_feature("relatorios/plan21-nger")
def partial_relatorios_plan21_nger():
    return render_template("partials/relatorios_plan21_nger.html")


@home_bp.route("/api/permissoes/<int:perfil_id>", methods=["GET", "POST"])
@login_required
def api_permissoes(perfil_id):
    perfil = db.session.get(Perfil, perfil_id)
    if not perfil:
        return jsonify({"error": "Perfil nao encontrado."}), 404

    if not (has_permission("painel") or getattr(g, "user_nivel", None) == 1):
        return jsonify({"error": "Sem permissao."}), 403

    if request.method == "GET":
        try:
            feats = _load_permissoes_perfil(perfil_id)
            nivel_feats = _load_permissoes_nivel(perfil.nivel)
            return jsonify({"features": feats, "nivel_features": nivel_feats, "nivel": perfil.nivel})
        except ProgrammingError:
            db.session.rollback()
            return jsonify({"features": [], "nivel_features": [], "nivel": perfil.nivel})

    data = request.get_json() or {}
    feats = data.get("features") or []
    if not isinstance(feats, list):
        return jsonify({"error": "Formato invalido."}), 400
    # limpa valores vazios/None e remove duplicados
    clean_feats = []
    seen = set()
    for f in feats:
        if not isinstance(f, str):
            continue
        fid = f.strip()
        if not fid or fid in seen:
            continue
        seen.add(fid)
        clean_feats.append(fid)
    locked_feats = {f["id"] for f in FEATURES if f.get("locked")}
    clean_feats = [f for f in clean_feats if f not in locked_feats]
    try:
        # desativa anteriores
        PerfilPermissao.query.filter_by(perfil_id=perfil_id).update({"ativo": False, "updated_at": datetime.utcnow()})
        # remove qualquer registro sem feature
        PerfilPermissao.query.filter(PerfilPermissao.feature == None).delete(synchronize_session=False)  # noqa: E711
        for f in clean_feats:
            db.session.add(
                PerfilPermissao(
                    perfil_id=perfil_id,
                    feature=f,
                    ativo=True,
                    created_at=datetime.utcnow(),
                )
            )
        db.session.commit()
    except ProgrammingError:
        db.session.rollback()
        return jsonify({"error": "Tabela perfil_permissoes inexistente. Crie a tabela antes de salvar."}), 500
    return jsonify({"ok": True, "message": "Permissoes atualizadas."})


@home_bp.route("/api/permissoes/nivel/<int:nivel>", methods=["GET", "POST"])
@login_required
def api_permissoes_nivel(nivel):
    if not (has_permission("painel") or getattr(g, "user_nivel", None) == 1):
        return jsonify({"error": "Sem permissao."}), 403
    if nivel < 1 or nivel > 5:
        return jsonify({"error": "Nivel invalido."}), 400

    if request.method == "GET":
        try:
            feats = _load_permissoes_nivel(nivel)
            perfil_feats = _load_permissoes_por_nivel_perfis(nivel)
            return jsonify({"features": feats, "perfil_features": perfil_feats})
        except ProgrammingError:
            db.session.rollback()
            return jsonify({"features": [], "perfil_features": []})

    data = request.get_json() or {}
    feats = data.get("features") or []
    if not isinstance(feats, list):
        return jsonify({"error": "Formato invalido."}), 400
    clean_feats = []
    seen = set()
    for f in feats:
        if not isinstance(f, str):
            continue
        fid = f.strip()
        if not fid or fid in seen:
            continue
        seen.add(fid)
        clean_feats.append(fid)
    locked_feats = {f["id"] for f in FEATURES if f.get("locked")}
    clean_feats = [f for f in clean_feats if f not in locked_feats]
    try:
        NivelPermissao.query.filter_by(nivel=nivel).update({"ativo": False, "updated_at": datetime.utcnow()})
        NivelPermissao.query.filter(NivelPermissao.feature == None).delete(synchronize_session=False)  # noqa: E711
        for f in clean_feats:
            db.session.add(NivelPermissao(nivel=nivel, feature=f, ativo=True))
        db.session.commit()
    except ProgrammingError:
        db.session.rollback()
        return jsonify({"error": "Tabela nivel_permissoes inexistente. Crie a tabela antes de salvar."}), 500
    return jsonify({"ok": True, "message": "Permissoes atualizadas."})


@home_bp.route("/api/permissoes/current", methods=["GET"])
@login_required
def api_permissoes_current():
    user_session = session.get("user") or {}
    perfil_id = getattr(g, "user_perfil_id", None) or user_session.get("perfil_id")
    _debug_probe(
        "permissoes_current_start",
        path=request.path,
        perfil_id=perfil_id,
        nivel=getattr(g, "user_nivel", None),
        has_user=bool(session.get("user")),
    )
    try:
        feats = _permissoes_with_parents(perfil_id, getattr(g, "user_nivel", None))
    except SQLAlchemyError as exc:
        current_app.logger.warning(
            "Falha de banco ao carregar permissoes (perfil_id=%s, path=%s): %s",
            perfil_id,
            request.path,
            exc,
            exc_info=True,
        )
        _debug_probe("permissoes_current_db_error", path=request.path, perfil_id=perfil_id, exc=str(exc))
        return jsonify({"error": "Permissoes indisponiveis temporariamente."}), 503
    except Exception as exc:
        current_app.logger.warning("Falha ao carregar permissoes: %s", exc)
        _debug_probe("permissoes_current_generic_error", path=request.path, perfil_id=perfil_id, exc=str(exc))
        feats = []
    _debug_probe("permissoes_current_success", path=request.path, perfil_id=perfil_id, features_count=len(feats))
    return jsonify({"features": feats})


def _parse_decimal(raw_val):
    if raw_val is None:
        return None
    if isinstance(raw_val, (list, tuple)):
        if not raw_val:
            return None
        return _parse_decimal(raw_val[0])
    raw = str(raw_val).replace("\xa0", " ").strip()
    if not raw:
        return None
    if raw.startswith("[") and raw.endswith("]"):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list) and parsed:
                return _parse_decimal(parsed[0])
        except (TypeError, ValueError):
            pass
    raw = raw.strip("\"'").strip()
    if not raw:
        return None
    cleaned = raw.replace(" ", "")
    has_dot = "." in cleaned
    has_comma = "," in cleaned
    if has_dot and has_comma:
        # Decide decimal separator by the last symbol position.
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif has_comma:
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+", cleaned):
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
    elif has_dot:
        if re.fullmatch(r"-?\d+\.\d{1,6}", cleaned):
            pass
        elif re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+\.?", cleaned):
            cleaned = cleaned.replace(".", "")
        else:
            cleaned = cleaned.replace(".", "")
    try:
        return Decimal(cleaned)
    except Exception:
        return None


def _dec_or_zero(value):
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    parsed = _parse_decimal(value)
    return parsed if parsed is not None else Decimal("0")


def _extract_justificativa_text(raw: str) -> str:
    if not raw:
        return ""
    text = str(raw).strip()
    match = re.search(r"^DOT\\.[^.]*\\.[^.]*\\.\\d+(?:\\s+(.*))?$", text)
    if match:
        return (match.group(1) or "").strip()
    return text


def _natureza_prefix(value: str) -> str:
    if not value:
        return ""
    parts = [p for p in str(value).split(".") if p]
    if len(parts) >= 3:
        return ".".join(parts[:3])
    return str(value).strip()


def _dotacao_payload(registro: Dotacao, adj_label: str) -> dict:
    aprovado_nome = ""
    aprovado_perfil = ""
    aprovado_id = getattr(registro, "aprovado_por", None)
    if aprovado_id:
        try:
            aprovado_id = int(aprovado_id)
        except Exception:
            aprovado_id = None
    if aprovado_id:
        usuario_aprov = Usuario.query.filter_by(id=aprovado_id).first()
        aprovado_nome = (usuario_aprov.nome or "").strip() if usuario_aprov else ""
        aprovado_perfil = (usuario_aprov.perfil or "").strip() if usuario_aprov else ""
    usuario_perfil = ""
    usuarios_id = getattr(registro, "usuarios_id", None)
    if usuarios_id:
        usuario_row = Usuario.query.filter_by(id=usuarios_id).first()
        usuario_perfil = (usuario_row.perfil or "").strip() if usuario_row else ""
    return {
        "id": registro.id,
        "exercicio": registro.exercicio,
        "perfil_id": getattr(registro, "perfil_id", None),
        "adjunta": adj_label,
        "chave_planejamento": registro.chave_planejamento,
        "adj_concedente": getattr(registro, "adj_concedente", "") or "",
        "adj_concedente_id": _perfil_id_by_nome(getattr(registro, "adj_concedente", "") or ""),
        "status_aprovacao": getattr(registro, "status_aprovacao", "") or "",
        "aprovado_por": getattr(registro, "aprovado_por", "") or "",
        "aprovado_por_nome": aprovado_nome,
        "aprovado_por_perfil": aprovado_perfil,
        "data_aprovacao": registro.data_aprovacao.isoformat() if getattr(registro, "data_aprovacao", None) else "",
        "motivo_rejeicao": getattr(registro, "motivo_rejeicao", "") or "",
        "uo": registro.uo,
        "programa": registro.programa,
        "acao_paoe": registro.acao_paoe,
        "produto": registro.produto,
        "ug": registro.ug,
        "regiao": registro.regiao,
        "subacao_entrega": registro.subacao_entrega,
        "etapa": registro.etapa,
        "natureza_despesa": registro.natureza_despesa,
        "elemento": registro.elemento,
        "subelemento": registro.subelemento,
        "fonte": registro.fonte,
        "iduso": registro.iduso,
        "justificativa_historico": registro.justificativa_historico,
        "valor_dotacao": str(registro.valor_dotacao or ""),
        "chave_dotacao": registro.chave_dotacao,
        "usuario_nome": getattr(registro, "usuario_nome", ""),
        "usuario_perfil": usuario_perfil,
        "criado_em": registro.criado_em.isoformat() if registro.criado_em else "",
        "alterado_em": registro.alterado_em.isoformat() if registro.alterado_em else "",
    }


def _resolve_usuario_id():
    user = session.get("user") or {}
    email = (user.get("email") or "").strip()
    if not email:
        return None
    usuario = Usuario.query.filter_by(email=email).first()
    return getattr(usuario, "id", None) if usuario else None


def _resolve_usuario_perfil_id():
    perfil_id = getattr(g, "user_perfil_id", None)
    if perfil_id:
        return int(perfil_id)
    user = session.get("user") or {}
    perfil_id = user.get("perfil_id")
    if perfil_id:
        try:
            return int(perfil_id)
        except (TypeError, ValueError):
            return None
    return None


def _current_user_matches_perfil(perfil_id) -> bool:
    current_perfil_id = _resolve_usuario_perfil_id()
    if not current_perfil_id or not perfil_id:
        return False
    try:
        return int(current_perfil_id) == int(perfil_id)
    except (TypeError, ValueError):
        return False


def _now_local():
    tz = pytz.timezone("America/Manaus")
    return datetime.now(tz).replace(tzinfo=None)


def _attach_usuario_nome(registro: Dotacao) -> Dotacao:
    usuarios_id = getattr(registro, "usuarios_id", None)
    if not usuarios_id:
        registro.usuario_nome = ""
        return registro
    usuario = Usuario.query.filter_by(id=usuarios_id).first()
    registro.usuario_nome = (getattr(usuario, "nome", "") or getattr(usuario, "email", "") or "").strip() if usuario else ""
    return registro


def _leading_token(value: str) -> str:
    if not value:
        return ""
    return str(value).strip().split(" ", 1)[0]


def _normalize_controle_token(value: str) -> str:
    if not value:
        return ""
    text_val = unicodedata.normalize("NFKD", str(value))
    text_val = "".join(ch for ch in text_val if not unicodedata.combining(ch))
    text_val = re.sub(r"[^A-Za-z0-9]+", "", text_val).upper()
    return text_val


def _normalize_codigo_num(value: str) -> str:
    if not value:
        return ""
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    if digits:
        return str(int(digits))
    return _leading_token(value)


def _normalize_ug(value: str) -> str:
    return _normalize_codigo_num(value)


def _regiao_produto_exists(
    exercicio: str,
    uo: str,
    programa: str,
    acao_paoe: str,
    produto_acao: str,
    regiao_subacao: str,
) -> bool:
    regiao_key = _normalize_codigo_num(regiao_subacao)
    if not regiao_key:
        return True
    query = Plan21Nger.query.with_entities(Plan21Nger.regiao_produto)
    query = query.filter(Plan21Nger.exercicio == str(exercicio))
    query = query.filter(Plan21Nger.uo == uo)
    query = query.filter(Plan21Nger.programa == programa)
    query = query.filter(Plan21Nger.acao_paoe == acao_paoe)
    query = query.filter(Plan21Nger.produto == produto_acao)
    rows = query.all()
    for (regiao_produto,) in rows:
        if regiao_produto is None:
            continue
        rp_str = str(regiao_produto).strip()
        if not rp_str:
            continue
        if rp_str == regiao_subacao or rp_str == regiao_key:
            return True
        rp_key = _normalize_codigo_num(rp_str)
        if rp_key == regiao_key:
            return True
        if rp_str.startswith(regiao_key) or rp_str.startswith(f"R{regiao_key}"):
            return True
    return False


def _normalize_uo(value: str) -> str:
    if not value:
        return ""
    token = _leading_token(value)
    digits = "".join(ch for ch in token if ch.isdigit())
    return digits or token


def _is_valid_cpf(value: str) -> bool:
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 11:
        return False
    if digits == digits[0] * 11:
        return False
    nums = [int(d) for d in digits]
    sum1 = sum((10 - i) * nums[i] for i in range(9))
    d1 = 0 if (sum1 % 11) < 2 else 11 - (sum1 % 11)
    if nums[9] != d1:
        return False
    sum2 = sum((11 - i) * nums[i] for i in range(10))
    d2 = 0 if (sum2 % 11) < 2 else 11 - (sum2 % 11)
    return nums[10] == d2


def _safe_query_mappings(sql: str, params: dict | None = None):
    try:
        return db.session.execute(text(sql), params or {}).mappings().all()
    except Exception:
        db.session.rollback()
        return []


def _options_code_name(rows, code_key: str, name_key: str, value_key: str = "value"):
    options = []
    for row in rows:
        code = row.get(code_key)
        name = row.get(name_key)
        if code is None or name is None:
            continue
        label = f"{code} - {name}".strip()
        value = code if value_key == "value" else row.get(value_key)
        options.append({"value": str(value), "label": label})
    return options


def _options_abrev_nome(rows, abrev_key: str, name_key: str):
    options = []
    for row in rows:
        abrev = row.get(abrev_key)
        name = row.get(name_key)
        if abrev is None:
            continue
        label = f"{abrev} - {name}".strip() if name else str(abrev)
        options.append({"value": str(abrev), "label": label})
    return options


def _find_plan21_nger_row(
    exercicio: str,
    uo: str,
    programa: str,
    acao_paoe: str,
    responsavel_acao: str,
    produto_acao: str,
    chave_planejamento: str | None = None,
    regiao_produto: str | None = None,
    subacao_entrega: str | None = None,
    etapa: str | None = None,
    responsavel_etapa: str | None = None,
    regiao_etapa: str | None = None,
    natureza: str | None = None,
    descricao_item_despesa: str | None = None,
):
    query = Plan21Nger.query
    if exercicio:
        query = query.filter(Plan21Nger.exercicio == exercicio)
    if uo:
        query = query.filter(Plan21Nger.uo == uo)
    if programa:
        query = query.filter(Plan21Nger.programa == programa)
    if acao_paoe:
        query = query.filter(Plan21Nger.acao_paoe == acao_paoe)
    if responsavel_acao:
        query = query.filter(Plan21Nger.responsavel_acao == responsavel_acao)
    if produto_acao:
        query = query.filter(Plan21Nger.produto == produto_acao)
    if chave_planejamento:
        query = query.filter(Plan21Nger.chave_planejamento == chave_planejamento)
    if regiao_produto:
        query = query.filter(Plan21Nger.regiao_produto == regiao_produto)
    if subacao_entrega:
        query = query.filter(Plan21Nger.subacao_entrega == subacao_entrega)
    if etapa:
        query = query.filter(Plan21Nger.etapa == etapa)
    if responsavel_etapa:
        query = query.filter(Plan21Nger.responsavel_etapa == responsavel_etapa)
    if regiao_etapa:
        query = query.filter(Plan21Nger.regiao_etapa == regiao_etapa)
    if natureza:
        query = query.filter(Plan21Nger.natureza == natureza)
    if descricao_item_despesa:
        query = query.filter(Plan21Nger.descricao_item_despesa == descricao_item_despesa)
    rows = query.order_by(Plan21Nger.id.desc()).all()
    if not rows:
        return None, "Nenhum registro encontrado com os filtros informados."
    if len(rows) > 1:
        # Seleciona o registro mais recente quando houver ambiguidade.
        return rows[0], None
    return rows[0], None


def _normalize_iduso(value: str) -> str:
    if not value:
        return ""
    token = str(value).strip()
    digits = "".join(ch for ch in token if ch.isdigit())
    if digits:
        return str(int(digits))
    return _leading_token(token)


def _iduso_variants(value: str) -> list[str]:
    base = _normalize_iduso(value)
    if base == "":
        return []
    variants = {base}
    for width in (2, 3, 4):
        variants.add(base.zfill(width))
    return sorted(variants)


def _codigo_variants(value: str) -> list[str]:
    if not value:
        return []
    raw = str(value).strip()
    digits = "".join(ch for ch in raw if ch.isdigit())
    variants = [raw]
    if digits:
        norm = str(int(digits))
        variants.append(norm)
        for width in (2, 3):
            variants.append(norm.zfill(width))
    return list(dict.fromkeys([v for v in variants if v]))


def _normalize_chave(value: str) -> str:
    if not value:
        return ""
    value = str(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return "".join(ch for ch in value if ch.isalnum() or ch == "*").upper()


def _normalize_dotacao_key(value: str) -> str:
    if not value:
        return ""
    cleaned = str(value).strip()
    cleaned = re.sub(r"\s+", "", cleaned)
    cleaned = cleaned.rstrip("*")
    return cleaned.upper()


def _calc_ped_sum_for_dotacao(chave_dotacao: str) -> Decimal:
    key_norm = _normalize_dotacao_key(chave_dotacao)
    if not key_norm:
        return Decimal("0")
    rows = (
        PedRegistro.query.with_entities(PedRegistro.valor_ped, PedRegistro.chave)
        .filter(PedRegistro.ativo == True)  # noqa: E712
        .all()
    )
    total = Decimal("0")
    for row in rows:
        if _normalize_dotacao_key(row.chave) == key_norm:
            total += _dec_or_zero(row.valor_ped)
    return total


def _calc_emp_sum_for_dotacao(chave_dotacao: str) -> Decimal:
    key_norm = _normalize_dotacao_key(chave_dotacao)
    if not key_norm:
        return Decimal("0")
    rows = (
        EmpRegistro.query.with_entities(EmpRegistro.valor_emp_devolucao_gcv, EmpRegistro.chave)
        .filter(EmpRegistro.ativo == True)  # noqa: E712
        .all()
    )
    total = Decimal("0")
    for row in rows:
        if _normalize_dotacao_key(row.chave) == key_norm:
            total += _dec_or_zero(row.valor_emp_devolucao_gcv)
    return total


def _parse_decimal_value(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return _dec_or_zero(value)
    try:
        return _dec_or_zero(_parse_decimal(value))
    except Exception:
        return Decimal("0")


def _fetch_estorno_rows() -> list[tuple[str, Decimal]]:
    try:
        rows = db.session.execute(
            text("SELECT chave_dotacao, valor_a_ser_est FROM est_dotacao WHERE ativo = 1")
        ).fetchall()
        return [(r[0], _parse_decimal_value(r[1])) for r in rows]
    except Exception:
        try:
            rows = db.session.execute(
                text("SELECT chave_dotacao, valor_estorno FROM est_dotacao WHERE ativo = 1")
            ).fetchall()
            return [(r[0], _parse_decimal_value(r[1])) for r in rows]
        except Exception:
            try:
                rows = db.session.execute(
                    text("SELECT chave, valor_a_ser_est FROM est_dotacao WHERE ativo = 1")
                ).fetchall()
                return [(r[0], _parse_decimal_value(r[1])) for r in rows]
            except Exception:
                try:
                    rows = db.session.execute(
                        text("SELECT chave, valor_estorno FROM est_dotacao WHERE ativo = 1")
                    ).fetchall()
                    return [(r[0], _parse_decimal_value(r[1])) for r in rows]
                except Exception:
                    return []


def _fetch_estorno_rows_full() -> list[tuple[str, Decimal, str]]:
    try:
        rows = db.session.execute(
            text(
                "SELECT chave_dotacao, valor_a_ser_est, situacao FROM est_dotacao WHERE ativo = 1"
            )
        ).fetchall()
        return [(r[0], _parse_decimal_value(r[1]), r[2] if len(r) > 2 else "") for r in rows]
    except Exception:
        try:
            rows = db.session.execute(
                text("SELECT chave_dotacao, valor_estorno, situacao FROM est_dotacao WHERE ativo = 1")
            ).fetchall()
            return [(r[0], _parse_decimal_value(r[1]), r[2] if len(r) > 2 else "") for r in rows]
        except Exception:
            try:
                rows = db.session.execute(
                    text("SELECT chave, valor_a_ser_est, situacao FROM est_dotacao WHERE ativo = 1")
                ).fetchall()
                return [(r[0], _parse_decimal_value(r[1]), r[2] if len(r) > 2 else "") for r in rows]
            except Exception:
                try:
                    rows = db.session.execute(
                        text("SELECT chave, valor_estorno, situacao FROM est_dotacao WHERE ativo = 1")
                    ).fetchall()
                    return [(r[0], _parse_decimal_value(r[1]), r[2] if len(r) > 2 else "") for r in rows]
                except Exception:
                    return []


def _calc_estorno_sum_for_dotacao(chave_dotacao: str) -> Decimal:
    key_norm = _normalize_dotacao_key(chave_dotacao)
    if not key_norm:
        return Decimal("0")
    total = Decimal("0")
    for chave, valor in _fetch_estorno_rows():
        if _normalize_dotacao_key(chave) == key_norm:
            total += _dec_or_zero(valor)
    return total


def _build_estorno_maps() -> tuple[dict[str, Decimal], dict[str, str]]:
    est_map: dict[str, Decimal] = {}
    situacao_map: dict[str, str] = {}
    for chave, valor, situacao in _fetch_estorno_rows_full():
        key_norm = _normalize_dotacao_key(chave)
        if not key_norm:
            continue
        est_map[key_norm] = _dec_or_zero(est_map.get(key_norm)) + _dec_or_zero(valor)
        if situacao:
            situacao_map[key_norm] = str(situacao).strip()
    return est_map, situacao_map


def _build_estorno_map() -> dict[str, Decimal]:
    est_map, _ = _build_estorno_maps()
    return est_map


def _calc_ped_sum_for_dotacao_keys(keys: set[str]) -> tuple[Decimal, int]:
    if not keys:
        return Decimal("0"), 0
    rows = (
        PedRegistro.query.with_entities(PedRegistro.valor_ped, PedRegistro.chave)
        .filter(PedRegistro.ativo == True)  # noqa: E712
        .all()
    )
    total = Decimal("0")
    count = 0
    for row in rows:
        if _normalize_dotacao_key(row.chave) in keys:
            total += _dec_or_zero(row.valor_ped)
            count += 1
    return total, count


def _collect_ped_rows_for_dotacao_keys(keys: set[str]) -> dict[int, Decimal]:
    if not keys:
        return {}
    rows = (
        PedRegistro.query.with_entities(PedRegistro.id, PedRegistro.valor_ped, PedRegistro.chave)
        .filter(PedRegistro.ativo == True)  # noqa: E712
        .all()
    )
    matched: dict[int, Decimal] = {}
    for row in rows:
        row_id = _row_value(row, "id", 0)
        row_chave = _row_value(row, "chave", 2)
        row_valor = _row_value(row, "valor_ped", 1)
        if row_id is None:
            continue
        if _normalize_dotacao_key(row_chave) in keys:
            matched[int(row_id)] = _dec_or_zero(row_valor)
    return matched


def _calc_emp_sum_for_dotacao_keys(keys: set[str]) -> tuple[Decimal, int]:
    if not keys:
        return Decimal("0"), 0
    rows = (
        EmpRegistro.query.with_entities(
            EmpRegistro.numero_emp, EmpRegistro.valor_emp_devolucao_gcv, EmpRegistro.chave
        )
        .filter(EmpRegistro.ativo == True)  # noqa: E712
        .all()
    )
    total = Decimal("0")
    emp_nums = []
    for row in rows:
        if _normalize_dotacao_key(row.chave) in keys:
            total += _dec_or_zero(row.valor_emp_devolucao_gcv)
            if row.numero_emp:
                emp_nums.append(row.numero_emp)
    emp_nums = list(dict.fromkeys(emp_nums))
    return total, len(emp_nums)


def _collect_emp_rows_for_dotacao_keys(keys: set[str]) -> dict[int, tuple[Decimal, str]]:
    if not keys:
        return {}
    rows = (
        EmpRegistro.query.with_entities(
            EmpRegistro.id,
            EmpRegistro.numero_emp,
            EmpRegistro.valor_emp_devolucao_gcv,
            EmpRegistro.chave,
        )
        .filter(EmpRegistro.ativo == True)  # noqa: E712
        .all()
    )
    matched: dict[int, tuple[Decimal, str]] = {}
    for row in rows:
        row_id = _row_value(row, "id", 0)
        row_num = _row_value(row, "numero_emp", 1)
        row_valor = _row_value(row, "valor_emp_devolucao_gcv", 2)
        row_chave = _row_value(row, "chave", 3)
        if row_id is None:
            continue
        if _normalize_dotacao_key(row_chave) in keys:
            matched[int(row_id)] = (_dec_or_zero(row_valor), row_num or "")
    return matched


@home_bp.route("/api/dotacao/options", methods=["GET"])
@login_required
@require_feature("cadastrar/dotacao")
def api_dotacao_options():
    current_year = str(_now_local().year)
    fields = {
        "exercicio": Plan21Nger.exercicio,
        "chave_planejamento": Plan21Nger.chave_planejamento,
        "uo": Plan21Nger.uo,
        "programa": Plan21Nger.programa,
        "acao_paoe": Plan21Nger.acao_paoe,
        "produto": Plan21Nger.produto,
        "ug": Plan21Nger.ug,
        "regiao": Plan21Nger.regiao_etapa,
        "subacao_entrega": Plan21Nger.subacao_entrega,
        "etapa": Plan21Nger.etapa,
        "natureza_despesa": Plan21Nger.natureza,
        "elemento": Plan21Nger.elemento,
        "subelemento": Plan21Nger.subelemento,
        "fonte": Plan21Nger.fonte,
        "iduso": Plan21Nger.idu,
    }
    selected = {}
    for key in fields:
        val = (request.args.get(key) or "").strip()
        if val:
            selected[key] = val
    if "exercicio" not in selected:
        selected["exercicio"] = current_year

    options = {}
    for key, col in fields.items():
        query = db.session.query(col).distinct().filter(Plan21Nger.ativo == True)  # noqa: E712
        for s_key, s_val in selected.items():
            if s_key == key:
                continue
            if s_key == "natureza_despesa":
                query = query.filter(fields[s_key].like(f"{s_val}%"))
            else:
                query = query.filter(fields[s_key] == s_val)
        rows = query.all()
        values = []
        for (val,) in rows:
            if val is None:
                continue
            s = str(val).strip()
            if s == "":
                continue
            if key == "natureza_despesa":
                s = _natureza_prefix(s)
            values.append(s)
        if key == "exercicio":
            options[key] = [current_year]
        else:
            options[key] = sorted(set(values), key=lambda v: v.lower())

    perfis_raw = (
        Perfil.query.filter(Perfil.ativo == True)  # noqa: E712
        .order_by(Perfil.nome)
        .all()
    )
    adj_options = []
    for p in perfis_raw:
        nome = (p.nome or "").strip()
        if not nome:
            continue
        if nome.lower() in {"admin", "consultor"}:
            continue
        adj_options.append({"id": p.id, "label": nome})
    perfis = [o["label"] for o in adj_options]
    return jsonify({"options": options, "adj": adj_options, "perfis": perfis})


@home_bp.route("/api/meta-fisica/options", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/meta_fisica")
def api_meta_fisica_options():
    def _arg_as_bool(name: str, default: bool = False) -> bool:
        raw = (request.args.get(name) or "").strip().lower()
        if not raw:
            return default
        return raw in {"1", "true", "t", "yes", "y", "sim", "s"}

    current_year = str(_now_local().year)
    include_rows = _arg_as_bool("include_rows", default=True)
    include_history = _arg_as_bool("include_history", default=True)
    include_option_rows = _arg_as_bool("include_option_rows", default=False)
    field_cols = {
        "exercicio": "exercicio",
        "unidade_orcamentaria": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "adj_solicitante": "adj",
        "produto_acao": "produto_acao",
        "unid_medida_produto": "unid_medida_produto",
    }

    selected = {}
    for key in field_cols:
        val = (request.args.get(key) or "").strip()
        if val:
            selected[key] = val
    selected["exercicio"] = current_year
    ignore_meta_id_for_cache = 0
    try:
        ignore_meta_id_for_cache = int((request.args.get("meta_id") or "").strip() or 0)
    except Exception:
        ignore_meta_id_for_cache = 0

    cache_key = (
        "meta_fisica_options_v3",
        include_rows,
        include_history,
        include_option_rows,
        ignore_meta_id_for_cache,
        tuple(sorted((k, str(v)) for k, v in selected.items())),
    )
    cached_payload = _cache_get_meta_fisica_options(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    selected_for_cascade = {k: v for k, v in selected.items() if k != "adj_solicitante"}

    # Opções em uma única consulta (evita N queries por troca de filtro).
    options = {k: [] for k in field_cols.keys()}
    options["exercicio"] = [current_year]
    where = ["ativo = 1", "exercicio = :exercicio"]
    params = {"exercicio": current_year}
    for s_key, s_val in selected_for_cascade.items():
        if s_key == "exercicio":
            continue
        if not s_val:
            continue
        where.append(f"{field_cols[s_key]} = :{s_key}")
        params[s_key] = s_val
    options_sql = f"""
        SELECT DISTINCT
            exercicio,
            unidade_orcamentaria,
            programa,
            acao_paoe,
            produto_acao,
            unid_medida_produto
        FROM plan21_nger
        WHERE {' AND '.join(where)}
    """
    options_rows = _safe_query_mappings(options_sql, params)
    for row in options_rows:
        for key in (
            "unidade_orcamentaria",
            "programa",
            "acao_paoe",
            "produto_acao",
            "unid_medida_produto",
        ):
            val = row.get(field_cols[key])
            if val is None:
                continue
            txt = str(val).strip()
            if txt:
                options[key].append(txt)
    for key in (
        "unidade_orcamentaria",
        "programa",
        "acao_paoe",
        "produto_acao",
        "unid_medida_produto",
    ):
        options[key] = sorted(set(options[key]), key=lambda x: x.lower())

    # adj_solicitante permanece catálogo separado da tabela adj.
    adj_rows = _safe_query_mappings(
        """
        SELECT abreviacao AS value
        FROM adj
        WHERE ativo = 1
        ORDER BY abreviacao
        """
    )
    adj_vals = []
    for row in adj_rows:
        raw_val = row.get("value")
        if raw_val is None:
            continue
        txt = str(raw_val).strip()
        if txt:
            adj_vals.append(txt)
    options["adj_solicitante"] = sorted(set(adj_vals), key=lambda x: x.lower())

    option_rows = []
    if include_option_rows:
        option_rows_sql = """
            SELECT DISTINCT
                unidade_orcamentaria,
                programa,
                acao_paoe,
                produto_acao,
                unid_medida_produto
            FROM plan21_nger
            WHERE ativo = 1
              AND exercicio = :exercicio
        """
        option_rows_raw = _safe_query_mappings(option_rows_sql, {"exercicio": current_year})
        for row in option_rows_raw:
            option_rows.append(
                {
                    "unidade_orcamentaria": str(row.get("unidade_orcamentaria") or "").strip(),
                    "programa": str(row.get("programa") or "").strip(),
                    "acao_paoe": str(row.get("acao_paoe") or "").strip(),
                    "produto_acao": str(row.get("produto_acao") or "").strip(),
                    "unid_medida_produto": str(row.get("unid_medida_produto") or "").strip(),
                }
            )

    required_for_rows = (
        "unidade_orcamentaria",
        "programa",
        "acao_paoe",
        "produto_acao",
        "unid_medida_produto",
    )
    has_all_required = all((selected.get(k) or "").strip() for k in required_for_rows)
    pending_meta = None
    if include_rows and include_history and has_all_required:
        ignore_meta_id = ignore_meta_id_for_cache
        pending_query = (
            AlterarMeta.query.filter(AlterarMeta.ativo == True)  # noqa: E712
            .filter(func.lower(func.coalesce(AlterarMeta.status_aprovacao, "")) == "aguardando")
            .filter(AlterarMeta.exercicio == int(current_year))
            .filter(AlterarMeta.unidade_orcamentaria == (selected.get("unidade_orcamentaria") or "").strip())
            .filter(AlterarMeta.programa == (selected.get("programa") or "").strip())
            .filter(AlterarMeta.acao_paoe == (selected.get("acao_paoe") or "").strip())
            .filter(AlterarMeta.produto_acao == (selected.get("produto_acao") or "").strip())
            .filter(AlterarMeta.unid_medida_produto == (selected.get("unid_medida_produto") or "").strip())
        )
        if ignore_meta_id > 0:
            pending_query = pending_query.filter(AlterarMeta.id != ignore_meta_id)
        pending_registro = pending_query.order_by(AlterarMeta.id.desc()).first()
        if pending_registro:
            adj_token = _normalize_controle_token(getattr(pending_registro, "responsavel_acao", "") or "") or "SEMADJ"
            pending_exercicio = str(getattr(pending_registro, "exercicio", "") or current_year)
            controle = f"META.{pending_exercicio}.{adj_token}.{pending_registro.id}"
            pending_meta = {
                "id": pending_registro.id,
                "controle": controle,
            }
    plan_rows = []
    if include_rows and has_all_required and not pending_meta:
        where_rows = ["ativo = 1", "exercicio = :exercicio"]
        params_rows = {"exercicio": current_year}
        for key in required_for_rows:
            where_rows.append(f"{field_cols[key]} = :{key}")
            params_rows[key] = (selected.get(key) or "").strip()
        rows_sql = f"""
            SELECT id, regiao_produto, meta_produto, meta_credito, meta_anulada, meta_atual
            FROM plan21_nger
            WHERE {' AND '.join(where_rows)}
            ORDER BY id
        """
        plan_rows = _safe_query_mappings(rows_sql, params_rows)
    grouped = {}
    for row in plan_rows:
        regiao = str(row.get("regiao_produto") or "").strip()
        if not regiao:
            continue
        meta_val = _parse_decimal(row.get("meta_produto"))
        if regiao not in grouped:
            grouped[regiao] = {
                "regiao_produto": regiao,
                "meta_produto": "",
                "meta_produto_max": None,
                "meta_credito": "",
                "meta_anulada": "",
                "meta_atual": "",
                "latest_id": 0,
                "plan21_ids": [],
            }
        # Para a mesma região, mantém o maior valor de meta_produto.
        if meta_val is not None:
            current_max = grouped[regiao].get("meta_produto_max")
            if current_max is None or meta_val > current_max:
                grouped[regiao]["meta_produto_max"] = meta_val
                grouped[regiao]["meta_produto"] = str(meta_val)
        row_id = row.get("id")
        try:
            row_id_int = int(row_id)
        except Exception:
            row_id_int = 0
        if row_id_int > 0 and row_id_int >= int(grouped[regiao].get("latest_id") or 0):
            credito_val = _parse_decimal(row.get("meta_credito"))
            anulada_val = _parse_decimal(row.get("meta_anulada"))
            atual_val = _parse_decimal(row.get("meta_atual"))
            grouped[regiao]["meta_credito"] = str(credito_val) if credito_val is not None else ""
            grouped[regiao]["meta_anulada"] = str(anulada_val) if anulada_val is not None else ""
            grouped[regiao]["meta_atual"] = str(atual_val) if atual_val is not None else ""
            grouped[regiao]["latest_id"] = row_id_int
        if row_id_int > 0 and row_id_int not in grouped[regiao]["plan21_ids"]:
            grouped[regiao]["plan21_ids"].append(row_id_int)

    historico_map = {}
    movimento_map = {}
    if include_rows and include_history and has_all_required:
        hist_where = [
            "am.ativo = 1",
            "LOWER(COALESCE(am.status_aprovacao, '')) = 'aprovado'",
            "am.exercicio = :exercicio",
            "am.unidade_orcamentaria = :unidade_orcamentaria",
            "am.programa = :programa",
            "am.acao_paoe = :acao_paoe",
            "am.produto_acao = :produto_acao",
            "am.unid_medida_produto = :unid_medida_produto",
        ]
        hist_params = {
            "exercicio": current_year,
            "unidade_orcamentaria": (selected.get("unidade_orcamentaria") or "").strip(),
            "programa": (selected.get("programa") or "").strip(),
            "acao_paoe": (selected.get("acao_paoe") or "").strip(),
            "produto_acao": (selected.get("produto_acao") or "").strip(),
            "unid_medida_produto": (selected.get("unid_medida_produto") or "").strip(),
        }
        hist_sql = f"""
            SELECT am.id, am.regiao_produto
            FROM alterar_meta am
            WHERE {' AND '.join(hist_where)}
            ORDER BY am.id ASC
        """
        hist_meta_rows = _safe_query_mappings(hist_sql, hist_params)

        # Regra temporal para novo cadastro: todo movimento aprovado anterior e
        # ainda não contabilizado em uma linha anterior entra como histórico.
        approved_ids = []
        for meta_row in hist_meta_rows:
            try:
                mid = int(meta_row.get("id") or 0)
            except Exception:
                mid = 0
            if mid > 0 and mid not in approved_ids:
                approved_ids.append(mid)
        if approved_ids:
            approved_ids_csv = ",".join(str(int(x)) for x in approved_ids)
            temporal_items_rows = _safe_query_mappings(
                """
                SELECT alterar_meta_id, regiao_codigo, tipo, valor, ordem, id
                FROM alterar_meta_item
                WHERE ativo = 1
                  AND alterar_meta_id IN (""" + approved_ids_csv + """)
                ORDER BY alterar_meta_id ASC, ordem ASC, id ASC
                """,
                {},
            )
            items_by_meta_region = {}
            for it in temporal_items_rows:
                try:
                    meta_id = int(it.get("alterar_meta_id") or 0)
                except Exception:
                    meta_id = 0
                reg = str(it.get("regiao_codigo") or "").strip()
                reg_key = _normalize_codigo_num(reg) or reg.lower()
                if meta_id <= 0 or not reg_key:
                    continue
                items_by_meta_region.setdefault(meta_id, {}).setdefault(reg_key, []).append(it)

            def _same_pair_for_options(left: tuple[str, str], right: tuple[str, str]) -> bool:
                for idx in (0, 1):
                    l_val = _parse_decimal(left[idx])
                    r_val = _parse_decimal(right[idx])
                    l_num = l_val if l_val is not None else Decimal("0")
                    r_num = r_val if r_val is not None else Decimal("0")
                    if abs(l_num - r_num) >= Decimal("0.000001"):
                        return False
                return True

            def _strip_prefix_for_options(
                hist_pairs: list[tuple[str, str]],
                pairs: list[tuple[str, str]],
            ) -> list[tuple[str, str]]:
                if not hist_pairs or not pairs or len(pairs) < len(hist_pairs):
                    return pairs
                for idx, hist_pair in enumerate(hist_pairs):
                    if not _same_pair_for_options(hist_pair, pairs[idx]):
                        return pairs
                return pairs[len(hist_pairs):]

            temporal_history_by_region = {}
            for mid in approved_ids:
                by_region = items_by_meta_region.get(int(mid), {})
                for reg_key, reg_rows in by_region.items():
                    cred_list, anu_list = _build_aligned_lists_from_items(reg_rows)
                    size = max(len(cred_list), len(anu_list))
                    if size <= 0:
                        continue
                    pairs = []
                    for i in range(size):
                        c = str(cred_list[i] if i < len(cred_list) else "" or "").strip()
                        a = str(anu_list[i] if i < len(anu_list) else "" or "").strip()
                        if c or a:
                            pairs.append((c, a))
                    if not pairs:
                        continue
                    current_hist = list(temporal_history_by_region.get(reg_key) or [])
                    hist_delta = _strip_prefix_for_options(current_hist, pairs)
                    if hist_delta:
                        temporal_history_by_region.setdefault(reg_key, []).extend(hist_delta)

            historico_map = {}
            movimento_map = {}
            for reg_key, pairs in temporal_history_by_region.items():
                hist_credito_items = [p[0] for p in pairs]
                hist_anulada_items = [p[1] for p in pairs]
                historico_map[reg_key] = {
                    "meta_credito_historico_items": hist_credito_items,
                    "meta_anulada_historico_items": hist_anulada_items,
                }
                movimento_map[reg_key] = {
                    "meta_credito_mov_items": [],
                    "meta_anulada_mov_items": [],
                }

    def _regiao_sort_key(val: str):
        txt = str(val or "").strip()
        num_txt = _normalize_codigo_num(txt)
        if num_txt:
            try:
                return (0, int(num_txt), txt.lower())
            except Exception:
                pass
        return (1, txt.lower(), txt.lower())

    rows_out = []
    if include_rows:
        for regiao in sorted(grouped.keys(), key=_regiao_sort_key):
            item = grouped[regiao]
            ids = item["plan21_ids"]
            regiao_key = _normalize_codigo_num(regiao) or regiao.lower()
            hist_item = historico_map.get(
                regiao_key,
                {
                    "meta_credito_historico_items": [],
                    "meta_anulada_historico_items": [],
                },
            )
            mov_item = movimento_map.get(
                regiao_key,
                {
                    "meta_credito_mov_items": [],
                    "meta_anulada_mov_items": [],
                },
            )
            hist_credito_items = hist_item.get("meta_credito_historico_items") or []
            hist_anulada_items = hist_item.get("meta_anulada_historico_items") or []
            mov_credito_items = mov_item.get("meta_credito_mov_items") or []
            mov_anulada_items = mov_item.get("meta_anulada_mov_items") or []
            meta_credito_historico = Decimal("0")
            for v in hist_credito_items:
                parsed = _parse_decimal(v)
                if parsed is not None:
                    meta_credito_historico += parsed
            meta_anulada_historico = Decimal("0")
            for v in hist_anulada_items:
                parsed = _parse_decimal(v)
                if parsed is not None:
                    meta_anulada_historico += parsed
            rows_out.append(
                {
                    "regiao_produto": item["regiao_produto"],
                    "meta_produto": item["meta_produto"],
                    "meta_credito": item["meta_credito"],
                    "meta_anulada": item["meta_anulada"],
                    "meta_atual": item["meta_atual"],
                    "meta_credito_historico": str(meta_credito_historico),
                    "meta_anulada_historico": str(meta_anulada_historico),
                    "meta_credito_historico_items": hist_credito_items,
                    "meta_anulada_historico_items": hist_anulada_items,
                    "meta_credito_mov_items": mov_credito_items,
                    "meta_anulada_mov_items": mov_anulada_items,
                    "plan21_nger_id": ids[0] if ids else None,
                    "plan21_ids": ids,
                }
            )

    regioes_catalog_rows = _safe_query_mappings(
        """
        SELECT codigo, nome
        FROM regiao
        WHERE ativo = 1
        ORDER BY CAST(codigo AS UNSIGNED), codigo
        """
    )
    regioes_catalog = []
    for row in regioes_catalog_rows:
        codigo_raw = row.get("codigo")
        nome_raw = row.get("nome")
        if codigo_raw is None:
            continue
        codigo = str(codigo_raw).strip()
        if not codigo:
            continue
        nome = str(nome_raw or "").strip()
        label = f"{codigo} - {nome}" if nome else codigo
        regioes_catalog.append(
            {
                "codigo": codigo,
                "nome": nome,
                "label": label,
            }
        )

    payload = {
        "current_year": current_year,
        "options": options,
        "rows": rows_out,
        "regioes_catalog": regioes_catalog,
        "option_rows": option_rows,
        "pending_meta": pending_meta,
    }
    _cache_set_meta_fisica_options(cache_key, payload)
    return jsonify(payload)


@home_bp.route("/api/meta-fisica", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/meta_fisica")
def api_meta_fisica_create():
    data = request.get_json() or {}
    current_year = str(_now_local().year)
    meta_id_raw = str(data.get("meta_id") or "").strip()

    exercicio_raw = str(data.get("exercicio") or "").strip()
    unidade_orcamentaria = str(data.get("unidade_orcamentaria") or "").strip()
    programa = str(data.get("programa") or "").strip()
    acao_paoe = str(data.get("acao_paoe") or "").strip()
    adj_solicitante = str(data.get("adj_solicitante") or "").strip()
    produto_acao = str(data.get("produto_acao") or "").strip()
    unid_medida_produto = str(data.get("unid_medida_produto") or "").strip()
    justificativa = re.sub(
        r"^META\.[^*\r\n]+\*\s*",
        "",
        str(data.get("justificativa") or "").strip(),
        flags=re.IGNORECASE,
    ).strip()
    rows = data.get("rows") or []

    registro_existente = None
    criador_registro = None
    if meta_id_raw:
        try:
            meta_id = int(meta_id_raw)
        except Exception:
            return jsonify({"error": "ID de registro inválido para edição."}), 400
        registro_existente = (
            AlterarMeta.query.filter(AlterarMeta.id == meta_id)
            .filter(AlterarMeta.ativo == True)  # noqa: E712
            .first()
        )
        if not registro_existente:
            return jsonify({"error": "Registro de meta não encontrado para edição."}), 404
        criador_registro = None
        if getattr(registro_existente, "usuario_id", None):
            criador_registro = Usuario.query.filter_by(id=registro_existente.usuario_id).first()
        exercicio_ctrl = str(getattr(registro_existente, "exercicio", "") or current_year)
        adj_nome_ctrl = (getattr(registro_existente, "responsavel_acao", "") or "").strip()
        if not adj_nome_ctrl and criador_registro:
            perfil_id_ctrl = getattr(criador_registro, "perfil_id", None)
            if perfil_id_ctrl:
                perfil_row_ctrl = Perfil.query.filter_by(id=perfil_id_ctrl).first()
                adj_nome_ctrl = (getattr(perfil_row_ctrl, "nome", "") or "").strip()
        controle_meta = f"META.{exercicio_ctrl}.{_normalize_controle_token(adj_nome_ctrl) or 'SEMADJ'}.{registro_existente.id}"
        status_atual = str(getattr(registro_existente, "status_aprovacao", "") or "").strip().lower()
        if status_atual != "aguardando":
            return jsonify({"error": f"Somente registros com status Aguardando podem ser editados ({controle_meta})."}), 400
        perfil_criador = getattr(criador_registro, "perfil_id", None) if criador_registro else None
        if not _current_user_matches_perfil(perfil_criador):
            return jsonify({"error": f"Usuário sem permissão para editar registro {controle_meta}."}), 403

    if exercicio_raw != current_year:
        return jsonify({"error": "Exercicio invalido para o ano corrente."}), 400

    required = {
        "unidade_orcamentaria": unidade_orcamentaria,
        "programa": programa,
        "acao_paoe": acao_paoe,
        "adj_solicitante": adj_solicitante,
        "produto_acao": produto_acao,
        "unid_medida_produto": unid_medida_produto,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({"error": f"Campos obrigatorios ausentes: {', '.join(missing)}."}), 400
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "Adicione ao menos uma linha de meta fisica."}), 400

    seen_regioes = set()
    parsed_rows = []

    unidade_tipo = (unid_medida_produto or "").strip().lower()

    def _parse_item_list(raw_items):
        if not isinstance(raw_items, list):
            return []
        out = []
        for raw in raw_items:
            txt = str(raw or "").strip()
            if txt:
                out.append(txt)
        return out
    def _parse_item_list_aligned(raw_items):
        if not isinstance(raw_items, list):
            return []
        out = []
        for raw in raw_items:
            out.append(str(raw or "").strip())
        return out

    def _validate_item_value(val: Decimal, regiao_produto: str, field_name: str):
        if val == Decimal("0"):
            return None
        if unidade_tipo == "percentual" and val > Decimal("0") and val < Decimal("0.1"):
            return jsonify({"error": f"{field_name} minimo e 0,1 para a regiao {regiao_produto}."}), 400
        if unidade_tipo == "unidade" and val > Decimal("0") and val != val.to_integral_value():
            return jsonify({"error": f"{field_name} deve ser inteiro para a regiao {regiao_produto}."}), 400
        return None

    for row in rows:
        regiao_produto = str(row.get("regiao_produto") or "").strip()
        if not regiao_produto:
            return jsonify({"error": "Regiao PTA/LOA e obrigatoria em todas as linhas."}), 400
        regiao_key = _normalize_codigo_num(regiao_produto) or regiao_produto.lower()
        if regiao_key in seen_regioes:
            return jsonify({"error": f"Regiao repetida na tabela: {regiao_produto}."}), 400
        seen_regioes.add(regiao_key)

        is_novo = bool(row.get("is_novo"))
        meta_produto = _parse_decimal(row.get("meta_produto"))
        credito_items_raw = _parse_item_list(row.get("meta_credito_items"))
        anulada_items_raw = _parse_item_list(row.get("meta_anulada_items"))
        credito_mov_items_aligned_raw = _parse_item_list_aligned(row.get("meta_credito_mov_items"))
        anulada_mov_items_aligned_raw = _parse_item_list_aligned(row.get("meta_anulada_mov_items"))
        credito_hist_items_aligned_raw = _parse_item_list_aligned(row.get("meta_credito_historico_items"))
        anulada_hist_items_aligned_raw = _parse_item_list_aligned(row.get("meta_anulada_historico_items"))
        meta_credito_historico = _parse_decimal(row.get("meta_credito_historico"))
        meta_anulada_historico = _parse_decimal(row.get("meta_anulada_historico"))
        has_historico_movimento = bool(row.get("has_historico_movimento"))
        if not credito_items_raw:
            fallback_credito = str(row.get("meta_credito") or "").strip()
            if fallback_credito:
                credito_items_raw = [fallback_credito]
        if not anulada_items_raw:
            fallback_anulada = str(row.get("meta_anulada") or "").strip()
            if fallback_anulada:
                anulada_items_raw = [fallback_anulada]

        credito_items = []
        meta_credito = Decimal("0")
        for idx_item, item_raw in enumerate(credito_items_raw, start=1):
            item_val = _parse_decimal(item_raw)
            if item_val is None:
                return jsonify({"error": f"Acrescimo invalido para a regiao {regiao_produto}."}), 400
            validate_err = _validate_item_value(item_val, regiao_produto, "Acrescimo")
            if validate_err:
                return validate_err
            meta_credito += item_val
            if item_val > Decimal("0"):
                credito_items.append({"ordem": idx_item, "valor": item_val})

        anulada_items = []
        meta_anulada = Decimal("0")
        for idx_item, item_raw in enumerate(anulada_items_raw, start=1):
            item_val = _parse_decimal(item_raw)
            if item_val is None:
                return jsonify({"error": f"Reducao invalida para a regiao {regiao_produto}."}), 400
            validate_err = _validate_item_value(item_val, regiao_produto, "Reducao")
            if validate_err:
                return validate_err
            meta_anulada += item_val
            if item_val > Decimal("0"):
                anulada_items.append({"ordem": idx_item, "valor": item_val})

        has_credito_mov_key = isinstance(row, dict) and ("meta_credito_mov_items" in row)
        has_anulada_mov_key = isinstance(row, dict) and ("meta_anulada_mov_items" in row)
        credito_mov_source = credito_mov_items_aligned_raw if has_credito_mov_key else credito_items_raw
        anulada_mov_source = anulada_mov_items_aligned_raw if has_anulada_mov_key else anulada_items_raw

        credito_hist_items = []
        for item_raw in credito_hist_items_aligned_raw:
            item_val = _parse_decimal(item_raw)
            if item_val is not None and item_val > Decimal("0"):
                credito_hist_items.append(item_val)
        anulada_hist_items = []
        for item_raw in anulada_hist_items_aligned_raw:
            item_val = _parse_decimal(item_raw)
            if item_val is not None and item_val > Decimal("0"):
                anulada_hist_items.append(item_val)

        credito_mov_items = []
        for item_raw in credito_mov_source:
            if str(item_raw or "").strip() == "":
                continue
            item_val = _parse_decimal(item_raw)
            if item_val is None:
                return jsonify({"error": f"Acrescimo invalido para a regiao {regiao_produto}."}), 400
            validate_err = _validate_item_value(item_val, regiao_produto, "Acrescimo")
            if validate_err:
                return validate_err
            if item_val > Decimal("0"):
                credito_mov_items.append({"ordem": len(credito_mov_items) + 1, "valor": item_val})

        anulada_mov_items = []
        for item_raw in anulada_mov_source:
            if str(item_raw or "").strip() == "":
                continue
            item_val = _parse_decimal(item_raw)
            if item_val is None:
                return jsonify({"error": f"Reducao invalida para a regiao {regiao_produto}."}), 400
            validate_err = _validate_item_value(item_val, regiao_produto, "Reducao")
            if validate_err:
                return validate_err
            if item_val > Decimal("0"):
                anulada_mov_items.append({"ordem": len(anulada_mov_items) + 1, "valor": item_val})

        if not is_novo and meta_produto is None:
            return jsonify({"error": f"Meta PTA/LOA invalida para a regiao {regiao_produto}."}), 400

        meta_base = meta_produto if meta_produto is not None else Decimal("0")
        meta_atual = meta_base + meta_credito - meta_anulada

        parsed_rows.append(
            {
                "regiao_produto": regiao_produto,
                "meta_produto": None if is_novo else meta_produto,
                "meta_credito": meta_credito,
                "meta_anulada": meta_anulada,
                "meta_atual": meta_atual,
                "is_novo": is_novo,
                "credito_items": credito_items,
                "anulada_items": anulada_items,
                "credito_mov_items": credito_mov_items,
                "anulada_mov_items": anulada_mov_items,
                "meta_credito_mov_items": [str(i.get("valor")) for i in credito_mov_items],
                "meta_anulada_mov_items": [str(i.get("valor")) for i in anulada_mov_items],
                "meta_credito_mov_items_aligned": credito_mov_items_aligned_raw,
                "meta_anulada_mov_items_aligned": anulada_mov_items_aligned_raw,
                "meta_credito_historico_items_aligned": credito_hist_items_aligned_raw,
                "meta_anulada_historico_items_aligned": anulada_hist_items_aligned_raw,
                "meta_credito_historico_items": [str(v) for v in credito_hist_items],
                "meta_anulada_historico_items": [str(v) for v in anulada_hist_items],
                "meta_credito_historico": meta_credito_historico,
                "meta_anulada_historico": meta_anulada_historico,
                "has_historico_movimento": has_historico_movimento,
                "plan21_ids_raw": row.get("plan21_ids") if isinstance(row.get("plan21_ids"), list) else [],
            }
        )

    # IDs do plan21_nger para os filtros do formulário (ordem crescente)
    filters_where_base = [
        "ativo = 1",
        "CAST(exercicio AS CHAR) = :exercicio",
        "unidade_orcamentaria = :uo",
        "programa = :programa",
        "acao_paoe = :acao_paoe",
        "adj = :adj_solicitante",
        "produto_acao = :produto_acao",
        "unid_medida_produto = :unid_medida_produto",
    ]
    filters_params_base = {
        "exercicio": exercicio_raw,
        "uo": unidade_orcamentaria,
        "programa": programa,
        "acao_paoe": acao_paoe,
        "adj_solicitante": adj_solicitante,
        "produto_acao": produto_acao,
        "unid_medida_produto": unid_medida_produto,
    }
    all_plan_ids_set = set()
    for row in parsed_rows:
        regiao = str(row.get("regiao_produto") or "").strip()
        row_plan_ids = []
        for val in row.get("plan21_ids_raw", []) or []:
            try:
                rid = int(val)
            except Exception:
                rid = 0
            if rid > 0 and rid not in row_plan_ids:
                row_plan_ids.append(rid)
        if not row_plan_ids and regiao:
            where = list(filters_where_base) + ["regiao_produto = :regiao_produto"]
            params = dict(filters_params_base)
            params["regiao_produto"] = regiao
            ids_sql = f"SELECT id FROM plan21_nger WHERE {' AND '.join(where)} ORDER BY id ASC"
            id_rows = _safe_query_mappings(ids_sql, params)
            for r in id_rows:
                try:
                    rid = int(r.get("id"))
                except Exception:
                    rid = 0
                if rid > 0 and rid not in row_plan_ids:
                    row_plan_ids.append(rid)
        row["plan21_ids"] = row_plan_ids
        for rid in row_plan_ids:
            all_plan_ids_set.add(rid)
    all_plan_ids = sorted(all_plan_ids_set)
    plan21_nger_id_latest = all_plan_ids[-1] if all_plan_ids else None

    # Salvar como registro único com dados da tabela em JSON (fácil reconstrução)
    meta_produto_total = Decimal("0")
    meta_credito_total = Decimal("0")
    meta_anulada_total = Decimal("0")
    meta_atual_total = Decimal("0")
    rows_serialized = []
    for row in parsed_rows:
        mp = row["meta_produto"] if row["meta_produto"] is not None else Decimal("0")
        mc = row["meta_credito"] if row["meta_credito"] is not None else Decimal("0")
        ma = row["meta_anulada"] if row["meta_anulada"] is not None else Decimal("0")
        mt = row["meta_atual"] if row["meta_atual"] is not None else Decimal("0")
        meta_produto_total += mp
        meta_credito_total += mc
        meta_anulada_total += ma
        meta_atual_total += mt
        rows_serialized.append(
            {
                "regiao_produto": row["regiao_produto"],
                "meta_produto": str(row["meta_produto"]) if row["meta_produto"] is not None else "",
                "meta_credito": str(row["meta_credito"]) if row["meta_credito"] is not None else "",
                "meta_anulada": str(row["meta_anulada"]) if row["meta_anulada"] is not None else "",
                "meta_atual": str(row["meta_atual"]) if row["meta_atual"] is not None else "",
                "is_novo": bool(row.get("is_novo")),
                "meta_credito_items": [str(i.get("valor")) for i in row.get("credito_items", [])],
                "meta_anulada_items": [str(i.get("valor")) for i in row.get("anulada_items", [])],
                "meta_credito_mov_items": [str(v) for v in row.get("meta_credito_mov_items_aligned", row.get("meta_credito_mov_items", []))],
                "meta_anulada_mov_items": [str(v) for v in row.get("meta_anulada_mov_items_aligned", row.get("meta_anulada_mov_items", []))],
                "meta_credito_historico_items": [str(v) for v in row.get("meta_credito_historico_items_aligned", row.get("meta_credito_historico_items", []))],
                "meta_anulada_historico_items": [str(v) for v in row.get("meta_anulada_historico_items_aligned", row.get("meta_anulada_historico_items", []))],
                "meta_credito_historico": str(row.get("meta_credito_historico")) if row.get("meta_credito_historico") is not None else "",
                "meta_anulada_historico": str(row.get("meta_anulada_historico")) if row.get("meta_anulada_historico") is not None else "",
                "has_historico_movimento": bool(row.get("has_historico_movimento")),
                "plan21_ids": row.get("plan21_ids") or [],
            }
        )

    user_id = _resolve_usuario_id()
    now = _now_local()
    try:
        if registro_existente:
            registro = registro_existente
            registro.exercicio = int(exercicio_raw)
            registro.unidade_orcamentaria = unidade_orcamentaria
            registro.programa = programa
            registro.acao_paoe = acao_paoe
            registro.responsavel_acao = adj_solicitante
            registro.produto_acao = produto_acao
            registro.unid_medida_produto = unid_medida_produto
            registro.regiao_produto = json.dumps(rows_serialized, ensure_ascii=False)
            registro.meta_produto = meta_produto_total
            registro.meta_credito = meta_credito_total
            registro.meta_anulada = meta_anulada_total
            registro.meta_atual = meta_atual_total
            registro.justificativa = justificativa
            registro.plan21_nger_id = plan21_nger_id_latest
            registro.plan21_nger_ids = json.dumps(all_plan_ids, ensure_ascii=False) if all_plan_ids else None
            registro.status_aprovacao = "aguardando"
            registro.situacao = "ativo"
            registro.alterado_em = now
            registro.ativo = True
            (
                db.session.query(AlterarMetaItem)
                .filter(AlterarMetaItem.alterar_meta_id == registro.id)
                .filter(AlterarMetaItem.ativo == True)  # noqa: E712
                .update(
                    {
                        "ativo": False,
                        "alterado_em": now,
                        "excluido_em": now,
                    },
                    synchronize_session=False,
                )
            )
            db.session.flush()
        else:
            registro = AlterarMeta(
                exercicio=int(exercicio_raw),
                unidade_orcamentaria=unidade_orcamentaria,
                programa=programa,
                acao_paoe=acao_paoe,
                responsavel_acao=adj_solicitante,
                produto_acao=produto_acao,
                unid_medida_produto=unid_medida_produto,
                regiao_produto=json.dumps(rows_serialized, ensure_ascii=False),
                meta_produto=meta_produto_total,
                meta_credito=meta_credito_total,
                meta_anulada=meta_anulada_total,
                meta_atual=meta_atual_total,
                justificativa=justificativa,
                plan21_nger_id=plan21_nger_id_latest,
                plan21_nger_ids=json.dumps(all_plan_ids, ensure_ascii=False) if all_plan_ids else None,
                usuario_id=user_id,
                status_aprovacao="aguardando",
                situacao="ativo",
                criado_em=now,
                ativo=True,
            )
            db.session.add(registro)
            db.session.flush()
        for row in parsed_rows:
            regiao_codigo = _normalize_codigo_num(row.get("regiao_produto")) or str(row.get("regiao_produto") or "").strip()
            for item in row.get("credito_mov_items", []):
                db.session.add(
                    AlterarMetaItem(
                        alterar_meta_id=registro.id,
                        regiao_codigo=regiao_codigo,
                        tipo="acrescimo",
                        valor=item.get("valor") or Decimal("0"),
                        ordem=item.get("ordem") or 1,
                        usuario_id=user_id,
                        ativo=True,
                        criado_em=now,
                    )
                )
            for item in row.get("anulada_mov_items", []):
                db.session.add(
                    AlterarMetaItem(
                        alterar_meta_id=registro.id,
                        regiao_codigo=regiao_codigo,
                        tipo="reducao",
                        valor=item.get("valor") or Decimal("0"),
                        ordem=item.get("ordem") or 1,
                        usuario_id=user_id,
                        ativo=True,
                        criado_em=now,
                    )
                )
        db.session.commit()
        _cache_invalidate_meta_fisica_options()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar meta fisica: {exc}"}), 500

    adj_token = _normalize_controle_token(adj_solicitante) or "SEMADJ"
    controle = f"META.{exercicio_raw}.{adj_token}.{registro.id}"
    return jsonify(
        {
            "ok": True,
            "message": "Meta fisica atualizada." if registro_existente else "Meta fisica salva.",
            "count": 1,
            "meta_fisica": {
                "id": registro.id,
                "controle": controle,
                "status_aprovacao": registro.status_aprovacao or "aguardando",
                "criado_em": now.isoformat(),
                "usuario_nome": (session.get("user") or {}).get("nome") or "",
                "usuario_id": user_id or "",
                "exercicio": exercicio_raw,
                "unidade_orcamentaria": unidade_orcamentaria,
                "programa": programa,
                "acao_paoe": acao_paoe,
                "adj_solicitante": adj_solicitante,
                "produto_acao": produto_acao,
                "unid_medida_produto": unid_medida_produto,
                "justificativa": justificativa,
                "linhas": rows_serialized,
                "plan21_nger_id": plan21_nger_id_latest,
                "plan21_nger_ids": all_plan_ids,
            },
        }
    )


def _parse_int_list_unique(raw_value) -> list[int]:
    values = raw_value
    if isinstance(raw_value, str):
        txt = raw_value.strip()
        if not txt:
            values = []
        else:
            try:
                values = json.loads(txt)
            except Exception:
                values = []
    if not isinstance(values, list):
        values = []
    out = []
    seen = set()
    for item in values:
        try:
            num = int(item)
        except Exception:
            num = 0
        if num > 0 and num not in seen:
            seen.add(num)
            out.append(num)
    return out


def _parse_meta_rows_for_apply(registro: AlterarMeta) -> list[dict]:
    reg_raw = (getattr(registro, "regiao_produto", "") or "").strip()
    if not reg_raw:
        return []
    try:
        parsed = json.loads(reg_raw)
    except Exception:
        parsed = []
    if not isinstance(parsed, list):
        parsed = []
    rows = []
    for item in parsed:
        if not isinstance(item, dict):
            continue
        regiao = str(item.get("regiao_produto") or "").strip()
        if not regiao:
            continue
        meta_produto = _parse_decimal(item.get("meta_produto")) or Decimal("0")
        meta_credito = _parse_decimal(item.get("meta_credito")) or Decimal("0")
        meta_anulada = _parse_decimal(item.get("meta_anulada")) or Decimal("0")
        meta_atual = _parse_decimal(item.get("meta_atual"))
        if meta_atual is None:
            meta_atual = meta_produto + meta_credito - meta_anulada
        plan21_ids = _parse_int_list_unique(item.get("plan21_ids"))
        rows.append(
            {
                "regiao_produto": regiao,
                "meta_produto": meta_produto,
                "meta_credito": meta_credito,
                "meta_anulada": meta_anulada,
                "meta_atual": meta_atual,
                "meta_credito_items": item.get("meta_credito_items") if isinstance(item.get("meta_credito_items"), list) else [],
                "meta_anulada_items": item.get("meta_anulada_items") if isinstance(item.get("meta_anulada_items"), list) else [],
                "meta_credito_historico_items": item.get("meta_credito_historico_items")
                if isinstance(item.get("meta_credito_historico_items"), list)
                else [],
                "meta_anulada_historico_items": item.get("meta_anulada_historico_items")
                if isinstance(item.get("meta_anulada_historico_items"), list)
                else [],
                "meta_credito_mov_items": item.get("meta_credito_mov_items")
                if isinstance(item.get("meta_credito_mov_items"), list)
                else [],
                "meta_anulada_mov_items": item.get("meta_anulada_mov_items")
                if isinstance(item.get("meta_anulada_mov_items"), list)
                else [],
                "plan21_ids": plan21_ids,
            }
        )
    return rows


def _build_aligned_lists_from_items(item_rows: list[dict]) -> tuple[list[str], list[str]]:
    # Preserva o pareamento contábil por "ordem":
    # itens de acréscimo/redução com a mesma ordem devem ocupar a mesma linha.
    grouped = {}
    order_keys = []
    for it in item_rows or []:
        tipo = str(it.get("tipo") or "").strip().lower()
        val = _parse_decimal(it.get("valor"))
        if val is None or val <= Decimal("0"):
            continue
        ordem = it.get("ordem")
        try:
            ordem_key = int(ordem) if ordem is not None else None
        except Exception:
            ordem_key = None
        if ordem_key is None:
            try:
                ordem_key = int(it.get("id") or 0)
            except Exception:
                ordem_key = 0
        if ordem_key not in grouped:
            grouped[ordem_key] = {"credito": "", "anulada": ""}
            order_keys.append(ordem_key)
        if tipo == "acrescimo":
            grouped[ordem_key]["credito"] = str(val)
        elif tipo == "reducao":
            grouped[ordem_key]["anulada"] = str(val)

    credito = []
    anulada = []
    for key in sorted(order_keys):
        row = grouped.get(key) or {}
        c = str(row.get("credito") or "").strip()
        a = str(row.get("anulada") or "").strip()
        if not c and not a:
            continue
        credito.append(c)
        anulada.append(a)
    return credito, anulada


def _rebuild_meta_rows_from_alterar_meta_item(registro: AlterarMeta) -> list[dict]:
    base_rows = _parse_meta_rows_for_apply(registro)
    if not base_rows:
        return []

    registro_id = int(getattr(registro, "id", 0) or 0)
    if registro_id <= 0:
        return []

    ctx = {
        "exercicio": getattr(registro, "exercicio", None),
        "unidade_orcamentaria": str(getattr(registro, "unidade_orcamentaria", "") or "").strip(),
        "programa": str(getattr(registro, "programa", "") or "").strip(),
        "acao_paoe": str(getattr(registro, "acao_paoe", "") or "").strip(),
        "produto_acao": str(getattr(registro, "produto_acao", "") or "").strip(),
        "unid_medida_produto": str(getattr(registro, "unid_medida_produto", "") or "").strip(),
    }
    region_keys_in_current = set()
    for row in base_rows:
        reg = str(row.get("regiao_produto") or "").strip()
        reg_key = _normalize_codigo_num(reg) or reg.lower()
        if reg_key:
            region_keys_in_current.add(reg_key)

    metas_ctx = (
        AlterarMeta.query.filter(AlterarMeta.ativo == True)  # noqa: E712
        .filter(AlterarMeta.exercicio == ctx["exercicio"])
        .filter(AlterarMeta.unidade_orcamentaria == ctx["unidade_orcamentaria"])
        .filter(AlterarMeta.programa == ctx["programa"])
        .filter(AlterarMeta.acao_paoe == ctx["acao_paoe"])
        .filter(AlterarMeta.produto_acao == ctx["produto_acao"])
        .filter(AlterarMeta.unid_medida_produto == ctx["unid_medida_produto"])
        .filter(AlterarMeta.id <= registro_id)
        .order_by(AlterarMeta.id.asc())
        .all()
    )
    metas_chain = []
    for m in metas_ctx:
        mid = int(getattr(m, "id", 0) or 0)
        if mid <= 0:
            continue
        if mid == registro_id:
            metas_chain.append(mid)
            continue
        st = str(getattr(m, "status_aprovacao", "") or "").strip().lower()
        if st == "aprovado":
            metas_chain.append(mid)

    if not metas_chain:
        metas_chain = [registro_id]

    status_registro_atual = str(getattr(registro, "status_aprovacao", "") or "").strip().lower()
    latest_approved_totals_by_region = {}
    if status_registro_atual != "aprovado":
        for m in metas_ctx:
            mid = int(getattr(m, "id", 0) or 0)
            if mid <= 0 or mid >= registro_id:
                continue
            st = str(getattr(m, "status_aprovacao", "") or "").strip().lower()
            if st != "aprovado":
                continue
            for prev_row in _parse_meta_rows_for_apply(m):
                prev_reg = str(prev_row.get("regiao_produto") or "").strip()
                prev_key = _normalize_codigo_num(prev_reg) or prev_reg.lower()
                if not prev_key or prev_key not in region_keys_in_current:
                    continue
                latest_approved_totals_by_region[prev_key] = {
                    "meta_credito": _parse_decimal(prev_row.get("meta_credito")) or Decimal("0"),
                    "meta_anulada": _parse_decimal(prev_row.get("meta_anulada")) or Decimal("0"),
                }

    meta_ids_csv = ",".join(str(int(x)) for x in metas_chain)
    chain_items_rows = _safe_query_mappings(
        """
        SELECT alterar_meta_id, regiao_codigo, tipo, valor, ordem, id
        FROM alterar_meta_item
        WHERE ativo = 1
          AND alterar_meta_id IN (""" + meta_ids_csv + """)
        ORDER BY alterar_meta_id ASC, ordem ASC, id ASC
        """,
        {},
    )
    items_by_meta_region = {}
    for it in chain_items_rows:
        meta_id = int(it.get("alterar_meta_id") or 0)
        reg = str(it.get("regiao_codigo") or "").strip()
        reg_key = _normalize_codigo_num(reg) or reg.lower()
        if not reg_key or reg_key not in region_keys_in_current:
            continue
        items_by_meta_region.setdefault(meta_id, {}).setdefault(reg_key, []).append(it)

    def _same_meta_pair(left: tuple[str, str], right: tuple[str, str]) -> bool:
        for idx in (0, 1):
            l_val = _parse_decimal(left[idx])
            r_val = _parse_decimal(right[idx])
            l_num = l_val if l_val is not None else Decimal("0")
            r_num = r_val if r_val is not None else Decimal("0")
            if abs(l_num - r_num) >= Decimal("0.000001"):
                return False
        return True

    def _strip_historico_prefix(
        hist_pairs: list[tuple[str, str]],
        mov_pairs: list[tuple[str, str]],
    ) -> list[tuple[str, str]]:
        if not hist_pairs or not mov_pairs or len(mov_pairs) < len(hist_pairs):
            return mov_pairs
        for idx, hist_pair in enumerate(hist_pairs):
            if not _same_meta_pair(hist_pair, mov_pairs[idx]):
                return mov_pairs
        return mov_pairs[len(hist_pairs):]

    historico_by_region = {}
    movimento_by_region = {}
    for mid in metas_chain:
        by_region = items_by_meta_region.get(int(mid), {})
        for reg_key in region_keys_in_current:
            reg_rows = by_region.get(reg_key, [])
            cred_list, anu_list = _build_aligned_lists_from_items(reg_rows)
            size = max(len(cred_list), len(anu_list))
            if size <= 0:
                continue
            pairs = []
            for i in range(size):
                c = str(cred_list[i] if i < len(cred_list) else "" or "").strip()
                a = str(anu_list[i] if i < len(anu_list) else "" or "").strip()
                if c or a:
                    pairs.append((c, a))
            if not pairs:
                continue
            if int(mid) == registro_id:
                current_hist = list(historico_by_region.get(reg_key) or [])
                current_mov = _strip_historico_prefix(current_hist, pairs)
                if current_mov:
                    movimento_by_region.setdefault(reg_key, []).extend(current_mov)
            else:
                current_hist = list(historico_by_region.get(reg_key) or [])
                hist_delta = _strip_historico_prefix(current_hist, pairs)
                if hist_delta:
                    historico_by_region.setdefault(reg_key, []).extend(hist_delta)

    def _fmt_meta_decimal(value: Decimal) -> str:
        if value <= Decimal("0"):
            return ""
        txt = format(value, "f")
        if "." in txt:
            txt = txt.rstrip("0").rstrip(".")
        return txt

    out = []
    for row in base_rows:
        reg = str(row.get("regiao_produto") or "").strip()
        reg_key = _normalize_codigo_num(reg) or reg.lower()
        if status_registro_atual != "aprovado":
            approved_totals = latest_approved_totals_by_region.get(reg_key) or {}
            hist_credito_total = approved_totals.get("meta_credito") or Decimal("0")
            hist_anulada_total = approved_totals.get("meta_anulada") or Decimal("0")
            current_credito_total = _parse_decimal(row.get("meta_credito")) or Decimal("0")
            current_anulada_total = _parse_decimal(row.get("meta_anulada")) or Decimal("0")
            mov_credito_total = current_credito_total - hist_credito_total
            mov_anulada_total = current_anulada_total - hist_anulada_total
            hist_pairs = list(historico_by_region.get(reg_key) or [])
            mov_pairs = []
            if mov_credito_total > Decimal("0") or mov_anulada_total > Decimal("0"):
                mov_pairs.append(
                    (
                        _fmt_meta_decimal(mov_credito_total),
                        _fmt_meta_decimal(mov_anulada_total),
                    )
                )
        else:
            hist_pairs = list(historico_by_region.get(reg_key) or [])
            timeline_mov_pairs = list(movimento_by_region.get(reg_key) or [])
            mov_pairs = _strip_historico_prefix(hist_pairs, timeline_mov_pairs)
        hist_credito_items = [p[0] for p in hist_pairs]
        hist_anulada_items = [p[1] for p in hist_pairs]
        mov_credito_items = [p[0] for p in mov_pairs]
        mov_anulada_items = [p[1] for p in mov_pairs]
        meta_credito_historico = Decimal("0")
        for v in hist_credito_items:
            p = _parse_decimal(v)
            if p is not None:
                meta_credito_historico += p
        meta_anulada_historico = Decimal("0")
        for v in hist_anulada_items:
            p = _parse_decimal(v)
            if p is not None:
                meta_anulada_historico += p
        out.append(
            {
                "regiao_produto": reg,
                "meta_produto": str(row.get("meta_produto") or Decimal("0")),
                "meta_credito": str(row.get("meta_credito") or Decimal("0")),
                "meta_anulada": str(row.get("meta_anulada") or Decimal("0")),
                "meta_atual": str(row.get("meta_atual") or Decimal("0")),
                "meta_credito_historico": str(meta_credito_historico),
                "meta_anulada_historico": str(meta_anulada_historico),
                "meta_credito_historico_items": hist_credito_items,
                "meta_anulada_historico_items": hist_anulada_items,
                "meta_credito_mov_items": mov_credito_items,
                "meta_anulada_mov_items": mov_anulada_items,
                "plan21_ids": row.get("plan21_ids") or [],
            }
        )
    return out


@home_bp.route("/api/meta-fisica/<int:meta_id>/linhas", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/meta_fisica")
def api_meta_fisica_linhas(meta_id: int):
    registro = (
        AlterarMeta.query.filter(AlterarMeta.id == meta_id)
        .filter(
            or_(
                AlterarMeta.ativo == True,  # noqa: E712
                func.lower(func.coalesce(AlterarMeta.status_aprovacao, "")) == "rejeitado",
            )
        )
        .first()
    )
    if not registro:
        return jsonify({"error": "Registro de meta não encontrado."}), 404
    linhas = _rebuild_meta_rows_from_alterar_meta_item(registro)
    return jsonify({"linhas": linhas})


def _apply_meta_fisica_to_plan21(registro: AlterarMeta) -> list[int]:
    rows = _parse_meta_rows_for_apply(registro)
    if not rows:
        return []

    impacted_ids: list[int] = []

    for row in rows:
        for plan_id in row.get("plan21_ids") or []:
            db.session.execute(
                text(
                    """
                    UPDATE plan21_nger
                    SET meta_credito = :meta_credito,
                        meta_anulada = :meta_anulada,
                        meta_atual = :meta_atual
                    WHERE id = :id
                    """
                ),
                {
                    "meta_credito": row["meta_credito"],
                    "meta_anulada": row["meta_anulada"],
                    "meta_atual": row["meta_atual"],
                    "id": int(plan_id),
                },
            )
            impacted_ids.append(int(plan_id))

    template_id = None
    try:
        template_id = int(getattr(registro, "plan21_nger_id", 0) or 0)
    except Exception:
        template_id = 0
    if template_id <= 0:
        plan_ids_fallback = _parse_int_list_unique(getattr(registro, "plan21_nger_ids", None))
        if plan_ids_fallback:
            template_id = plan_ids_fallback[-1]

    template_row = {}
    if template_id and template_id > 0:
        template_sql = """
            SELECT
                exercicio,
                programa,
                funcao,
                unidade_orcamentaria,
                acao_paoe,
                subfuncao,
                objetivo_especifico,
                esfera,
                responsavel_acao,
                produto_acao,
                unid_medida_produto,
                publico_transversal
            FROM plan21_nger
            WHERE id = :id
            LIMIT 1
        """
        template_row = db.session.execute(text(template_sql), {"id": template_id}).mappings().first() or {}

    for row in rows:
        if row.get("plan21_ids"):
            continue
        insert_sql = text(
            """
            INSERT INTO plan21_nger (
                exercicio,
                programa,
                funcao,
                unidade_orcamentaria,
                acao_paoe,
                subfuncao,
                objetivo_especifico,
                esfera,
                responsavel_acao,
                produto_acao,
                unid_medida_produto,
                regiao_produto,
                meta_produto,
                meta_credito,
                meta_anulada,
                meta_atual,
                publico_transversal,
                ativo
            )
            VALUES (
                :exercicio,
                :programa,
                :funcao,
                :unidade_orcamentaria,
                :acao_paoe,
                :subfuncao,
                :objetivo_especifico,
                :esfera,
                :responsavel_acao,
                :produto_acao,
                :unid_medida_produto,
                :regiao_produto,
                :meta_produto,
                :meta_credito,
                :meta_anulada,
                :meta_atual,
                :publico_transversal,
                :ativo
            )
            """
        )
        insert_params = {
            "exercicio": template_row.get("exercicio") or getattr(registro, "exercicio", None),
            "programa": template_row.get("programa") or getattr(registro, "programa", None),
            "funcao": template_row.get("funcao"),
            "unidade_orcamentaria": template_row.get("unidade_orcamentaria")
            or getattr(registro, "unidade_orcamentaria", None),
            "acao_paoe": template_row.get("acao_paoe") or getattr(registro, "acao_paoe", None),
            "subfuncao": template_row.get("subfuncao"),
            "objetivo_especifico": template_row.get("objetivo_especifico"),
            "esfera": template_row.get("esfera"),
            "responsavel_acao": template_row.get("responsavel_acao")
            or getattr(registro, "responsavel_acao", None),
            "produto_acao": template_row.get("produto_acao") or getattr(registro, "produto_acao", None),
            "unid_medida_produto": template_row.get("unid_medida_produto")
            or getattr(registro, "unid_medida_produto", None),
            "regiao_produto": row.get("regiao_produto"),
            "meta_produto": row.get("meta_produto") or Decimal("0"),
            "meta_credito": row.get("meta_credito") or Decimal("0"),
            "meta_anulada": row.get("meta_anulada") or Decimal("0"),
            "meta_atual": row.get("meta_atual") or Decimal("0"),
            "publico_transversal": template_row.get("publico_transversal"),
            "ativo": True,
        }
        result = db.session.execute(insert_sql, insert_params)
        new_id = 0
        try:
            new_id = int(getattr(result, "lastrowid", 0) or 0)
        except Exception:
            new_id = 0
        if new_id > 0:
            impacted_ids.append(new_id)

    impacted_unique = sorted(set(int(v) for v in impacted_ids if int(v) > 0))
    return impacted_unique


@home_bp.route("/api/meta-fisica/<int:meta_id>", methods=["DELETE"])
@login_required
@require_feature("cadastrar/plan_21-nger/meta_fisica")
def api_meta_fisica_delete(meta_id: int):
    registro = (
        AlterarMeta.query.filter(AlterarMeta.id == meta_id)
        .filter(AlterarMeta.ativo == True)  # noqa: E712
        .first()
    )
    if not registro:
        return jsonify({"error": "Registro de meta nao encontrado."}), 404

    criador = None
    if getattr(registro, "usuario_id", None):
        criador = Usuario.query.filter_by(id=registro.usuario_id).first()

    exercicio_ctrl = str(getattr(registro, "exercicio", "") or _now_local().year)
    adj_nome = (getattr(registro, "responsavel_acao", "") or "").strip()
    if not adj_nome and criador:
        perfil_id = getattr(criador, "perfil_id", None)
        if perfil_id:
            perfil_row = Perfil.query.filter_by(id=perfil_id).first()
            adj_nome = (getattr(perfil_row, "nome", "") or "").strip()
        if not adj_nome:
            adj_nome = (getattr(criador, "perfil", "") or "").strip()
    adj_token = _normalize_controle_token(adj_nome) or "SEMADJ"
    controle_meta = f"META.{exercicio_ctrl}.{adj_token}.{registro.id}"

    status = str(getattr(registro, "status_aprovacao", "") or "").strip().lower()
    if status != "aguardando":
        return jsonify({"error": f"Somente registros com status Aguardando podem ser excluidos ({controle_meta})."}), 400

    usuario_id = _resolve_usuario_id()
    if not usuario_id:
        return jsonify({"error": f"Usuario sem permissao para excluir registro {controle_meta}."}), 403
    usuario_atual = Usuario.query.filter_by(id=usuario_id).first()
    perfil_atual = str(getattr(usuario_atual, "perfil_id", "") or "").strip() if usuario_atual else ""
    perfil_criador = str(getattr(criador, "perfil_id", "") or "").strip() if criador else ""
    if not perfil_atual or perfil_atual != perfil_criador:
        return jsonify({"error": f"Usuario sem permissao para excluir registro {controle_meta}."}), 403

    now = _now_local()
    try:
        registro.ativo = False
        registro.situacao = "inativo"
        registro.excluido_em = now
        registro.alterado_em = now
        (
            db.session.query(AlterarMetaItem)
            .filter(AlterarMetaItem.alterar_meta_id == registro.id)
            .filter(AlterarMetaItem.ativo == True)  # noqa: E712
            .update(
                {
                    AlterarMetaItem.ativo: False,
                    AlterarMetaItem.excluido_em: now,
                    AlterarMetaItem.alterado_em: now,
                },
                synchronize_session=False,
            )
        )
        db.session.commit()
        _cache_invalidate_meta_fisica_options()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir registro {controle_meta}: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "message": f"Registro {controle_meta} excluido com sucesso.",
            "controle_meta": controle_meta,
            "id": registro.id,
        }
    )


@home_bp.route("/api/meta-fisica/<int:meta_id>/aprovar", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/meta_fisica")
def api_meta_fisica_aprovar(meta_id: int):
    registro = (
        AlterarMeta.query.filter(AlterarMeta.id == meta_id)
        .filter(AlterarMeta.ativo == True)  # noqa: E712
        .first()
    )
    if not registro:
        return jsonify({"error": "Registro de meta nao encontrado."}), 404

    status_atual = str(getattr(registro, "status_aprovacao", "") or "").strip().lower()
    if status_atual != "aguardando":
        return jsonify({"error": "Registro de meta ja foi processado."}), 400

    try:
        nivel_int = int(getattr(g, "user_nivel", 0) or 0)
    except (TypeError, ValueError):
        nivel_int = 0
    if nivel_int not in (1, 2):
        return jsonify({"error": "Usuario sem permissao para aprovar o registro atual."}), 403

    data = request.get_json() or {}
    aprovado_raw = str(data.get("meta_aprovada") or "").strip().lower()
    aprovado = aprovado_raw if aprovado_raw in ("sim", "nao") else ""
    if not aprovado:
        return jsonify({"error": "Informe se deseja aprovar o registro (Sim ou Nao)."}), 400
    motivo = str(data.get("motivo_rejeicao") or "").strip()
    if not motivo:
        return jsonify({"error": "Justificativa da decisao obrigatoria."}), 400

    usuario_id = _resolve_usuario_id()
    if usuario_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400

    now = _now_local()
    try:
        impacted_ids = []
        if aprovado == "sim":
            impacted_ids = _apply_meta_fisica_to_plan21(registro)
            registro.status_aprovacao = "Aprovado"
            registro.situacao = "ativo"
            registro.motivo_rejeicao = motivo
            if impacted_ids:
                registro.plan21_nger_id = impacted_ids[-1]
                registro.plan21_nger_ids = json.dumps(impacted_ids, ensure_ascii=False)
        else:
            registro.status_aprovacao = "Rejeitado"
            registro.situacao = "inativo"
            registro.ativo = False
            registro.excluido_em = now
            registro.motivo_rejeicao = motivo
            (
                db.session.query(AlterarMetaItem)
                .filter(AlterarMetaItem.alterar_meta_id == registro.id)
                .filter(AlterarMetaItem.ativo == True)  # noqa: E712
                .update(
                    {
                        AlterarMetaItem.ativo: False,
                        AlterarMetaItem.alterado_em: now,
                        AlterarMetaItem.excluido_em: now,
                    },
                    synchronize_session=False,
                )
            )

        registro.aprovado_por = str(usuario_id)
        registro.data_aprovacao = now
        registro.alterado_em = now

        db.session.commit()
        _cache_invalidate_meta_fisica_options()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao aprovar registro de meta: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "message": "Registro de meta aprovado com sucesso."
            if aprovado == "sim"
            else "Registro de meta rejeitado com sucesso.",
            "id": registro.id,
            "status_aprovacao": registro.status_aprovacao,
            "data_aprovacao": registro.data_aprovacao.isoformat() if getattr(registro, "data_aprovacao", None) else "",
            "aprovado_por": registro.aprovado_por or "",
        }
    )


@home_bp.route("/api/etapa/options", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def api_etapa_options():
    current_year = str(_now_local().year)
    cols = set(_plan21_columns())
    fields = {
        "exercicio": "exercicio",
        "unidade_orcamentaria": "unidade_orcamentaria",
        "programa": "programa",
        "produto_acao": "produto_acao",
        "acao_paoe": "acao_paoe",
        "subacao_entrega": "subacao_entrega",
    }
    selected = {}
    for key in fields:
        val = (request.args.get(key) or "").strip()
        if val:
            selected[key] = val
    selected["exercicio"] = selected.get("exercicio") or current_year
    cache_key = (
        "etapa_options_v2",
        tuple(sorted((k, str(v)) for k, v in selected.items())),
    )
    cached_payload = _cache_get_etapa_options(cache_key)
    if cached_payload is not None:
        return jsonify(cached_payload)

    payload = {key: [] for key in fields}
    payload["exercicio"] = [current_year]
    where = ["ativo = 1"] if "ativo" in cols else []
    params = {}
    for key, col in fields.items():
        if key == "exercicio":
            continue
        val = (selected.get(key) or "").strip()
        if val and col in cols:
            where.append(f"{col} = :{key}")
            params[key] = val
    if "exercicio" in cols:
        where.append("exercicio = :exercicio")
        params["exercicio"] = selected["exercicio"]
    select_cols = [f"{col} AS {key}" for key, col in fields.items() if key != "exercicio" and col in cols]
    if select_cols:
        sql = f"SELECT DISTINCT {', '.join(select_cols)} FROM plan21_nger"
        if where:
            sql += " WHERE " + " AND ".join(where)
        option_rows = _safe_query_mappings(sql, params)
        for row in option_rows:
            for key in ("produto_acao", "acao_paoe", "subacao_entrega"):
                val = row.get(key)
                if val is None:
                    continue
                text_val = str(val).strip()
                if text_val:
                    payload[key].append(text_val)
    for key in ("unidade_orcamentaria", "programa", "produto_acao", "acao_paoe", "subacao_entrega"):
        payload[key] = sorted(set(payload[key]), key=lambda v: v.lower())

    option_rows = []
    option_select_cols = [
        "id",
        "unidade_orcamentaria",
        "programa",
        "produto_acao",
        "acao_paoe",
        "subacao_entrega",
        "codigo",
        "municipios_entrega",
        "etapa",
        "responsavel_etapa",
    ]
    option_select_parts = [c for c in option_select_cols if c in cols]
    option_select_parts.append("prazo_etapa" if "prazo_etapa" in cols else "NULL AS prazo_etapa")
    option_where = ["ativo = 1"] if "ativo" in cols else []
    option_params = {}
    if "exercicio" in cols:
        option_where.append("exercicio = :exercicio")
        option_params["exercicio"] = selected["exercicio"]
    option_sql = f"SELECT {', '.join(option_select_parts)} FROM plan21_nger"
    if option_where:
        option_sql += " WHERE " + " AND ".join(option_where)
    option_sql += " ORDER BY id"
    for row in _safe_query_mappings(option_sql, option_params):
        option_rows.append(
            {
                "id": row.get("id"),
                "unidade_orcamentaria": str(row.get("unidade_orcamentaria") or "").strip(),
                "programa": str(row.get("programa") or "").strip(),
                "produto_acao": str(row.get("produto_acao") or "").strip(),
                "acao_paoe": str(row.get("acao_paoe") or "").strip(),
                "subacao_entrega": str(row.get("subacao_entrega") or "").strip(),
                "codigo": str(row.get("codigo") or "").strip(),
                "municipios_entrega": str(row.get("municipios_entrega") or "").strip(),
                "etapa": str(row.get("etapa") or "").strip(),
                "responsavel_etapa": str(row.get("responsavel_etapa") or "").strip(),
                "prazo_etapa": str(row.get("prazo_etapa") or "").strip(),
            }
        )
    payload["option_rows"] = option_rows

    adj_rows = _safe_query_mappings(
        """
        SELECT abreviacao AS value
        FROM adj
        WHERE ativo = 1
        ORDER BY abreviacao
        """
    )
    adj_vals = []
    for row in adj_rows:
        raw_val = row.get("value")
        if raw_val is None:
            continue
        txt = str(raw_val).strip()
        if txt:
            adj_vals.append(txt)
    payload["adj_solicitante"] = sorted(set(adj_vals), key=lambda v: v.lower())

    responsaveis_nger = []
    usuarios_nger = Usuario.query.filter(Usuario.perfil_id == 2).order_by(Usuario.nome).all()
    for usuario in usuarios_nger:
        nome = (getattr(usuario, "nome", "") or "").strip()
        email = (getattr(usuario, "email", "") or "").strip()
        label = nome or email
        if label:
            responsaveis_nger.append(label)
    payload["responsaveis_nger"] = responsaveis_nger

    etapa_filters = {
        "exercicio": selected.get("exercicio", ""),
        "unidade_orcamentaria": selected.get("unidade_orcamentaria", ""),
        "programa": selected.get("programa", ""),
        "produto_acao": selected.get("produto_acao", ""),
        "acao_paoe": selected.get("acao_paoe", ""),
        "subacao_entrega": selected.get("subacao_entrega", ""),
    }
    etapa_rows = _load_etapa_plan21_rows(etapa_filters)
    etapas = []
    seen = set()
    for row in etapa_rows:
        etapa = (row.get("etapa") or "").strip()
        if not etapa or etapa in seen:
            continue
        seen.add(etapa)
        etapas.append(
            {
                "id": row.get("id"),
                "value": etapa,
                "label": etapa,
                "responsavel_etapa": row.get("responsavel_etapa") or "",
                "prazo_etapa": row.get("prazo_etapa") or "",
            }
        )
    payload["etapas"] = etapas
    _cache_set_etapa_options(cache_key, payload)
    return jsonify(payload)


@home_bp.route("/api/etapa/plan21-rows", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def api_etapa_plan21_rows():
    filters = {
        "exercicio": (request.args.get("exercicio") or "").strip(),
        "unidade_orcamentaria": (request.args.get("unidade_orcamentaria") or "").strip(),
        "programa": (request.args.get("programa") or "").strip(),
        "produto_acao": (request.args.get("produto_acao") or "").strip(),
        "acao_paoe": (request.args.get("acao_paoe") or "").strip(),
        "subacao_entrega": (request.args.get("subacao_entrega") or "").strip(),
    }
    etapa = (request.args.get("etapa") or "").strip()
    validar_exclusao = (request.args.get("validar_exclusao") or "").strip().lower() in {"1", "true", "sim"}
    rows = _load_etapa_plan21_rows(filters, etapa=etapa or None)
    if validar_exclusao:
        err = _validate_etapa_exclusao_sem_memoria(rows, etapa)
        if err:
            return jsonify({"error": err}), 400
    return jsonify({"rows": rows})


def _apply_etapa_payload(registro: CadastrarEtapa, data: dict):
    tipo = (data.get("tipo_solicitacao") or "cadastrar").strip().lower()
    if tipo not in {"cadastrar", "alterar", "excluir"}:
        return "Tipo de solicitacao invalido."
    exercicio_raw = (data.get("exercicio") or "").strip()
    try:
        exercicio = int(exercicio_raw)
    except ValueError:
        return "Exercicio invalido."
    filters = {
        "exercicio": exercicio_raw,
        "unidade_orcamentaria": (data.get("unidade_orcamentaria") or "").strip(),
        "programa": (data.get("programa") or "").strip(),
        "produto_acao": (data.get("produto_acao") or "").strip(),
        "acao_paoe": (data.get("acao_paoe") or "").strip(),
        "subacao_entrega": (data.get("subacao_entrega") or "").strip(),
        "adj_solicitante": (data.get("adj_solicitante") or "").strip(),
    }
    missing_filter = [k for k, v in filters.items() if not v]
    if missing_filter:
        return f"Campos obrigatorios ausentes: {', '.join(missing_filter)}."
    etapa_nova = (data.get("etapa_nova") or "").strip()
    responsavel_novo = (data.get("responsavel_etapa_novo") or "").strip()
    cpf_novo = re.sub(r"\D", "", (data.get("cpf_responsavel_etapa_novo") or ""))
    email_novo = (data.get("email_responsavel_etapa_novo") or "").strip()
    data_inicio = _parse_etapa_date(data.get("data_inicio_novo") or "")
    data_fim = _parse_etapa_date(data.get("data_fim_novo") or "")
    justificativa = (data.get("justificativa") or "").strip()
    responsavel_nger = (data.get("responsavel_nger") or "").strip()
    municipio_filtro = (data.get("municipios_entrega") or "").strip()
    source_rows = []
    source_id_raw = data.get("plan21_nger_id")
    source_id = None
    if source_id_raw:
        try:
            source_id = int(source_id_raw)
        except (TypeError, ValueError):
            source_id = None
    if tipo != "excluir":
        required = {
            "etapa_nova": etapa_nova,
            "responsavel_etapa_novo": responsavel_novo,
            "cpf_responsavel_etapa_novo": cpf_novo,
            "email_responsavel_etapa_novo": email_novo,
            "data_inicio_novo": data_inicio,
            "data_fim_novo": data_fim,
            "justificativa": justificativa,
        }
    else:
        required = {
            "plan21_nger_id": source_id,
            "justificativa": justificativa,
        }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return f"Campos obrigatorios ausentes: {', '.join(missing)}."
    if cpf_novo and not _is_valid_cpf(cpf_novo):
        return "CPF do responsavel da etapa invalido."
    if email_novo and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email_novo):
        return "E-mail do responsavel da etapa invalido."
    if data_inicio and data_fim and data_fim < data_inicio:
        return "Data de fim nao pode ser anterior a data de inicio."

    if tipo in {"alterar", "excluir"}:
        if not source_id:
            return "Selecione uma etapa ja cadastrada."
        selected = _load_etapa_plan21_rows(filters, source_id=source_id)
        if not selected:
            return "Etapa selecionada nao encontrada na plan21_nger."
        etapa_origem = (selected[0].get("etapa") or "").strip()
        source_rows = _load_etapa_plan21_rows(filters, etapa=etapa_origem)
        if not source_rows:
            source_rows = selected
        if tipo == "excluir":
            err = _validate_etapa_exclusao_sem_memoria(source_rows, etapa_origem)
            if err:
                return err
    else:
        source_rows = _load_etapa_plan21_rows(filters)
        if municipio_filtro:
            filtered_source_rows = [
                row for row in source_rows if (row.get("municipios_entrega") or "").strip() == municipio_filtro
            ]
            if filtered_source_rows:
                source_rows = filtered_source_rows
        if not source_rows:
            return "Nenhuma linha base encontrada na plan21_nger para cadastrar a etapa."

    first = source_rows[0]
    origem_inicio, origem_fim = _split_etapa_prazo(first.get("prazo_etapa"))
    ids = []
    for row in source_rows:
        try:
            ids.append(int(row.get("id")))
        except (TypeError, ValueError):
            continue
    registro.exercicio = exercicio
    registro.unidade_orcamentaria = first.get("unidade_orcamentaria") or ""
    registro.programa = first.get("programa") or ""
    registro.acao_paoe = filters["acao_paoe"]
    registro.responsavel_acao = first.get("responsavel_acao") or filters["adj_solicitante"]
    registro.produto_acao = filters["produto_acao"]
    registro.adj_solicitante = filters["adj_solicitante"]
    registro.subacao_entrega = filters["subacao_entrega"]
    registro.chave_planejamento = first.get("chave_planejamento") or ""
    registro.tipo_solicitacao = tipo
    registro.plan21_nger_id = ids[0] if ids else None
    registro.plan21_nger_ids = json.dumps(ids) if len(ids) > 1 else None
    registro.etapa_origem = first.get("etapa") or ""
    registro.responsavel_etapa_origem = first.get("responsavel_etapa") or ""
    registro.data_inicio_origem = origem_inicio
    registro.data_fim_origem = origem_fim
    registro.etapa_nova = etapa_nova
    registro.responsavel_etapa_novo = responsavel_novo
    registro.cpf_responsavel_etapa_novo = cpf_novo
    registro.email_responsavel_etapa_novo = email_novo
    registro.data_inicio_novo = data_inicio
    registro.data_fim_novo = data_fim
    registro.justificativa = justificativa
    registro.responsavel_nger = responsavel_nger
    registro.status_aprovacao = "Aguardando"
    registro.situacao = "Aguardando"
    registro.alterado_em = _now_local()
    return None


@home_bp.route("/api/etapa", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def api_etapa_create():
    data = request.get_json() or {}
    registro = CadastrarEtapa(
        usuario_id=_resolve_usuario_id(),
        ativo=True,
        criado_em=_now_local(),
    )
    err = _apply_etapa_payload(registro, data)
    if err:
        return jsonify({"error": err}), 400
    db.session.add(registro)
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar etapa: {exc}"}), 500
    return jsonify({"etapa": _etapa_payload(registro)})


@home_bp.route("/api/etapa/<int:etapa_id>", methods=["PUT"])
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def api_etapa_update(etapa_id):
    registro = db.session.get(CadastrarEtapa, etapa_id)
    if not registro or registro.excluido_em is not None:
        return jsonify({"error": "Registro nao encontrado."}), 404
    status = (registro.status_aprovacao or "").strip().lower()
    if status and status != "aguardando":
        return jsonify({"error": "Somente registros com status Aguardando podem ser editados."}), 400
    criador = Usuario.query.filter_by(id=registro.usuario_id).first() if registro.usuario_id else None
    if not _current_user_matches_perfil(getattr(criador, "perfil_id", None)):
        return jsonify({"error": "Usuario sem permissao para editar o registro de etapa."}), 403
    err = _apply_etapa_payload(registro, request.get_json() or {})
    if err:
        return jsonify({"error": err}), 400
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar etapa: {exc}"}), 500
    return jsonify({"etapa": _etapa_payload(registro)})


@home_bp.route("/api/etapa/<int:etapa_id>", methods=["DELETE"])
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def api_etapa_delete(etapa_id):
    registro = db.session.get(CadastrarEtapa, etapa_id)
    if not registro or registro.ativo is False:
        return jsonify({"error": "Registro nao encontrado."}), 404
    status = (registro.status_aprovacao or "").strip().lower()
    if status and status != "aguardando":
        return jsonify({"error": "Somente registros com status Aguardando podem ser excluidos."}), 400
    criador = Usuario.query.filter_by(id=registro.usuario_id).first() if registro.usuario_id else None
    if not _current_user_matches_perfil(getattr(criador, "perfil_id", None)):
        return jsonify({"error": "Usuario sem permissao para excluir o registro de etapa."}), 403
    registro.ativo = False
    registro.excluido_em = _now_local()
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir etapa: {exc}"}), 500
    return jsonify({"ok": True})


def _apply_etapa_to_plan21(registro: CadastrarEtapa) -> list[int]:
    ids = _extract_etapa_plan21_ids(registro)
    if not ids:
        raise ValueError("Referencia plan21_nger nao encontrada para etapa.")
    prazo = _format_etapa_prazo(registro.data_inicio_novo, registro.data_fim_novo)
    overrides = {
        "etapa": registro.etapa_nova,
        "responsavel_etapa": registro.responsavel_etapa_novo,
        "prazo_etapa": prazo,
        "ativo": True,
    }
    tipo = (registro.tipo_solicitacao or "cadastrar").strip().lower()
    if tipo == "excluir":
        rows = _load_plan21_full_rows_by_ids(ids)
        err = _validate_etapa_exclusao_sem_memoria(rows, registro.etapa_origem)
        if err:
            raise ValueError(err)
        new_ids = _clone_etapa_exclusao_base_rows(rows)
        Plan21Nger.query.filter(Plan21Nger.id.in_(ids)).update({"ativo": False}, synchronize_session=False)
        return new_ids
    source_ids = [ids[0]] if tipo == "cadastrar" else ids
    nullify = _etapa_nullify_cols() if tipo == "cadastrar" else set()
    new_ids = []
    for source_id in source_ids:
        new_id = _clone_plan21_row(source_id, overrides=overrides, nullify=nullify)
        if new_id:
            new_ids.append(new_id)
    if tipo == "alterar":
        Plan21Nger.query.filter(Plan21Nger.id.in_(ids)).update({"ativo": False}, synchronize_session=False)
    return new_ids


@home_bp.route("/api/etapa/<int:etapa_id>/aprovar", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/etapa")
def api_etapa_aprovar(etapa_id):
    registro = db.session.get(CadastrarEtapa, etapa_id)
    if not registro or registro.ativo is False:
        return jsonify({"error": "Registro nao encontrado."}), 404
    status_atual = (registro.status_aprovacao or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Etapa ja foi processada."}), 400
    try:
        nivel_int = int(getattr(g, "user_nivel", 0) or 0)
    except (TypeError, ValueError):
        nivel_int = 0
    if nivel_int not in (1, 2):
        return jsonify({"error": "Usuario sem permissao para aprovar a etapa atual."}), 403
    data = request.get_json() or {}
    aprovado = (data.get("etapa_aprovada") or "sim").strip().lower()
    aprovado = aprovado if aprovado in {"sim", "nao"} else "sim"
    motivo = (data.get("motivo_rejeicao") or "").strip()
    if not motivo:
        return jsonify({"error": "Justificativa obrigatoria."}), 400
    novos_ids = []
    if aprovado == "sim":
        try:
            novos_ids = _apply_etapa_to_plan21(registro)
        except Exception as exc:
            db.session.rollback()
            return jsonify({"error": f"Falha ao aprovar etapa: {exc}"}), 500
        registro.status_aprovacao = "Aprovado"
        registro.situacao = "Aprovado"
        if novos_ids:
            registro.plan21_nger_id = novos_ids[-1]
            registro.plan21_nger_ids = json.dumps(novos_ids) if len(novos_ids) > 1 else None
    else:
        registro.status_aprovacao = "Rejeitado"
        registro.situacao = "Rejeitado"
        registro.ativo = False
        registro.excluido_em = _now_local()
    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400
    registro.aprovado_por = str(usuarios_id)
    registro.data_aprovacao = _now_local()
    registro.motivo_rejeicao = motivo
    registro.alterado_em = _now_local()
    try:
        db.session.commit()
        _cache_invalidate_etapa_options()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao aprovar etapa: {exc}"}), 500
    return jsonify({"ok": True, "message": "Etapa atualizada.", "etapa": _etapa_payload(registro)})


@home_bp.route("/api/subacao/options", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_options():
    current_year = str(_now_local().year)
    cols = set(_plan21_columns())
    plan_fields = {
        "exercicio": Plan21Nger.exercicio,
        "uo": Plan21Nger.uo,
        "programa": Plan21Nger.programa,
        "acao_paoe": Plan21Nger.acao_paoe,
        "responsavel_acao": Plan21Nger.responsavel_acao,
        "produto_acao": Plan21Nger.produto,
    }
    selected = {}
    for key in plan_fields:
        val = (request.args.get(key) or "").strip()
        if val:
            selected[key] = val
    # força o exercício ao ano atual
    selected["exercicio"] = current_year
    plan_options = {}
    for key, col in plan_fields.items():
        query = db.session.query(col).distinct()
        for s_key, s_val in selected.items():
            if s_key == key:
                continue
            query = query.filter(plan_fields[s_key] == s_val)
        rows = query.all()
        values = []
        for (val,) in rows:
            if val is None:
                continue
            s = str(val).strip()
            if s:
                values.append(s)
        if key == "exercicio":
            plan_options[key] = [current_year]
        else:
            plan_options[key] = sorted(set(values), key=lambda v: v.lower())

    option_rows = []
    option_select_cols = [
        "id",
        "exercicio",
        "unidade_orcamentaria",
        "programa",
        "acao_paoe",
        "responsavel_acao",
        "produto_acao",
        "subfuncao_ug",
        "adj",
        "macropolitica",
        "pilar",
        "eixo",
        "politica_decreto",
        "regiao_produto",
    ]
    option_select_parts = [c for c in option_select_cols if c in cols]
    option_where = ["ativo = 1"] if "ativo" in cols else []
    option_params = {}
    if "exercicio" in cols:
        option_where.append("exercicio = :exercicio")
        option_params["exercicio"] = current_year
    option_sql = f"SELECT {', '.join(option_select_parts)} FROM plan21_nger"
    if option_where:
        option_sql += " WHERE " + " AND ".join(option_where)
    option_sql += " ORDER BY id"
    for row in _safe_query_mappings(option_sql, option_params):
        option_rows.append(
            {
                "id": row.get("id"),
                "exercicio": str(row.get("exercicio") or "").strip(),
                "uo": str(row.get("unidade_orcamentaria") or "").strip(),
                "programa": str(row.get("programa") or "").strip(),
                "acao_paoe": str(row.get("acao_paoe") or "").strip(),
                "responsavel_acao": str(row.get("responsavel_acao") or "").strip(),
                "produto_acao": str(row.get("produto_acao") or "").strip(),
                "subfuncao_ug": str(row.get("subfuncao_ug") or "").strip(),
                "adj": str(row.get("adj") or "").strip(),
                "macropolitica": str(row.get("macropolitica") or "").strip(),
                "pilar": str(row.get("pilar") or "").strip(),
                "eixo": str(row.get("eixo") or "").strip(),
                "politica_decreto": str(row.get("politica_decreto") or "").strip(),
                "regiao_produto": str(row.get("regiao_produto") or "").strip(),
            }
        )

    # Ponte com plan21_nger para filtrar parte da chave do planejamento
    ponte_params = {}
    ponte_where = []
    ponte_map = {
        "exercicio": "exercicio",
        "uo": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "responsavel_acao": "responsavel_acao",
        "produto_acao": "produto_acao",
    }
    for key, col in ponte_map.items():
        val = (selected.get(key) or "").strip()
        if val:
            ponte_where.append(f"{col} = :{key}")
            ponte_params[key] = val
    ponte_sql = """
        SELECT DISTINCT subfuncao_ug, adj, macropolitica, pilar, eixo, politica_decreto
        FROM plan21_nger
    """
    if ponte_where:
        ponte_sql += " WHERE " + " AND ".join(ponte_where)
    ponte_rows = _safe_query_mappings(ponte_sql, ponte_params)
    allowed_subfuncao = set()
    allowed_ug = set()
    allowed_adj = set()
    allowed_macro = set()
    allowed_pilar = set()
    allowed_eixo = set()
    allowed_politica = set()
    for row in ponte_rows:
        sub_ug = (row.get("subfuncao_ug") or "").strip()
        if sub_ug:
            if "." in sub_ug:
                sub_part, ug_part = sub_ug.split(".", 1)
                sub_norm = _normalize_codigo_num(sub_part)
                ug_norm = _normalize_codigo_num(ug_part)
                if sub_norm:
                    allowed_subfuncao.add(sub_norm)
                if ug_norm:
                    allowed_ug.add(ug_norm)
            else:
                sub_norm = _normalize_codigo_num(sub_ug)
                if sub_norm:
                    allowed_subfuncao.add(sub_norm)
        adj_val = (row.get("adj") or "").strip()
        if adj_val:
            allowed_adj.add(adj_val)
        macro_val = (row.get("macropolitica") or "").strip()
        if macro_val:
            allowed_macro.add(macro_val)
        pilar_val = (row.get("pilar") or "").strip()
        if pilar_val:
            allowed_pilar.add(pilar_val)
        eixo_val = (row.get("eixo") or "").strip()
        if eixo_val:
            allowed_eixo.add(eixo_val)
        politica_val = (row.get("politica_decreto") or "").strip()
        if politica_val:
            allowed_politica.add(politica_val)

    regiao_where = []
    regiao_params = {}
    regiao_map = {
        "exercicio": "exercicio",
        "uo": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "produto_acao": "produto_acao",
    }
    for key, col in regiao_map.items():
        val = (selected.get(key) or "").strip()
        if val:
            regiao_where.append(f"{col} = :{key}")
            regiao_params[key] = val
    regiao_sql = "SELECT DISTINCT regiao_produto FROM plan21_nger"
    if regiao_where:
        regiao_sql += " WHERE " + " AND ".join(regiao_where)
    regiao_rows = _safe_query_mappings(regiao_sql, regiao_params)
    regioes = []
    for row in regiao_rows:
        val = row.get("regiao_produto")
        if val is None:
            continue
        text = str(val).strip()
        if not text:
            continue
        regioes.append({"value": text, "label": text})
    regioes = sorted(regioes, key=lambda r: r["label"].lower())

    subfuncao_rows = _safe_query_mappings(
        "SELECT codigo, nome FROM subfuncao ORDER BY codigo"
    )
    subfuncoes = _options_code_name(subfuncao_rows, "codigo", "nome")

    ug_rows = _safe_query_mappings("SELECT codigo, nome FROM ug ORDER BY codigo")
    ugs = _options_code_name(ug_rows, "codigo", "nome")

    subfuncao_selected = _normalize_codigo_num(request.args.get("subfuncao") or "")
    ug_selected = _normalize_codigo_num(request.args.get("ug") or "")
    if ug_selected:
        rows = _safe_query_mappings(
            "SELECT DISTINCT subfuncao FROM ug_subfuncao WHERE ug = :ug",
            {"ug": ug_selected},
        )
        allowed = {str(_normalize_codigo_num(r.get("subfuncao"))) for r in rows if r.get("subfuncao")}
        if allowed:
            subfuncoes = [s for s in subfuncoes if _normalize_codigo_num(s["value"]) in allowed]
    if subfuncao_selected:
        rows = _safe_query_mappings(
            "SELECT DISTINCT ug FROM ug_subfuncao WHERE subfuncao = :sub",
            {"sub": subfuncao_selected},
        )
        allowed = {str(_normalize_codigo_num(r.get("ug"))) for r in rows if r.get("ug")}
        if allowed:
            ugs = [u for u in ugs if _normalize_codigo_num(u["value"]) in allowed]
    if allowed_subfuncao:
        subfuncoes = [s for s in subfuncoes if _normalize_codigo_num(s["value"]) in allowed_subfuncao]
    if allowed_ug:
        ugs = [u for u in ugs if _normalize_codigo_num(u["value"]) in allowed_ug]

    adj_rows = _safe_query_mappings("SELECT abreviacao, nome FROM adj ORDER BY abreviacao")
    adjs = _options_abrev_nome(adj_rows, "abreviacao", "nome")
    if allowed_adj:
        adjs = [a for a in adjs if (a.get("value") or "") in allowed_adj]

    adj_selected = (request.args.get("adj") or "").strip()
    macropoliticas = []
    if adj_selected:
        macropoliticas = _options_abrev_nome(
            _safe_query_mappings(
                """
                SELECT m.abreviacao, m.nome
                FROM macropolitica m
                JOIN macropolitica_adj ma ON ma.macropolitica_id = m.id
                JOIN adj a ON a.id = ma.adj_id
                WHERE a.abreviacao = :adj
                ORDER BY m.abreviacao
                """,
                {"adj": adj_selected},
            ),
            "abreviacao",
            "nome",
        )
    if not macropoliticas:
        macropoliticas = _options_abrev_nome(
            _safe_query_mappings("SELECT abreviacao, nome FROM macropolitica ORDER BY abreviacao"),
            "abreviacao",
            "nome",
        )
    if allowed_macro:
        macropoliticas = [m for m in macropoliticas if (m.get("value") or "") in allowed_macro]

    pilar_rows = []
    macropolitica_selected = (request.args.get("macropolitica") or "").strip()
    if adj_selected or macropolitica_selected:
        base_sql = """
            SELECT DISTINCT p.abreviacao, p.nome
            FROM pilar p
            {joins}
            {where}
            ORDER BY p.abreviacao
        """
        joins = []
        wheres = []
        params = {}
        if adj_selected:
            joins.append("JOIN pilar_adj pa ON pa.pilar_id = p.id JOIN adj a ON a.id = pa.adj_id")
            wheres.append("a.abreviacao = :adj")
            params["adj"] = adj_selected
        if macropolitica_selected:
            joins.append("JOIN macropolitica_pilar mp ON mp.pilar_id = p.id JOIN macropolitica m ON m.id = mp.macropolitica_id")
            wheres.append("m.abreviacao = :macro")
            params["macro"] = macropolitica_selected
        sql = base_sql.format(
            joins=" ".join(joins),
            where=f"WHERE {' AND '.join(wheres)}" if wheres else "",
        )
        pilar_rows = _safe_query_mappings(sql, params)
    pilares = _options_abrev_nome(pilar_rows, "abreviacao", "nome")
    if not pilares:
        pilares = _options_abrev_nome(
            _safe_query_mappings("SELECT abreviacao, nome FROM pilar ORDER BY abreviacao"),
            "abreviacao",
            "nome",
        )
    if allowed_pilar:
        filtered_pilares = [p for p in pilares if (p.get("value") or "") in allowed_pilar]
        if filtered_pilares:
            pilares = filtered_pilares

    eixo_rows = []
    if adj_selected:
        eixo_rows = _safe_query_mappings(
            """
            SELECT e.abreviacao, e.nome
            FROM eixo e
            JOIN eixo_adj ea ON ea.eixo_id = e.id
            JOIN adj a ON a.id = ea.adj_id
            WHERE a.abreviacao = :adj
            ORDER BY e.abreviacao
            """,
            {"adj": adj_selected},
        )
    eixos = _options_abrev_nome(eixo_rows, "abreviacao", "nome")
    if not eixos:
        eixos = _options_abrev_nome(
            _safe_query_mappings("SELECT abreviacao, nome FROM eixo ORDER BY abreviacao"),
            "abreviacao",
            "nome",
        )
    if allowed_eixo:
        eixos = [e for e in eixos if (e.get("value") or "") in allowed_eixo]

    politica_rows = []
    eixo_selected = (request.args.get("eixo") or "").strip()
    if adj_selected or eixo_selected:
        base_sql = """
            SELECT DISTINCT p.abreviacao, p.nome
            FROM politica_decr p
            {joins}
            {where}
            ORDER BY p.abreviacao
        """
        joins = []
        wheres = []
        params = {}
        if adj_selected:
            joins.append("JOIN politica_decr_adj pa ON pa.politica_decr_id = p.id JOIN adj a ON a.id = pa.adj_id")
            wheres.append("a.abreviacao = :adj")
            params["adj"] = adj_selected
        if eixo_selected:
            joins.append("JOIN politica_decr_eixo pe ON pe.politica_decr_id = p.id JOIN eixo e ON e.id = pe.eixo_id")
            wheres.append("e.abreviacao = :eixo")
            params["eixo"] = eixo_selected
        sql = base_sql.format(
            joins=" ".join(joins),
            where=f"WHERE {' AND '.join(wheres)}" if wheres else "",
        )
        politica_rows = _safe_query_mappings(sql, params)
    politicas = _options_abrev_nome(politica_rows, "abreviacao", "nome")
    if not politicas:
        politicas = _options_abrev_nome(
            _safe_query_mappings("SELECT abreviacao, nome FROM politica_decr ORDER BY abreviacao"),
            "abreviacao",
            "nome",
        )
    if allowed_politica:
        politicas = [p for p in politicas if (p.get("value") or "") in allowed_politica]

    publico_rows = _safe_query_mappings(
        "SELECT codigo, nome FROM publico_transversal ORDER BY codigo"
    )
    publicos = _options_code_name(publico_rows, "codigo", "nome")

    municipio_params = {}
    municipio_where = []
    regiao_subacao = _normalize_codigo_num(request.args.get("regiao_subacao") or "")
    if regiao_subacao:
        reg_row = _safe_query_mappings(
            "SELECT id FROM regiao WHERE codigo = :codigo",
            {"codigo": regiao_subacao},
        )
        reg_id = reg_row[0]["id"] if reg_row else None
        if reg_id is not None:
            municipio_where.append("m.regiao_id = :regiao_id")
            municipio_params["regiao_id"] = reg_id
    codigo_municipio = _normalize_codigo_num(request.args.get("codigo") or "")
    if codigo_municipio:
        municipio_where.append("m.codigo_ibge = :codigo_ibge")
        municipio_params["codigo_ibge"] = codigo_municipio
    municipio_sql = "SELECT m.codigo_ibge, m.nome FROM municipio m"
    if municipio_where:
        municipio_sql += " WHERE " + " AND ".join(municipio_where)
    municipio_sql += " ORDER BY m.nome"
    municipio_rows = _safe_query_mappings(municipio_sql, municipio_params)
    municipios = []
    for row in municipio_rows:
        codigo = row.get("codigo_ibge")
        nome = row.get("nome")
        if codigo is None or nome is None:
            continue
        municipios.append(
            {
                "codigo": str(codigo),
                "nome": str(nome),
                "label": f"{codigo} - {nome}",
            }
        )

    responsaveis_nger = []
    usuarios_nger = Usuario.query.filter(Usuario.perfil_id == 2).order_by(Usuario.nome).all()
    for usuario in usuarios_nger:
        nome = (getattr(usuario, "nome", "") or "").strip()
        email = (getattr(usuario, "email", "") or "").strip()
        label = nome or email
        if label:
            responsaveis_nger.append(label)

    return jsonify(
        {
            "plan21": plan_options,
            "regioes": regioes,
            "subfuncoes": subfuncoes,
            "ugs": ugs,
            "adjs": adjs,
            "macropoliticas": macropoliticas,
            "pilares": pilares,
            "eixos": eixos,
            "politicas": politicas,
            "publicos": publicos,
            "municipios": municipios,
            "responsaveis_nger": responsaveis_nger,
            "option_rows": option_rows,
        }
    )


def _subacao_payload(registro: CadastrarSubacao) -> dict:
    return {
        "id": registro.id,
        "exercicio": registro.exercicio,
        "unidade_orcamentaria": registro.unidade_orcamentaria,
        "programa": registro.programa,
        "acao_paoe": registro.acao_paoe,
        "responsavel_acao": registro.responsavel_acao,
        "produto_acao": registro.produto_acao,
        "chave_planejamento": registro.chave_planejamento,
        "subacao_entrega": registro.subacao_entrega,
        "responsavel": registro.responsavel,
        "cpf_responsavel": registro.cpf_responsavel,
        "prazo": registro.prazo,
        "unid_gestora": registro.unid_gestora,
        "unidade_setorial_planejamento": registro.unidade_setorial_planejamento,
        "produto_subacao": registro.produto_subacao,
        "unidade_medida": registro.unidade_medida,
        "regiao_subacao": registro.regiao_subacao,
        "codigo": registro.codigo,
        "municipios_entrega": registro.municipios_entrega,
        "meta_subacao": str(registro.meta_subacao or ""),
        "detalhamento_produto": registro.detalhamento_produto,
        "etapa": registro.etapa,
        "responsavel_etapa": registro.responsavel_etapa,
        "cpf_responsavel_etapa": registro.cpf_responsavel_etapa,
        "justificativa": registro.justificativa,
        "responsavel_nger": registro.responsavel_nger,
        "tipo_solicitacao": registro.tipo_solicitacao,
        "status_aprovacao": registro.status_aprovacao,
        "tipo_edicao": registro.tipo_edicao,
    }


def _format_etapa_prazo(data_inicio, data_fim) -> str:
    inicio = data_inicio.strftime("%d/%m/%Y") if hasattr(data_inicio, "strftime") else str(data_inicio or "").strip()
    fim = data_fim.strftime("%d/%m/%Y") if hasattr(data_fim, "strftime") else str(data_fim or "").strip()
    return f"{inicio} à {fim}".strip()


def _parse_etapa_date(value: str):
    token = (value or "").strip()
    if not token:
        return None
    try:
        return datetime.strptime(token, "%Y-%m-%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(token, "%d/%m/%Y").date()
    except ValueError:
        return None


def _split_etapa_prazo(value: str | None) -> tuple:
    token = str(value or "").strip()
    if not token:
        return None, None
    parts = re.split(r"\s+(?:à|a|ate|até)\s+", token, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) != 2:
        return None, None
    return _parse_etapa_date(parts[0]), _parse_etapa_date(parts[1])


def _extract_etapa_plan21_ids(registro: CadastrarEtapa) -> list[int]:
    ids = []
    raw_ids = registro.plan21_nger_ids
    if raw_ids:
        try:
            parsed = json.loads(raw_ids) if isinstance(raw_ids, str) else raw_ids
            if isinstance(parsed, list):
                for val in parsed:
                    try:
                        num = int(val)
                        if num > 0:
                            ids.append(num)
                    except (TypeError, ValueError):
                        continue
        except (TypeError, ValueError):
            pass
    if not ids and registro.plan21_nger_id:
        try:
            num = int(registro.plan21_nger_id)
            if num > 0:
                ids.append(num)
        except (TypeError, ValueError):
            pass
    return ids


def _etapa_payload(registro: CadastrarEtapa) -> dict:
    return {
        "id": registro.id,
        "exercicio": registro.exercicio,
        "unidade_orcamentaria": registro.unidade_orcamentaria,
        "programa": registro.programa,
        "acao_paoe": registro.acao_paoe,
        "responsavel_acao": registro.responsavel_acao,
        "produto_acao": registro.produto_acao,
        "adj_solicitante": registro.adj_solicitante,
        "subacao_entrega": registro.subacao_entrega,
        "chave_planejamento": registro.chave_planejamento,
        "tipo_solicitacao": registro.tipo_solicitacao,
        "plan21_nger_id": registro.plan21_nger_id,
        "plan21_nger_ids": registro.plan21_nger_ids,
        "etapa_origem": registro.etapa_origem,
        "responsavel_etapa_origem": registro.responsavel_etapa_origem,
        "data_inicio_origem": registro.data_inicio_origem.isoformat() if registro.data_inicio_origem else "",
        "data_fim_origem": registro.data_fim_origem.isoformat() if registro.data_fim_origem else "",
        "etapa_nova": registro.etapa_nova,
        "responsavel_etapa_novo": registro.responsavel_etapa_novo,
        "cpf_responsavel_etapa_novo": registro.cpf_responsavel_etapa_novo,
        "email_responsavel_etapa_novo": registro.email_responsavel_etapa_novo,
        "data_inicio_novo": registro.data_inicio_novo.isoformat() if registro.data_inicio_novo else "",
        "data_fim_novo": registro.data_fim_novo.isoformat() if registro.data_fim_novo else "",
        "justificativa": registro.justificativa,
        "responsavel_nger": registro.responsavel_nger,
        "status_aprovacao": registro.status_aprovacao,
        "situacao": registro.situacao,
        "criado_em": registro.criado_em.isoformat() if getattr(registro, "criado_em", None) else "",
        "data_aprovacao": registro.data_aprovacao.isoformat() if getattr(registro, "data_aprovacao", None) else "",
        "motivo_rejeicao": registro.motivo_rejeicao or "",
    }


def _load_etapa_plan21_rows(filters: dict, etapa: str | None = None, source_id: int | None = None) -> list[dict]:
    cols = set(_plan21_columns())
    select_cols = [
        "id",
        "exercicio",
        "unidade_orcamentaria",
        "programa",
        "acao_paoe",
        "responsavel_acao",
        "produto_acao",
        "chave_planejamento",
        "subacao_entrega",
        "codigo",
        "municipios_entrega",
        "etapa",
        "responsavel_etapa",
    ]
    select_parts = [c for c in select_cols if c in cols]
    select_parts.append("prazo_etapa" if "prazo_etapa" in cols else "NULL AS prazo_etapa")
    where = ["ativo = 1"] if "ativo" in cols else []
    params = {}
    field_map = {
        "exercicio": "exercicio",
        "unidade_orcamentaria": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "responsavel_acao": "responsavel_acao",
        "produto_acao": "produto_acao",
        "subacao_entrega": "subacao_entrega",
        "chave_planejamento": "chave_planejamento",
    }
    for key, col in field_map.items():
        val = (filters.get(key) or "").strip()
        if val and col in cols:
            where.append(f"{col} = :{key}")
            params[key] = val
    if etapa and "etapa" in cols:
        where.append("etapa = :etapa")
        params["etapa"] = etapa
    if source_id:
        where.append("id = :source_id")
        params["source_id"] = source_id
    sql = f"SELECT {', '.join(select_parts)} FROM plan21_nger"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id"
    return [dict(r) for r in _safe_query_mappings(sql, params)]


def _etapa_nullify_cols() -> set:
    return {
        "regiao_etapa",
        "natureza",
        "fonte",
        "idu",
        "descricao_item_despesa",
        "unid_medida_item",
        "quantidade",
        "valor_unitario",
        "valor_total",
        "suplementacao",
        "anulacao",
        "valor_atual",
        "cat_econ",
        "grupo",
        "modalidade",
        "elemento",
        "subelemento",
    }


def _plan21_cols_before_etapa() -> list[str]:
    cols = _plan21_columns()
    if "etapa" not in cols:
        return [col for col in cols if col not in {"id", "ativo"}]
    return [col for col in cols[: cols.index("etapa")] if col not in {"id", "ativo"}]


def _etapa_clear_cols() -> set[str]:
    cols = set(_plan21_columns())
    etapa_cols = {
        "etapa",
        "responsavel_etapa",
        "cpf_responsavel_etapa",
        "email_responsavel_etapa",
        "prazo_etapa",
        "regiao_etapa",
    }
    return {col for col in etapa_cols if col in cols}.union(_etapa_memoria_cols())


def _normalize_plan21_signature_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        return " ".join(value.replace("\xa0", " ").split())
    return value


def _plan21_before_etapa_signature(row: dict, signature_cols: list[str]) -> tuple:
    return tuple((col, _normalize_plan21_signature_value(row.get(col))) for col in signature_cols)


def _load_plan21_full_rows_by_ids(ids: list[int]) -> list[dict]:
    clean_ids = []
    for val in ids or []:
        try:
            num = int(val)
            if num > 0:
                clean_ids.append(num)
        except (TypeError, ValueError):
            continue
    clean_ids = sorted(set(clean_ids))
    if not clean_ids:
        return []
    params = {}
    placeholders = []
    for idx, row_id in enumerate(clean_ids):
        key = f"id_{idx}"
        params[key] = row_id
        placeholders.append(f":{key}")
    cols = _plan21_columns()
    sql = (
        f"SELECT {', '.join(f'`{col}`' for col in cols)} "
        f"FROM plan21_nger WHERE id IN ({', '.join(placeholders)}) ORDER BY id"
    )
    return [dict(row) for row in db.session.execute(text(sql), params).mappings().all()]


def _plan21_empty_etapa_where(cols: set[str]) -> list[str]:
    conditions = []
    for col in ("etapa", "responsavel_etapa", "prazo_etapa", "regiao_etapa"):
        if col in cols:
            conditions.append(_plan21_empty_marker_sql(col))
    return conditions


def _plan21_empty_marker_sql(col: str) -> str:
    raw = f"TRIM(CAST(`{col}` AS CHAR))"
    normalized = (
        f"REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE({raw}, CHAR(160), ''), "
        "' ', ''), '.', ''), ',', ''), '/', ''), '-', '')"
    )
    return (
        f"(`{col}` IS NULL OR {raw} = '' OR {raw} IN ('-', '–', '—') "
        f"OR {normalized} REGEXP '^0+$')"
    )


def _plan21_empty_memory_where(cols: set[str]) -> list[str]:
    conditions = []
    for col in _etapa_memoria_cols():
        if col in cols:
            conditions.append(_plan21_empty_marker_sql(col))
    return conditions


def _active_base_row_without_etapa_exists(source_row: dict, signature_cols: list[str]) -> bool:
    cols = set(_plan21_columns())
    where = ["`ativo` = 1"] if "ativo" in cols else []
    where.extend(_plan21_empty_etapa_where(cols))
    where.extend(_plan21_empty_memory_where(cols))
    params = {}
    for idx, col in enumerate(signature_cols):
        key = f"sig_{idx}"
        value = source_row.get(col)
        if value is None:
            where.append(f"`{col}` IS NULL")
        else:
            where.append(f"`{col}` = :{key}")
            params[key] = value
    if source_row.get("id") is not None:
        where.append("`id` <> :source_id")
        params["source_id"] = source_row.get("id")
    sql = "SELECT id FROM plan21_nger"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " LIMIT 1"
    return db.session.execute(text(sql), params).first() is not None


def _clone_etapa_exclusao_base_rows(rows: list[dict]) -> list[int]:
    if not rows:
        return []
    signature_cols = _plan21_cols_before_etapa()
    grouped: dict[tuple, dict] = {}
    for row in rows:
        row_id = row.get("id")
        if not row_id:
            continue
        signature = _plan21_before_etapa_signature(row, signature_cols)
        if signature not in grouped:
            grouped[signature] = row
    nullify = _etapa_clear_cols()
    overrides = {"ativo": True}
    new_ids = []
    for row in grouped.values():
        if _active_base_row_without_etapa_exists(row, signature_cols):
            continue
        new_id = _clone_plan21_row(int(row["id"]), overrides=overrides, nullify=nullify)
        if new_id:
            new_ids.append(new_id)
    return new_ids


def _etapa_memoria_cols() -> list[str]:
    cols = _plan21_columns()
    memoria_cols = [
        "regiao_etapa",
        "natureza",
        "cat_econ",
        "grupo",
        "modalidade",
        "elemento",
        "subelemento",
        "fonte",
        "idu",
        "descricao_item_despesa",
        "unid_medida_item",
        "quantidade",
        "valor_unitario",
        "valor_total",
        "suplementacao",
        "anulacao",
        "valor_atual",
    ]
    return [col for col in memoria_cols if col in cols]


def _subacao_cols() -> list[str]:
    cols = set(_plan21_columns())
    subacao_cols = [
        "subacao_entrega",
        "responsavel",
        "prazo",
        "unid_gestora",
        "unidade_setorial_planejamento",
        "produto_subacao",
        "unidade_medida",
        "regiao_subacao",
        "codigo",
        "municipios_entrega",
        "meta_subacao",
        "detalhamento_produto",
    ]
    return [col for col in subacao_cols if col in cols]


def _subacao_etapa_cols() -> list[str]:
    cols = set(_plan21_columns())
    etapa_cols = ["etapa", "responsavel_etapa", "prazo_etapa"]
    return [col for col in etapa_cols if col in cols]


def _subacao_clear_cols() -> set[str]:
    return set(_subacao_cols()).union(_subacao_etapa_cols()).union(_etapa_memoria_cols())


def _plan21_cols_before_subacao() -> list[str]:
    cols = _plan21_columns()
    if "subacao_entrega" not in cols:
        return [col for col in cols if col not in {"id", "ativo"}]
    return [col for col in cols[: cols.index("subacao_entrega")] if col not in {"id", "ativo"}]


def _plan21_empty_subacao_where(cols: set[str]) -> list[str]:
    conditions = []
    for col in _subacao_cols():
        if col in cols:
            conditions.append(_plan21_empty_marker_sql(col))
    return conditions


def _active_base_row_without_subacao_exists(source_row: dict, signature_cols: list[str]) -> bool:
    cols = set(_plan21_columns())
    where = ["`ativo` = 1"] if "ativo" in cols else []
    where.extend(_plan21_empty_subacao_where(cols))
    where.extend(_plan21_empty_etapa_where(cols))
    where.extend(_plan21_empty_memory_where(cols))
    params = {}
    for idx, col in enumerate(signature_cols):
        key = f"sig_sub_{idx}"
        value = source_row.get(col)
        if value is None:
            where.append(f"`{col}` IS NULL")
        else:
            where.append(f"`{col}` = :{key}")
            params[key] = value
    if source_row.get("id") is not None:
        where.append("`id` <> :source_id")
        params["source_id"] = source_row.get("id")
    sql = "SELECT id FROM plan21_nger"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " LIMIT 1"
    return db.session.execute(text(sql), params).first() is not None


def _clone_subacao_exclusao_base_rows(rows: list[dict]) -> list[int]:
    if not rows:
        return []
    signature_cols = _plan21_cols_before_subacao()
    grouped: dict[tuple, dict] = {}
    for row in rows:
        row_id = row.get("id")
        if not row_id:
            continue
        signature = _plan21_before_etapa_signature(row, signature_cols)
        if signature not in grouped:
            grouped[signature] = row
    nullify = _subacao_clear_cols()
    overrides = {"ativo": True}
    new_ids = []
    for row in grouped.values():
        if _active_base_row_without_subacao_exists(row, signature_cols):
            continue
        new_id = _clone_plan21_row(int(row["id"]), overrides=overrides, nullify=nullify)
        if new_id:
            new_ids.append(new_id)
    return new_ids


def _value_has_content(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return value != 0
    if isinstance(value, (bytes, bytearray)):
        try:
            value = value.decode("utf-8", errors="ignore")
        except Exception:
            return bool(value)
    if isinstance(value, str):
        token = value.replace("\xa0", " ").strip()
        if not token:
            return False
        if token in {"-", "–", "—"} or set(token) <= {"-", "–", "—"}:
            return False
        compact = re.sub(r"[\s\.\,\-_/]", "", token)
        if compact and set(compact) == {"0"}:
            return False
        number_token = token.replace(".", "").replace(",", ".")
        try:
            return Decimal(number_token) != 0
        except Exception:
            return True
    return True


def _validate_etapa_exclusao_sem_memoria(rows: list[dict], etapa_nome: str | None = None) -> str | None:
    ids = []
    for row in rows or []:
        try:
            ids.append(int(row.get("id")))
        except (TypeError, ValueError):
            continue
    ids = sorted(set(i for i in ids if i > 0))
    if not ids:
        return None
    memoria_cols = _etapa_memoria_cols()
    if not memoria_cols:
        return None
    placeholders = []
    params = {}
    for idx, row_id in enumerate(ids):
        key = f"id_{idx}"
        placeholders.append(f":{key}")
        params[key] = row_id
    select_cols = ["id", *memoria_cols]
    sql = (
        f"SELECT {', '.join(f'`{col}`' for col in select_cols)} "
        f"FROM plan21_nger WHERE id IN ({', '.join(placeholders)})"
    )
    result_rows = db.session.execute(text(sql), params).mappings().all()
    for row in result_rows:
        if any(_value_has_content(row.get(col)) for col in memoria_cols):
            etapa = (etapa_nome or "").strip()
            if not etapa:
                etapa = next((str(src.get("etapa") or "").strip() for src in rows if src.get("etapa")), "")
            return (
                f'Não é possível excluir a Etapa "{etapa}", pois existem Memórias de Cálculo vinculadas a ela. '
                "Para realizar a exclusão, exclua previamente todas as Memórias de Cálculo associadas a esta Etapa."
            )
    return None


def _validate_subacao_exclusao_sem_vinculos(rows: list[dict], subacao_nome: str | None = None) -> str | None:
    ids = []
    for row in rows or []:
        try:
            ids.append(int(row.get("id")))
        except (TypeError, ValueError):
            continue
    ids = sorted(set(i for i in ids if i > 0))
    if not ids:
        return None
    check_cols = [*_subacao_etapa_cols(), *_etapa_memoria_cols()]
    if not check_cols:
        return None
    placeholders = []
    params = {}
    for idx, row_id in enumerate(ids):
        key = f"id_{idx}"
        placeholders.append(f":{key}")
        params[key] = row_id
    select_cols = ["id", *check_cols]
    sql = (
        f"SELECT {', '.join(f'`{col}`' for col in select_cols)} "
        f"FROM plan21_nger WHERE id IN ({', '.join(placeholders)})"
    )
    result_rows = db.session.execute(text(sql), params).mappings().all()
    etapa_cols = set(_subacao_etapa_cols())
    memoria_cols = set(_etapa_memoria_cols())
    subacao = (subacao_nome or "").strip()
    if not subacao:
        subacao = next((str(src.get("subacao_entrega") or "").strip() for src in rows if src.get("subacao_entrega")), "")
    for row in result_rows:
        if any(_value_has_content(row.get(col)) for col in etapa_cols):
            return (
                f'Não é possível excluir a Subação "{subacao}", pois existem Etapas vinculadas a ela. '
                "Para realizar a exclusão, exclua previamente todas as Etapas associadas a esta Subação."
            )
        if any(_value_has_content(row.get(col)) for col in memoria_cols):
            return (
                f'Não é possível excluir a Subação "{subacao}", pois existem Memórias de Cálculo vinculadas a ela. '
                "Para realizar a exclusão, exclua previamente todas as Memórias de Cálculo associadas a esta Subação."
            )
    return None


def _parse_json_list(value) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, (tuple, set)):
        return list(value)
    token = str(value).strip()
    if not token:
        return []
    if token.startswith("[") and token.endswith("]"):
        try:
            parsed = json.loads(token)
            if isinstance(parsed, list):
                return parsed
        except (TypeError, ValueError):
            pass
    if "*" in token:
        parts = [p.strip() for p in token.split("*")]
        return [p for p in parts if p]
    return [token]


def _pick_list_value(values: list, idx: int) -> str:
    if not values:
        return ""
    if idx < len(values):
        return "" if values[idx] is None else str(values[idx]).strip()
    return "" if values[0] is None else str(values[0]).strip()


def _extract_plan21_ids(registro: CadastrarSubacao) -> list[int]:
    ids = []
    raw_ids = registro.plan21_nger_ids
    if raw_ids:
        try:
            parsed = json.loads(raw_ids) if isinstance(raw_ids, str) else raw_ids
            if isinstance(parsed, list):
                for val in parsed:
                    try:
                        num = int(val)
                        if num > 0:
                            ids.append(num)
                    except (TypeError, ValueError):
                        continue
        except (TypeError, ValueError):
            pass
    if not ids and registro.plan21_nger_id:
        try:
            num = int(registro.plan21_nger_id)
            if num > 0:
                ids.append(num)
        except (TypeError, ValueError):
            pass
    return ids


@lru_cache(maxsize=1)
def _plan21_columns() -> list[str]:
    insp = inspect(db.engine)
    cols = [col["name"] for col in insp.get_columns("plan21_nger")]
    return cols


def _clone_plan21_row(source_id: int, overrides: dict | None = None, nullify: set | None = None) -> int:
    cols = [c for c in _plan21_columns() if c != "id"]
    overrides = overrides or {}
    nullify = nullify or set()
    override_keys = {k for k in overrides.keys() if k in cols}
    nullify_keys = {k for k in nullify if k in cols and k not in override_keys}
    select_parts = []
    params = {"source_id": source_id}
    for col in cols:
        if col in override_keys:
            select_parts.append(f":{col} AS {col}")
            params[col] = overrides[col]
        elif col in nullify_keys:
            select_parts.append(f"NULL AS {col}")
        else:
            select_parts.append(col)
    sql = f"INSERT INTO plan21_nger ({', '.join(cols)}) SELECT {', '.join(select_parts)} FROM plan21_nger WHERE id = :source_id"
    result = db.session.execute(text(sql), params)
    new_id = result.lastrowid
    if not new_id:
        try:
            new_id = db.session.execute(text("SELECT LAST_INSERT_ID()")).scalar()
        except Exception:
            new_id = None
    return int(new_id) if new_id else 0


def _subacao_list_payload(registro: CadastrarSubacao) -> dict:
    return {
        "codigos": _parse_json_list(registro.codigo),
        "municipios": _parse_json_list(registro.municipios_entrega),
        "metas": _parse_json_list(registro.meta_subacao),
        "etapas": _parse_json_list(registro.etapa),
        "responsaveis_etapa": _parse_json_list(registro.responsavel_etapa),
    }


def _parse_chave_planejamento_parts(chave_planejamento: str | None) -> dict:
    raw = str(chave_planejamento or "")
    parts = [part.replace("\xa0", " ").strip() for part in raw.split("*")]
    parts = [part for part in parts if part]
    columns = [
        "regiao",
        "subfuncao_ug",
        "adj",
        "macropolitica",
        "pilar",
        "eixo",
        "politica_decreto",
        "publico_transversal_chave",
    ]
    parsed = {}
    for idx, col in enumerate(columns):
        if idx >= len(parts):
            break
        value = (parts[idx] or "").strip()
        if value:
            parsed[col] = value
    return parsed


def _build_subacao_override(registro: CadastrarSubacao, idx: int) -> dict:
    lists = _subacao_list_payload(registro)
    meta_raw = _pick_list_value(lists["metas"], idx)
    meta_val = _parse_decimal(meta_raw) if meta_raw else None
    overrides = {
        "exercicio": registro.exercicio,
        "unidade_orcamentaria": registro.unidade_orcamentaria,
        "programa": registro.programa,
        "acao_paoe": registro.acao_paoe,
        "responsavel_acao": registro.responsavel_acao,
        "produto_acao": registro.produto_acao,
        "chave_planejamento": registro.chave_planejamento,
        "subacao_entrega": registro.subacao_entrega,
        "responsavel": registro.responsavel,
        "prazo": registro.prazo,
        "unid_gestora": registro.unid_gestora,
        "unidade_setorial_planejamento": registro.unidade_setorial_planejamento,
        "produto_subacao": registro.produto_subacao,
        "unidade_medida": registro.unidade_medida,
        "regiao_subacao": registro.regiao_subacao,
        "codigo": _pick_list_value(lists["codigos"], idx),
        "municipios_entrega": _pick_list_value(lists["municipios"], idx),
        "meta_subacao": meta_val,
        "detalhamento_produto": registro.detalhamento_produto,
        "etapa": _pick_list_value(lists["etapas"], idx) or registro.etapa,
        "responsavel_etapa": _pick_list_value(lists["responsaveis_etapa"], idx) or registro.responsavel_etapa,
        "ativo": True,
    }
    overrides.update(_parse_chave_planejamento_parts(registro.chave_planejamento))
    return overrides


def _apply_subacao_cadastrar(registro: CadastrarSubacao) -> list[int]:
    ids = _extract_plan21_ids(registro)
    if not ids:
        raise ValueError("Referencia plan21_nger nao encontrada para cadastrar.")
    lists = _subacao_list_payload(registro)
    list_sizes = [
        len(lists["codigos"]),
        len(lists["municipios"]),
        len(lists["metas"]),
        len(lists["etapas"]),
        len(lists["responsaveis_etapa"]),
    ]
    items_count = max([1] + list_sizes)
    source_ids = ids if len(ids) > 1 else [ids[0]] * items_count
    if len(source_ids) < items_count:
        source_ids = [source_ids[min(i, len(source_ids) - 1)] for i in range(items_count)]
    nullify_cols = {
        "regiao_etapa",
        "natureza",
        "fonte",
        "idu",
        "descricao_item_despesa",
        "unid_medida_item",
        "quantidade",
        "valor_unitario",
        "valor_total",
        "suplementacao",
        "anulacao",
        "valor_atual",
        "cat_econ",
        "grupo",
        "modalidade",
        "elemento",
        "subelemento",
    }
    new_ids = []
    for idx in range(items_count):
        overrides = _build_subacao_override(registro, idx)
        new_id = _clone_plan21_row(source_ids[idx], overrides=overrides, nullify=nullify_cols)
        if new_id:
            new_ids.append(new_id)
    return new_ids


def _apply_subacao_alterar(registro: CadastrarSubacao) -> list[int]:
    ids = _extract_plan21_ids(registro)
    if not ids:
        raise ValueError("Referencia plan21_nger nao encontrada para alterar.")
    tipo_edicao = (registro.tipo_edicao or "").strip().lower()
    lists = _subacao_list_payload(registro)
    list_sizes = [len(lists["codigos"]), len(lists["municipios"]), len(lists["metas"])]
    items_count = max([1] + list_sizes)
    if tipo_edicao in {"novo_municipio", "remover_municipio"} and items_count > 1 and len(ids) == 1:
        source_ids = [ids[0]] * items_count
    else:
        source_ids = ids
    if tipo_edicao in {"novo_municipio", "remover_municipio"} and items_count > len(source_ids):
        source_ids = [source_ids[min(i, len(source_ids) - 1)] for i in range(items_count)]

    def build_overrides(idx: int) -> dict:
        overrides = {"ativo": True}
        if tipo_edicao == "subacao_name":
            overrides.update(
                {
                    "subacao_entrega": registro.subacao_entrega,
                    "responsavel": registro.responsavel,
                }
            )
        elif tipo_edicao == "responsavel_name":
            overrides.update({"responsavel": registro.responsavel})
        elif tipo_edicao == "produto_subacao":
            overrides.update(
                {
                    "produto_subacao": registro.produto_subacao,
                    "unidade_medida": registro.unidade_medida,
                }
            )
        elif tipo_edicao in {"novo_municipio", "remover_municipio"}:
            meta_raw = _pick_list_value(lists["metas"], idx)
            meta_val = _parse_decimal(meta_raw) if meta_raw else None
            overrides.update(
                {
                    "codigo": _pick_list_value(lists["codigos"], idx),
                    "municipios_entrega": _pick_list_value(lists["municipios"], idx),
                    "meta_subacao": meta_val,
                }
            )
        return overrides

    new_ids = []
    for idx, source_id in enumerate(source_ids):
        overrides = build_overrides(idx)
        new_id = _clone_plan21_row(source_id, overrides=overrides, nullify=set())
        if new_id:
            new_ids.append(new_id)
    if ids:
        Plan21Nger.query.filter(Plan21Nger.id.in_(ids)).update(
            {"ativo": False}, synchronize_session=False
        )
    return new_ids


def _apply_subacao_excluir(registro: CadastrarSubacao) -> list[int]:
    ids = _extract_plan21_ids(registro)
    if not ids:
        raise ValueError("Referencia plan21_nger nao encontrada para excluir.")
    rows = _load_plan21_full_rows_by_ids(ids)
    err = _validate_subacao_exclusao_sem_vinculos(rows, registro.subacao_entrega)
    if err:
        raise ValueError(err)
    new_ids = _clone_subacao_exclusao_base_rows(rows)
    Plan21Nger.query.filter(Plan21Nger.id.in_(ids)).update(
        {"ativo": False}, synchronize_session=False
    )
    return new_ids


def _load_plan21_edit_rows(filters: dict) -> list[dict]:
    plan_fields = {
        "exercicio": "exercicio",
        "uo": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "responsavel_acao": "responsavel_acao",
        "produto_acao": "produto_acao",
    }
    edit_fields = {
        "chave_planejamento": "chave_planejamento",
        "subacao_entrega": "subacao_entrega",
        "responsavel": "responsavel",
        "prazo": "prazo",
        "unid_gestora": "unid_gestora",
        "unidade_setorial_planejamento": "unidade_setorial_planejamento",
        "produto_subacao": "produto_subacao",
        "unidade_medida": "unidade_medida",
        "regiao_subacao": "regiao_subacao",
        "codigo": "codigo",
        "municipios_entrega": "municipios_entrega",
        "meta_subacao": "meta_subacao",
        "detalhamento_produto": "detalhamento_produto",
    }
    where = ["ativo = 1"]
    params = {}
    for key, col in plan_fields.items():
        val = (filters.get(key) or "").strip()
        if val:
            where.append(f"{col} = :{key}")
            params[key] = val
    for key, col in edit_fields.items():
        val = (filters.get(key) or "").strip()
        if val:
            where.append(f"{col} = :{key}")
            params[key] = val
    sql = """
        SELECT
          id,
          exercicio,
          unidade_orcamentaria,
          programa,
          acao_paoe,
          responsavel_acao,
          produto_acao,
          chave_planejamento,
          subacao_entrega,
          responsavel,
          prazo,
          unid_gestora,
          unidade_setorial_planejamento,
          produto_subacao,
          unidade_medida,
          regiao_subacao,
          codigo,
          municipios_entrega,
          meta_subacao,
          etapa,
          responsavel_etapa,
          prazo_etapa,
          detalhamento_produto,
          valor_atual
        FROM plan21_nger
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"
    return _safe_query_mappings(sql, params)


@home_bp.route("/api/subacao/plan21-edit-options", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_plan21_edit_options():
    current_year = str(_now_local().year)
    plan_fields = {
        "exercicio": "exercicio",
        "uo": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "responsavel_acao": "responsavel_acao",
        "produto_acao": "produto_acao",
    }
    edit_fields = {
        "chave_planejamento": "chave_planejamento",
        "subacao_entrega": "subacao_entrega",
        "responsavel": "responsavel",
        "prazo": "prazo",
        "unid_gestora": "unid_gestora",
        "unidade_setorial_planejamento": "unidade_setorial_planejamento",
        "produto_subacao": "produto_subacao",
        "unidade_medida": "unidade_medida",
        "regiao_subacao": "regiao_subacao",
        "codigo": "codigo",
        "municipios_entrega": "municipios_entrega",
        "meta_subacao": "meta_subacao",
        "detalhamento_produto": "detalhamento_produto",
    }
    selected = {}
    for key in list(plan_fields.keys()) + list(edit_fields.keys()):
        val = (request.args.get(key) or "").strip()
        if val:
            selected[key] = val
    if "exercicio" not in selected:
        selected["exercicio"] = current_year

    def build_options(field_key: str) -> list[str]:
        col = edit_fields[field_key]
        where = ["ativo = 1"]
        params = {}
        for k, col_name in plan_fields.items():
            if k == field_key:
                continue
            val = selected.get(k, "")
            if val:
                where.append(f"{col_name} = :{k}")
                params[k] = val
        for k, col_name in edit_fields.items():
            if k == field_key:
                continue
            val = selected.get(k, "")
            if val:
                where.append(f"{col_name} = :{k}")
                params[k] = val
        sql = f"SELECT DISTINCT {col} AS value FROM plan21_nger"
        if where:
            sql += " WHERE " + " AND ".join(where)
        rows = _safe_query_mappings(sql, params)
        def _normalize_spaces(value: str) -> str:
            return " ".join(value.replace("\xa0", " ").split())
        def _normalize_meta(value: str) -> str:
            parts = [_normalize_spaces(p) for p in value.split("*")]
            parts = [p for p in parts if p]
            return " * ".join(parts) if parts else ""
        values = []
        for row in rows:
            val = row.get("value")
            if val is None:
                continue
            s = str(val).strip()
            if s:
                if field_key == "meta_subacao":
                    s = _normalize_meta(s)
                else:
                    s = _normalize_spaces(s)
                values.append(s)
        return sorted(set(values), key=lambda v: v.lower())

    options = {key: build_options(key) for key in edit_fields.keys()}

    option_where = ["ativo = 1"]
    option_params = {}
    for k, col_name in plan_fields.items():
        val = selected.get(k, "")
        if val:
            option_where.append(f"{col_name} = :{k}")
            option_params[k] = val
    option_cols = ["id", *plan_fields.values(), *edit_fields.values()]
    option_sql = f"SELECT {', '.join(option_cols)} FROM plan21_nger"
    if option_where:
        option_sql += " WHERE " + " AND ".join(option_where)
    option_sql += " ORDER BY id"
    option_rows = []
    for row in _safe_query_mappings(option_sql, option_params):
        item = {"id": row.get("id")}
        for key, col_name in plan_fields.items():
            item[key] = "" if row.get(col_name) is None else str(row.get(col_name)).strip()
        for key, col_name in edit_fields.items():
            item[key] = "" if row.get(col_name) is None else str(row.get(col_name)).strip()
        option_rows.append(item)

    pair_where = ["ativo = 1"]
    pair_params = {}
    for k, col_name in plan_fields.items():
        val = selected.get(k, "")
        if val:
            pair_where.append(f"{col_name} = :{k}")
            pair_params[k] = val
    for k, col_name in edit_fields.items():
        val = selected.get(k, "")
        if not val:
            continue
        if k in {"codigo", "municipios_entrega", "meta_subacao"}:
            pair_where.append(f"{col_name} = :{k}")
            pair_params[k] = val
            continue
        pair_where.append(f"{col_name} = :{k}")
        pair_params[k] = val
    pair_sql = "SELECT DISTINCT codigo, municipios_entrega, meta_subacao FROM plan21_nger"
    if pair_where:
        pair_sql += " WHERE " + " AND ".join(pair_where)
    pair_rows = _safe_query_mappings(pair_sql, pair_params)
    pairs = []
    def _split_parts(value):
        if value is None:
            return []
        parts = [part.replace("\xa0", " ").strip() for part in str(value).split("*")]
        return [part for part in parts if part]
    def _normalize_meta(value):
        parts = _split_parts(value)
        normalized = []
        for part in parts:
            s = part.replace("\xa0", " ").strip()
            if not s:
                continue
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "." in s and re.match(r"^\d+\.\d{1,2}$", s):
                normalized.append(s.replace(".", ","))
                continue
            elif "," in s:
                s = s.replace(",", ".")
            elif "." in s and re.match(r"^\d+\.\d{3}$", s):
                s = s.replace(".", "")
            try:
                num = float(s)
                s = f"{num:.2f}".replace(".", ",")
            except ValueError:
                pass
            normalized.append(s)
        return " * ".join(normalized) if normalized else ""

    pair_entries = set()

    for row in pair_rows:
        codigo = row.get("codigo")
        municipio = row.get("municipios_entrega")
        meta = row.get("meta_subacao")
        if codigo is None or municipio is None:
            continue
        raw_codigo = str(codigo).strip()
        raw_municipio = str(municipio).strip()
        raw_meta = "" if meta is None else _normalize_meta(meta)
        codigos = _split_parts(codigo)
        municipios = _split_parts(municipio)
        metas = _split_parts(meta)
        if raw_codigo and raw_municipio:
            combined_meta = raw_meta
            if "*" not in combined_meta and len(metas) > 1:
                combined_meta = " * ".join(metas)
            pair_entries.add((raw_codigo, raw_municipio, combined_meta))
        if not codigos or not municipios:
            continue
        max_len = max(len(codigos), len(municipios), len(metas) if metas else 0)
        for idx in range(max_len):
            codigo_s = codigos[idx] if idx < len(codigos) else ""
            municipio_s = municipios[idx] if idx < len(municipios) else ""
            meta_s = metas[idx] if idx < len(metas) else ""
            if not codigo_s or not municipio_s:
                continue
            meta_value = "" if meta_s is None else _normalize_meta(meta_s)
            pair_entries.add((codigo_s, municipio_s, meta_value))

    for codigo_s, municipio_s, meta_value in pair_entries:
        pairs.append({"codigo": codigo_s, "municipio": municipio_s, "meta": meta_value})

    return jsonify({"options": options, "pairs": pairs, "option_rows": option_rows})


@home_bp.route("/api/subacao/plan21-municipios", methods=["GET"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_plan21_municipios():
    plan_fields = {
        "exercicio": "exercicio",
        "uo": "unidade_orcamentaria",
        "programa": "programa",
        "acao_paoe": "acao_paoe",
        "responsavel_acao": "responsavel_acao",
        "produto_acao": "produto_acao",
    }
    filters = {}
    for key, col in plan_fields.items():
        val = (request.args.get(key) or "").strip()
        if val:
            filters[col] = val
    chave = (request.args.get("chave_planejamento") or "").strip()
    subacao = (request.args.get("subacao_entrega") or "").strip()
    if not chave or not subacao:
        return jsonify({"municipios": [], "has_etapas": False})
    where = ["ativo = 1", "chave_planejamento = :chave", "subacao_entrega = :subacao"]
    params = {"chave": chave, "subacao": subacao}
    for col, val in filters.items():
        where.append(f"{col} = :{col}")
        params[col] = val
    sql = """
        SELECT codigo, municipios_entrega, meta_subacao, etapa
        FROM plan21_nger
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    rows = _safe_query_mappings(sql, params)
    def _split_parts(value):
        if value is None:
            return []
        parts = [part.replace("\xa0", " ").strip() for part in str(value).split("*")]
        return [part for part in parts if part]

    seen = set()
    municipios = []
    has_etapas = False
    for row in rows:
        codigo_raw = (row.get("codigo") or "").strip()
        municipio_raw = (row.get("municipios_entrega") or "").strip()
        meta_raw = row.get("meta_subacao")
        etapa = (row.get("etapa") or "").strip()
        if etapa:
            has_etapas = True
        if not codigo_raw or not municipio_raw:
            continue

        codigos = _split_parts(codigo_raw)
        municipios_raw = _split_parts(municipio_raw)
        metas = _split_parts(meta_raw)
        if not codigos or not municipios_raw:
            key = f"{codigo_raw}::{municipio_raw}"
            if key in seen:
                continue
            seen.add(key)
            meta_str = "" if meta_raw is None else str(meta_raw)
            municipios.append(
                {
                    "codigo": codigo_raw,
                    "municipio": municipio_raw,
                    "meta": meta_str,
                    "codigo_label": codigo_raw,
                    "municipio_label": municipio_raw,
                }
            )
            continue

        max_len = max(len(codigos), len(municipios_raw), len(metas) if metas else 0)
        for idx in range(max_len):
            codigo = codigos[idx] if idx < len(codigos) else ""
            municipio = municipios_raw[idx] if idx < len(municipios_raw) else ""
            meta = metas[idx] if idx < len(metas) else ""
            if not codigo or not municipio:
                continue
            key = f"{codigo}::{municipio}"
            if key in seen:
                continue
            seen.add(key)
            municipios.append(
                {
                    "codigo": codigo,
                    "municipio": municipio,
                    "meta": meta,
                    "codigo_label": codigo,
                    "municipio_label": municipio,
                }
            )
    return jsonify({"municipios": municipios, "has_etapas": has_etapas})


@home_bp.route("/api/subacao", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_create():
    data = request.get_json() or {}
    exercicio_raw = (data.get("exercicio") or "").strip()
    unidade_orcamentaria = (data.get("unidade_orcamentaria") or "").strip()
    programa = (data.get("programa") or "").strip()
    acao_paoe = (data.get("acao_paoe") or "").strip()
    responsavel_acao = (data.get("responsavel_acao") or "").strip()
    produto_acao = (data.get("produto_acao") or "").strip()
    chave_planejamento = (data.get("chave_planejamento") or "").strip()
    subacao_entrega_raw = (data.get("subacao_entrega") or "").strip()
    responsavel = (data.get("responsavel") or "").strip()
    cpf_responsavel_raw = (data.get("cpf_responsavel") or "").strip()
    cpf_responsavel = re.sub(r"\D", "", cpf_responsavel_raw)
    prazo_inicio = (data.get("prazo_inicio") or "").strip()
    prazo_fim = (data.get("prazo_fim") or "").strip()
    unid_gestora = (data.get("unid_gestora") or "").strip()
    unidade_setorial_planejamento = (data.get("unidade_setorial_planejamento") or "").strip()
    produto_subacao = (data.get("produto_subacao") or "").strip()
    unidade_medida = (data.get("unidade_medida") or "").strip()
    regiao_subacao = (data.get("regiao_subacao") or "").strip()
    codigo = (data.get("codigo") or "").strip()
    municipios_entrega = (data.get("municipios_entrega") or "").strip()
    meta_raw = (data.get("meta_subacao") or "").strip()
    municipios_items = data.get("municipios_items") or []
    municipios_items = data.get("municipios_items") or []
    detalhamento_produto = (data.get("detalhamento_produto") or "").strip()
    etapa = (data.get("etapa") or "").strip()
    responsavel_etapa = (data.get("responsavel_etapa") or "").strip()
    cpf_responsavel_etapa_raw = (data.get("cpf_responsavel_etapa") or "").strip()
    cpf_responsavel_etapa = re.sub(r"\D", "", cpf_responsavel_etapa_raw)
    justificativa = (data.get("justificativa") or "").strip()
    responsavel_nger = (data.get("responsavel_nger") or "").strip()
    tipo_solicitacao = (data.get("tipo_solicitacao") or "cadastrar").strip().lower()
    etapas_items = data.get("etapas_items") or []

    required = {
        "exercicio": exercicio_raw,
        "unidade_orcamentaria": unidade_orcamentaria,
        "programa": programa,
        "acao_paoe": acao_paoe,
        "responsavel_acao": responsavel_acao,
        "produto_acao": produto_acao,
        "chave_planejamento": chave_planejamento,
        "subacao_entrega": subacao_entrega_raw,
        "responsavel": responsavel,
        "cpf_responsavel": cpf_responsavel,
        "prazo_inicio": prazo_inicio,
        "prazo_fim": prazo_fim,
        "unid_gestora": unid_gestora,
        "unidade_setorial_planejamento": unidade_setorial_planejamento,
        "produto_subacao": produto_subacao,
        "unidade_medida": unidade_medida,
        "regiao_subacao": regiao_subacao,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({"error": f"Campos obrigatorios ausentes: {', '.join(missing)}."}), 400
    if not _is_valid_cpf(cpf_responsavel):
        return jsonify({"error": "CPF do responsavel invalido."}), 400
    if cpf_responsavel_etapa and not _is_valid_cpf(cpf_responsavel_etapa):
        return jsonify({"error": "CPF do responsavel da etapa invalido."}), 400
    if (not isinstance(etapas_items, list) or not etapas_items) and not cpf_responsavel_etapa:
        return jsonify({"error": "CPF do responsÃƒÂ¡vel da etapa ÃƒÂ© obrigatÃƒÂ³rio."}), 400
    if cpf_responsavel_etapa and not _is_valid_cpf(cpf_responsavel_etapa):
        return jsonify({"error": "CPF do responsavel da etapa invalido."}), 400
    if cpf_responsavel_etapa and not _is_valid_cpf(cpf_responsavel_etapa):
        return jsonify({"error": "CPF do responsavel da etapa invalido."}), 400
    if (not isinstance(etapas_items, list) or not etapas_items) and not cpf_responsavel_etapa:
        return jsonify({"error": "CPF do responsÃƒÂ¡vel da etapa ÃƒÂ© obrigatÃƒÂ³rio."}), 400

    try:
        exercicio = int(exercicio_raw)
    except ValueError:
        return jsonify({"error": "Exercicio invalido."}), 400

    items: list[dict] = []
    if isinstance(municipios_items, list) and municipios_items:
        for item in municipios_items:
            codigo_i = (item.get("codigo") or "").strip()
            municipio_i = (item.get("municipios_entrega") or "").strip()
            meta_i = (item.get("meta_subacao") or "").strip()
            if not codigo_i or not municipio_i or not meta_i:
                return jsonify({"error": "Informe codigo, municipio e meta para todos os itens."}), 400
            meta_i_dec = _parse_decimal(meta_i)
            if meta_i_dec is None:
                return jsonify({"error": "Meta da subacao invalida."}), 400
            items.append(
                {
                    "codigo": codigo_i,
                    "municipios_entrega": municipio_i,
                    "meta_subacao": meta_i_dec,
                }
            )
    else:
        if not codigo or not municipios_entrega or not meta_raw:
            return jsonify({"error": "Codigo, municipio e meta sao obrigatorios."}), 400
        meta_subacao = _parse_decimal(meta_raw)
        if meta_subacao is None:
            return jsonify({"error": "Meta da subacao invalida."}), 400
        items.append(
            {
                "codigo": codigo,
                "municipios_entrega": municipios_entrega,
                "meta_subacao": meta_subacao,
            }
        )

    etapa_by_municipio = {}
    if isinstance(etapas_items, list) and etapas_items:
        for item in etapas_items:
            muni = (item.get("municipio") or "").strip()
            nome = (item.get("nome_etapa") or "").strip()
            if not muni or not nome:
                return jsonify({"error": "Etapas incompletas para municipios."}), 400
            cpf_etapa = re.sub(r"\D", "", (item.get("cpf_responsavel_etapa") or ""))
            if not cpf_etapa:
                return jsonify({"error": "CPF do responsÃƒÂ¡vel da etapa ÃƒÂ© obrigatÃƒÂ³rio."}), 400
            if not _is_valid_cpf(cpf_etapa):
                return jsonify({"error": "CPF do responsÃƒÂ¡vel da etapa invÃƒÂ¡lido."}), 400
            etapa_by_municipio[muni] = {
                "etapa": nome,
                "responsavel_etapa": (item.get("responsavel_etapa") or "").strip(),
                "cpf_responsavel_etapa": cpf_etapa,
                "justificativa": (item.get("justificativa") or "").strip(),
                "responsavel_nger": (item.get("responsavel_nger") or "").strip(),
            }

    if not _regiao_produto_exists(
        str(exercicio),
        unidade_orcamentaria,
        programa,
        acao_paoe,
        produto_acao,
        regiao_subacao,
    ):
        return (
            jsonify(
                {
                    "error": "RegiÃƒÂ£o nÃƒÂ£o encontrada no PTA. Por favor, antes de criar esta SubaÃƒÂ§ÃƒÂ£o, cadastre uma nova regiÃƒÂ£o do produto e sua respectiva meta fÃƒÂ­sica."
                }
            ),
            400,
        )

    plan_row, plan_err = _find_plan21_nger_row(
        str(exercicio),
        unidade_orcamentaria,
        programa,
        acao_paoe,
        responsavel_acao,
        produto_acao,
        regiao_produto=regiao_subacao,
    )
    if plan_err:
        return jsonify({"error": plan_err}), 400

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400

    prazo = f"{prazo_inicio} a {prazo_fim}".strip()
    subacao_entrega = f"{chave_planejamento} {subacao_entrega_raw}".strip()

    registros = []
    codigos = []
    municipios = []
    metas = []
    etapas = []
    responsaveis_etapa = []
    cpfs_etapa = []
    plan21_ids = []

    for item in items:
        etapa_data = etapa_by_municipio.get(item["municipios_entrega"], {})
        etapa_value = (etapa_data.get("etapa") or etapa).strip()
        responsavel_etapa_value = (etapa_data.get("responsavel_etapa") or responsavel_etapa).strip()
        cpf_responsavel_etapa_value = (
            etapa_data.get("cpf_responsavel_etapa") or cpf_responsavel_etapa
        )

        plan_row_item, _ = _find_plan21_nger_row(
            str(exercicio),
            unidade_orcamentaria,
            programa,
            acao_paoe,
            responsavel_acao,
            produto_acao,
            regiao_produto=regiao_subacao,
            subacao_entrega=subacao_entrega,
            etapa=etapa_value,
            responsavel_etapa=responsavel_etapa_value,
        )
        if not plan_row_item and subacao_entrega_raw:
            plan_row_item, _ = _find_plan21_nger_row(
                str(exercicio),
                unidade_orcamentaria,
                programa,
                acao_paoe,
                responsavel_acao,
                produto_acao,
                regiao_produto=regiao_subacao,
                subacao_entrega=subacao_entrega_raw,
                etapa=etapa_value,
                responsavel_etapa=responsavel_etapa_value,
            )
        if plan_row_item:
            plan21_ids.append(plan_row_item.id)
        elif plan_row:
            plan21_ids.append(plan_row.id)

        codigos.append(item["codigo"])
        municipios.append(item["municipios_entrega"])
        meta_val = item["meta_subacao"]
        if isinstance(meta_val, Decimal):
            meta_val = format(meta_val, "f")
        metas.append(str(meta_val))
        etapas.append(etapa_value or etapa)
        responsaveis_etapa.append(responsavel_etapa_value)
        cpfs_etapa.append(cpf_responsavel_etapa_value)

    plan21_id = plan21_ids[-1] if plan21_ids else (plan_row.id if plan_row else None)
    registro = CadastrarSubacao(
        exercicio=exercicio,
        unidade_orcamentaria=unidade_orcamentaria,
        programa=programa,
        acao_paoe=acao_paoe,
        responsavel_acao=responsavel_acao,
        produto_acao=produto_acao,
        chave_planejamento=chave_planejamento,
        subacao_entrega=subacao_entrega,
        responsavel=responsavel,
        cpf_responsavel=cpf_responsavel,
        prazo=prazo,
        unid_gestora=unid_gestora,
        unidade_setorial_planejamento=unidade_setorial_planejamento,
        produto_subacao=produto_subacao,
        unidade_medida=unidade_medida,
        regiao_subacao=regiao_subacao,
        codigo=json.dumps(codigos),
        municipios_entrega=json.dumps(municipios),
        meta_subacao=json.dumps(metas),
        etapa=json.dumps(etapas),
        responsavel_etapa=json.dumps(responsaveis_etapa),
        cpf_responsavel_etapa=json.dumps(cpfs_etapa),
        detalhamento_produto=detalhamento_produto,
        justificativa=justificativa,
        responsavel_nger=responsavel_nger,
        tipo_solicitacao=tipo_solicitacao,
        ativo=True,
        plan21_nger_id=plan21_id,
        plan21_nger_ids=json.dumps(plan21_ids) if plan21_ids else None,
        usuario_id=usuarios_id,
        status_aprovacao="aguardando",
        criado_em=_now_local(),
    )
    registros.append(registro)
    db.session.add(registro)
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar subacao: {exc}"}), 500
    return jsonify({"subacoes": [_subacao_payload(r) for r in registros]})


@home_bp.route("/api/subacao/editar", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_editar():
    data = request.get_json() or {}
    edit_mode = (data.get("edit_mode") or "").strip()
    plan_fields = [
        "exercicio",
        "uo",
        "programa",
        "acao_paoe",
        "responsavel_acao",
        "produto_acao",
    ]
    edit_fields = [
        "chave_planejamento",
        "subacao_entrega",
        "responsavel",
        "prazo",
        "unid_gestora",
        "unidade_setorial_planejamento",
        "produto_subacao",
        "unidade_medida",
        "regiao_subacao",
        "codigo",
        "municipios_entrega",
        "meta_subacao",
        "detalhamento_produto",
    ]
    filters = {}
    for key in plan_fields + edit_fields:
        val = (data.get(key) or "").strip()
        if val:
            filters[key] = val
    if "exercicio" not in filters:
        filters["exercicio"] = str(_now_local().year)

    subacao_edit = (data.get("subacao_entrega_edit") or "").strip()
    responsavel_edit = (data.get("responsavel_edit") or "").strip()
    cpf_responsavel = re.sub(r"\D", "", (data.get("cpf_responsavel") or "").strip())
    produto_subacao_edit = (data.get("produto_subacao_edit") or "").strip()
    justificativa = (data.get("justificativa") or "").strip()
    responsavel_nger = (data.get("responsavel_nger") or "").strip()
    tipo_solicitacao = (data.get("tipo_solicitacao") or edit_mode or "alterar").strip().lower()
    if tipo_solicitacao == "editar":
        tipo_solicitacao = "alterar"
    tipo_edicao = (data.get("tipo_edicao") or edit_mode or "").strip().lower()
    plan21_nger_id_raw = (data.get("plan21_nger_id") or "").strip()
    plan21_nger_ids_raw = data.get("plan21_nger_ids")
    registro_id_raw = (data.get("registro_id") or data.get("id") or "").strip()

    required_common = [
        "exercicio",
        "uo",
        "programa",
        "acao_paoe",
        "responsavel_acao",
        "produto_acao",
        "chave_planejamento",
        "subacao_entrega",
        "responsavel",
    ]
    required_map = {
        "subacao_name": required_common
        + ["subacao_entrega_edit", "responsavel_edit", "cpf_responsavel"],
        "responsavel_name": required_common + ["responsavel_edit", "cpf_responsavel"],
        "produto_subacao": required_common
        + [
            "responsavel_edit",
            "prazo",
            "unid_gestora",
            "unidade_setorial_planejamento",
            "produto_subacao_edit",
        ],
        "remover_municipio": required_common
        + [
            "responsavel_edit",
            "prazo",
            "unid_gestora",
            "unidade_setorial_planejamento",
            "produto_subacao",
            "regiao_subacao",
        ],
        "novo_municipio": required_common
        + [
            "responsavel_edit",
            "prazo",
            "unid_gestora",
            "unidade_setorial_planejamento",
            "produto_subacao",
            "regiao_subacao",
        ],
        "excluir": required_common + ["justificativa", "responsavel_nger"],
    }
    if edit_mode not in required_map:
        return jsonify({"error": "Modo de edicao invalido."}), 400
    required_values = {
        **filters,
        "subacao_entrega_edit": subacao_edit,
        "responsavel_edit": responsavel_edit,
        "cpf_responsavel": cpf_responsavel,
        "produto_subacao_edit": produto_subacao_edit,
        "justificativa": justificativa,
        "responsavel_nger": responsavel_nger,
    }
    missing = [k for k in required_map[edit_mode] if not required_values.get(k)]
    if missing:
        return jsonify({"error": f"Campos obrigatorios ausentes: {', '.join(missing)}."}), 400
    if not justificativa or not responsavel_nger:
        return jsonify({"error": "Justificativa e responsavel NGER sao obrigatorios."}), 400
    if edit_mode in {"subacao_name", "responsavel_name"} and cpf_responsavel:
        if not _is_valid_cpf(cpf_responsavel):
            return jsonify({"error": "CPF do responsavel invalido."}), 400

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400

    registro_id = None
    registro_row = None
    if registro_id_raw:
        try:
            registro_id = int(registro_id_raw)
        except ValueError:
            registro_id = None
    if registro_id:
        registro_row = CadastrarSubacao.query.get(registro_id)

    ids: list[int] = []
    if plan21_nger_ids_raw:
        try:
            if isinstance(plan21_nger_ids_raw, str):
                parsed = json.loads(plan21_nger_ids_raw)
            else:
                parsed = plan21_nger_ids_raw
            if isinstance(parsed, list):
                for val in parsed:
                    try:
                        ids.append(int(val))
                    except (TypeError, ValueError):
                        continue
        except (TypeError, ValueError):
            pass
    if plan21_nger_id_raw:
        try:
            ids.append(int(plan21_nger_id_raw))
        except ValueError:
            pass
    ids = [i for i in ids if i > 0]
    rows: list[dict] = []
    if ids:
        plan_rows = Plan21Nger.query.filter(Plan21Nger.id.in_(ids)).all()
        for r in plan_rows:
            rows.append(
                {
                    "id": r.id,
                    "exercicio": getattr(r, "exercicio", None),
                    "unidade_orcamentaria": getattr(r, "uo", None),
                    "programa": getattr(r, "programa", None),
                    "acao_paoe": getattr(r, "acao_paoe", None),
                    "responsavel_acao": getattr(r, "responsavel_acao", None),
                    "produto_acao": getattr(r, "produto", None),
                    "chave_planejamento": getattr(r, "chave_planejamento", None),
                    "subacao_entrega": getattr(r, "subacao_entrega", None),
                    "responsavel": getattr(r, "responsavel", None),
                    "cpf_responsavel": getattr(r, "cpf_responsavel", None),
                    "prazo": getattr(r, "prazo", None),
                    "unid_gestora": getattr(r, "unid_gestora", None),
                    "unidade_setorial_planejamento": getattr(
                        r, "unidade_setorial_planejamento", None
                    ),
                    "produto_subacao": getattr(r, "produto_subacao", None),
                    "unidade_medida": getattr(r, "unidade_medida", None),
                    "regiao_subacao": getattr(r, "regiao_subacao", None),
                    "codigo": getattr(r, "codigo", None),
                    "municipios_entrega": getattr(r, "municipios_entrega", None),
                    "meta_subacao": getattr(r, "meta_subacao", None),
                    "etapa": getattr(r, "etapa", None),
                    "responsavel_etapa": getattr(r, "responsavel_etapa", None),
                    "prazo_etapa": getattr(r, "prazo_etapa", None),
                    "detalhamento_produto": getattr(r, "detalhamento_produto", None),
                    "valor_atual": getattr(r, "valor_atual", None),
                }
            )
    else:
        rows = _load_plan21_edit_rows(filters)
    if not rows:
        return jsonify({"error": "Nenhum registro encontrado com os filtros informados."}), 404

    if edit_mode == "remover_municipio":
        filter_map = {
            "exercicio": "exercicio",
            "uo": "unidade_orcamentaria",
            "programa": "programa",
            "acao_paoe": "acao_paoe",
            "responsavel_acao": "responsavel_acao",
            "produto_acao": "produto_acao",
            "chave_planejamento": "chave_planejamento",
            "subacao_entrega": "subacao_entrega",
            "responsavel": "responsavel",
            "prazo": "prazo",
            "unid_gestora": "unid_gestora",
            "unidade_setorial_planejamento": "unidade_setorial_planejamento",
            "produto_subacao": "produto_subacao",
            "unidade_medida": "unidade_medida",
            "regiao_subacao": "regiao_subacao",
            "codigo": "codigo",
            "municipios_entrega": "municipios_entrega",
            "meta_subacao": "meta_subacao",
            "detalhamento_produto": "detalhamento_produto",
        }
        where = ["ativo = 1", "etapa IS NOT NULL", "TRIM(etapa) <> ''"]
        params = {}
        for key, col in filter_map.items():
            val = filters.get(key)
            if val:
                where.append(f"{col} = :{key}")
                params[key] = val
        etapa_sql = "SELECT 1 AS has_etapa FROM plan21_nger"
        if where:
            etapa_sql += " WHERE " + " AND ".join(where)
        etapa_sql += " LIMIT 1"
        etapa_rows = _safe_query_mappings(etapa_sql, params)
        if etapa_rows:
            return (
                jsonify(
                    {
                        "error": "Antes de remover um municÃƒÂ­pio da SubaÃƒÂ§ÃƒÂ£o, por favor, exclua as etapas vinculadas."
                    }
                ),
                400,
            )

    if edit_mode == "excluir":
        for row in rows:
            etapa_val = (str(row.get("etapa") or "").strip())
            responsavel_etapa_val = (str(row.get("responsavel_etapa") or "").strip())
            prazo_etapa_val = (str(row.get("prazo_etapa") or "").strip())
            if etapa_val or responsavel_etapa_val or prazo_etapa_val:
                subacao_nome = (filters.get("subacao_entrega") or row.get("subacao_entrega") or "").strip()
                return (
                    jsonify(
                        {
                            "error": (
                                f'Não é possível excluir a Subação "{subacao_nome}", pois existem Etapas vinculadas a ela. '
                                "Para realizar a exclusão, exclua previamente todas as Etapas associadas a esta Subação."
                            )
                        }
                    ),
                    400,
                )
        for row in rows:
            val = row.get("valor_atual")
            if val is None or val == "":
                continue
            dec = _parse_decimal(val)
            if dec is None or dec != 0:
                return (
                    jsonify(
                        {
                            "error": "Por favor, antes de excluir esta SubaÃƒÂ§ÃƒÂ£o, remaneje os valores da MemÃƒÂ³ria de CÃƒÂ¡lculo."
                        }
                    ),
                    400,
                )

    row_ids = [str(r.get("id")) for r in rows if r.get("id") is not None]
    rows = rows[:1]

    def pick_row(row: dict, key: str) -> str:
        val = row.get(key)
        return "" if val is None else str(val).strip()

    if registro_id_raw and not registro_row:
        return jsonify({"error": "Registro de ediÃƒÂ§ÃƒÂ£o nÃƒÂ£o encontrado."}), 404

    def _coalesce_update(target, key, value):
        if value is None:
            return
        if isinstance(value, str) and value.strip() == "":
            return
        setattr(target, key, value)

    registros = []
    for row in rows:
        subacao_value = pick_row(row, "subacao_entrega")
        if edit_mode == "subacao_name" and subacao_edit:
            subacao_value = f"{filters.get('chave_planejamento', '')} {subacao_edit}".strip()
        responsavel_value = responsavel_edit or pick_row(row, "responsavel") or filters.get("responsavel", "")
        produto_value = (
            produto_subacao_edit
            if edit_mode == "produto_subacao"
            else pick_row(row, "produto_subacao") or filters.get("produto_subacao", "")
        )
        prazo_value = filters.get("prazo") or pick_row(row, "prazo")
        unid_gestora = filters.get("unid_gestora") or pick_row(row, "unid_gestora")
        unidade_setorial = (
            filters.get("unidade_setorial_planejamento")
            or pick_row(row, "unidade_setorial_planejamento")
        )
        unidade_medida = filters.get("unidade_medida") or pick_row(row, "unidade_medida")
        regiao_subacao = filters.get("regiao_subacao") or pick_row(row, "regiao_subacao")
        detalhamento = filters.get("detalhamento_produto") or pick_row(row, "detalhamento_produto")

        base_kwargs = {
            "exercicio": filters.get("exercicio") or pick_row(row, "exercicio"),
            "unidade_orcamentaria": filters.get("uo") or pick_row(row, "unidade_orcamentaria"),
            "programa": filters.get("programa") or pick_row(row, "programa"),
            "acao_paoe": filters.get("acao_paoe") or pick_row(row, "acao_paoe"),
            "responsavel_acao": filters.get("responsavel_acao") or pick_row(row, "responsavel_acao"),
            "produto_acao": filters.get("produto_acao") or pick_row(row, "produto_acao"),
            "chave_planejamento": filters.get("chave_planejamento") or pick_row(row, "chave_planejamento"),
            "subacao_entrega": subacao_value,
            "responsavel": responsavel_value,
            "cpf_responsavel": cpf_responsavel if cpf_responsavel else "",
            "prazo": prazo_value,
            "unid_gestora": unid_gestora,
            "unidade_setorial_planejamento": unidade_setorial,
            "produto_subacao": produto_value,
            "unidade_medida": unidade_medida,
            "regiao_subacao": regiao_subacao,
            "detalhamento_produto": detalhamento,
            "tipo_solicitacao": tipo_solicitacao,
            "tipo_edicao": tipo_edicao if edit_mode != "excluir" else None,
            "ativo": True,
            "plan21_nger_id": (row_ids[-1] if row_ids else row.get("id")),
            "plan21_nger_ids": None
            if edit_mode in {"novo_municipio", "remover_municipio"}
            else json.dumps(row_ids)
            if row_ids
            else None,
            "usuario_id": usuarios_id,
            "status_aprovacao": "aguardando",
            "justificativa": justificativa,
            "responsavel_nger": responsavel_nger,
        }

        if edit_mode in {"novo_municipio", "remover_municipio"}:
            municipios_items = data.get("municipios_items") or []
            if not isinstance(municipios_items, list) or not municipios_items:
                return jsonify({"error": "Informe municipios para adicionar."}), 400
            codigos = []
            municipios = []
            metas = []
            for item in municipios_items:
                codigo = (item.get("codigo") or "").strip()
                municipio = (item.get("municipios_entrega") or "").strip()
                meta_raw = (item.get("meta_subacao") or "").strip()
                if not codigo or not municipio or not meta_raw:
                    return jsonify({"error": "Informe codigo, municipio e meta para todos os itens."}), 400
                meta_dec = _parse_decimal(meta_raw)
                if meta_dec is None:
                    return jsonify({"error": "Meta da subacao invalida."}), 400
                codigos.append(codigo)
                municipios.append(municipio)
                metas.append(meta_raw)
            if registro_row:
                for key, value in base_kwargs.items():
                    _coalesce_update(registro_row, key, value)
                _coalesce_update(registro_row, "codigo", json.dumps(codigos))
                _coalesce_update(registro_row, "municipios_entrega", json.dumps(municipios))
                _coalesce_update(registro_row, "meta_subacao", json.dumps(metas))
                registros.append(registro_row)
            else:
                registro = CadastrarSubacao(
                    **base_kwargs,
                    codigo=json.dumps(codigos),
                    municipios_entrega=json.dumps(municipios),
                    meta_subacao=json.dumps(metas),
                    criado_em=_now_local(),
                )
                registros.append(registro)
                db.session.add(registro)
        else:
            meta_val = pick_row(row, "meta_subacao")
            meta_dec = _parse_decimal(meta_val) if meta_val else None
            if registro_row:
                for key, value in base_kwargs.items():
                    _coalesce_update(registro_row, key, value)
                _coalesce_update(registro_row, "codigo", pick_row(row, "codigo"))
                _coalesce_update(
                    registro_row, "municipios_entrega", pick_row(row, "municipios_entrega")
                )
                if meta_dec is not None:
                    registro_row.meta_subacao = meta_dec
                registros.append(registro_row)
            else:
                registro = CadastrarSubacao(
                    **base_kwargs,
                    codigo=pick_row(row, "codigo"),
                    municipios_entrega=pick_row(row, "municipios_entrega"),
                    meta_subacao=meta_dec,
                    criado_em=_now_local(),
                )
                registros.append(registro)
                db.session.add(registro)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar subacao: {exc}"}), 500

    warning = ""
    if edit_mode == "novo_municipio":
        warning = "Novo municipio adicionado. Cadastre uma nova etapa em seguida."
    if edit_mode == "remover_municipio":
        warning = "Municipios atualizados com sucesso."
    return jsonify(
        {
            "subacoes": [_subacao_payload(r) for r in registros],
            "count": len(registros),
            "warning": warning or None,
        }
    )


@home_bp.route("/api/subacao/<int:subacao_id>", methods=["PUT"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_update(subacao_id):
    registro = db.session.get(CadastrarSubacao, subacao_id)
    if not registro or registro.excluido_em is not None:
        return jsonify({"error": "Registro nao encontrado."}), 404
    status_atual = (registro.status_aprovacao or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Somente registros com status Aguardando podem ser editados."}), 400
    criador = None
    if getattr(registro, "usuario_id", None):
        criador = Usuario.query.filter_by(id=registro.usuario_id).first()
    perfil_criador = getattr(criador, "perfil_id", None) if criador else None
    if not _current_user_matches_perfil(perfil_criador):
        return jsonify({"error": "Usuario sem permissao para editar o registro de controle de subacao."}), 403
    data = request.get_json() or {}
    exercicio_raw = (data.get("exercicio") or "").strip()
    unidade_orcamentaria = (data.get("unidade_orcamentaria") or "").strip()
    programa = (data.get("programa") or "").strip()
    acao_paoe = (data.get("acao_paoe") or "").strip()
    responsavel_acao = (data.get("responsavel_acao") or "").strip()
    produto_acao = (data.get("produto_acao") or "").strip()
    chave_planejamento = (data.get("chave_planejamento") or "").strip()
    subacao_entrega_raw = (data.get("subacao_entrega") or "").strip()
    responsavel = (data.get("responsavel") or "").strip()
    cpf_responsavel_raw = (data.get("cpf_responsavel") or "").strip()
    cpf_responsavel = re.sub(r"\D", "", cpf_responsavel_raw)
    prazo_inicio = (data.get("prazo_inicio") or "").strip()
    prazo_fim = (data.get("prazo_fim") or "").strip()
    unid_gestora = (data.get("unid_gestora") or "").strip()
    unidade_setorial_planejamento = (data.get("unidade_setorial_planejamento") or "").strip()
    produto_subacao = (data.get("produto_subacao") or "").strip()
    unidade_medida = (data.get("unidade_medida") or "").strip()
    regiao_subacao = (data.get("regiao_subacao") or "").strip()
    codigo = (data.get("codigo") or "").strip()
    municipios_entrega = (data.get("municipios_entrega") or "").strip()
    meta_raw = (data.get("meta_subacao") or "").strip()
    municipios_items = data.get("municipios_items") or []
    detalhamento_produto = (data.get("detalhamento_produto") or "").strip()
    etapa = (data.get("etapa") or "").strip()
    responsavel_etapa = (data.get("responsavel_etapa") or "").strip()
    cpf_responsavel_etapa_raw = (data.get("cpf_responsavel_etapa") or "").strip()
    cpf_responsavel_etapa = re.sub(r"\D", "", cpf_responsavel_etapa_raw)
    justificativa = (data.get("justificativa") or "").strip()
    responsavel_nger = (data.get("responsavel_nger") or "").strip()
    etapas_items = data.get("etapas_items") or []

    has_municipios_items = isinstance(municipios_items, list) and bool(municipios_items)
    required = {
        "exercicio": exercicio_raw,
        "unidade_orcamentaria": unidade_orcamentaria,
        "programa": programa,
        "acao_paoe": acao_paoe,
        "responsavel_acao": responsavel_acao,
        "produto_acao": produto_acao,
        "chave_planejamento": chave_planejamento,
        "subacao_entrega": subacao_entrega_raw,
        "responsavel": responsavel,
        "cpf_responsavel": cpf_responsavel,
        "prazo_inicio": prazo_inicio,
        "prazo_fim": prazo_fim,
        "unid_gestora": unid_gestora,
        "unidade_setorial_planejamento": unidade_setorial_planejamento,
        "produto_subacao": produto_subacao,
        "unidade_medida": unidade_medida,
        "regiao_subacao": regiao_subacao,
    }
    if not has_municipios_items:
        required["codigo"] = codigo
        required["municipios_entrega"] = municipios_entrega
        required["meta_subacao"] = meta_raw
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({"error": f"Campos obrigatorios ausentes: {', '.join(missing)}."}), 400
    if not _is_valid_cpf(cpf_responsavel):
        return jsonify({"error": "CPF do responsavel invalido."}), 400
    if cpf_responsavel_etapa and not _is_valid_cpf(cpf_responsavel_etapa):
        return jsonify({"error": "CPF do responsavel da etapa invalido."}), 400

    try:
        exercicio = int(exercicio_raw)
    except ValueError:
        return jsonify({"error": "Exercicio invalido."}), 400

    items: list[dict] = []
    if has_municipios_items:
        for item in municipios_items:
            codigo_i = (item.get("codigo") or "").strip()
            municipio_i = (item.get("municipios_entrega") or "").strip()
            meta_i = (item.get("meta_subacao") or "").strip()
            if not codigo_i or not municipio_i or not meta_i:
                return jsonify({"error": "Informe codigo, municipio e meta para todos os itens."}), 400
            meta_i_dec = _parse_decimal(meta_i)
            if meta_i_dec is None:
                return jsonify({"error": "Meta da subacao invalida."}), 400
            items.append(
                {
                    "codigo": codigo_i,
                    "municipios_entrega": municipio_i,
                    "meta_subacao": meta_i_dec,
                }
            )
    else:
        meta_subacao = _parse_decimal(meta_raw)
        if meta_subacao is None:
            return jsonify({"error": "Meta da subacao invalida."}), 400
        items.append(
            {
                "codigo": codigo,
                "municipios_entrega": municipios_entrega,
                "meta_subacao": meta_subacao,
            }
        )

    etapa_by_municipio = {}
    if isinstance(etapas_items, list) and etapas_items:
        for item in etapas_items:
            muni = (item.get("municipio") or "").strip()
            nome = (item.get("nome_etapa") or "").strip()
            if not muni or not nome:
                return jsonify({"error": "Etapas incompletas para municipios."}), 400
            cpf_etapa = re.sub(r"\D", "", (item.get("cpf_responsavel_etapa") or ""))
            if not cpf_etapa:
                return jsonify({"error": "CPF do responsÃƒÂ¡vel da etapa ÃƒÂ© obrigatÃƒÂ³rio."}), 400
            if not _is_valid_cpf(cpf_etapa):
                return jsonify({"error": "CPF do responsÃƒÂ¡vel da etapa invÃƒÂ¡lido."}), 400
            etapa_by_municipio[muni] = {
                "etapa": nome,
                "responsavel_etapa": (item.get("responsavel_etapa") or "").strip(),
                "cpf_responsavel_etapa": cpf_etapa,
            }

    prazo = f"{prazo_inicio} a {prazo_fim}".strip()
    subacao_entrega = f"{chave_planejamento} {subacao_entrega_raw}".strip()

    if not _regiao_produto_exists(
        str(exercicio),
        unidade_orcamentaria,
        programa,
        acao_paoe,
        produto_acao,
        regiao_subacao,
    ):
        return (
            jsonify(
                {
                    "error": "RegiÃƒÂ£o nÃƒÂ£o encontrada no PTA. Por favor, antes de criar esta SubaÃƒÂ§ÃƒÂ£o, cadastre uma nova regiÃƒÂ£o do produto e sua respectiva meta fÃƒÂ­sica."
                }
            ),
            400,
        )

    plan_row_base, plan_err = _find_plan21_nger_row(
        str(exercicio),
        unidade_orcamentaria,
        programa,
        acao_paoe,
        responsavel_acao,
        produto_acao,
        chave_planejamento,
    )
    existing_ids = _extract_plan21_ids(registro)
    if plan_err and not existing_ids:
        return jsonify({"error": plan_err}), 400

    codigos = []
    municipios = []
    metas = []
    etapas = []
    responsaveis_etapa = []
    cpfs_etapa = []
    plan21_ids = []

    for idx, item in enumerate(items):
        etapa_data = etapa_by_municipio.get(item["municipios_entrega"], {})
        etapa_value = (etapa_data.get("etapa") or etapa).strip()
        responsavel_etapa_value = (etapa_data.get("responsavel_etapa") or responsavel_etapa).strip()
        cpf_responsavel_etapa_value = (
            etapa_data.get("cpf_responsavel_etapa") or cpf_responsavel_etapa
        )

        plan_row_item, _ = _find_plan21_nger_row(
            str(exercicio),
            unidade_orcamentaria,
            programa,
            acao_paoe,
            responsavel_acao,
            produto_acao,
            chave_planejamento,
            subacao_entrega=subacao_entrega,
            etapa=etapa_value,
            responsavel_etapa=responsavel_etapa_value,
        )
        if not plan_row_item and subacao_entrega_raw:
            plan_row_item, _ = _find_plan21_nger_row(
                str(exercicio),
                unidade_orcamentaria,
                programa,
                acao_paoe,
                responsavel_acao,
                produto_acao,
                chave_planejamento,
                subacao_entrega=subacao_entrega_raw,
                etapa=etapa_value,
                responsavel_etapa=responsavel_etapa_value,
            )
        if plan_row_item:
            plan21_ids.append(plan_row_item.id)
        elif plan_row_base:
            plan21_ids.append(plan_row_base.id)
        elif existing_ids:
            plan21_ids.append(existing_ids[min(idx, len(existing_ids) - 1)])

        meta_val = item["meta_subacao"]
        if isinstance(meta_val, Decimal):
            meta_val = format(meta_val, "f")
        codigos.append(item["codigo"])
        municipios.append(item["municipios_entrega"])
        metas.append(str(meta_val))
        etapas.append(etapa_value or etapa)
        responsaveis_etapa.append(responsavel_etapa_value)
        cpfs_etapa.append(cpf_responsavel_etapa_value)

    use_grouped = has_municipios_items or len(items) > 1

    registro.exercicio = exercicio
    registro.unidade_orcamentaria = unidade_orcamentaria
    registro.programa = programa
    registro.acao_paoe = acao_paoe
    registro.responsavel_acao = responsavel_acao
    registro.produto_acao = produto_acao
    registro.chave_planejamento = chave_planejamento
    registro.subacao_entrega = subacao_entrega
    registro.responsavel = responsavel
    registro.cpf_responsavel = cpf_responsavel
    registro.prazo = prazo
    registro.unid_gestora = unid_gestora
    registro.unidade_setorial_planejamento = unidade_setorial_planejamento
    registro.produto_subacao = produto_subacao
    registro.unidade_medida = unidade_medida
    registro.regiao_subacao = regiao_subacao
    if use_grouped:
        registro.codigo = json.dumps(codigos)
        registro.municipios_entrega = json.dumps(municipios)
        registro.meta_subacao = json.dumps(metas)
        registro.etapa = json.dumps(etapas)
        registro.responsavel_etapa = json.dumps(responsaveis_etapa)
        registro.cpf_responsavel_etapa = json.dumps(cpfs_etapa)
        registro.plan21_nger_ids = json.dumps(plan21_ids) if plan21_ids else None
    else:
        registro.codigo = codigo
        registro.municipios_entrega = municipios_entrega
        registro.meta_subacao = metas[0] if metas else meta_raw
        registro.etapa = etapa
        registro.responsavel_etapa = responsavel_etapa
        registro.cpf_responsavel_etapa = cpf_responsavel_etapa
    registro.detalhamento_produto = detalhamento_produto
    registro.justificativa = justificativa
    registro.responsavel_nger = responsavel_nger
    registro.plan21_nger_id = plan21_ids[-1] if plan21_ids else plan_row_base.id if plan_row_base else None
    registro.alterado_em = _now_local()
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar subacao: {exc}"}), 500
    return jsonify({"subacao": _subacao_payload(registro)})


@home_bp.route("/api/subacao/<int:subacao_id>", methods=["DELETE"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_delete(subacao_id):
    registro = db.session.get(CadastrarSubacao, subacao_id)
    if not registro or registro.ativo is False:
        return jsonify({"error": "Registro nao encontrado."}), 404
    status = (registro.status_aprovacao or "").strip().lower()
    if status and status != "aguardando":
        return jsonify({"error": "Somente registros com status Aguardando podem ser excluÃƒÂ­dos."}), 400
    criador = None
    if getattr(registro, "usuario_id", None):
        criador = Usuario.query.filter_by(id=registro.usuario_id).first()
    perfil_criador = getattr(criador, "perfil_id", None) if criador else None
    if not _current_user_matches_perfil(perfil_criador):
        return jsonify({"error": "UsuÃ¡rio sem permissÃ£o para excluir o registro de controle de subaÃ§Ã£o."}), 403
    registro.ativo = False
    registro.excluido_em = _now_local()
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir subacao: {exc}"}), 500
    return jsonify({"ok": True})


@home_bp.route("/api/subacao/<int:subacao_id>/aprovar", methods=["POST"])
@login_required
@require_feature("cadastrar/plan_21-nger/subacao")
def api_subacao_aprovar(subacao_id):
    registro = db.session.get(CadastrarSubacao, subacao_id)
    if not registro or registro.ativo is False:
        return jsonify({"error": "Registro nao encontrado."}), 404

    status_atual = (registro.status_aprovacao or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Subacao ja foi processada."}), 400

    try:
        nivel_int = int(getattr(g, "user_nivel", 0) or 0)
    except (TypeError, ValueError):
        nivel_int = 0
    if nivel_int not in (1, 2):
        return jsonify({"error": "Usuario sem permissao para aprovar a subacao atual."}), 403

    data = request.get_json() or {}
    aprovado_raw = (data.get("subacao_aprovada") or "").strip().lower()
    motivo = (data.get("motivo_rejeicao") or "").strip()
    if not motivo:
        return jsonify({"error": "Justificativa obrigatoria."}), 400
    aprovado = aprovado_raw if aprovado_raw in ("sim", "nao") else "sim"

    novos_ids = []
    if aprovado == "sim":
        tipo = (registro.tipo_solicitacao or "cadastrar").strip().lower()
        if tipo == "editar":
            tipo = "alterar"
        try:
            if tipo == "cadastrar":
                novos_ids = _apply_subacao_cadastrar(registro)
            elif tipo == "alterar":
                novos_ids = _apply_subacao_alterar(registro)
            elif tipo == "excluir":
                novos_ids = _apply_subacao_excluir(registro)
            else:
                return jsonify({"error": "Tipo de solicitacao invalido."}), 400
        except Exception as exc:
            db.session.rollback()
            return jsonify({"error": f"Falha ao aprovar subacao: {exc}"}), 500

        registro.status_aprovacao = "Aprovado"
        registro.situacao = "Aprovado"
        registro.motivo_rejeicao = motivo
        if novos_ids:
            registro.plan21_nger_id = novos_ids[-1]
            registro.plan21_nger_ids = json.dumps(novos_ids) if len(novos_ids) > 1 else None
    else:
        registro.status_aprovacao = "Rejeitado"
        registro.situacao = "Rejeitado"
        registro.ativo = False
        registro.excluido_em = _now_local()
        registro.motivo_rejeicao = motivo

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400
    registro.aprovado_por = str(usuarios_id)
    registro.data_aprovacao = _now_local()
    registro.alterado_em = _now_local()

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao aprovar subacao: {exc}"}), 500

    return jsonify({"ok": True, "message": "Subacao atualizada.", "subacao": _subacao_payload(registro)})


@home_bp.route("/api/dotacao", methods=["POST"])
@login_required
@require_feature("cadastrar/dotacao")
def api_dotacao_create():
    data = request.get_json() or {}
    exercicio = (data.get("exercicio") or "").strip()
    chave_planejamento = (data.get("chave_planejamento") or "").strip()
    uo = (data.get("uo") or "").strip()
    programa = (data.get("programa") or "").strip()
    acao_paoe = (data.get("acao_paoe") or "").strip()
    produto = (data.get("produto") or "").strip()
    ug = (data.get("ug") or "").strip()
    regiao = (data.get("regiao") or "").strip()
    subacao_entrega = (data.get("subacao_entrega") or "").strip()
    etapa = (data.get("etapa") or "").strip()
    natureza_despesa = (data.get("natureza_despesa") or "").strip()
    fonte = (data.get("fonte") or "").strip()
    iduso = (data.get("iduso") or "").strip()
    adj_raw = (data.get("perfil_id") or "").strip()
    emprestada_raw = (data.get("dotacao_emprestada") or "").strip().lower()
    adj_concedente_raw = (data.get("adj_concedente") or "").strip()
    elemento_raw = (data.get("elemento") or "").strip()
    subelemento = (data.get("subelemento") or "").strip()
    valor_raw = (data.get("valor_dotacao") or "").strip()
    justificativa_raw = (data.get("justificativa_historico") or "").strip()
    justificativa = _extract_justificativa_text(justificativa_raw)

    required = {
        "exercicio": exercicio,
        "chave_planejamento": chave_planejamento,
        "uo": uo,
        "programa": programa,
        "acao_paoe": acao_paoe,
        "produto": produto,
        "ug": ug,
        "regiao": regiao,
        "subacao_entrega": subacao_entrega,
        "etapa": etapa,
        "natureza_despesa": natureza_despesa,
        "elemento": elemento_raw,
        "subelemento": subelemento,
        "fonte": fonte,
        "iduso": iduso,
        "perfil_id": adj_raw,
        "valor_dotacao": valor_raw,
        "justificativa_historico": justificativa,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({"error": f"Campos obrigatorios ausentes: {', '.join(missing)}."}), 400

    try:
        perfil_id = int(adj_raw)
    except ValueError:
        return jsonify({"error": "Adjunta Responsavel invalida."}), 400
    adj_row = db.session.get(Perfil, perfil_id)
    if not adj_row or (adj_row.nome or "").strip().lower() in {"admin", "consultor"}:
        return jsonify({"error": "Adjunta Responsavel nao encontrada."}), 400

    emprestada = emprestada_raw == "sim"
    adj_concedente = ""
    if emprestada:
        if not adj_concedente_raw:
            return jsonify({"error": "Adjunta Concedente obrigatoria para dotacao emprestada."}), 400
        perfil = (
            Perfil.query.filter(Perfil.ativo == True)  # noqa: E712
            .filter(Perfil.nome == adj_concedente_raw)
            .first()
        )
        if not perfil or adj_concedente_raw.lower() in {"admin", "consultor"}:
            return jsonify({"error": "Adjunta Concedente invalida."}), 400
        adj_concedente = adj_concedente_raw
    else:
        adj_concedente = (adj_row.nome or str(perfil_id)).strip()

    try:
        elemento = int(elemento_raw)
    except ValueError:
        return jsonify({"error": "Elemento invalido."}), 400

    valor_dotacao = _parse_decimal(valor_raw)
    if valor_dotacao is None:
        return jsonify({"error": "Valor da dotacao invalido."}), 400

    saldo_info = _calc_dotacao_saldo(
        exercicio,
        programa,
        acao_paoe,
        produto,
        ug,
        uo,
        regiao,
        subacao_entrega,
        etapa,
        natureza_despesa,
        elemento_raw,
        subelemento,
        fonte,
        iduso,
        chave_planejamento,
    )
    saldo_disponivel = saldo_info["saldo"]
    saldo_disponivel = _dec_or_zero(saldo_disponivel).quantize(Decimal("0.01"))
    valor_dotacao = _dec_or_zero(valor_dotacao).quantize(Decimal("0.01"))
    if valor_dotacao <= 0 or valor_dotacao > saldo_disponivel:
        return jsonify({"error": "Valor da DotaÃƒÂ§ÃƒÂ£o deve ser menor ou igual ao Saldo da DotaÃƒÂ§ÃƒÂ£o"}), 400

    query = Plan21Nger.query
    query = query.filter(Plan21Nger.exercicio == exercicio)
    query = query.filter(Plan21Nger.chave_planejamento == chave_planejamento)
    query = query.filter(Plan21Nger.uo == uo)
    query = query.filter(Plan21Nger.programa == programa)
    query = query.filter(Plan21Nger.acao_paoe == acao_paoe)
    query = query.filter(Plan21Nger.produto == produto)
    query = query.filter(Plan21Nger.ug == ug)
    query = query.filter(Plan21Nger.regiao_etapa == regiao)
    if subacao_entrega:
        query = query.filter(Plan21Nger.subacao_entrega == subacao_entrega)
    if etapa:
        query = query.filter(Plan21Nger.etapa == etapa)
    if natureza_despesa:
        query = query.filter(Plan21Nger.natureza.like(f"{natureza_despesa}%"))
    query = query.filter(Plan21Nger.elemento == elemento_raw)
    if subelemento:
        query = query.filter(Plan21Nger.subelemento == subelemento)
    query = query.filter(Plan21Nger.fonte == fonte)
    query = query.filter(Plan21Nger.idu == iduso)
    rows = query.limit(2).all()
    if not rows:
        return jsonify({"error": "Regi\u00e3o n\u00e3o encontrada no PTA. Por favor, antes de criar esta Suba\u00e7\u00e3o, cadastre uma nova regi\u00e3o do produto e sua respectiva meta f\u00edsica."}), 400
    if len(rows) > 1:
        return jsonify({"error": "Selecao ambigua no plan21_nger. Ajuste os filtros."}), 400
    plan = rows[0]

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400

    registro = Dotacao(
        plan21_nger_id=plan.id,
        exercicio=exercicio,
        perfil_id=perfil_id,
        chave_planejamento=chave_planejamento,
        uo=uo,
        programa=getattr(plan, "programa", None),
        acao_paoe=getattr(plan, "acao_paoe", None),
        produto=getattr(plan, "produto", None),
        ug=getattr(plan, "ug", None),
        regiao=regiao,
        subacao_entrega=subacao_entrega,
        etapa=etapa,
        natureza_despesa=natureza_despesa,
        elemento=elemento,
        subelemento=subelemento,
        fonte=fonte,
        iduso=iduso,
        valor_dotacao=valor_dotacao,
        adj_concedente=adj_concedente,
        status_aprovacao="Aguardando",
        aprovado_por=None,
        data_aprovacao=None,
        justificativa_historico="",
        chave_dotacao="",
        usuarios_id=usuarios_id,
        criado_em=_now_local(),
        alterado_em=None,
        ativo=True,
    )
    db.session.add(registro)
    try:
        db.session.flush()
        adj_label = (adj_row.nome or str(perfil_id)).strip()
        chave_dotacao = f"DOT.{exercicio}.{adj_label}.{registro.id}*"
        justificativa_full = f"{chave_dotacao} {justificativa}".strip()
        ped_sum = _calc_ped_sum_for_dotacao(chave_dotacao)
        emp_sum = _calc_emp_sum_for_dotacao(chave_dotacao)
        est_sum = _calc_estorno_sum_for_dotacao(chave_dotacao)
        _, situacao_map = _build_estorno_maps()
        est_situacao = situacao_map.get(_normalize_dotacao_key(chave_dotacao), "")
        ped_emp_sum = _dec_or_zero(ped_sum) + _dec_or_zero(emp_sum)
        valor_atual = _dec_or_zero(valor_dotacao) - _dec_or_zero(est_sum) - ped_emp_sum
        db.session.execute(
            Dotacao.__table__.update()
            .where(Dotacao.id == registro.id)
            .values(
                chave_dotacao=chave_dotacao,
                justificativa_historico=justificativa_full,
                valor_ped_emp=ped_emp_sum,
                valor_estorno=_dec_or_zero(est_sum),
                situacao=est_situacao,
                valor_atual=valor_atual,
                alterado_em=None,
            )
        )
        db.session.commit()
        db.session.refresh(registro)
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar dotacao: {exc}"}), 500

    return (
        jsonify(
            {
                "ok": True,
                "message": "Dotacao cadastrada.",
                "dotacao": _dotacao_payload(_attach_usuario_nome(registro), adj_label),
            }
        ),
        201,
    )


@home_bp.route("/api/dotacao/<int:dotacao_id>", methods=["PUT"])
@login_required
@require_feature("cadastrar/dotacao")
def api_dotacao_update(dotacao_id):
    registro = db.session.get(Dotacao, dotacao_id)
    if not registro:
        return jsonify({"error": "Dotacao nao encontrada."}), 404

    status_atual = (registro.status_aprovacao or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Somente dota\u00e7\u00f5es com status Aguardando podem ser editadas."}), 400
    criador = None
    if getattr(registro, "usuarios_id", None):
        criador = Usuario.query.filter_by(id=registro.usuarios_id).first()
    perfil_criador = getattr(criador, "perfil_id", None) if criador else None
    if not _current_user_matches_perfil(perfil_criador):
        return jsonify({"error": "Usu\u00e1rio sem permiss\u00e3o para editar a dota\u00e7\u00e3o atual."}), 403

    data = request.get_json() or {}
    exercicio = (data.get("exercicio") or "").strip()
    chave_planejamento = (data.get("chave_planejamento") or "").strip()
    uo = (data.get("uo") or "").strip()
    programa = (data.get("programa") or "").strip()
    acao_paoe = (data.get("acao_paoe") or "").strip()
    produto = (data.get("produto") or "").strip()
    ug = (data.get("ug") or "").strip()
    regiao = (data.get("regiao") or "").strip()
    subacao_entrega = (data.get("subacao_entrega") or "").strip()
    etapa = (data.get("etapa") or "").strip()
    natureza_despesa = (data.get("natureza_despesa") or "").strip()
    fonte = (data.get("fonte") or "").strip()
    iduso = (data.get("iduso") or "").strip()
    adj_raw = (data.get("perfil_id") or "").strip()
    emprestada_raw = (data.get("dotacao_emprestada") or "").strip().lower()
    adj_concedente_raw = (data.get("adj_concedente") or "").strip()
    elemento_raw = (data.get("elemento") or "").strip()
    subelemento = (data.get("subelemento") or "").strip()
    valor_raw = (data.get("valor_dotacao") or "").strip()
    justificativa_raw = (data.get("justificativa_historico") or "").strip()
    justificativa = _extract_justificativa_text(justificativa_raw)

    required = {
        "exercicio": exercicio,
        "chave_planejamento": chave_planejamento,
        "uo": uo,
        "programa": programa,
        "acao_paoe": acao_paoe,
        "produto": produto,
        "ug": ug,
        "regiao": regiao,
        "subacao_entrega": subacao_entrega,
        "etapa": etapa,
        "natureza_despesa": natureza_despesa,
        "elemento": elemento_raw,
        "subelemento": subelemento,
        "fonte": fonte,
        "iduso": iduso,
        "perfil_id": adj_raw,
        "valor_dotacao": valor_raw,
        "justificativa_historico": justificativa,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({"error": f"Campos obrigatorios ausentes: {', '.join(missing)}."}), 400

    try:
        perfil_id = int(adj_raw)
    except ValueError:
        return jsonify({"error": "Adjunta Responsavel invalida."}), 400
    adj_row = db.session.get(Perfil, perfil_id)
    if not adj_row or (adj_row.nome or "").strip().lower() in {"admin", "consultor"}:
        return jsonify({"error": "Adjunta Responsavel nao encontrada."}), 400

    emprestada = emprestada_raw == "sim"
    adj_concedente = ""
    if emprestada:
        if not adj_concedente_raw:
            return jsonify({"error": "Adjunta Concedente obrigatoria para dotacao emprestada."}), 400
        perfil = (
            Perfil.query.filter(Perfil.ativo == True)  # noqa: E712
            .filter(Perfil.nome == adj_concedente_raw)
            .first()
        )
        if not perfil or adj_concedente_raw.lower() in {"admin", "consultor"}:
            return jsonify({"error": "Adjunta Concedente invalida."}), 400
        adj_concedente = adj_concedente_raw
    else:
        adj_concedente = (adj_row.nome or str(perfil_id)).strip()

    try:
        elemento = int(elemento_raw)
    except ValueError:
        return jsonify({"error": "Elemento invalido."}), 400

    valor_dotacao = _parse_decimal(valor_raw)
    if valor_dotacao is None:
        return jsonify({"error": "Valor da dotacao invalido."}), 400

    saldo_info = _calc_dotacao_saldo(
        exercicio,
        programa,
        acao_paoe,
        produto,
        ug,
        uo,
        regiao,
        subacao_entrega,
        etapa,
        natureza_despesa,
        elemento_raw,
        subelemento,
        fonte,
        iduso,
        chave_planejamento,
    )
    saldo_disponivel = saldo_info["saldo"]
    saldo_disponivel = _dec_or_zero(saldo_disponivel).quantize(Decimal("0.01"))
    valor_dotacao = _dec_or_zero(valor_dotacao).quantize(Decimal("0.01"))
    saldo_disponivel += _dec_or_zero(registro.valor_dotacao)
    if valor_dotacao <= 0 or valor_dotacao > saldo_disponivel:
        return jsonify({"error": "Valor da DotaÃƒÂ§ÃƒÂ£o deve ser menor ou igual ao Saldo da DotaÃƒÂ§ÃƒÂ£o"}), 400

    query = Plan21Nger.query
    query = query.filter(Plan21Nger.exercicio == exercicio)
    query = query.filter(Plan21Nger.chave_planejamento == chave_planejamento)
    query = query.filter(Plan21Nger.uo == uo)
    query = query.filter(Plan21Nger.programa == programa)
    query = query.filter(Plan21Nger.acao_paoe == acao_paoe)
    query = query.filter(Plan21Nger.produto == produto)
    query = query.filter(Plan21Nger.ug == ug)
    query = query.filter(Plan21Nger.regiao_etapa == regiao)
    if subacao_entrega:
        query = query.filter(Plan21Nger.subacao_entrega == subacao_entrega)
    if etapa:
        query = query.filter(Plan21Nger.etapa == etapa)
    if natureza_despesa:
        query = query.filter(Plan21Nger.natureza.like(f"{natureza_despesa}%"))
    query = query.filter(Plan21Nger.elemento == elemento_raw)
    if subelemento:
        query = query.filter(Plan21Nger.subelemento == subelemento)
    query = query.filter(Plan21Nger.fonte == fonte)
    query = query.filter(Plan21Nger.idu == iduso)
    rows = query.limit(2).all()
    if not rows:
        return jsonify({"error": "Regi\u00e3o n\u00e3o encontrada no PTA. Por favor, antes de criar esta Suba\u00e7\u00e3o, cadastre uma nova regi\u00e3o do produto e sua respectiva meta f\u00edsica."}), 400
    if len(rows) > 1:
        return jsonify({"error": "Selecao ambigua no plan21_nger. Ajuste os filtros."}), 400
    plan = rows[0]

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400

    registro.plan21_nger_id = plan.id
    registro.exercicio = exercicio
    registro.perfil_id = perfil_id
    registro.chave_planejamento = chave_planejamento
    registro.uo = uo
    registro.programa = getattr(plan, "programa", None)
    registro.acao_paoe = getattr(plan, "acao_paoe", None)
    registro.produto = getattr(plan, "produto", None)
    registro.ug = getattr(plan, "ug", None)
    registro.regiao = regiao
    registro.subacao_entrega = subacao_entrega
    registro.etapa = etapa
    registro.natureza_despesa = natureza_despesa
    registro.elemento = elemento
    registro.subelemento = subelemento
    registro.fonte = fonte
    registro.iduso = iduso
    registro.valor_dotacao = valor_dotacao
    registro.usuarios_id = usuarios_id
    registro.alterado_em = _now_local()
    registro.adj_concedente = adj_concedente
    if not getattr(registro, "status_aprovacao", None):
        registro.status_aprovacao = "Aguardando"
    adj_label = (adj_row.nome or str(perfil_id)).strip()
    chave_dotacao = f"DOT.{exercicio}.{adj_label}.{registro.id}*"
    registro.chave_dotacao = chave_dotacao
    registro.justificativa_historico = f"{chave_dotacao} {justificativa}".strip()
    ped_sum = _calc_ped_sum_for_dotacao(chave_dotacao)
    emp_sum = _calc_emp_sum_for_dotacao(chave_dotacao)
    est_sum = _calc_estorno_sum_for_dotacao(chave_dotacao)
    _, situacao_map = _build_estorno_maps()
    est_situacao = situacao_map.get(_normalize_dotacao_key(chave_dotacao), "")
    ped_emp_sum = _dec_or_zero(ped_sum) + _dec_or_zero(emp_sum)
    registro.valor_ped_emp = ped_emp_sum
    registro.valor_estorno = _dec_or_zero(est_sum)
    registro.situacao = est_situacao
    registro.valor_atual = _dec_or_zero(valor_dotacao) - _dec_or_zero(est_sum) - ped_emp_sum
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar dotacao: {exc}"}), 500
    return jsonify(
        {
            "ok": True,
            "message": "Dotacao atualizada.",
            "dotacao": _dotacao_payload(_attach_usuario_nome(registro), adj_label),
        }
    )


@home_bp.route("/api/dotacao/<int:dotacao_id>", methods=["DELETE"])
@login_required
@require_feature("cadastrar/dotacao")
def api_dotacao_delete(dotacao_id):
    registro = db.session.get(Dotacao, dotacao_id)
    if not registro:
        return jsonify({"error": "Dotacao nao encontrada."}), 404

    status_atual = (registro.status_aprovacao or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Somente dota\u00e7\u00f5es com status Aguardando podem ser exclu\u00eddas."}), 400
    criador = None
    if getattr(registro, "usuarios_id", None):
        criador = Usuario.query.filter_by(id=registro.usuarios_id).first()
    perfil_criador = getattr(criador, "perfil_id", None) if criador else None
    if not _current_user_matches_perfil(perfil_criador):
        return jsonify({"error": "Usu\u00e1rio sem permiss\u00e3o para excluir a dota\u00e7\u00e3o atual."}), 403

    registro.ativo = False
    registro.excluido_em = _now_local()
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir dotacao: {exc}"}), 500

    return jsonify({"ok": True, "message": "Dotacao excluida."})


@home_bp.route("/api/dotacao/<int:dotacao_id>/aprovar", methods=["POST"])
@login_required
@require_feature("cadastrar/dotacao")
def api_dotacao_aprovar(dotacao_id):
    registro = db.session.get(Dotacao, dotacao_id)
    if not registro:
        return jsonify({"error": "Dotacao nao encontrada."}), 404

    status_atual = (registro.status_aprovacao or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Dotacao ja foi processada."}), 400

    adj_concedente = (getattr(registro, "adj_concedente", "") or "").strip()
    if not adj_concedente:
        return jsonify({"error": "Adjunta Concedente nao definida."}), 400
    adj_concedente_id = _perfil_id_by_nome(adj_concedente)
    if not adj_concedente_id:
        return jsonify({"error": "Adjunta Concedente nao definida."}), 400
    if not _current_user_matches_perfil(adj_concedente_id):
        return jsonify({"error": "Usuario sem permissao para aprovar a dotacao atual."}), 403

    data = request.get_json() or {}
    aprovado_raw = (data.get("dotacao_aprovada") or "").strip().lower()
    motivo = (data.get("motivo_rejeicao") or "").strip()
    if not motivo:
        return jsonify({"error": "Justificativa obrigatoria."}), 400

    if aprovado_raw == "sim":
        registro.status_aprovacao = "Aprovado"
    else:
        registro.status_aprovacao = "Rejeitado"
        registro.ativo = False
        registro.excluido_em = _now_local()
    registro.motivo_rejeicao = motivo

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usuario nao encontrado."}), 400
    registro.aprovado_por = str(usuarios_id)
    registro.data_aprovacao = _now_local()
    registro.alterado_em = _now_local()

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao aprovar dotacao: {exc}"}), 500

    adj_label = ""
    if registro.perfil_id:
        adj_row = db.session.get(Perfil, registro.perfil_id)
        adj_label = (adj_row.nome or str(registro.perfil_id)).strip() if adj_row else ""

    return jsonify(
        {
            "ok": True,
            "message": "Dotacao atualizada.",
            "dotacao": _dotacao_payload(_attach_usuario_nome(registro), adj_label),
        }
    )


@home_bp.route("/api/est-dotacao", methods=["POST"])
@login_required
@require_feature("cadastrar/est-dotacao")
def api_est_dotacao_create():
    data = request.get_json() or {}
    exercicio = (data.get("exercicio") or "").strip()
    adjunta = (data.get("adjunta") or "").strip()
    chave_planejamento = (data.get("chave_planejamento") or "").strip()
    chave_dotacao = (data.get("chave_dotacao") or "").strip()
    uo = (data.get("uo") or "").strip()
    programa = (data.get("programa") or "").strip()
    acao_paoe = (data.get("acao_paoe") or "").strip()
    produto = (data.get("produto") or "").strip()
    ug = (data.get("ug") or "").strip()
    regiao = (data.get("regiao") or "").strip()
    subacao_entrega = (data.get("subacao_entrega") or "").strip()
    etapa = (data.get("etapa") or "").strip()
    natureza_despesa = (data.get("natureza_despesa") or "").strip()
    elemento = (data.get("elemento") or "").strip()
    subelemento = (data.get("subelemento") or "").strip()
    fonte = (data.get("fonte") or "").strip()
    iduso = (data.get("iduso") or "").strip()
    valor_dotacao_raw = (data.get("valor_dotacao") or "").strip()
    valor_est_raw = (data.get("valor_a_ser_est") or "").strip()
    saldo_raw = (data.get("saldo_dotacao_apos") or "").strip()
    justificativa = (data.get("justificativa") or "").strip()
    situacao = (data.get("situacao") or "").strip()

    required = {
        "exercicio": exercicio,
        "adjunta": adjunta,
        "chave_planejamento": chave_planejamento,
        "chave_dotacao": chave_dotacao,
        "valor_dotacao": valor_dotacao_raw,
        "valor_a_ser_est": valor_est_raw,
        "saldo_dotacao_apos": saldo_raw,
        "justificativa": justificativa,
        "situacao": situacao,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({"error": f"Campos obrigatÃƒÂ³rios ausentes: {', '.join(missing)}."}), 400

    if not _current_user_matches_perfil(_perfil_id_by_nome(adjunta)):
        return jsonify({"error": "UsuÃƒÂ¡rio sem permissÃƒÂ£o de cadastrar estorno."}), 403

    valor_dotacao = _parse_decimal(valor_dotacao_raw)
    valor_est = _parse_decimal(valor_est_raw)
    saldo = _parse_decimal(saldo_raw)
    if valor_dotacao is None or valor_est is None or saldo is None:
        return jsonify({"error": "Valores monetÃƒÂ¡rios invÃƒÂ¡lidos."}), 400

    adj_row = Perfil.query.filter(Perfil.nome == adjunta).first()
    if not adj_row:
        return jsonify({"error": "Adjunta Solicitante nÃƒÂ£o encontrada."}), 400

    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "UsuÃƒÂ¡rio nÃƒÂ£o encontrado."}), 400

    now = _now_local()
    try:
        db.session.execute(
            text(
                """
                INSERT INTO est_dotacao (
                    exercicio, perfil_id, chave_planejamento, chave_dotacao, uo, programa, acao_paoe, produto,
                    ug, regiao, subacao_entrega, etapa, natureza_despesa, elemento, subelemento, fonte, iduso,
                    valor_dotacao, valor_a_ser_est, saldo_dotacao_apos, justificativa, usuarios_id, ativo,
                    status_aprovacao, situacao, aprovado_por, data_aprovacao, motivo_rejeicao, alterado_em,
                    excluido_em, criado_em
                )
                VALUES (
                    :exercicio, :perfil_id, :chave_planejamento, :chave_dotacao, :uo, :programa, :acao_paoe, :produto,
                    :ug, :regiao, :subacao_entrega, :etapa, :natureza_despesa, :elemento, :subelemento, :fonte, :iduso,
                    :valor_dotacao, :valor_a_ser_est, :saldo_dotacao_apos, :justificativa, :usuarios_id, :ativo,
                    :status_aprovacao, :situacao, :aprovado_por, :data_aprovacao, :motivo_rejeicao, :alterado_em,
                    :excluido_em, :criado_em
                )
                """
            ),
            {
                "exercicio": exercicio,
                "perfil_id": adj_row.id,
                "chave_planejamento": chave_planejamento,
                "chave_dotacao": chave_dotacao,
                "uo": uo,
                "programa": programa,
                "acao_paoe": acao_paoe,
                "produto": produto,
                "ug": ug,
                "regiao": regiao,
                "subacao_entrega": subacao_entrega,
                "etapa": etapa,
                "natureza_despesa": natureza_despesa,
                "elemento": elemento,
                "subelemento": subelemento,
                "fonte": fonte,
                "iduso": iduso,
                "valor_dotacao": valor_dotacao,
                "valor_a_ser_est": valor_est,
                "saldo_dotacao_apos": saldo,
                "justificativa": justificativa,
                "usuarios_id": usuarios_id,
                "ativo": True,
                "status_aprovacao": "Aguardando",
                "situacao": situacao,
                "aprovado_por": None,
                "data_aprovacao": None,
                "motivo_rejeicao": None,
                "alterado_em": None,
                "excluido_em": None,
                "criado_em": now,
            },
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao salvar estorno: {exc}"}), 500

    try:
        dotacao_row = Dotacao.query.filter(Dotacao.chave_dotacao == chave_dotacao).first()
        if not dotacao_row and chave_dotacao:
            key_norm = _normalize_dotacao_key(chave_dotacao)
            dotacao_row = next(
                (
                    d
                    for d in Dotacao.query.filter(Dotacao.chave_dotacao.isnot(None)).all()
                    if _normalize_dotacao_key(d.chave_dotacao) == key_norm
                ),
                None,
            )
        if dotacao_row:
            est_sum = _calc_estorno_sum_for_dotacao(chave_dotacao)
            ped_sum = _calc_ped_sum_for_dotacao(dotacao_row.chave_dotacao or chave_dotacao)
            emp_sum = _calc_emp_sum_for_dotacao(dotacao_row.chave_dotacao or chave_dotacao)
            dotacao_row.valor_estorno = _dec_or_zero(est_sum)
            dotacao_row.situacao = situacao
            dotacao_row.valor_atual = (
                _dec_or_zero(dotacao_row.valor_dotacao)
                - _dec_or_zero(est_sum)
                - _dec_or_zero(ped_sum)
                - _dec_or_zero(emp_sum)
            )
            dotacao_row.alterado_em = _now_local()
            db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({"ok": True, "message": "Estorno cadastrado."}), 201


@home_bp.route("/api/est-dotacao/<int:est_id>", methods=["PUT"])
@login_required
@require_feature("cadastrar/est-dotacao")
def api_est_dotacao_update(est_id):
    row = db.session.execute(
        text("SELECT * FROM est_dotacao WHERE id = :id"),
        {"id": est_id},
    ).mappings().first()
    if not row:
        return jsonify({"error": "Estorno n\u00e3o encontrado."}), 404
    if row.get("ativo") is not None and int(row.get("ativo") or 0) != 1:
        return jsonify({"error": "Estorno inativo."}), 400

    status_atual = str(row.get("status_aprovacao") or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Somente estornos com status Aguardando podem ser alterados."}), 400

    criador = None
    if row.get("usuarios_id"):
        criador = Usuario.query.filter_by(id=row.get("usuarios_id")).first()
    perfil_criador = getattr(criador, "perfil_id", None) if criador else None
    if not _current_user_matches_perfil(perfil_criador):
        return jsonify({"error": "Usu\u00e1rio sem permiss\u00e3o para editar o estorno atual."}), 403

    data = request.get_json() or {}
    exercicio = (data.get("exercicio") or "").strip()
    situacao = (data.get("situacao") or "").strip()
    justificativa = (data.get("justificativa") or "").strip()
    valor_est_raw = (data.get("valor_a_ser_est") or "").strip()

    if not exercicio or not situacao or not justificativa or not valor_est_raw:
        return jsonify({"error": "Campos obrigat\u00f3rios ausentes."}), 400

    valor_est = _parse_decimal(valor_est_raw)
    if valor_est is None:
        return jsonify({"error": "Valor do Estorno inv\u00e1lido."}), 400

    valor_dotacao = row.get("valor_dotacao")
    if valor_dotacao is None:
        valor_dotacao = _parse_decimal(data.get("valor_dotacao") or "")
    valor_dotacao = _dec_or_zero(valor_dotacao)
    saldo = _dec_or_zero(valor_dotacao) - _dec_or_zero(valor_est)

    try:
        db.session.execute(
            text(
                """
                UPDATE est_dotacao
                SET exercicio = :exercicio,
                    valor_a_ser_est = :valor_est,
                    saldo_dotacao_apos = :saldo,
                    justificativa = :justificativa,
                    situacao = :situacao,
                    alterado_em = :alterado_em
                WHERE id = :id
                """
            ),
            {
                "exercicio": exercicio,
                "valor_est": valor_est,
                "saldo": saldo,
                "justificativa": justificativa,
                "situacao": situacao,
                "alterado_em": _now_local(),
                "id": est_id,
            },
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao atualizar estorno: {exc}"}), 500

    chave_dotacao = (row.get("chave_dotacao") or "").strip()
    if chave_dotacao:
        try:
            dotacao_row = Dotacao.query.filter(Dotacao.chave_dotacao == chave_dotacao).first()
            if dotacao_row:
                est_sum = _calc_estorno_sum_for_dotacao(chave_dotacao)
                ped_sum = _calc_ped_sum_for_dotacao(dotacao_row.chave_dotacao or chave_dotacao)
                emp_sum = _calc_emp_sum_for_dotacao(dotacao_row.chave_dotacao or chave_dotacao)
                dotacao_row.valor_estorno = _dec_or_zero(est_sum)
                dotacao_row.situacao = situacao
                dotacao_row.valor_atual = (
                    _dec_or_zero(dotacao_row.valor_dotacao)
                    - _dec_or_zero(est_sum)
                    - _dec_or_zero(ped_sum)
                    - _dec_or_zero(emp_sum)
                )
                dotacao_row.alterado_em = _now_local()
                db.session.commit()
        except Exception:
            db.session.rollback()

    return jsonify({"ok": True, "message": "Estorno atualizado."}), 200


@home_bp.route("/api/est-dotacao/<int:est_id>", methods=["DELETE"])
@login_required
@require_feature("cadastrar/est-dotacao")
def api_est_dotacao_delete(est_id):
    row = db.session.execute(
        text("SELECT * FROM est_dotacao WHERE id = :id"),
        {"id": est_id},
    ).mappings().first()
    if not row:
        return jsonify({"error": "Estorno n\u00e3o encontrado."}), 404
    if row.get("ativo") is not None and int(row.get("ativo") or 0) != 1:
        return jsonify({"error": "Estorno inativo."}), 400

    status_atual = str(row.get("status_aprovacao") or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Somente estornos com status Aguardando podem ser exclu\u00eddos."}), 400

    criador = None
    if row.get("usuarios_id"):
        criador = Usuario.query.filter_by(id=row.get("usuarios_id")).first()
    perfil_criador = getattr(criador, "perfil_id", None) if criador else None
    if not _current_user_matches_perfil(perfil_criador):
        return jsonify({"error": "Usu\u00e1rio sem permiss\u00e3o para excluir o estorno atual."}), 403

    try:
        db.session.execute(
            text("UPDATE est_dotacao SET ativo = 0, excluido_em = :excluido_em WHERE id = :id"),
            {"excluido_em": _now_local(), "id": est_id},
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao excluir estorno: {exc}"}), 500

    chave_dotacao = (row.get("chave_dotacao") or "").strip()
    if chave_dotacao:
        try:
            dotacao_row = Dotacao.query.filter(Dotacao.chave_dotacao == chave_dotacao).first()
            if dotacao_row:
                est_sum = _calc_estorno_sum_for_dotacao(chave_dotacao)
                ped_sum = _calc_ped_sum_for_dotacao(dotacao_row.chave_dotacao or chave_dotacao)
                emp_sum = _calc_emp_sum_for_dotacao(dotacao_row.chave_dotacao or chave_dotacao)
                dotacao_row.valor_estorno = _dec_or_zero(est_sum)
                dotacao_row.valor_atual = (
                    _dec_or_zero(dotacao_row.valor_dotacao)
                    - _dec_or_zero(est_sum)
                    - _dec_or_zero(ped_sum)
                    - _dec_or_zero(emp_sum)
                )
                dotacao_row.alterado_em = _now_local()
                db.session.commit()
        except Exception:
            db.session.rollback()

    return jsonify({"ok": True, "message": "Estorno exclu\u00eddo."}), 200


@home_bp.route("/api/est-dotacao/<int:est_id>/aprovar", methods=["POST"])
@login_required
@require_feature("cadastrar/est-dotacao")
def api_est_dotacao_aprovar(est_id):
    row = db.session.execute(
        text("SELECT * FROM est_dotacao WHERE id = :id"),
        {"id": est_id},
    ).mappings().first()
    if not row:
        return jsonify({"error": "Estorno n\u00e3o encontrado."}), 404
    if row.get("ativo") is not None and int(row.get("ativo") or 0) != 1:
        return jsonify({"error": "Estorno inativo."}), 400

    status_atual = str(row.get("status_aprovacao") or "").strip().lower()
    if status_atual and status_atual != "aguardando":
        return jsonify({"error": "Estorno j\u00e1 foi processado."}), 400

    perfil_id = row.get("perfil_id")
    if not _current_user_matches_perfil(perfil_id):
        return jsonify({"error": "Usu\u00e1rio sem permiss\u00e3o para aprovar o estorno atual."}), 403

    data = request.get_json() or {}
    aprovado_raw = (data.get("estorno_aprovado") or "").strip().lower()
    motivo = (data.get("motivo_rejeicao") or "").strip()
    if not motivo:
        return jsonify({"error": "Justificativa obrigat\u00f3ria."}), 400

    status_novo = "Aprovado" if aprovado_raw == "sim" else "Rejeitado"
    usuarios_id = _resolve_usuario_id()
    if usuarios_id is None:
        return jsonify({"error": "Usu\u00e1rio n\u00e3o encontrado."}), 400

    try:
        db.session.execute(
            text(
                """
                UPDATE est_dotacao
                SET status_aprovacao = :status,
                    aprovado_por = :aprovado_por,
                    data_aprovacao = :data_aprovacao,
                    motivo_rejeicao = :motivo,
                    alterado_em = :alterado_em
                WHERE id = :id
                """
            ),
            {
                "status": status_novo,
                "aprovado_por": str(usuarios_id),
                "data_aprovacao": _now_local(),
                "motivo": motivo,
                "alterado_em": _now_local(),
                "id": est_id,
            },
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao aprovar estorno: {exc}"}), 500

    return jsonify({"ok": True, "message": "Estorno atualizado."}), 200


def _calc_dotacao_saldo(
    exercicio,
    programa,
    acao_paoe,
    produto,
    ug,
    uo,
    regiao,
    subacao_entrega,
    etapa,
    natureza,
    elemento,
    subelemento,
    fonte,
    iduso,
    chave_planejamento,
):
    if not exercicio or not chave_planejamento:
        return {
            "saldo": Decimal("0"),
            "valor_atual": Decimal("0"),
            "valor_dotacao": Decimal("0"),
            "valor_ped": Decimal("0"),
            "valor_emp_liquido": Decimal("0"),
            "plan21_count": 0,
            "dotacao_count": 0,
            "ped_count": 0,
            "emp_count": 0,
        }

    elemento_int = None
    if elemento:
        try:
            elemento_int = int(elemento)
        except ValueError:
            elemento_int = None

    elemento_variants = _codigo_variants(elemento)
    subelemento_variants = _codigo_variants(subelemento)

    ug_norm = _normalize_ug(ug)
    plan21_filters = [Plan21Nger.ativo == True]  # noqa: E712
    if exercicio:
        plan21_filters.append(Plan21Nger.exercicio == exercicio)
    if programa:
        plan21_filters.append(Plan21Nger.programa == programa)
    if acao_paoe:
        plan21_filters.append(Plan21Nger.acao_paoe == acao_paoe)
    if produto:
        plan21_filters.append(Plan21Nger.produto == produto)
    if ug:
        if ug_norm and ug_norm != ug:
            plan21_filters.append(or_(Plan21Nger.ug == ug, Plan21Nger.ug == ug_norm))
        else:
            plan21_filters.append(Plan21Nger.ug == ug)
    if uo:
        plan21_filters.append(Plan21Nger.uo == uo)
    if regiao:
        plan21_filters.append(Plan21Nger.regiao_etapa == regiao)
    if subacao_entrega:
        plan21_filters.append(Plan21Nger.subacao_entrega == subacao_entrega)
    if etapa:
        plan21_filters.append(Plan21Nger.etapa == etapa)
    if natureza:
        plan21_filters.append(Plan21Nger.natureza.like(f"{natureza}%"))
    if elemento_variants:
        plan21_filters.append(Plan21Nger.elemento.in_(elemento_variants))
    if subelemento_variants:
        plan21_filters.append(Plan21Nger.subelemento.in_(subelemento_variants))
    if fonte:
        plan21_filters.append(Plan21Nger.fonte == fonte)
    if iduso:
        plan21_filters.append(Plan21Nger.idu == iduso)
    if chave_planejamento:
        plan21_filters.append(Plan21Nger.chave_planejamento == chave_planejamento)

    plan21_query = db.session.query(Plan21Nger.valor_atual).filter(*plan21_filters)
    plan21_count = plan21_query.count()
    valor_atual = (
        db.session.query(func.coalesce(func.sum(Plan21Nger.valor_atual), 0))
        .filter(*plan21_filters)
        .scalar()
    )
    valor_atual = _dec_or_zero(valor_atual)

    dot_filters = [Dotacao.ativo == True]  # noqa: E712
    if exercicio:
        dot_filters.append(Dotacao.exercicio == exercicio)
    if programa:
        dot_filters.append(Dotacao.programa == programa)
    if acao_paoe:
        dot_filters.append(Dotacao.acao_paoe == acao_paoe)
    if produto:
        dot_filters.append(Dotacao.produto == produto)
    if etapa:
        dot_filters.append(Dotacao.etapa == etapa)
    if natureza:
        dot_filters.append(Dotacao.natureza_despesa.like(f"{natureza}%"))
    if uo:
        dot_filters.append(Dotacao.uo == uo)
    if regiao:
        dot_filters.append(Dotacao.regiao == regiao)
    if elemento_int is not None:
        dot_filters.append(Dotacao.elemento == elemento_int)
    if subelemento_variants:
        dot_filters.append(Dotacao.subelemento.in_(subelemento_variants))
    if fonte:
        dot_filters.append(Dotacao.fonte == fonte)
    if iduso:
        dot_filters.append(Dotacao.iduso == iduso)
    if chave_planejamento:
        dot_filters.append(Dotacao.chave_planejamento == chave_planejamento)

    dot_rows = (
        Dotacao.query.with_entities(
            Dotacao.valor_dotacao,
            Dotacao.valor_atual,
            Dotacao.valor_estorno,
            Dotacao.valor_ped_emp,
            Dotacao.subacao_entrega,
            Dotacao.chave_dotacao,
            Dotacao.ug,
        )
        .filter(*dot_filters)
        .all()
    )
    est_map, _ = _build_estorno_maps()

    if subacao_entrega:
        subacao_norm = _normalize_chave(subacao_entrega)
        dot_rows = [r for r in dot_rows if _normalize_chave(r.subacao_entrega) == subacao_norm]
    if ug_norm:
        dot_rows = [r for r in dot_rows if _normalize_ug(r.ug or "") == ug_norm]

    valor_dotacao = sum(
        (
            _dec_or_zero(
                r.valor_atual
                if r.valor_atual is not None
                else _dec_or_zero(r.valor_dotacao)
                - _dec_or_zero(est_map.get(_normalize_dotacao_key(r.chave_dotacao), r.valor_estorno))
                - _dec_or_zero(r.valor_ped_emp)
            )
            for r in dot_rows
        ),
        Decimal("0"),
    )
    valor_dotacao = _dec_or_zero(valor_dotacao)

    def _count_chave_parts(value: str) -> int:
        if not value:
            return 0
        return len([p for p in str(value).split("*") if p.strip()])

    programa_key = _normalize_codigo_num(programa)
    acao_paoe_key = _normalize_codigo_num(acao_paoe)
    uo_norm = _normalize_uo(uo)
    try:
        exercicio_int = int(str(exercicio).split(".")[0])
    except ValueError:
        exercicio_int = None
    chave_parts = _count_chave_parts(chave_planejamento)
    if exercicio_int and exercicio_int <= 2025:
        chave_field = "chave_planejamento"
    elif chave_parts >= 8:
        chave_field = "chave_planejamento"
    elif chave_parts == 4:
        chave_field = "chave"
    else:
        chave_field = "chave"
    chave_norm = _normalize_chave(chave_planejamento)

    ped_base_common = [PedRegistro.ativo == True]  # noqa: E712
    if exercicio:
        ped_base_common.append(PedRegistro.exercicio == exercicio)
    if programa_key:
        ped_base_common.append(PedRegistro.programa_governo == programa_key)
    if acao_paoe_key:
        ped_base_common.append(PedRegistro.paoe == acao_paoe_key)
    if fonte:
        ped_base_common.append(PedRegistro.fonte == fonte)
    if iduso:
        variants = _iduso_variants(iduso)
        if variants:
            ped_base_common.append(PedRegistro.iduso.in_(variants))
    if elemento_variants:
        ped_base_common.append(PedRegistro.elemento.in_(elemento_variants))
    if uo_norm:
        ped_base_common.append(PedRegistro.uo == uo_norm)
    if ug_norm:
        ped_base_common.append(PedRegistro.subfuncao_ug.like(f"%.{ug_norm}"))
    if regiao:
        regiao_key = _normalize_codigo_num(regiao)
        if regiao_key:
            ped_base_common.append(
                or_(
                    PedRegistro.regiao == regiao,
                    PedRegistro.regiao == regiao_key,
                    PedRegistro.regiao == f"R{regiao_key}",
                )
            )
        else:
            ped_base_common.append(PedRegistro.regiao == regiao)

    emp_base_common = [EmpRegistro.ativo == True]  # noqa: E712
    if exercicio:
        emp_base_common.append(EmpRegistro.exercicio == exercicio)
    if programa_key:
        emp_base_common.append(EmpRegistro.programa_governo == programa_key)
    if acao_paoe_key:
        emp_base_common.append(EmpRegistro.paoe == acao_paoe_key)
    if fonte:
        emp_base_common.append(EmpRegistro.fonte == fonte)
    if iduso:
        variants = _iduso_variants(iduso)
        if variants:
            emp_base_common.append(EmpRegistro.iduso.in_(variants))
    if elemento_variants:
        emp_base_common.append(EmpRegistro.elemento.in_(elemento_variants))
    if uo_norm:
        emp_base_common.append(EmpRegistro.uo == uo_norm)
    if ug_norm:
        emp_base_common.append(EmpRegistro.subfuncao_ug.like(f"%.{ug_norm}"))
    if regiao:
        regiao_key = _normalize_codigo_num(regiao)
        if regiao_key:
            emp_base_common.append(
                or_(
                    EmpRegistro.regiao == regiao,
                    EmpRegistro.regiao == regiao_key,
                    EmpRegistro.regiao == f"R{regiao_key}",
                )
            )
        else:
            emp_base_common.append(EmpRegistro.regiao == regiao)

    dotacao_keys = {
        _normalize_dotacao_key(r.chave_dotacao) for r in dot_rows if _normalize_dotacao_key(r.chave_dotacao)
    }
    ped_base = list(ped_base_common)
    if chave_planejamento:
        if chave_field == "chave_planejamento":
            ped_base.append(PedRegistro.chave_planejamento == chave_planejamento)
        else:
            ped_base.append(PedRegistro.chave == chave_planejamento)
    ped_rows = (
        PedRegistro.query.with_entities(
            PedRegistro.id, PedRegistro.valor_ped, PedRegistro.chave_planejamento, PedRegistro.chave
        )
        .filter(*ped_base)
        .all()
    )
    if not ped_rows and chave_planejamento:
        ped_rows = (
            PedRegistro.query.with_entities(
                PedRegistro.id, PedRegistro.valor_ped, PedRegistro.chave_planejamento, PedRegistro.chave
            )
            .filter(*ped_base_common)
            .all()
        )
    plan_map: dict[int, Decimal] = {}
    for row in ped_rows:
        if not chave_planejamento:
            continue
        row_id = _row_value(row, "id", 0)
        row_valor = _row_value(row, "valor_ped", 1)
        row_chave_planejamento = _row_value(row, "chave_planejamento", 2)
        row_chave = _row_value(row, "chave", 3)
        if row_id is None:
            continue
        chave_val = row_chave_planejamento if chave_field == "chave_planejamento" else row_chave
        if _normalize_chave(chave_val) == chave_norm:
            plan_map[int(row_id)] = _dec_or_zero(row_valor)

    dot_map = _collect_ped_rows_for_dotacao_keys(dotacao_keys) if dotacao_keys else {}
    merged = dict(dot_map)
    for row_id, valor in plan_map.items():
        if row_id not in merged:
            merged[row_id] = valor

    valor_ped = sum(merged.values(), Decimal("0"))
    ped_count = len(merged)
    if ped_count == 0 and chave_planejamento:
        ped_fallback = [PedRegistro.ativo == True]  # noqa: E712
        if exercicio:
            ped_fallback.append(PedRegistro.exercicio == exercicio)
        ped_rows = (
            PedRegistro.query.with_entities(
                PedRegistro.id, PedRegistro.valor_ped, PedRegistro.chave_planejamento, PedRegistro.chave
            )
            .filter(*ped_fallback)
            .all()
        )
        for row in ped_rows:
            row_id = _row_value(row, "id", 0)
            row_valor = _row_value(row, "valor_ped", 1)
            row_chave_planejamento = _row_value(row, "chave_planejamento", 2)
            row_chave = _row_value(row, "chave", 3)
            if row_id is None:
                continue
            if _normalize_chave(row_chave_planejamento) == chave_norm or _normalize_chave(row_chave) == chave_norm:
                merged[int(row_id)] = _dec_or_zero(row_valor)
        valor_ped = sum(merged.values(), Decimal("0"))
        ped_count = len(merged)

    plan_emp_map: dict[int, tuple[Decimal, str]] = {}
    emp_base = list(emp_base_common)
    if chave_planejamento:
        if chave_field == "chave_planejamento":
            emp_base.append(EmpRegistro.chave_planejamento == chave_planejamento)
        else:
            emp_base.append(EmpRegistro.chave == chave_planejamento)
    emp_rows = (
        EmpRegistro.query.with_entities(
            EmpRegistro.id,
            EmpRegistro.numero_emp,
            EmpRegistro.chave_planejamento,
            EmpRegistro.chave,
            EmpRegistro.valor_emp_devolucao_gcv,
        )
        .filter(*emp_base)
        .all()
    )
    if not emp_rows and chave_planejamento:
        emp_rows = (
            EmpRegistro.query.with_entities(
                EmpRegistro.id,
                EmpRegistro.numero_emp,
                EmpRegistro.chave_planejamento,
                EmpRegistro.chave,
                EmpRegistro.valor_emp_devolucao_gcv,
            )
            .filter(*emp_base_common)
            .all()
        )
    for row in emp_rows:
        if not chave_planejamento:
            continue
        row_id = _row_value(row, "id", 0)
        row_num = _row_value(row, "numero_emp", 1)
        row_chave_planejamento = _row_value(row, "chave_planejamento", 2)
        row_chave = _row_value(row, "chave", 3)
        row_valor = _row_value(row, "valor_emp_devolucao_gcv", 4)
        if row_id is None:
            continue
        chave_val = row_chave_planejamento if chave_field == "chave_planejamento" else row_chave
        if _normalize_chave(chave_val) == chave_norm:
            plan_emp_map[int(row_id)] = (_dec_or_zero(row_valor), row_num or "")

    dot_emp_map = _collect_emp_rows_for_dotacao_keys(dotacao_keys) if dotacao_keys else {}
    merged_emp = dict(dot_emp_map)
    for row_id, data in plan_emp_map.items():
        if row_id not in merged_emp:
            merged_emp[row_id] = data

    valor_emp_liquido = sum((v for v, _ in merged_emp.values()), Decimal("0"))
    emp_nums = [n for _, n in merged_emp.values() if n]
    emp_nums = list(dict.fromkeys(emp_nums))
    emp_count = len(emp_nums)

    saldo = valor_atual - valor_dotacao - valor_ped - valor_emp_liquido
    dotacao_count = len(dot_rows)
    return {
        "saldo": saldo,
        "valor_atual": valor_atual,
        "valor_dotacao": valor_dotacao,
        "valor_ped": valor_ped,
        "valor_emp_liquido": valor_emp_liquido,
        "plan21_count": plan21_count,
        "dotacao_count": dotacao_count,
        "ped_count": ped_count,
        "emp_count": emp_count,
    }


@home_bp.route("/api/dotacao/saldo", methods=["GET"])
@login_required
@require_feature("cadastrar/dotacao")
def api_dotacao_saldo():
    exercicio = (request.args.get("exercicio") or "").strip()
    programa = (request.args.get("programa") or "").strip()
    acao_paoe = (request.args.get("acao_paoe") or "").strip()
    produto = (request.args.get("produto") or "").strip()
    ug = (request.args.get("ug") or "").strip()
    uo = (request.args.get("uo") or "").strip()
    regiao = (request.args.get("regiao") or "").strip()
    subacao_entrega = (request.args.get("subacao_entrega") or "").strip()
    etapa = (request.args.get("etapa") or "").strip()
    natureza = (request.args.get("natureza_despesa") or "").strip()
    elemento = (request.args.get("elemento") or "").strip()
    subelemento = (request.args.get("subelemento") or "").strip()
    fonte = (request.args.get("fonte") or "").strip()
    iduso = (request.args.get("iduso") or "").strip()
    chave_planejamento = (request.args.get("chave_planejamento") or "").strip()

    result = _calc_dotacao_saldo(
        exercicio,
        programa,
        acao_paoe,
        produto,
        ug,
        uo,
        regiao,
        subacao_entrega,
        etapa,
        natureza,
        elemento,
        subelemento,
        fonte,
        iduso,
        chave_planejamento,
    )

    return jsonify(
        {
            "saldo": float(result["saldo"]),
            "valor_atual": float(result["valor_atual"]),
            "valor_dotacao": float(result["valor_dotacao"]),
            "valor_ped": float(result["valor_ped"]),
            "valor_emp_liquido": float(result["valor_emp_liquido"]),
            "plan21_count": result["plan21_count"],
            "dotacao_count": result["dotacao_count"],
            "ped_count": result["ped_count"],
            "emp_count": result["emp_count"],
        }
    )


@home_bp.route("/api/fip613/status", methods=["GET"])
@login_required
@require_feature("atualizar/fip613")
def api_fip613_status():
    def _as_iso(value):
        if not value:
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        try:
            # tenta converter string para datetime
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    last = Fip613Upload.query.order_by(Fip613Upload.uploaded_at.desc()).first()
    if not last:
        return jsonify({"ok": True, "last": None})
    return jsonify(
        {
            "ok": True,
            "last": {
                "user_email": last.user_email,
                "uploaded_at": _as_iso(last.uploaded_at),
                "data_arquivo": _as_iso(last.data_arquivo),
                "original_filename": last.original_filename,
                "output_filename": last.output_filename,
            },
        }
    )


@home_bp.route("/api/fip613/upload", methods=["POST"])
@login_required
@require_feature("atualizar/fip613")
def api_fip613_upload():
    if "arquivo" not in request.files:
        return jsonify({"error": "Arquivo ÃƒÂ© obrigatÃƒÂ³rio."}), 400
    arquivo = request.files["arquivo"]
    data_arquivo_raw = request.form.get("data_arquivo")
    if not data_arquivo_raw:
        return jsonify({"error": "Data do download ÃƒÂ© obrigatÃƒÂ³ria."}), 400
    try:
        data_arquivo = datetime.fromisoformat(data_arquivo_raw)
    except ValueError:
        return jsonify({"error": "Data do download invÃƒÂ¡lida."}), 400

    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400

    user = session.get("user") or {}
    user_email = user.get("email") or "desconhecido"

    try:
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        (UPLOAD_DIR / "tmp").mkdir(parents=True, exist_ok=True)
        for f in UPLOAD_DIR.glob("*.xlsx"):
            dest = UPLOAD_DIR / "tmp" / f"{f.stem}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{f.suffix}"
            try:
                f.rename(dest)
            except OSError:
                pass
        stored_name = f"fip613_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = UPLOAD_DIR / stored_name
        arquivo.save(save_path)

        layout_error = _validate_upload_layout(save_path, "fip613", "FIP613")
        if layout_error:
            try:
                save_path.unlink(missing_ok=True)
            except Exception:
                pass
            return jsonify({"error": layout_error}), 400

        registro = Fip613Upload(
            user_email=user_email,
            original_filename=arquivo.filename,
            stored_filename=stored_name,
            data_arquivo=data_arquivo,
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(registro)
        db.session.commit()

        total, output_path = run_fip613(save_path, data_arquivo, user_email, registro.id)

        registro.output_filename = str(output_path.name)
        db.session.commit()

        return jsonify(
            {
                "ok": True,
                "message": f"Processado com sucesso. Registros inseridos: {total}.",
                "output": output_path.name,
            }
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao processar: {exc}"}), 500


@home_bp.route("/api/relatorios/fip613", methods=["GET"])
@login_required
@require_feature("relatorios/fip613")
def api_relatorio_fip613():
    def _as_iso(value):
        if not value:
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    try:
        rows = Fip613Registro.query.filter_by(ativo=True).all()
        last_upload = Fip613Upload.query.order_by(Fip613Upload.uploaded_at.desc()).first()
        data_arquivo = _as_iso(last_upload.data_arquivo) if last_upload else None
        uploaded_at = _as_iso(last_upload.uploaded_at) if last_upload else None
        user_email = last_upload.user_email if last_upload else None
        data = []
        for r in rows:
            data.append(
                {
                    "uo": r.uo,
                    "ug": r.ug,
                    "funcao": r.funcao,
                    "subfuncao": r.subfuncao,
                    "programa": r.programa,
                    "projeto_atividade": r.projeto_atividade,
                    "regional": r.regional,
                    "natureza_despesa": str(r.natureza_despesa or ""),
                    "fonte_recurso": str(r.fonte_recurso or ""),
                    "iduso": r.iduso,
                    "tipo_recurso": r.tipo_recurso,
                    "dotacao_inicial": float(r.dotacao_inicial or 0),
                    "cred_suplementar": float(r.cred_suplementar or 0),
                    "cred_especial": float(r.cred_especial or 0),
                    "cred_extraordinario": float(r.cred_extraordinario or 0),
                    "reducao": float(r.reducao or 0),
                    "cred_autorizado": float(r.cred_autorizado or 0),
                    "bloqueado_conting": float(r.bloqueado_conting or 0),
                    "reserva_empenho": float(r.reserva_empenho or 0),
                    "saldo_destaque": float(r.saldo_destaque or 0),
                    "saldo_dotacao": float(r.saldo_dotacao or 0),
                    "empenhado": float(r.empenhado or 0),
                    "liquidado": float(r.liquidado or 0),
                    "a_liquidar": float(r.a_liquidar or 0),
                    "valor_pago": float(r.valor_pago or 0),
                    "valor_a_pagar": float(r.valor_a_pagar or 0),
                }
            )
        return jsonify({"ok": True, "data": data, "data_arquivo": data_arquivo, "uploaded_at": uploaded_at, "user_email": user_email})
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados: {exc}"}), 500


@home_bp.route("/api/relatorios/plan21-nger", methods=["GET"])
@login_required
@require_feature("relatorios/plan21-nger")
def api_relatorio_plan21_nger():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    def _to_float(val):
        dec = _parse_decimal(val)
        return float(dec) if dec is not None else 0.0

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        publico_transversal_chave,
                        programa,
                        funcao,
                        unidade_orcamentaria,
                        acao_paoe,
                        subfuncao,
                        objetivo_especifico,
                        esfera,
                        responsavel_acao,
                        produto_acao,
                        unid_medida_produto,
                        regiao_produto,
                        meta_produto,
                        saldo_meta_produto,
                        meta_credito,
                        meta_anulada,
                        meta_atual,
                        publico_transversal,
                        subacao_entrega,
                        responsavel,
                        prazo,
                        unid_gestora,
                        unidade_setorial_planejamento,
                        produto_subacao,
                        unidade_medida,
                        regiao_subacao,
                        codigo,
                        municipios_entrega,
                        meta_subacao,
                        detalhamento_produto,
                        etapa,
                        responsavel_etapa,
                        prazo_etapa,
                        regiao_etapa,
                        natureza,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        subelemento,
                        fonte,
                        idu,
                        descricao_item_despesa,
                        unid_medida_item,
                        quantidade,
                        valor_unitario,
                        valor_total,
                        suplementacao,
                        anulacao,
                        valor_atual
                    FROM plan21_nger
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )

        last_row = (
            db.session.execute(
                text(
                    """
                    SELECT user_email, data_arquivo, data_atualizacao
                    FROM plan21_nger
                    WHERE data_arquivo IS NOT NULL
                    ORDER BY data_arquivo DESC
                    LIMIT 1
                    """
                )
            )
            .mappings()
            .first()
        )
        data_arquivo = _as_iso(last_row.get("data_arquivo")) if last_row else None
        uploaded_at = _as_iso(last_row.get("data_atualizacao")) if last_row else None
        user_email = last_row.get("user_email") if last_row else None

        data = []
        for r in rows:
            data.append(
                {
                    "exercicio": r.get("exercicio"),
                    "chave_planejamento": r.get("chave_planejamento"),
                    "regiao": r.get("regiao"),
                    "subfuncao_ug": r.get("subfuncao_ug"),
                    "adj": r.get("adj"),
                    "macropolitica": r.get("macropolitica"),
                    "pilar": r.get("pilar"),
                    "eixo": r.get("eixo"),
                    "politica_decreto": r.get("politica_decreto"),
                    "publico_transversal_chave": r.get("publico_transversal_chave"),
                    "programa": r.get("programa"),
                    "funcao": r.get("funcao"),
                    "unidade_orcamentaria": r.get("unidade_orcamentaria"),
                    "acao_paoe": r.get("acao_paoe"),
                    "subfuncao": r.get("subfuncao"),
                    "objetivo_especifico": r.get("objetivo_especifico"),
                    "esfera": r.get("esfera"),
                    "responsavel_acao": r.get("responsavel_acao"),
                    "produto_acao": r.get("produto_acao"),
                    "unid_medida_produto": r.get("unid_medida_produto"),
                    "regiao_produto": r.get("regiao_produto"),
                    "meta_produto": _to_float(r.get("meta_produto")),
                    "saldo_meta_produto": _to_float(r.get("saldo_meta_produto")),
                    "meta_credito": _to_float(r.get("meta_credito")),
                    "meta_anulada": _to_float(r.get("meta_anulada")),
                    "meta_atual": _to_float(r.get("meta_atual")),
                    "publico_transversal": r.get("publico_transversal"),
                    "subacao_entrega": r.get("subacao_entrega"),
                    "responsavel": r.get("responsavel"),
                    "prazo": r.get("prazo"),
                    "unid_gestora": r.get("unid_gestora"),
                    "unidade_setorial_planejamento": r.get("unidade_setorial_planejamento"),
                    "produto_subacao": r.get("produto_subacao"),
                    "unidade_medida": r.get("unidade_medida"),
                    "regiao_subacao": r.get("regiao_subacao"),
                    "codigo": r.get("codigo"),
                    "municipios_entrega": r.get("municipios_entrega"),
                    "meta_subacao": _to_float(r.get("meta_subacao")),
                    "detalhamento_produto": r.get("detalhamento_produto"),
                    "etapa": r.get("etapa"),
                    "responsavel_etapa": r.get("responsavel_etapa"),
                    "prazo_etapa": r.get("prazo_etapa"),
                    "regiao_etapa": r.get("regiao_etapa"),
                    "natureza": r.get("natureza"),
                    "cat_econ": r.get("cat_econ"),
                    "grupo": r.get("grupo"),
                    "modalidade": r.get("modalidade"),
                    "elemento": r.get("elemento"),
                    "subelemento": r.get("subelemento"),
                    "fonte": r.get("fonte"),
                    "idu": r.get("idu"),
                    "descricao_item_despesa": r.get("descricao_item_despesa"),
                    "unid_medida_item": r.get("unid_medida_item"),
                    "quantidade": _to_float(r.get("quantidade")),
                    "valor_unitario": _to_float(r.get("valor_unitario")),
                    "valor_total": _to_float(r.get("valor_total")),
                    "suplementacao": _to_float(r.get("suplementacao")),
                    "anulacao": _to_float(r.get("anulacao")),
                    "valor_atual": _to_float(r.get("valor_atual")),
                }
            )

        return jsonify(
            {
                "ok": True,
                "data": data,
                "data_arquivo": data_arquivo,
                "uploaded_at": uploaded_at,
                "user_email": user_email,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados: {exc}"}), 500


@home_bp.route("/api/relatorios/fip613/download", methods=["GET"])
@login_required
@require_feature("relatorios/fip613")
def api_relatorio_fip613_download():
    try:
        rows = Fip613Registro.query.filter_by(ativo=True).all()
        data = []
        for r in rows:
            data.append(
                {
                    "UO": r.uo,
                    "UG": r.ug,
                    "Fun\u00e7\u00e3o": r.funcao,
                    "Subfun\u00e7\u00e3o": r.subfuncao,
                    "Programa": r.programa,
                    "Projeto/Atividade": r.projeto_atividade,
                    "Regional": r.regional,
                    "Natureza de Despesa": str(r.natureza_despesa or ""),
                    "Fonte de Recurso": str(r.fonte_recurso or ""),
                    "Iduso": r.iduso,
                    "Tipo de Recurso": r.tipo_recurso,
                    "Dota\u00e7\u00e3o Inicial": float(r.dotacao_inicial or 0),
                    "Cr\u00e9d. Suplementar": float(r.cred_suplementar or 0),
                    "Cr\u00e9d. Especial": float(r.cred_especial or 0),
                    "Cr\u00e9d. Extraordin\u00e1rio": float(r.cred_extraordinario or 0),
                    "Redu\u00e7\u00e3o": float(r.reducao or 0),
                    "Cr\u00e9d. Autorizado": float(r.cred_autorizado or 0),
                    "Bloqueado/Conting.": float(r.bloqueado_conting or 0),
                    "Reserva Empenho": float(r.reserva_empenho or 0),
                    "Saldo de Destaque": float(r.saldo_destaque or 0),
                    "Saldo Dota\u00e7\u00e3o": float(r.saldo_dotacao or 0),
                    "Empenhado": float(r.empenhado or 0),
                    "Liquidado": float(r.liquidado or 0),
                    "A liquidar": float(r.a_liquidar or 0),
                    "Valor Pago": float(r.valor_pago or 0),
                    "Valor a Pagar": float(r.valor_a_pagar or 0),
                }
            )
        db.session.close()
        df = None
        try:
            import pandas as pd
            from io import BytesIO
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            import unicodedata

            df = pd.DataFrame(data)

            invert_targets = {"reducao", "bloqueadoconting", "reservaempenho", "empenhado"}

            def _norm_col(name: str) -> str:
                base = unicodedata.normalize("NFKD", str(name or ""))
                ascii_only = "".join(ch for ch in base if not unicodedata.combining(ch))
                return (
                    ascii_only.lower()
                    .replace(" ", "")
                    .replace(".", "")
                    .replace("/", "")
                    .replace("_", "")
                )

            for col in list(df.columns):
                if _norm_col(col) in invert_targets:
                    df[col] = df[col].apply(lambda x: -(x or 0))
            output = BytesIO()
            df.to_excel(output, index=False)
            output.seek(0)

            # aplica fonte e formato num\u00e9rico no Excel
            wb = load_workbook(output)
            ws = wb.active
            font = Font(name="Helvetica", size=8)
            number_format = "[Blue]#,##0.00;[Red]-#,##0.00;0"
            # colunas num\u00e9ricas come\u00e7am em 12 (1-based) at\u00e9 o final
            numeric_cols = set(range(12, ws.max_column + 1))
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    if cell.col_idx in numeric_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = number_format

            styled = BytesIO()
            wb.save(styled)
            styled.seek(0)

            filename = f"fip613_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return _send_excel_bytes(styled, filename)
        except Exception as exc:
            return jsonify({"error": f"Falha ao preparar planilha: {exc}"}), 500
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


# PED


@home_bp.route("/api/ped/status", methods=["GET"])
@login_required
@require_feature("atualizar/ped")
def api_ped_status():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    PED_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    PED_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    last = PedUpload.query.order_by(PedUpload.uploaded_at.desc()).first()
    if not last:
        return jsonify({"ok": True, "last": None})
    return jsonify(
        {
            "ok": True,
            "last": {
                "user_email": last.user_email,
                "uploaded_at": _as_iso(last.uploaded_at),
                "data_arquivo": _as_iso(last.data_arquivo),
                "original_filename": last.original_filename,
                "output_filename": last.output_filename,
            },
        }
    )


@home_bp.route("/api/ped/upload", methods=["POST"])
@login_required
@require_feature("atualizar/ped")
def api_ped_upload():
    if "arquivo" not in request.files:
        return jsonify({"error": "Arquivo é obrigatório."}), 400
    arquivo = request.files["arquivo"]
    data_arquivo_raw = request.form.get("data_arquivo")
    if not data_arquivo_raw:
        return jsonify({"error": "Data do download é obrigatória."}), 400
    try:
        data_arquivo = datetime.fromisoformat(data_arquivo_raw)
    except ValueError:
        return jsonify({"error": "Data do download inválida."}), 400

    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400

    user = session.get("user") or {}
    user_email = user.get("email") or "desconhecido"

    try:
        PED_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        move_existing_to_tmp(PED_UPLOAD_DIR)
        stored_name = f"ped_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = PED_UPLOAD_DIR / stored_name
        arquivo.save(save_path)

        layout_error = _validate_upload_layout(save_path, "ped", "PED")
        if layout_error:
            try:
                save_path.unlink(missing_ok=True)
            except Exception:
                pass
            return jsonify({"error": layout_error}), 400

        registro = PedUpload(
            user_email=user_email,
            original_filename=arquivo.filename,
            stored_filename=stored_name,
            data_arquivo=data_arquivo,
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(registro)
        db.session.commit()

        total, output_path, missing_dotacao_keys, missing_planejamento_lines = run_ped(
            save_path, data_arquivo, user_email, registro.id
        )

        if missing_dotacao_keys:
            session["ped_dotacao_missing"] = missing_dotacao_keys
            session.modified = True
        else:
            if "ped_dotacao_missing" in session:
                session["ped_dotacao_missing"] = []
                session.modified = True
        if missing_planejamento_lines:
            session["ped_planejamento_missing_lines"] = missing_planejamento_lines
            session.modified = True
        else:
            if "ped_planejamento_missing_lines" in session:
                session["ped_planejamento_missing_lines"] = []
                session.modified = True

        registro.output_filename = str(output_path.name)
        db.session.commit()

        return jsonify(
            {
                "ok": True,
                "message": f"Processado com sucesso. Registros inseridos: {total}.",
                "output": output_path.name,
            }
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao processar: {exc}"}), 500


# EMP


@home_bp.route("/api/emp/status", methods=["GET"])
@login_required
@require_feature("atualizar/emp")
def api_emp_status():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    try:
        EMP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        EMP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        last = EmpUpload.query.order_by(EmpUpload.uploaded_at.desc()).first()
        if not last:
            return jsonify({"ok": True, "last": None})
        status_data = read_status("emp", last.id)
        output_name = last.output_filename
        if not output_name and status_data:
            output_name = status_data.get("output_filename") or output_name
        return jsonify(
            {
                "ok": True,
                "last": {
                    "user_email": last.user_email,
                    "uploaded_at": _as_iso(last.uploaded_at),
                    "data_arquivo": _as_iso(last.data_arquivo),
                    "original_filename": last.original_filename,
                    "output_filename": output_name,
                    "status": status_data.get("state") if status_data else None,
                    "status_message": status_data.get("message") if status_data else None,
                    "status_updated_at": status_data.get("updated_at") if status_data else None,
                    "status_progress": status_data.get("progress") if status_data else None,
                    "status_pid": status_data.get("pid") if status_data else None,
                },
            }
        )
    except Exception:
        return jsonify({"ok": True, "last": None, "status_error": True})


@home_bp.route("/api/est-emp/status", methods=["GET"])
@login_required
@require_feature("atualizar/est-emp")
def api_est_emp_status():
    def _as_iso(value):
        if not value:
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    EST_EMP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    EST_EMP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    last = EstEmpUpload.query.order_by(EstEmpUpload.uploaded_at.desc()).first()
    if not last:
        return jsonify({"ok": True, "last": None})
    return jsonify(
        {
            "ok": True,
            "last": {
                "user_email": last.user_email,
                "uploaded_at": _as_iso(last.uploaded_at),
                "data_arquivo": _as_iso(last.data_arquivo),
                "original_filename": last.original_filename,
                "output_filename": last.output_filename,
            },
        }
    )


@home_bp.route("/api/nob/status", methods=["GET"])
@login_required
@require_feature("atualizar/nob")
def api_nob_status():
    def _as_iso(value):
        if not value:
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    try:
        NOB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        NOB_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        last = NobUpload.query.order_by(NobUpload.uploaded_at.desc()).first()
        if not last:
            return jsonify({"ok": True, "last": None})
        status_data = read_status("nob", last.id)
        output_name = last.output_filename
        if not output_name and status_data:
            output_name = status_data.get("output_filename") or output_name
        return jsonify(
            {
                "ok": True,
                "last": {
                    "user_email": last.user_email,
                    "uploaded_at": _as_iso(last.uploaded_at),
                    "data_arquivo": _as_iso(last.data_arquivo),
                    "original_filename": last.original_filename,
                    "output_filename": output_name,
                    "status": status_data.get("state") if status_data else None,
                    "status_message": status_data.get("message") if status_data else None,
                    "status_updated_at": status_data.get("updated_at") if status_data else None,
                    "status_progress": status_data.get("progress") if status_data else None,
                    "status_pid": status_data.get("pid") if status_data else None,
                },
            }
        )
    except Exception:
        return jsonify({"ok": True, "last": None, "status_error": True})


@home_bp.route("/api/emp/upload", methods=["POST"])
@login_required
@require_feature("atualizar/emp")
def api_emp_upload():
    if "arquivo" not in request.files:
        return jsonify({"error": "Arquivo obrigatorio."}), 400
    arquivo = request.files["arquivo"]
    data_arquivo_raw = request.form.get("data_arquivo")
    if not data_arquivo_raw:
        return jsonify({"error": "Data do download obrigatoria."}), 400
    try:
        data_arquivo = datetime.fromisoformat(data_arquivo_raw)
    except ValueError:
        return jsonify({"error": "Data do download invalida."}), 400

    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400

    user = session.get("user") or {}
    user_email = user.get("email") or "desconhecido"

    try:
        EMP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        _move_existing_to_tmp(EMP_UPLOAD_DIR)
        stored_name = f"emp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = EMP_UPLOAD_DIR / stored_name
        arquivo.save(save_path)

        layout_error = _validate_upload_layout(save_path, "emp", "EMP")
        if layout_error:
            try:
                save_path.unlink(missing_ok=True)
            except Exception:
                pass
            return jsonify({"error": layout_error}), 400

        registro = EmpUpload(
            user_email=user_email,
            original_filename=arquivo.filename,
            stored_filename=stored_name,
            data_arquivo=data_arquivo,
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(registro)
        db.session.commit()

        write_status("emp", registro.id, "em processamento", "Arquivo recebido. Processamento em background.")
        _start_worker("emp", registro.id)
        return jsonify(
            {
                "ok": True,
                "message": "Arquivo recebido. O processamento ocorrer\u00e1 em segundo plano.",
                "job_id": registro.id,
            }
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao processar: {exc}"}), 500


@home_bp.route("/api/emp/reprocess", methods=["POST"])
@login_required
@require_feature("atualizar/emp")
def api_emp_reprocess():
    payload = request.get_json(silent=True) or {}
    upload_id = payload.get("upload_id")
    if upload_id:
        registro = db.session.get(EmpUpload, upload_id)
    else:
        registro = EmpUpload.query.order_by(EmpUpload.uploaded_at.desc()).first()
    if not registro:
        return jsonify({"error": "Nenhum upload encontrado para reprocessar."}), 404
    try:
        file_path = _find_upload_path(EMP_UPLOAD_DIR, registro.stored_filename)
        if not file_path:
            return jsonify({"error": "Arquivo do upload nao encontrado."}), 404
        # garante o caminho correto para o worker
        if file_path.parent.name == "tmp":
            registro.stored_filename = f"tmp/{file_path.name}"
        registro.output_filename = None
        db.session.commit()
        write_status("emp", registro.id, "em processamento", "Reprocessamento iniciado.")
        _start_worker("emp", registro.id)
        return jsonify({"ok": True, "message": "Reprocessamento iniciado.", "job_id": registro.id})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao reprocessar: {exc}"}), 500


@home_bp.route("/api/emp/cancel", methods=["POST"])
@login_required
@require_feature("atualizar/emp")
def api_emp_cancel():
    payload = request.get_json(silent=True) or {}
    upload_id = payload.get("upload_id")
    if upload_id:
        registro = db.session.get(EmpUpload, upload_id)
    else:
        registro = EmpUpload.query.order_by(EmpUpload.uploaded_at.desc()).first()
    if not registro:
        return jsonify({"error": "Nenhum upload encontrado para cancelar."}), 404
    set_cancel_flag("emp", registro.id)
    update_status_fields("emp", registro.id, message="Cancelamento solicitado.")
    return jsonify({"ok": True, "message": "Cancelamento solicitado.", "job_id": registro.id})


@home_bp.route("/api/est-emp/upload", methods=["POST"])
@login_required
@require_feature("atualizar/est-emp")
def api_est_emp_upload():
    if "arquivo" not in request.files:
        return jsonify({"error": "Arquivo obrigatorio."}), 400
    arquivo = request.files["arquivo"]
    data_arquivo_raw = request.form.get("data_arquivo")
    if not data_arquivo_raw:
        return jsonify({"error": "Data do download obrigatoria."}), 400
    try:
        data_arquivo = datetime.fromisoformat(data_arquivo_raw)
    except ValueError:
        return jsonify({"error": "Data do download invalida."}), 400

    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400

    user = session.get("user") or {}
    user_email = user.get("email") or "desconhecido"

    try:
        EST_EMP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        move_est_emp_existing_to_tmp(EST_EMP_UPLOAD_DIR)
        stored_name = f"est_emp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = EST_EMP_UPLOAD_DIR / stored_name
        arquivo.save(save_path)

        layout_error = _validate_upload_layout(save_path, "est_emp", "EST_EMP")
        if layout_error:
            try:
                save_path.unlink(missing_ok=True)
            except Exception:
                pass
            return jsonify({"error": layout_error}), 400

        registro = EstEmpUpload(
            user_email=user_email,
            original_filename=arquivo.filename,
            stored_filename=stored_name,
            data_arquivo=data_arquivo,
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(registro)
        db.session.commit()

        total, output_path = run_est_emp(save_path, data_arquivo, user_email, registro.id)

        registro.output_filename = str(output_path.name)
        db.session.commit()

        return jsonify(
            {
                "ok": True,
                "message": f"Processado com sucesso. Registros inseridos: {total}.",
                "output": output_path.name,
            }
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao processar: {exc}"}), 500


@home_bp.route("/api/nob/upload", methods=["POST"])
@login_required
@require_feature("atualizar/nob")
def api_nob_upload():
    if "arquivo" not in request.files:
        return jsonify({"error": "Arquivo obrigatorio."}), 400
    arquivo = request.files["arquivo"]
    data_arquivo_raw = request.form.get("data_arquivo")
    if not data_arquivo_raw:
        return jsonify({"error": "Data do download obrigatoria."}), 400
    try:
        data_arquivo = datetime.fromisoformat(data_arquivo_raw)
    except ValueError:
        return jsonify({"error": "Data do download invalida."}), 400

    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400

    user = session.get("user") or {}
    user_email = user.get("email") or "desconhecido"

    try:
        NOB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        _move_existing_to_tmp(NOB_UPLOAD_DIR)
        stored_name = f"nob_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = NOB_UPLOAD_DIR / stored_name
        arquivo.save(save_path)

        layout_error = _validate_upload_layout(save_path, "nob", "NOB")
        if layout_error:
            try:
                save_path.unlink(missing_ok=True)
            except Exception:
                pass
            return jsonify({"error": layout_error}), 400

        registro = NobUpload(
            user_email=user_email,
            original_filename=arquivo.filename,
            stored_filename=stored_name,
            data_arquivo=data_arquivo,
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(registro)
        db.session.commit()

        write_status("nob", registro.id, "em processamento", "Arquivo recebido. Processamento em background.")
        _start_worker("nob", registro.id)
        return jsonify(
            {
                "ok": True,
                "message": "Arquivo recebido. O processamento ocorrer\u00e1 em segundo plano.",
                "job_id": registro.id,
            }
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao processar: {exc}"}), 500


@home_bp.route("/api/nob/reprocess", methods=["POST"])
@login_required
@require_feature("atualizar/nob")
def api_nob_reprocess():
    payload = request.get_json(silent=True) or {}
    upload_id = payload.get("upload_id")
    if upload_id:
        registro = db.session.get(NobUpload, upload_id)
    else:
        registro = NobUpload.query.order_by(NobUpload.uploaded_at.desc()).first()
    if not registro:
        return jsonify({"error": "Nenhum upload encontrado para reprocessar."}), 404
    try:
        file_path = _find_upload_path(NOB_UPLOAD_DIR, registro.stored_filename)
        if not file_path:
            return jsonify({"error": "Arquivo do upload nao encontrado."}), 404
        if file_path.parent.name == "tmp":
            registro.stored_filename = f"tmp/{file_path.name}"
        registro.output_filename = None
        db.session.commit()
        write_status("nob", registro.id, "em processamento", "Reprocessamento iniciado.")
        _start_worker("nob", registro.id)
        return jsonify({"ok": True, "message": "Reprocessamento iniciado.", "job_id": registro.id})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao reprocessar: {exc}"}), 500


@home_bp.route("/api/nob/cancel", methods=["POST"])
@login_required
@require_feature("atualizar/nob")
def api_nob_cancel():
    payload = request.get_json(silent=True) or {}
    upload_id = payload.get("upload_id")
    if upload_id:
        registro = db.session.get(NobUpload, upload_id)
    else:
        registro = NobUpload.query.order_by(NobUpload.uploaded_at.desc()).first()
    if not registro:
        return jsonify({"error": "Nenhum upload encontrado para cancelar."}), 404
    set_cancel_flag("nob", registro.id)
    update_status_fields("nob", registro.id, message="Cancelamento solicitado.")
    return jsonify({"ok": True, "message": "Cancelamento solicitado.", "job_id": registro.id})


@home_bp.route("/api/ped/download/<path:filename>", methods=["GET"])
@login_required
@require_feature("atualizar/ped")
def api_ped_download(filename):
    target = PED_OUTPUT_DIR / filename
    if not target.exists():
        abort(404)
    return send_file(target, as_attachment=True)


@home_bp.route("/api/relatorios/ped", methods=["GET"])
@login_required
@require_feature("relatorios/ped")
def api_relatorio_ped():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        chave,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        exercicio,
                        numero_ped,
                        numero_ped_estorno,
                        numero_emp,
                        numero_cad,
                        numero_noblist,
                        numero_os,
                        convenio,
                        numero_processo_orcamentario_pagamento,
                        valor_ped,
                        valor_estorno,
                        indicativo_licitacao_exercicios_anteriores,
                        data_licitacao,
                        liberado_fisco_estadual,
                        situacao,
                        uo,
                        nome_unidade_orcamentaria,
                        ug,
                        nome_unidade_gestora,
                        data_solicitacao,
                        data_criacao,
                        tipo_empenho,
                        dotacao_orcamentaria,
                        funcao,
                        subfuncao,
                        programa_governo,
                        paoe,
                        natureza_despesa,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        nome_elemento,
                        fonte,
                        iduso,
                        numero_emenda_ep,
                        autor_emenda_ep,
                        numero_cac,
                        licitacao,
                        usuario_responsavel,
                        historico,
                        credor,
                        nome_credor,
                        data_autorizacao,
                        data_hora_cadastro_autorizacao,
                        tipo_despesa,
                        numero_abj,
                        numero_processo_sequestro_judicial,
                        indicativo_entrega_imediata,
                        indicativo_contrato,
                        codigo_uo_extinta,
                        devolucao_gcv,
                        mes_competencia_folha_pagamento,
                        exercicio_competencia_folha,
                        obrigacao_patronal,
                        tipo_obrigacao_patronal,
                        numero_nla
                    FROM ped
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )

        last_upload = PedUpload.query.order_by(PedUpload.uploaded_at.desc()).first()
        data_arquivo = _as_iso(getattr(last_upload, "data_arquivo", None)) if last_upload else None
        uploaded_at = _as_iso(getattr(last_upload, "uploaded_at", None)) if last_upload else None
        user_email = last_upload.user_email if last_upload else None

        data = []
        for r in rows:
            data.append(
                {
                    "chave": r.get("chave"),
                    "chave_planejamento": r.get("chave_planejamento"),
                    "regiao": r.get("regiao"),
                    "subfuncao_ug": r.get("subfuncao_ug"),
                    "adj": r.get("adj"),
                    "macropolitica": r.get("macropolitica"),
                    "pilar": r.get("pilar"),
                    "eixo": r.get("eixo"),
                    "politica_decreto": r.get("politica_decreto"),
                    "exercicio": r.get("exercicio"),
                    "numero_ped": r.get("numero_ped"),
                    "numero_ped_estorno": r.get("numero_ped_estorno"),
                    "numero_emp": r.get("numero_emp"),
                    "numero_cad": r.get("numero_cad"),
                    "numero_noblist": r.get("numero_noblist"),
                    "numero_os": r.get("numero_os"),
                    "convenio": r.get("convenio"),
                    "numero_processo_orcamentario_pagamento": r.get("numero_processo_orcamentario_pagamento"),
                    "valor_ped": _to_float(r.get("valor_ped")),
                    "valor_estorno": _to_float(r.get("valor_estorno")),
                    "indicativo_licitacao_exercicios_anteriores": r.get("indicativo_licitacao_exercicios_anteriores"),
                    "data_licitacao": r.get("data_licitacao"),
                    "liberado_fisco_estadual": r.get("liberado_fisco_estadual"),
                    "situacao": r.get("situacao"),
                    "uo": r.get("uo"),
                    "nome_unidade_orcamentaria": r.get("nome_unidade_orcamentaria"),
                    "ug": r.get("ug"),
                    "nome_unidade_gestora": r.get("nome_unidade_gestora"),
                    "data_solicitacao": r.get("data_solicitacao"),
                    "data_criacao": r.get("data_criacao"),
                    "tipo_empenho": r.get("tipo_empenho"),
                    "dotacao_orcamentaria": r.get("dotacao_orcamentaria"),
                    "funcao": r.get("funcao"),
                    "subfuncao": r.get("subfuncao"),
                    "programa_governo": r.get("programa_governo"),
                    "paoe": r.get("paoe"),
                    "natureza_despesa": r.get("natureza_despesa"),
                    "cat_econ": r.get("cat_econ"),
                    "grupo": r.get("grupo"),
                    "modalidade": r.get("modalidade"),
                    "elemento": r.get("elemento"),
                    "nome_elemento": r.get("nome_elemento"),
                    "fonte": r.get("fonte"),
                    "iduso": r.get("iduso"),
                    "numero_emenda_ep": r.get("numero_emenda_ep"),
                    "autor_emenda_ep": r.get("autor_emenda_ep"),
                    "numero_cac": r.get("numero_cac"),
                    "licitacao": r.get("licitacao"),
                    "usuario_responsavel": r.get("usuario_responsavel"),
                    "historico": r.get("historico"),
                    "credor": r.get("credor"),
                    "nome_credor": r.get("nome_credor"),
                    "data_autorizacao": r.get("data_autorizacao"),
                    "data_hora_cadastro_autorizacao": r.get("data_hora_cadastro_autorizacao"),
                    "tipo_despesa": r.get("tipo_despesa"),
                    "numero_abj": r.get("numero_abj"),
                    "numero_processo_sequestro_judicial": r.get("numero_processo_sequestro_judicial"),
                    "indicativo_entrega_imediata": r.get("indicativo_entrega_imediata"),
                    "indicativo_contrato": r.get("indicativo_contrato"),
                    "codigo_uo_extinta": r.get("codigo_uo_extinta"),
                    "devolucao_gcv": r.get("devolucao_gcv"),
                    "mes_competencia_folha_pagamento": r.get("mes_competencia_folha_pagamento"),
                    "exercicio_competencia_folha": r.get("exercicio_competencia_folha"),
                    "obrigacao_patronal": r.get("obrigacao_patronal"),
                    "tipo_obrigacao_patronal": r.get("tipo_obrigacao_patronal"),
                    "numero_nla": r.get("numero_nla"),
                }
            )

        return jsonify(
            {
                "ok": True,
                "data": data,
                "data_arquivo": data_arquivo,
                "uploaded_at": uploaded_at,
                "user_email": user_email,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados do PED: {exc}"}), 500


@home_bp.route("/api/relatorios/ped/download", methods=["GET"])
@login_required
@require_feature("relatorios/ped")
def api_relatorio_ped_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        chave,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        exercicio,
                        numero_ped,
                        numero_ped_estorno,
                        numero_emp,
                        numero_cad,
                        numero_noblist,
                        numero_os,
                        convenio,
                        numero_processo_orcamentario_pagamento,
                        valor_ped,
                        valor_estorno,
                        indicativo_licitacao_exercicios_anteriores,
                        data_licitacao,
                        liberado_fisco_estadual,
                        situacao,
                        uo,
                        nome_unidade_orcamentaria,
                        ug,
                        nome_unidade_gestora,
                        data_solicitacao,
                        data_criacao,
                        tipo_empenho,
                        dotacao_orcamentaria,
                        funcao,
                        subfuncao,
                        programa_governo,
                        paoe,
                        natureza_despesa,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        nome_elemento,
                        fonte,
                        iduso,
                        numero_emenda_ep,
                        autor_emenda_ep,
                        numero_cac,
                        licitacao,
                        usuario_responsavel,
                        historico,
                        credor,
                        nome_credor,
                        data_autorizacao,
                        data_hora_cadastro_autorizacao,
                        tipo_despesa,
                        numero_abj,
                        numero_processo_sequestro_judicial,
                        indicativo_entrega_imediata,
                        indicativo_contrato,
                        codigo_uo_extinta,
                        devolucao_gcv,
                        mes_competencia_folha_pagamento,
                        exercicio_competencia_folha,
                        obrigacao_patronal,
                        tipo_obrigacao_patronal,
                        numero_nla
                    FROM ped
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404
        db.session.close()

        df = pd.DataFrame(rows)

        def _chave_display(row):
            try:
                ex = int(str(row.get("exercicio") or 0)[:4])
            except Exception:
                ex = 0
            chave = row.get("chave") or ""
            chave_plan = row.get("chave_planejamento") or ""
            if ex >= 2026 and chave:
                return chave
            return chave_plan or chave

        df["Chave / Chave de Planejamento"] = df.apply(_chave_display, axis=1)
        df.drop(columns=["chave", "chave_planejamento"], inplace=True, errors="ignore")
        if "valor_ped" in df.columns:
            df["valor_ped"] = df["valor_ped"].apply(_to_float)
        if "valor_estorno" in df.columns:
            df["valor_estorno"] = df["valor_estorno"].apply(_to_float)

        rename_map = {
            "regiao": "Região",
            "subfuncao_ug": "Subfunção + UG",
            "adj": "ADJ",
            "macropolitica": "Macropolítica",
            "pilar": "Pilar",
            "eixo": "Eixo",
            "politica_decreto": "Política_Decreto",
            "exercicio": "Exercício",
            "numero_ped": "Nº PED",
            "numero_ped_estorno": "Nº PED Estorno/Estornado",
            "numero_emp": "Nº EMP",
            "numero_cad": "Nº CAD",
            "numero_noblist": "Nº NOBLIST",
            "numero_os": "Nº OS",
            "convenio": "Convênio",
            "numero_processo_orcamentario_pagamento": "Nº Processo Orçamentário de Pagamento",
            "valor_ped": "Valor PED",
            "valor_estorno": "Valor do Estorno",
            "indicativo_licitacao_exercicios_anteriores": "Indicativo de Licitação de Exercícios Anteriores",
            "data_licitacao": "Data da Licitação",
            "liberado_fisco_estadual": "Liberado Fisco Estadual",
            "situacao": "Situação",
            "uo": "UO",
            "nome_unidade_orcamentaria": "Nome da Unidade Orçamentária",
            "ug": "UG",
            "nome_unidade_gestora": "Nome da Unidade Gestora",
            "data_solicitacao": "Data Solicitação",
            "data_criacao": "Data Criação",
            "tipo_empenho": "Tipo Empenho",
            "dotacao_orcamentaria": "Dotação Orçamentária",
            "funcao": "Função",
            "subfuncao": "Subfunção",
            "programa_governo": "Programa de Governo",
            "paoe": "PAOE",
            "natureza_despesa": "Natureza de Despesa",
            "cat_econ": "Cat.Econ",
            "grupo": "Grupo",
            "modalidade": "Modalidade",
            "elemento": "Elemento",
            "nome_elemento": "Nome do Elemento",
            "fonte": "Fonte",
            "iduso": "Iduso",
            "numero_emenda_ep": "Nº Emenda (EP)",
            "autor_emenda_ep": "Autor da Emenda (EP)",
            "numero_cac": "Nº CAC",
            "licitacao": "Licitação",
            "usuario_responsavel": "Usuário Responsável",
            "historico": "Histórico",
            "credor": "Credor",
            "nome_credor": "Nome do Credor",
            "data_autorizacao": "Data Autorização",
            "data_hora_cadastro_autorizacao": "Data/Hora Cadastro Autorização",
            "tipo_despesa": "Tipo de Despesa",
            "numero_abj": "Nº ABJ",
            "numero_processo_sequestro_judicial": "Nº Processo do Sequestro Judicial",
            "indicativo_entrega_imediata": "Indicativo de Entrega imediata - § 4º Art. 62 Lei 8.666",
            "indicativo_contrato": "Indicativo de contrato",
            "codigo_uo_extinta": "Código UO Extinta",
            "devolucao_gcv": "Devolução GCV",
            "mes_competencia_folha_pagamento": "Mês de Competência da Folha de Pagamento",
            "exercicio_competencia_folha": "Exercício de Competência da Folha de Pagamento",
            "obrigacao_patronal": "Obrigação Patronal",
            "tipo_obrigacao_patronal": "Tipo de Obrigação Patronal",
            "numero_nla": "Nº NLA",
        }
        df.rename(columns=rename_map, inplace=True)

        col_order = [
            "Chave / Chave de Planejamento",
            "Região",
            "Subfunção + UG",
            "ADJ",
            "Macropolítica",
            "Pilar",
            "Eixo",
            "Política_Decreto",
            "Exercício",
            "Nº PED",
            "Nº PED Estorno/Estornado",
            "Nº EMP",
            "Nº CAD",
            "Nº NOBLIST",
            "Nº OS",
            "Convênio",
            "Nº Processo Orçamentário de Pagamento",
            "Valor PED",
            "Valor do Estorno",
            "Indicativo de Licitação de Exercícios Anteriores",
            "Data da Licitação",
            "Liberado Fisco Estadual",
            "Situação",
            "UO",
            "Nome da Unidade Orçamentária",
            "UG",
            "Nome da Unidade Gestora",
            "Data Solicitação",
            "Data Criação",
            "Tipo Empenho",
            "Dotação Orçamentária",
            "Função",
            "Subfunção",
            "Programa de Governo",
            "PAOE",
            "Natureza de Despesa",
            "Cat.Econ",
            "Grupo",
            "Modalidade",
            "Elemento",
            "Nome do Elemento",
            "Fonte",
            "Iduso",
            "Nº Emenda (EP)",
            "Autor da Emenda (EP)",
            "Nº CAC",
            "Licitação",
            "Usuário Responsável",
            "Histórico",
            "Credor",
            "Nome do Credor",
            "Data Autorização",
            "Data/Hora Cadastro Autorização",
            "Tipo de Despesa",
            "Nº ABJ",
            "Nº Processo do Sequestro Judicial",
            "Indicativo de Entrega imediata - § 4º Art. 62 Lei 8.666",
            "Indicativo de contrato",
            "Código UO Extinta",
            "Devolução GCV",
            "Mês de Competência da Folha de Pagamento",
            "Exercício de Competência da Folha de Pagamento",
            "Obrigação Patronal",
            "Tipo de Obrigação Patronal",
            "Nº NLA",
        ]
        col_order = [c for c in col_order if c in df.columns]
        if col_order:
            df = df[col_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="PED", header=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["PED"]
            cell_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            header_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            worksheet.set_default_row(12, cell_fmt)
            if len(df.columns) > 0:
                worksheet.set_column(0, len(df.columns) - 1, None, cell_fmt)
                worksheet.write_row(0, 0, df.columns, header_fmt)
                worksheet.set_row(0, None, header_fmt)
        output.seek(0)
        filename = f"ped_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_bytes(output, filename)
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


# Plan20 SEDUC
PLAN20_UPLOAD_DIR = Path("upload/plan20_seduc")
PLAN20_OUTPUT_DIR = Path("outputs/plan20_seduc")


@home_bp.route("/api/plan20/status", methods=["GET"])
@login_required
@require_feature("atualizar/plan20-seduc")
def api_plan20_status():
    def _as_iso(value):
        """
        Converte datetime ou string em isoformat; se jÃƒÂ¡ for string que nÃƒÂ£o
        parseia, devolve a string mesmo.
        """
        if value in (None, ""):
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    PLAN20_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    PLAN20_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    registro = Plan20Upload.query.order_by(Plan20Upload.uploaded_at.desc()).first()
    if not registro:
        return jsonify({"ok": True, "last": None})
    last = {
        "user_email": registro.user_email,
        "uploaded_at": _as_iso(registro.uploaded_at),
        "data_arquivo": _as_iso(registro.data_arquivo),
        "original_filename": registro.original_filename,
        "output_filename": registro.output_filename,
    }
    return jsonify({"ok": True, "last": last})


@home_bp.route("/api/plan20/upload", methods=["POST"])
@login_required
@require_feature("atualizar/plan20-seduc")
def api_plan20_upload():
    def _as_iso(value):
        """
        Converte datetime ou string em isoformat; se jÃƒÂ¡ for string que nÃƒÂ£o
        parseia, devolve a string mesmo.
        """
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    if "arquivo" not in request.files:
        return jsonify({"error": "Arquivo ÃƒÂ© obrigatÃƒÂ³rio."}), 400
    arquivo = request.files["arquivo"]
    data_arquivo_raw = request.form.get("data_arquivo")
    if not data_arquivo_raw:
        return jsonify({"error": "Data do download ÃƒÂ© obrigatÃƒÂ³ria."}), 400
    try:
        data_arquivo = datetime.fromisoformat(data_arquivo_raw)
    except ValueError:
        return jsonify({"error": "Data do download invÃƒÂ¡lida."}), 400

    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx."}), 400

    user = session.get("user") or {}
    user_email = user.get("email") or "desconhecido"

    try:
        PLAN20_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        PLAN20_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        (PLAN20_UPLOAD_DIR / "tmp").mkdir(parents=True, exist_ok=True)
        (PLAN20_OUTPUT_DIR / "tmp").mkdir(parents=True, exist_ok=True)

        for f in PLAN20_UPLOAD_DIR.glob("*.xlsx"):
            dest = PLAN20_UPLOAD_DIR / "tmp" / f"{f.stem}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{f.suffix}"
            try:
                f.rename(dest)
            except OSError:
                pass
        for f in PLAN20_OUTPUT_DIR.glob("*.xlsx"):
            dest = PLAN20_OUTPUT_DIR / "tmp" / f"{f.stem}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{f.suffix}"
            try:
                f.rename(dest)
            except OSError:
                pass

        stored_name = f"plan20_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = PLAN20_UPLOAD_DIR / stored_name
        arquivo.save(save_path)

        layout_error = _validate_upload_layout(save_path, "plan20", "PLAN20")
        if layout_error:
            try:
                save_path.unlink(missing_ok=True)
            except Exception:
                pass
            return jsonify({"error": layout_error}), 400

        output_path = run_plan20(save_path, PLAN20_OUTPUT_DIR)

        registro = Plan20Upload(
            user_email=user_email,
            original_filename=arquivo.filename,
            stored_filename=stored_name,
            output_filename=output_path.name if output_path else None,
            data_arquivo=data_arquivo,
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(registro)
        db.session.commit()

        # Insere dados na tabela plan20_seduc a partir do arquivo processado
        try:
            df_out = pd.read_excel(output_path, sheet_name="Plan20_SEDUC")
            if not df_out.empty:
                col_map = {
                    "ExercÃƒÂ­cio": "exercicio",
                    "Programa": "programa",
                    "FunÃƒÂ§ÃƒÂ£o": "funcao",
                    "Unidade OrÃƒÂ§amentÃƒÂ¡ria": "unidade_orcamentaria",
                    "AÃƒÂ§ÃƒÂ£o (P/A/OE)": "acao_paoe",
                    "SubfunÃƒÂ§ÃƒÂ£o": "subfuncao",
                    "Objetivo EspecÃƒÂ­fico": "objetivo_especifico",
                    "Esfera": "esfera",
                    "ResponsÃƒÂ¡vel pela AÃƒÂ§ÃƒÂ£o": "responsavel_acao",
                    "Produto(s) da AÃƒÂ§ÃƒÂ£o": "produto_acao",
                    "Unidade de Medida do Produto": "unid_medida_produto",
                    "RegiÃƒÂ£o do Produto": "regiao_produto",
                    "Meta do Produto": "meta_produto",
                    "Saldo Meta do Produto": "saldo_meta_produto",
                    "PÃƒÂºblico Transversal": "publico_transversal",
                    "SubAÃƒÂ§ÃƒÂ£o/entrega": "subacao_entrega",
                    "ResponsÃƒÂ¡vel": "responsavel",
                    "Prazo": "prazo",
                    "Unid. Gestora": "unid_gestora",
                    "Unidade Setorial de Planejamento": "unidade_setorial_planejamento",
                    "Produto da SubAÃƒÂ§ÃƒÂ£o": "produto_subacao",
                    "Unidade de Medida": "unidade_medida",
                    "RegiÃƒÂ£o da SubAÃƒÂ§ÃƒÂ£o": "regiao_subacao",
                    "CÃƒÂ³digo": "codigo",
                    "MunicÃƒÂ­pio(s) da entrega": "municipios_entrega",
                    "Meta da SubAÃƒÂ§ÃƒÂ£o": "meta_subacao",
                    "Detalhamento do produto": "detalhamento_produto",
                    "Etapa": "etapa",
                    "ResponsÃƒÂ¡vel da Etapa": "responsavel_etapa",
                    "Prazo da Etapa": "prazo_etapa",
                    "RegiÃƒÂ£o da Etapa": "regiao_etapa",
                    "Natureza": "natureza",
                    "Fonte": "fonte",
                    "IDU": "idu",
                    "DescriÃƒÂ§ÃƒÂ£o do Item de Despesa": "descricao_item_despesa",
                    "Unid. Medida": "unid_medida_item",
                    "Quantidade": "quantidade",
                    "Valor UnitÃƒÂ¡rio": "valor_unitario",
                    "Valor Total": "valor_total",
                    "Chave de Planejamento": "chave_planejamento",
                    "RegiÃƒÂ£o": "regiao",
                    "SubfunÃƒÂ§ÃƒÂ£o + UG": "subfuncao_ug",
                    "ADJ": "adj",
                    "Macropolitica": "macropolitica",
                    "Pilar": "pilar",
                    "Eixo": "eixo",
                    "Politica_Decreto": "politica_decreto",
                    "PÃƒÂºblico Transversal (chave)": "publico_transversal_chave",
                    "Cat.Econ": "cat_econ",
                    "Grupo": "grupo",
                    "Modalidade": "modalidade",
                    "Elemento": "elemento",
                    "Subelemento": "subelemento",
                }

                def _norm_col(name: str) -> str:
                    base = unicodedata.normalize("NFKD", str(name or ""))
                    ascii_only = "".join(ch for ch in base if not unicodedata.combining(ch))
                    return ascii_only.lower().replace(" ", "").replace("_", "").replace(".", "").replace("/", "")

                norm_map = {_norm_col(src): dst for src, dst in col_map.items()}
                rename_dict = {}
                for col in df_out.columns:
                    norm = _norm_col(col)
                    if norm in norm_map:
                        rename_dict[col] = norm_map[norm]
                df_out = df_out.rename(columns=rename_dict)

                meta_cols = {
                    "data_atualizacao",
                    "ano",
                    "data_arquivo",
                    "user_email",
                    "ativo",
                }
                keep_cols = list(col_map.values()) + list(meta_cols)
                for col in keep_cols:
                    if col not in df_out.columns:
                        df_out[col] = None
                df_out = df_out[[c for c in keep_cols if c in df_out.columns]]

                # Converte colunas numericas para evitar erro de cast (usa formato pt-BR)
                def _to_numeric_br(series):
                    return pd.to_numeric(
                        series.astype(str)
                        .str.replace(".", "", regex=False)
                        .str.replace(",", ".", regex=False),
                        errors="coerce",
                    )

                # Apenas colunas realmente numÃƒÂ©ricas no banco
                numeric_cols = [
                    "exercicio",
                    "quantidade",
                    "valor_unitario",
                    "valor_total",
                ]
                for col in numeric_cols:
                    if col in df_out.columns:
                        df_out[col] = _to_numeric_br(df_out[col])

                now = datetime.utcnow()
                df_out["data_atualizacao"] = now
                df_out["data_arquivo"] = data_arquivo
                df_out["user_email"] = user_email
                df_out["ativo"] = True
                if "exercicio" in df_out.columns:
                    df_out["ano"] = pd.to_numeric(df_out["exercicio"], errors="coerce")
                else:
                    df_out["ano"] = None
                # Desativa somente registros do mesmo exercicio+unidade_orcamentaria
                combos = set()
                if "unidade_orcamentaria" in df_out.columns and "exercicio" in df_out.columns:
                    for _, uo, ex in df_out[["unidade_orcamentaria", "exercicio"]].dropna().itertuples():
                        try:
                            ex_int = int(ex)
                        except (TypeError, ValueError):
                            continue
                        combos.add((str(uo).strip(), ex_int))
                for uo, ex in combos:
                    db.session.execute(
                        text(
                            "UPDATE plan20_seduc SET ativo = 0 WHERE unidade_orcamentaria = :uo AND exercicio = :ex"
                        ),
                        {"uo": uo, "ex": ex},
                    )
                db.session.commit()
                df_out.to_sql("plan20_seduc", db.engine, if_exists="append", index=False)
        except Exception as exc:
            db.session.rollback()
            return jsonify({"error": f"Plan20 processado, mas falha ao gravar no banco: {exc}"}), 500

        return jsonify(
            {
                "ok": True,
                "message": "Plan20 processado com sucesso.",
                "output": output_path.name,
                "last": {
                    "user_email": user_email,
                    "uploaded_at": _as_iso(registro.uploaded_at),
                    "data_arquivo": _as_iso(data_arquivo),
                    "original_filename": arquivo.filename,
                    "output_filename": output_path.name,
                },
            }
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao processar: {exc}"}), 500


@home_bp.route("/api/plan20/download/<path:filename>", methods=["GET"])
@login_required
@require_feature("atualizar/plan20-seduc")
def api_plan20_download(filename):
    target = PLAN20_OUTPUT_DIR / filename
    if not target.exists():
        abort(404)
    return send_file(target, as_attachment=True, download_name=target.name)


@home_bp.route("/api/relatorios/plan20-seduc", methods=["GET"])
@login_required
@require_feature("relatorios/plan20-seduc")
def api_relatorio_plan20():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        publico_transversal_chave,
                        programa,
                        funcao,
                        unidade_orcamentaria,
                        acao_paoe,
                        subfuncao,
                        objetivo_especifico,
                        esfera,
                        responsavel_acao,
                        produto_acao,
                        unid_medida_produto,
                        regiao_produto,
                        meta_produto,
                        saldo_meta_produto,
                        publico_transversal,
                        subacao_entrega,
                        responsavel,
                        prazo,
                        unid_gestora,
                        unidade_setorial_planejamento,
                        produto_subacao,
                        unidade_medida,
                        regiao_subacao,
                        codigo,
                        municipios_entrega,
                        meta_subacao,
                        detalhamento_produto,
                        etapa,
                        responsavel_etapa,
                        prazo_etapa,
                        regiao_etapa,
                        natureza,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        subelemento,
                        fonte,
                        idu,
                        descricao_item_despesa,
                        unid_medida_item,
                        quantidade,
                        valor_unitario,
                        valor_total
                    FROM plan20_seduc
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )

        last_upload = Plan20Upload.query.order_by(Plan20Upload.uploaded_at.desc()).first()
        data_arquivo = _as_iso(getattr(last_upload, "data_arquivo", None)) if last_upload else None
        uploaded_at = _as_iso(getattr(last_upload, "uploaded_at", None)) if last_upload else None
        user_email = last_upload.user_email if last_upload else None

        data = []
        for r in rows:
            data.append(
                {
                    "exercicio": r.get("exercicio"),
                    "chave_planejamento": r.get("chave_planejamento"),
                    "regiao": r.get("regiao"),
                    "subfuncao_ug": r.get("subfuncao_ug"),
                    "adj": r.get("adj"),
                    "macropolitica": r.get("macropolitica"),
                    "pilar": r.get("pilar"),
                    "eixo": r.get("eixo"),
                    "politica_decreto": r.get("politica_decreto"),
                    "publico_transversal_chave": r.get("publico_transversal_chave"),
                    "programa": r.get("programa"),
                    "funcao": r.get("funcao"),
                    "unidade_orcamentaria": r.get("unidade_orcamentaria"),
                    "acao_paoe": r.get("acao_paoe"),
                    "subfuncao": r.get("subfuncao"),
                    "objetivo_especifico": r.get("objetivo_especifico"),
                    "esfera": r.get("esfera"),
                    "responsavel_acao": r.get("responsavel_acao"),
                    "produto_acao": r.get("produto_acao"),
                    "unid_medida_produto": r.get("unid_medida_produto"),
                    "regiao_produto": r.get("regiao_produto"),
                    "meta_produto": r.get("meta_produto"),
                    "saldo_meta_produto": r.get("saldo_meta_produto"),
                    "publico_transversal": r.get("publico_transversal"),
                    "subacao_entrega": r.get("subacao_entrega"),
                    "responsavel": r.get("responsavel"),
                    "prazo": r.get("prazo"),
                    "unid_gestora": r.get("unid_gestora"),
                    "unidade_setorial_planejamento": r.get("unidade_setorial_planejamento"),
                    "produto_subacao": r.get("produto_subacao"),
                    "unidade_medida": r.get("unidade_medida"),
                    "regiao_subacao": r.get("regiao_subacao"),
                    "codigo": r.get("codigo"),
                    "municipios_entrega": r.get("municipios_entrega"),
                    "meta_subacao": r.get("meta_subacao"),
                    "detalhamento_produto": r.get("detalhamento_produto"),
                    "etapa": r.get("etapa"),
                    "responsavel_etapa": r.get("responsavel_etapa"),
                    "prazo_etapa": r.get("prazo_etapa"),
                    "regiao_etapa": r.get("regiao_etapa"),
                    "natureza": r.get("natureza"),
                    "cat_econ": r.get("cat_econ"),
                    "grupo": r.get("grupo"),
                    "modalidade": r.get("modalidade"),
                    "elemento": r.get("elemento"),
                    "subelemento": r.get("subelemento"),
                    "fonte": r.get("fonte"),
                    "idu": r.get("idu"),
                    "descricao_item_despesa": r.get("descricao_item_despesa"),
                    "unid_medida_item": r.get("unid_medida_item"),
                    "quantidade": _to_float(r.get("quantidade")),
                    "valor_unitario": _to_float(r.get("valor_unitario")),
                    "valor_total": _to_float(r.get("valor_total")),
                }
            )

        return jsonify(
            {
                "ok": True,
                "data": data,
                "data_arquivo": data_arquivo,
                "uploaded_at": uploaded_at,
                "user_email": user_email,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados: {exc}"}), 500


@home_bp.route("/api/relatorios/emp", methods=["GET"])
@login_required
@require_feature("relatorios/emp")
def api_relatorio_emp():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_date(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y")
        return str(val)

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        chave,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        exercicio,
                        numero_emp,
                        numero_ped,
                        valor_emp,
                        devolucao_gcv,
                        valor_emp_devolucao_gcv,
                        uo,
                        nome_unidade_orcamentaria,
                        ug,
                        nome_unidade_gestora,
                        dotacao_orcamentaria,
                        funcao,
                        subfuncao,
                        programa_governo,
                        paoe,
                        natureza_despesa,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        fonte,
                        iduso,
                        historico,
                        tipo_despesa,
                        credor,
                        nome_credor,
                        cpf_cnpj_credor,
                        categoria_credor,
                        tipo_empenho,
                        situacao,
                        data_emissao,
                        data_criacao,
                        numero_contrato,
                        numero_convenio
                    FROM emp
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        last_upload = EmpUpload.query.order_by(EmpUpload.uploaded_at.desc()).first()
        data_arquivo = _as_iso(last_upload.data_arquivo) if last_upload else None
        uploaded_at = _as_iso(last_upload.uploaded_at) if last_upload else None
        user_email = last_upload.user_email if last_upload else None
        data = []
        for r in rows:
            data.append(
                {
                    "chave": r.get("chave"),
                    "chave_planejamento": r.get("chave_planejamento"),
                    "regiao": r.get("regiao"),
                    "subfuncao_ug": r.get("subfuncao_ug"),
                    "adj": r.get("adj"),
                    "macropolitica": r.get("macropolitica"),
                    "pilar": r.get("pilar"),
                    "eixo": r.get("eixo"),
                    "politica_decreto": r.get("politica_decreto"),
                    "exercicio": r.get("exercicio"),
                    "numero_emp": r.get("numero_emp"),
                    "numero_ped": r.get("numero_ped"),
                    "valor_emp": _to_float(r.get("valor_emp")),
                    "devolucao_gcv": _to_float(r.get("devolucao_gcv")),
                    "valor_emp_devolucao_gcv": _to_float(r.get("valor_emp_devolucao_gcv")),
                    "uo": r.get("uo"),
                    "nome_unidade_orcamentaria": r.get("nome_unidade_orcamentaria"),
                    "ug": r.get("ug"),
                    "nome_unidade_gestora": r.get("nome_unidade_gestora"),
                    "dotacao_orcamentaria": r.get("dotacao_orcamentaria"),
                    "funcao": r.get("funcao"),
                    "subfuncao": r.get("subfuncao"),
                    "programa_governo": r.get("programa_governo"),
                    "paoe": r.get("paoe"),
                    "natureza_despesa": r.get("natureza_despesa"),
                    "cat_econ": r.get("cat_econ"),
                    "grupo": r.get("grupo"),
                    "modalidade": r.get("modalidade"),
                    "elemento": r.get("elemento"),
                    "fonte": r.get("fonte"),
                    "iduso": r.get("iduso"),
                    "historico": r.get("historico"),
                    "tipo_despesa": r.get("tipo_despesa"),
                    "credor": r.get("credor"),
                    "nome_credor": r.get("nome_credor"),
                    "cpf_cnpj_credor": r.get("cpf_cnpj_credor"),
                    "categoria_credor": r.get("categoria_credor"),
                    "tipo_empenho": r.get("tipo_empenho"),
                    "situacao": r.get("situacao"),
                    "data_emissao": _format_date(r.get("data_emissao")),
                    "data_criacao": _format_date(r.get("data_criacao")),
                    "numero_contrato": r.get("numero_contrato"),
                    "numero_convenio": r.get("numero_convenio"),
                }
            )
        return jsonify(
            {
                "ok": True,
                "data": data,
                "data_arquivo": data_arquivo,
                "uploaded_at": uploaded_at,
                "user_email": user_email,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados do EMP: {exc}"}), 500


@home_bp.route("/api/relatorios/dotacao", methods=["GET"])
@login_required
@require_feature("relatorios/dotacao")
def api_relatorio_dotacao():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        return str(value)

    try:
        rows = (
            Dotacao.query.filter(
                or_(
                    Dotacao.ativo == True,  # noqa: E712
                    func.lower(Dotacao.status_aprovacao) == "rejeitado",
                )
            )
            .order_by(Dotacao.id.desc())
            .all()
        )
        if not rows:
            return jsonify({"ok": True, "data": []})

        perfil_ids = [r.perfil_id for r in rows if getattr(r, "perfil_id", None)]
        adj_map = {}
        if perfil_ids:
            perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
            adj_map = {p.id: p.nome for p in perfis if p and p.nome}

        user_ids = [r.usuarios_id for r in rows if getattr(r, "usuarios_id", None)]
        aprov_ids = []
        for r in rows:
            try:
                if getattr(r, "aprovado_por", None):
                    aprov_ids.append(int(r.aprovado_por))
            except Exception:
                pass
        user_ids = list({*user_ids, *aprov_ids})
        user_map = {}
        if user_ids:
            usuarios = Usuario.query.filter(Usuario.id.in_(user_ids)).all()
            user_map = {u.id: (u.nome or "", u.perfil or "") for u in usuarios}

        data = []
        for r in rows:
            adj_nome = (adj_map.get(r.perfil_id) or "").strip()
            criado_nome, criado_perfil = user_map.get(getattr(r, "usuarios_id", None), ("", ""))
            aprov_nome, aprov_perfil = ("", "")
            try:
                aprov_nome, aprov_perfil = user_map.get(int(r.aprovado_por), ("", ""))
            except Exception:
                aprov_nome, aprov_perfil = ("", "")
            usuario_nome_perfil = ""
            if criado_nome:
                usuario_nome_perfil = f"{criado_nome} - {criado_perfil}".strip(" -")
            aprovado_nome_perfil = ""
            if aprov_nome:
                aprovado_nome_perfil = f"{aprov_nome} - {aprov_perfil}".strip(" -")
            data.append(
                {
                    "exercicio": r.exercicio,
                    "status_aprovacao": r.status_aprovacao,
                    "adjunta_solicitante": adj_nome,
                    "adj_concedente": r.adj_concedente,
                    "chave_dotacao": r.chave_dotacao,
                    "chave_planejamento": r.chave_planejamento,
                    "valor_dotacao": _to_float(r.valor_dotacao),
                    "valor_estorno": _to_float(r.valor_estorno),
                    "valor_ped_emp": _to_float(r.valor_ped_emp),
                    "valor_atual": _to_float(r.valor_atual),
                    "situacao": r.situacao,
                    "uo": r.uo,
                    "programa": r.programa,
                    "acao_paoe": r.acao_paoe,
                    "produto": r.produto,
                    "ug": r.ug,
                    "regiao": r.regiao,
                    "subacao_entrega": r.subacao_entrega,
                    "etapa": r.etapa,
                    "natureza_despesa": r.natureza_despesa,
                    "elemento": r.elemento,
                    "subelemento": r.subelemento,
                    "fonte": r.fonte,
                    "iduso": r.iduso,
                    "justificativa_historico": r.justificativa_historico,
                    "usuario_nome_perfil": usuario_nome_perfil,
                    "criado_em": _as_iso(r.criado_em),
                    "alterado_em": _as_iso(r.alterado_em),
                    "aprovado_por_nome_perfil": aprovado_nome_perfil,
                    "data_aprovacao": _as_iso(r.data_aprovacao),
                    "motivo_rejeicao": r.motivo_rejeicao,
                }
            )

        return jsonify({"ok": True, "data": data})
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados da Dota\u00e7\u00e3o: {exc}"}), 500


@home_bp.route("/api/relatorios/est-dotacao", methods=["GET"])
@login_required
@require_feature("relatorios/est-dotacao")
def api_relatorio_est_dotacao():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        return str(value)

    try:
        rows = db.session.execute(
            text(
                """
                SELECT *
                FROM est_dotacao
                WHERE ativo = 1 OR lower(status_aprovacao) = 'rejeitado'
                ORDER BY id DESC
                """
            )
        ).mappings().all()
        if not rows:
            return jsonify({"ok": True, "data": []})

        perfil_ids = [r.get("perfil_id") for r in rows if r.get("perfil_id")]
        adj_map = {}
        if perfil_ids:
            perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
            adj_map = {p.id: p.nome for p in perfis if p and p.nome}

        user_ids = [r.get("usuarios_id") for r in rows if r.get("usuarios_id")]
        aprov_ids = []
        for r in rows:
            try:
                if r.get("aprovado_por"):
                    aprov_ids.append(int(r.get("aprovado_por")))
            except Exception:
                pass
        user_ids = list({*user_ids, *aprov_ids})
        user_map = {}
        if user_ids:
            usuarios = Usuario.query.filter(Usuario.id.in_(user_ids)).all()
            user_map = {u.id: (u.nome or "", u.perfil or "") for u in usuarios}

        data = []
        for r in rows:
            adj_nome = (adj_map.get(r.get("perfil_id")) or "").strip()
            criado_nome, criado_perfil = user_map.get(r.get("usuarios_id"), ("", ""))
            aprov_nome, aprov_perfil = ("", "")
            try:
                aprov_nome, aprov_perfil = user_map.get(int(r.get("aprovado_por")), ("", ""))
            except Exception:
                aprov_nome, aprov_perfil = ("", "")
            usuario_nome_perfil = ""
            if criado_nome:
                usuario_nome_perfil = f"{criado_nome} - {criado_perfil}".strip(" -")
            aprovado_nome_perfil = ""
            if aprov_nome:
                aprovado_nome_perfil = f"{aprov_nome} - {aprov_perfil}".strip(" -")
            data.append(
                {
                    "exercicio": r.get("exercicio"),
                    "status_aprovacao": r.get("status_aprovacao"),
                    "adjunta_solicitante": adj_nome,
                    "chave_dotacao": r.get("chave_dotacao"),
                    "chave_planejamento": r.get("chave_planejamento"),
                    "valor_dotacao": _to_float(r.get("valor_dotacao")),
                    "valor_a_ser_est": _to_float(r.get("valor_a_ser_est")),
                    "saldo_dotacao_apos": _to_float(r.get("saldo_dotacao_apos")),
                    "situacao": r.get("situacao"),
                    "uo": r.get("uo"),
                    "programa": r.get("programa"),
                    "acao_paoe": r.get("acao_paoe"),
                    "produto": r.get("produto"),
                    "ug": r.get("ug"),
                    "regiao": r.get("regiao"),
                    "subacao_entrega": r.get("subacao_entrega"),
                    "etapa": r.get("etapa"),
                    "natureza_despesa": r.get("natureza_despesa"),
                    "elemento": r.get("elemento"),
                    "subelemento": r.get("subelemento"),
                    "fonte": r.get("fonte"),
                    "iduso": r.get("iduso"),
                    "justificativa": r.get("justificativa"),
                    "usuario_nome_perfil": usuario_nome_perfil,
                    "criado_em": _as_iso(r.get("criado_em")),
                    "alterado_em": _as_iso(r.get("alterado_em")),
                    "aprovado_por_nome_perfil": aprovado_nome_perfil,
                    "data_aprovacao": _as_iso(r.get("data_aprovacao")),
                    "motivo_rejeicao": r.get("motivo_rejeicao"),
                }
            )

        return jsonify({"ok": True, "data": data})
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados do Estorno: {exc}"}), 500


@home_bp.route("/api/relatorios/est-emp", methods=["GET"])
@login_required
@require_feature("relatorios/est-emp")
def api_relatorio_est_emp():
    def _as_iso(value):
        if value in (None, ""):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            from datetime import datetime

            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_date(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y")
        return str(val)

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        numero_est,
                        numero_emp,
                        empenho_atual,
                        empenho_rp,
                        numero_ped,
                        valor_emp,
                        valor_est_emp_sem_aqs,
                        valor_est_emp_com_aqs,
                        valor_emp_liquido,
                        uo,
                        nome_unidade_orcamentaria,
                        ug,
                        nome_unidade_gestora,
                        dotacao_orcamentaria,
                        historico,
                        credor,
                        nome_credor,
                        cpf_cnpj_credor,
                        data_criacao,
                        data_emissao,
                        situacao,
                        rp
                    FROM est_emp
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        last_upload = EstEmpUpload.query.order_by(EstEmpUpload.uploaded_at.desc()).first()
        data_arquivo = _as_iso(getattr(last_upload, "data_arquivo", None)) if last_upload else None
        uploaded_at = _as_iso(getattr(last_upload, "uploaded_at", None)) if last_upload else None
        user_email = last_upload.user_email if last_upload else None
        data = []
        for r in rows:
            data.append(
                {
                    "exercicio": r.get("exercicio"),
                    "numero_est": r.get("numero_est"),
                    "numero_emp": r.get("numero_emp"),
                    "empenho_atual": r.get("empenho_atual"),
                    "empenho_rp": r.get("empenho_rp"),
                    "numero_ped": r.get("numero_ped"),
                    "valor_emp": _to_float(r.get("valor_emp")),
                    "valor_est_emp_sem_aqs": _to_float(r.get("valor_est_emp_sem_aqs")),
                    "valor_est_emp_com_aqs": _to_float(r.get("valor_est_emp_com_aqs")),
                    "valor_emp_liquido": _to_float(r.get("valor_emp_liquido")),
                    "uo": r.get("uo"),
                    "nome_unidade_orcamentaria": r.get("nome_unidade_orcamentaria"),
                    "ug": r.get("ug"),
                    "nome_unidade_gestora": r.get("nome_unidade_gestora"),
                    "dotacao_orcamentaria": r.get("dotacao_orcamentaria"),
                    "historico": r.get("historico"),
                    "credor": r.get("credor"),
                    "nome_credor": r.get("nome_credor"),
                    "cpf_cnpj_credor": r.get("cpf_cnpj_credor"),
                    "data_criacao": _format_date(r.get("data_criacao")),
                    "data_emissao": _format_date(r.get("data_emissao")),
                    "situacao": r.get("situacao"),
                    "rp": r.get("rp"),
                }
            )
        return jsonify(
            {
                "ok": True,
                "data": data,
                "data_arquivo": data_arquivo,
                "uploaded_at": uploaded_at,
                "user_email": user_email,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados do Est EMP: {exc}"}), 500


@home_bp.route("/api/relatorios/nob", methods=["GET"])
@login_required
@require_feature("relatorios/nob")
def api_relatorio_nob():
    def _as_iso(value):
        if not value:
            return None
        if isinstance(value, str) and value.startswith("0000-00-00"):
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        try:
            return datetime.fromisoformat(str(value)).isoformat()
        except Exception:
            return str(value)

    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_date(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y")
        return str(val)

    def _format_datetime(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y %H:%M:%S")
        return str(val)

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        numero_nob,
                        numero_nob_estorno,
                        numero_liq,
                        numero_emp,
                        empenho_atual,
                        empenho_rp,
                        numero_ped,
                        valor_nob,
                        devolucao_gcv,
                        valor_nob_gcv,
                        uo,
                        ug,
                        dotacao_orcamentaria,
                        funcao,
                        subfuncao,
                        programa_governo,
                        paoe,
                        natureza_despesa,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        nome_elemento_despesa,
                        fonte,
                        nome_fonte_recurso,
                        iduso,
                        historico_liq,
                        nome_credor_principal,
                        cpf_cnpj_credor_principal,
                        credor,
                        nome_credor,
                        cpf_cnpj_credor,
                        data_nob,
                        data_cadastro_nob,
                        data_hora_cadastro_liq
                    FROM nob
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        last_upload = NobUpload.query.order_by(NobUpload.uploaded_at.desc()).first()
        data_arquivo = _as_iso(getattr(last_upload, "data_arquivo", None)) if last_upload else None
        uploaded_at = _as_iso(getattr(last_upload, "uploaded_at", None)) if last_upload else None
        user_email = last_upload.user_email if last_upload else None
        data = []
        for r in rows:
            data.append(
                {
                    "exercicio": r.get("exercicio"),
                    "numero_nob": r.get("numero_nob"),
                    "numero_nob_estorno": r.get("numero_nob_estorno"),
                    "numero_liq": r.get("numero_liq"),
                    "numero_emp": r.get("numero_emp"),
                    "empenho_atual": r.get("empenho_atual"),
                    "empenho_rp": r.get("empenho_rp"),
                    "numero_ped": r.get("numero_ped"),
                    "valor_nob": _to_float(r.get("valor_nob")),
                    "devolucao_gcv": _to_float(r.get("devolucao_gcv")),
                    "valor_nob_gcv": _to_float(r.get("valor_nob_gcv")),
                    "uo": r.get("uo"),
                    "ug": r.get("ug"),
                    "dotacao_orcamentaria": r.get("dotacao_orcamentaria"),
                    "funcao": r.get("funcao"),
                    "subfuncao": r.get("subfuncao"),
                    "programa_governo": r.get("programa_governo"),
                    "paoe": r.get("paoe"),
                    "natureza_despesa": r.get("natureza_despesa"),
                    "cat_econ": r.get("cat_econ"),
                    "grupo": r.get("grupo"),
                    "modalidade": r.get("modalidade"),
                    "elemento": r.get("elemento"),
                    "nome_elemento_despesa": r.get("nome_elemento_despesa"),
                    "fonte": r.get("fonte"),
                    "nome_fonte_recurso": r.get("nome_fonte_recurso"),
                    "iduso": r.get("iduso"),
                    "historico_liq": r.get("historico_liq"),
                    "nome_credor_principal": r.get("nome_credor_principal"),
                    "cpf_cnpj_credor_principal": r.get("cpf_cnpj_credor_principal"),
                    "credor": r.get("credor"),
                    "nome_credor": r.get("nome_credor"),
                    "cpf_cnpj_credor": r.get("cpf_cnpj_credor"),
                    "data_nob": _format_date(r.get("data_nob")),
                    "data_cadastro_nob": _format_date(r.get("data_cadastro_nob")),
                    "data_hora_cadastro_liq": _format_datetime(r.get("data_hora_cadastro_liq")),
                }
            )
        return jsonify(
            {
                "ok": True,
                "data": data,
                "data_arquivo": data_arquivo,
                "uploaded_at": uploaded_at,
                "user_email": user_email,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar dados do NOB: {exc}"}), 500


@home_bp.route("/api/relatorios/nob/download", methods=["GET"])
@login_required
@require_feature("relatorios/nob")
def api_relatorio_nob_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_date(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y")
        return str(val)

    def _format_datetime(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y %H:%M:%S")
        return str(val)

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        numero_nob,
                        numero_nob_estorno,
                        numero_liq,
                        numero_emp,
                        empenho_atual,
                        empenho_rp,
                        numero_ped,
                        valor_nob,
                        devolucao_gcv,
                        valor_nob_gcv,
                        uo,
                        ug,
                        dotacao_orcamentaria,
                        funcao,
                        subfuncao,
                        programa_governo,
                        paoe,
                        natureza_despesa,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        nome_elemento_despesa,
                        fonte,
                        nome_fonte_recurso,
                        iduso,
                        historico_liq,
                        nome_credor_principal,
                        cpf_cnpj_credor_principal,
                        credor,
                        nome_credor,
                        cpf_cnpj_credor,
                        data_nob,
                        data_cadastro_nob,
                        data_hora_cadastro_liq
                    FROM nob
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404
        db.session.close()

        df = pd.DataFrame(rows)
        for col in ("valor_nob", "devolucao_gcv", "valor_nob_gcv"):
            if col in df.columns:
                df[col] = df[col].apply(_to_float)
        for col in ("data_nob", "data_cadastro_nob"):
            if col in df.columns:
                df[col] = df[col].apply(_format_date)
        if "data_hora_cadastro_liq" in df.columns:
            df["data_hora_cadastro_liq"] = df["data_hora_cadastro_liq"].apply(_format_datetime)

        rename_map = {
            "exercicio": "Exercicio",
            "numero_nob": "N\u00ba NOB",
            "numero_nob_estorno": "N\u00ba NOB Estorno/Estornado",
            "numero_liq": "N\u00ba LIQ",
            "numero_emp": "N\u00ba EMP",
            "empenho_atual": "Empenho Atual",
            "empenho_rp": "Empenho RP",
            "numero_ped": "N\u00ba PED",
            "valor_nob": "Valor NOB",
            "devolucao_gcv": "Devolu\u00e7\u00e3o GCV",
            "valor_nob_gcv": "Valor NOB - GCV",
            "uo": "UO",
            "ug": "UG",
            "dotacao_orcamentaria": "Dota\u00e7\u00e3o Or\u00e7ament\u00e1ria",
            "funcao": "Fun\u00e7\u00e3o",
            "subfuncao": "Subfun\u00e7\u00e3o",
            "programa_governo": "Programa de Governo",
            "paoe": "PAOE",
            "natureza_despesa": "Natureza de Despesa",
            "cat_econ": "Cat.Econ",
            "grupo": "Grupo",
            "modalidade": "Modalidade",
            "elemento": "Elemento",
            "nome_elemento_despesa": "Nome do Elemento da Despesa",
            "fonte": "Fonte",
            "nome_fonte_recurso": "Nome da Fonte de Recurso",
            "iduso": "Iduso",
            "historico_liq": "Hist\u00f3rico LIQ",
            "nome_credor_principal": "Nome do Credor Principal",
            "cpf_cnpj_credor_principal": "CPF/CNPJ do Credor Principal",
            "credor": "Credor",
            "nome_credor": "Nome do Credor",
            "cpf_cnpj_credor": "CPF/CNPJ do Credor",
            "data_nob": "Data NOB",
            "data_cadastro_nob": "Data Cadastro NOB",
            "data_hora_cadastro_liq": "Data/Hora de Cadastro da LIQ",
        }
        df.rename(columns=rename_map, inplace=True)

        col_order = [
            "Exercicio",
            "N\u00ba NOB",
            "N\u00ba NOB Estorno/Estornado",
            "N\u00ba LIQ",
            "N\u00ba EMP",
            "Empenho Atual",
            "Empenho RP",
            "N\u00ba PED",
            "Valor NOB",
            "Devolu\u00e7\u00e3o GCV",
            "Valor NOB - GCV",
            "UO",
            "UG",
            "Dota\u00e7\u00e3o Or\u00e7ament\u00e1ria",
            "Fun\u00e7\u00e3o",
            "Subfun\u00e7\u00e3o",
            "Programa de Governo",
            "PAOE",
            "Natureza de Despesa",
            "Cat.Econ",
            "Grupo",
            "Modalidade",
            "Elemento",
            "Nome do Elemento da Despesa",
            "Fonte",
            "Nome da Fonte de Recurso",
            "Iduso",
            "Hist\u00f3rico LIQ",
            "Nome do Credor Principal",
            "CPF/CNPJ do Credor Principal",
            "Credor",
            "Nome do Credor",
            "CPF/CNPJ do Credor",
            "Data NOB",
            "Data Cadastro NOB",
            "Data/Hora de Cadastro da LIQ",
        ]
        col_order = [c for c in col_order if c in df.columns]
        if col_order:
            df = df[col_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="NOB", header=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["NOB"]
            cell_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            header_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            worksheet.set_default_row(12, cell_fmt)
            if len(df.columns) > 0:
                worksheet.set_column(0, len(df.columns) - 1, None, cell_fmt)
                worksheet.write_row(0, 0, df.columns, header_fmt)
                worksheet.set_row(0, None, header_fmt)
        output.seek(0)
        filename = f"nob_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_bytes(output, filename)
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/relatorios/emp/download", methods=["GET"])
@login_required
@require_feature("relatorios/emp")
def api_relatorio_emp_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_date(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y")
        return str(val)

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        chave,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        exercicio,
                        numero_emp,
                        numero_ped,
                        valor_emp,
                        devolucao_gcv,
                        valor_emp_devolucao_gcv,
                        uo,
                        nome_unidade_orcamentaria,
                        ug,
                        nome_unidade_gestora,
                        dotacao_orcamentaria,
                        funcao,
                        subfuncao,
                        programa_governo,
                        paoe,
                        natureza_despesa,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        fonte,
                        iduso,
                        historico,
                        tipo_despesa,
                        credor,
                        nome_credor,
                        cpf_cnpj_credor,
                        categoria_credor,
                        tipo_empenho,
                        situacao,
                        data_emissao,
                        data_criacao,
                        numero_contrato,
                        numero_convenio
                    FROM emp
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404
        db.session.close()

        df = pd.DataFrame(rows)

        def _chave_display(row):
            try:
                ex = int(str(row.get("exercicio") or 0)[:4])
            except Exception:
                ex = 0
            chave = row.get("chave") or ""
            chave_plan = row.get("chave_planejamento") or ""
            if ex >= 2026 and chave:
                return chave
            return chave_plan or chave

        df["Chave / Chave de Planejamento"] = df.apply(_chave_display, axis=1)
        df.drop(columns=["chave", "chave_planejamento"], inplace=True, errors="ignore")
        for col in ("valor_emp", "devolucao_gcv", "valor_emp_devolucao_gcv"):
            if col in df.columns:
                df[col] = df[col].apply(_to_float)
        for col in ("data_emissao", "data_criacao"):
            if col in df.columns:
                df[col] = df[col].apply(_format_date)

        rename_map = {
            "regiao": "Região",
            "subfuncao_ug": "Subfunção + UG",
            "adj": "ADJ",
            "macropolitica": "Macropolítica",
            "pilar": "Pilar",
            "eixo": "Eixo",
            "politica_decreto": "Política_Decreto",
            "exercicio": "Exercício",
            "numero_emp": "Nº EMP",
            "numero_ped": "Nº PED",
            "valor_emp": "Valor EMP",
            "devolucao_gcv": "Devolução GCV",
            "valor_emp_devolucao_gcv": "Valor EMP-Devolução GCV",
            "uo": "UO",
            "nome_unidade_orcamentaria": "Nome da Unidade Orçamentária",
            "ug": "UG",
            "nome_unidade_gestora": "Nome da Unidade Gestora",
            "dotacao_orcamentaria": "Dotação Orçamentária",
            "funcao": "Função",
            "subfuncao": "Subfunção",
            "programa_governo": "Programa de Governo",
            "paoe": "PAOE",
            "natureza_despesa": "Natureza de Despesa",
            "cat_econ": "Cat.Econ",
            "grupo": "Grupo",
            "modalidade": "Modalidade",
            "elemento": "Elemento",
            "fonte": "Fonte",
            "iduso": "Iduso",
            "historico": "Histórico",
            "tipo_despesa": "Tipo de Despesa",
            "credor": "Credor",
            "nome_credor": "Nome do Credor",
            "cpf_cnpj_credor": "CPF/CNPJ do Credor",
            "categoria_credor": "Categoria do Credor",
            "tipo_empenho": "Tipo Empenho",
            "situacao": "Situação",
            "data_emissao": "Data emissão",
            "data_criacao": "Data criação",
            "numero_contrato": "Nº Contrato",
            "numero_convenio": "Nº Convênio",
        }
        df.rename(columns=rename_map, inplace=True)

        col_order = [
            "Chave / Chave de Planejamento",
            "Região",
            "Subfunção + UG",
            "ADJ",
            "Macropolítica",
            "Pilar",
            "Eixo",
            "Política_Decreto",
            "Exercício",
            "Nº EMP",
            "Nº PED",
            "Valor EMP",
            "Devolução GCV",
            "Valor EMP-Devolução GCV",
            "UO",
            "Nome da Unidade Orçamentária",
            "UG",
            "Nome da Unidade Gestora",
            "Dotação Orçamentária",
            "Função",
            "Subfunção",
            "Programa de Governo",
            "PAOE",
            "Natureza de Despesa",
            "Cat.Econ",
            "Grupo",
            "Modalidade",
            "Elemento",
            "Fonte",
            "Iduso",
            "Histórico",
            "Tipo de Despesa",
            "Credor",
            "Nome do Credor",
            "CPF/CNPJ do Credor",
            "Categoria do Credor",
            "Tipo Empenho",
            "Situação",
            "Data emissão",
            "Data criação",
            "Nº Contrato",
            "Nº Convênio",
        ]
        col_order = [c for c in col_order if c in df.columns]
        if col_order:
            df = df[col_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="EMP", header=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["EMP"]
            cell_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            header_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            worksheet.set_default_row(12, cell_fmt)
            if len(df.columns) > 0:
                worksheet.set_column(0, len(df.columns) - 1, None, cell_fmt)
                worksheet.write_row(0, 0, df.columns, header_fmt)
                worksheet.set_row(0, None, header_fmt)
        output.seek(0)
        filename = f"emp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_bytes(output, filename)
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/relatorios/dotacao/download", methods=["GET"])
@login_required
@require_feature("relatorios/dotacao")
def api_relatorio_dotacao_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_dt(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y %H:%M:%S")
        return str(val)

    try:
        rows = (
            Dotacao.query.filter(
                or_(
                    Dotacao.ativo == True,  # noqa: E712
                    func.lower(Dotacao.status_aprovacao) == "rejeitado",
                )
            )
            .order_by(Dotacao.id.desc())
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404

        perfil_ids = [r.perfil_id for r in rows if getattr(r, "perfil_id", None)]
        adj_map = {}
        if perfil_ids:
            perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
            adj_map = {p.id: p.nome for p in perfis if p and p.nome}

        user_ids = [r.usuarios_id for r in rows if getattr(r, "usuarios_id", None)]
        aprov_ids = []
        for r in rows:
            try:
                if getattr(r, "aprovado_por", None):
                    aprov_ids.append(int(r.aprovado_por))
            except Exception:
                pass
        user_ids = list({*user_ids, *aprov_ids})
        user_map = {}
        if user_ids:
            usuarios = Usuario.query.filter(Usuario.id.in_(user_ids)).all()
            user_map = {u.id: (u.nome or "", u.perfil or "") for u in usuarios}

        data = []
        for r in rows:
            adj_nome = (adj_map.get(r.perfil_id) or "").strip()
            criado_nome, criado_perfil = user_map.get(getattr(r, "usuarios_id", None), ("", ""))
            aprov_nome, aprov_perfil = ("", "")
            try:
                aprov_nome, aprov_perfil = user_map.get(int(r.aprovado_por), ("", ""))
            except Exception:
                aprov_nome, aprov_perfil = ("", "")
            usuario_nome_perfil = ""
            if criado_nome:
                usuario_nome_perfil = f"{criado_nome} - {criado_perfil}".strip(" -")
            aprovado_nome_perfil = ""
            if aprov_nome:
                aprovado_nome_perfil = f"{aprov_nome} - {aprov_perfil}".strip(" -")
            data.append(
                {
                    "exercicio": r.exercicio,
                    "status_aprovacao": r.status_aprovacao,
                    "adjunta_solicitante": adj_nome,
                    "adj_concedente": r.adj_concedente,
                    "chave_dotacao": r.chave_dotacao,
                    "chave_planejamento": r.chave_planejamento,
                    "valor_dotacao": _to_float(r.valor_dotacao),
                    "valor_estorno": _to_float(r.valor_estorno),
                    "valor_ped_emp": _to_float(r.valor_ped_emp),
                    "valor_atual": _to_float(r.valor_atual),
                    "situacao": r.situacao,
                    "uo": r.uo,
                    "programa": r.programa,
                    "acao_paoe": r.acao_paoe,
                    "produto": r.produto,
                    "ug": r.ug,
                    "regiao": r.regiao,
                    "subacao_entrega": r.subacao_entrega,
                    "etapa": r.etapa,
                    "natureza_despesa": r.natureza_despesa,
                    "elemento": r.elemento,
                    "subelemento": r.subelemento,
                    "fonte": r.fonte,
                    "iduso": r.iduso,
                    "justificativa_historico": r.justificativa_historico,
                    "usuario_nome_perfil": usuario_nome_perfil,
                    "criado_em": _format_dt(r.criado_em),
                    "alterado_em": _format_dt(r.alterado_em),
                    "aprovado_por_nome_perfil": aprovado_nome_perfil,
                    "data_aprovacao": _format_dt(r.data_aprovacao),
                    "motivo_rejeicao": r.motivo_rejeicao,
                }
            )

        df = pd.DataFrame(data)
        rename_map = {
            "exercicio": "Exerc\u00edcio",
            "status_aprovacao": "Status",
            "adjunta_solicitante": "Adjunta Solicitante",
            "adj_concedente": "Adjunta Concedente",
            "chave_dotacao": "Controle de Dota\u00e7\u00e3o",
            "chave_planejamento": "Chave de Planejamento",
            "valor_dotacao": "Valor da Dota\u00e7\u00e3o",
            "valor_estorno": "Valor do Estorno",
            "valor_ped_emp": "Valor do PED/EMP",
            "valor_atual": "Valor da Dota\u00e7\u00e3o Atualizada",
            "situacao": "Situa\u00e7\u00e3o",
            "uo": "UO",
            "programa": "Programa",
            "acao_paoe": "A\u00e7\u00e3o/PAOE",
            "produto": "Produto",
            "ug": "UG",
            "regiao": "Regi\u00e3o",
            "subacao_entrega": "SubA\u00e7\u00e3o/Entrega",
            "etapa": "Etapa",
            "natureza_despesa": "Natureza de Despesa",
            "elemento": "Elemento",
            "subelemento": "Subelemento",
            "fonte": "Fonte",
            "iduso": "Iduso",
            "justificativa_historico": "Justificativa/Hist\u00f3rico",
            "usuario_nome_perfil": "Criado/Alterado por",
            "criado_em": "Criado em",
            "alterado_em": "Alterado em",
            "aprovado_por_nome_perfil": "Aprovado por",
            "data_aprovacao": "Data da Aprova\u00e7\u00e3o",
            "motivo_rejeicao": "Justificativa da Aprova\u00e7\u00e3o/Rejei\u00e7\u00e3o",
        }
        df.rename(columns=rename_map, inplace=True)

        col_order = [
            "Exerc\u00edcio",
            "Status",
            "Adjunta Solicitante",
            "Adjunta Concedente",
            "Controle de Dota\u00e7\u00e3o",
            "Chave de Planejamento",
            "Valor da Dota\u00e7\u00e3o",
            "Valor do Estorno",
            "Valor do PED/EMP",
            "Valor da Dota\u00e7\u00e3o Atualizada",
            "Situa\u00e7\u00e3o",
            "UO",
            "Programa",
            "A\u00e7\u00e3o/PAOE",
            "Produto",
            "UG",
            "Regi\u00e3o",
            "SubA\u00e7\u00e3o/Entrega",
            "Etapa",
            "Natureza de Despesa",
            "Elemento",
            "Subelemento",
            "Fonte",
            "Iduso",
            "Justificativa/Hist\u00f3rico",
            "Criado/Alterado por",
            "Criado em",
            "Alterado em",
            "Aprovado por",
            "Data da Aprova\u00e7\u00e3o",
            "Justificativa da Aprova\u00e7\u00e3o/Rejei\u00e7\u00e3o",
        ]
        col_order = [c for c in col_order if c in df.columns]
        if col_order:
            df = df[col_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Dotacao", header=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["Dotacao"]
            cell_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            header_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            worksheet.set_default_row(12, cell_fmt)
            if len(df.columns) > 0:
                worksheet.set_column(0, len(df.columns) - 1, None, cell_fmt)
                worksheet.write_row(0, 0, df.columns, header_fmt)
                worksheet.set_row(0, None, header_fmt)
        output.seek(0)
        filename = f"dotacao_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_bytes(output, filename)
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/relatorios/est-dotacao/download", methods=["GET"])
@login_required
@require_feature("relatorios/est-dotacao")
def api_relatorio_est_dotacao_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_dt(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y %H:%M:%S")
        return str(val)

    try:
        rows = db.session.execute(
            text(
                """
                SELECT *
                FROM est_dotacao
                WHERE ativo = 1 OR lower(status_aprovacao) = 'rejeitado'
                ORDER BY id DESC
                """
            )
        ).mappings().all()
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404

        perfil_ids = [r.get("perfil_id") for r in rows if r.get("perfil_id")]
        adj_map = {}
        if perfil_ids:
            perfis = Perfil.query.filter(Perfil.id.in_(perfil_ids)).all()
            adj_map = {p.id: p.nome for p in perfis if p and p.nome}

        user_ids = [r.get("usuarios_id") for r in rows if r.get("usuarios_id")]
        aprov_ids = []
        for r in rows:
            try:
                if r.get("aprovado_por"):
                    aprov_ids.append(int(r.get("aprovado_por")))
            except Exception:
                pass
        user_ids = list({*user_ids, *aprov_ids})
        user_map = {}
        if user_ids:
            usuarios = Usuario.query.filter(Usuario.id.in_(user_ids)).all()
            user_map = {u.id: (u.nome or "", u.perfil or "") for u in usuarios}

        data = []
        for r in rows:
            adj_nome = (adj_map.get(r.get("perfil_id")) or "").strip()
            criado_nome, criado_perfil = user_map.get(r.get("usuarios_id"), ("", ""))
            aprov_nome, aprov_perfil = ("", "")
            try:
                aprov_nome, aprov_perfil = user_map.get(int(r.get("aprovado_por")), ("", ""))
            except Exception:
                aprov_nome, aprov_perfil = ("", "")
            usuario_nome_perfil = ""
            if criado_nome:
                usuario_nome_perfil = f"{criado_nome} - {criado_perfil}".strip(" -")
            aprovado_nome_perfil = ""
            if aprov_nome:
                aprovado_nome_perfil = f"{aprov_nome} - {aprov_perfil}".strip(" -")
            data.append(
                {
                    "exercicio": r.get("exercicio"),
                    "status_aprovacao": r.get("status_aprovacao"),
                    "adjunta_solicitante": adj_nome,
                    "chave_dotacao": r.get("chave_dotacao"),
                    "chave_planejamento": r.get("chave_planejamento"),
                    "valor_dotacao": _to_float(r.get("valor_dotacao")),
                    "valor_a_ser_est": _to_float(r.get("valor_a_ser_est")),
                    "saldo_dotacao_apos": _to_float(r.get("saldo_dotacao_apos")),
                    "situacao": r.get("situacao"),
                    "uo": r.get("uo"),
                    "programa": r.get("programa"),
                    "acao_paoe": r.get("acao_paoe"),
                    "produto": r.get("produto"),
                    "ug": r.get("ug"),
                    "regiao": r.get("regiao"),
                    "subacao_entrega": r.get("subacao_entrega"),
                    "etapa": r.get("etapa"),
                    "natureza_despesa": r.get("natureza_despesa"),
                    "elemento": r.get("elemento"),
                    "subelemento": r.get("subelemento"),
                    "fonte": r.get("fonte"),
                    "iduso": r.get("iduso"),
                    "justificativa": r.get("justificativa"),
                    "usuario_nome_perfil": usuario_nome_perfil,
                    "criado_em": _format_dt(r.get("criado_em")),
                    "alterado_em": _format_dt(r.get("alterado_em")),
                    "aprovado_por_nome_perfil": aprovado_nome_perfil,
                    "data_aprovacao": _format_dt(r.get("data_aprovacao")),
                    "motivo_rejeicao": r.get("motivo_rejeicao"),
                }
            )

        df = pd.DataFrame(data)
        rename_map = {
            "exercicio": "Exerc\u00edcio",
            "status_aprovacao": "Status",
            "adjunta_solicitante": "Adjunta Solicitante",
            "chave_dotacao": "Controle de Dota\u00e7\u00e3o",
            "chave_planejamento": "Chave de Planejamento",
            "valor_dotacao": "Valor da Dota\u00e7\u00e3o",
            "valor_a_ser_est": "Valor do Estorno",
            "saldo_dotacao_apos": "Saldo de Dota\u00e7\u00e3o",
            "situacao": "Situa\u00e7\u00e3o",
            "uo": "UO",
            "programa": "Programa",
            "acao_paoe": "A\u00e7\u00e3o/PAOE",
            "produto": "Produto",
            "ug": "UG",
            "regiao": "Regi\u00e3o",
            "subacao_entrega": "SubA\u00e7\u00e3o/Entrega",
            "etapa": "Etapa",
            "natureza_despesa": "Natureza de Despesa",
            "elemento": "Elemento",
            "subelemento": "Subelemento",
            "fonte": "Fonte",
            "iduso": "Iduso",
            "justificativa": "Justificativa do Estorno",
            "usuario_nome_perfil": "Criado/Alterado por",
            "criado_em": "Criado em",
            "alterado_em": "Alterado em",
            "aprovado_por_nome_perfil": "Aprovado por",
            "data_aprovacao": "Data da Aprova\u00e7\u00e3o",
            "motivo_rejeicao": "Justificativa da Aprova\u00e7\u00e3o/Rejei\u00e7\u00e3o",
        }
        df.rename(columns=rename_map, inplace=True)

        col_order = [
            "Exerc\u00edcio",
            "Status",
            "Adjunta Solicitante",
            "Controle de Dota\u00e7\u00e3o",
            "Chave de Planejamento",
            "Valor da Dota\u00e7\u00e3o",
            "Valor do Estorno",
            "Saldo de Dota\u00e7\u00e3o",
            "Situa\u00e7\u00e3o",
            "UO",
            "Programa",
            "A\u00e7\u00e3o/PAOE",
            "Produto",
            "UG",
            "Regi\u00e3o",
            "SubA\u00e7\u00e3o/Entrega",
            "Etapa",
            "Natureza de Despesa",
            "Elemento",
            "Subelemento",
            "Fonte",
            "Iduso",
            "Justificativa do Estorno",
            "Criado/Alterado por",
            "Criado em",
            "Alterado em",
            "Aprovado por",
            "Data da Aprova\u00e7\u00e3o",
            "Justificativa da Aprova\u00e7\u00e3o/Rejei\u00e7\u00e3o",
        ]
        col_order = [c for c in col_order if c in df.columns]
        if col_order:
            df = df[col_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Estorno", header=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["Estorno"]
            cell_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            header_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            worksheet.set_default_row(12, cell_fmt)
            if len(df.columns) > 0:
                worksheet.set_column(0, len(df.columns) - 1, None, cell_fmt)
                worksheet.write_row(0, 0, df.columns, header_fmt)
                worksheet.set_row(0, None, header_fmt)
        output.seek(0)
        filename = f"estorno_dotacao_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_bytes(output, filename)
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/relatorios/est-emp/download", methods=["GET"])
@login_required
@require_feature("relatorios/est-emp")
def api_relatorio_est_emp_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def _format_date(val):
        if not val:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y")
        return str(val)

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        numero_est,
                        numero_emp,
                        empenho_atual,
                        empenho_rp,
                        numero_ped,
                        valor_emp,
                        valor_est_emp_sem_aqs,
                        valor_est_emp_com_aqs,
                        valor_emp_liquido,
                        uo,
                        nome_unidade_orcamentaria,
                        ug,
                        nome_unidade_gestora,
                        dotacao_orcamentaria,
                        historico,
                        credor,
                        nome_credor,
                        cpf_cnpj_credor,
                        data_criacao,
                        data_emissao,
                        situacao,
                        rp
                    FROM est_emp
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404
        db.session.close()

        df = pd.DataFrame(rows)
        for col in (
            "valor_emp",
            "valor_est_emp_sem_aqs",
            "valor_est_emp_com_aqs",
            "valor_emp_liquido",
        ):
            if col in df.columns:
                df[col] = df[col].apply(_to_float)
        for col in ("data_criacao", "data_emissao"):
            if col in df.columns:
                df[col] = df[col].apply(_format_date)

        rename_map = {
            "exercicio": "Exerc\u00edcio",
            "numero_est": "N\u00ba EST",
            "numero_emp": "N\u00ba EMP",
            "empenho_atual": "Empenho Atual",
            "empenho_rp": "Empenho RP",
            "numero_ped": "N\u00ba PED",
            "valor_emp": "Valor EMP",
            "valor_est_emp_sem_aqs": "Valor Est EMP (A LIQ/Em LIQ sem AQS)",
            "valor_est_emp_com_aqs": "Valor Est EMP (Em LIQ com AQS)",
            "valor_emp_liquido": "Valor EMP - (A LIQ/Em LIQ sem AQS) - (Em LIQ com AQS)",
            "uo": "UO",
            "nome_unidade_orcamentaria": "Nome da Unidade Or\u00e7ament\u00e1ria",
            "ug": "UG",
            "nome_unidade_gestora": "Nome da Unidade Gestora",
            "dotacao_orcamentaria": "Dota\u00e7\u00e3o Or\u00e7ament\u00e1ria",
            "historico": "Hist\u00f3rico",
            "credor": "Credor",
            "nome_credor": "Nome do Credor",
            "cpf_cnpj_credor": "CPF/CNPJ do Credor",
            "data_criacao": "Data Cria\u00e7\u00e3o",
            "data_emissao": "Data Emiss\u00e3o",
            "situacao": "Situa\u00e7\u00e3o",
            "rp": "RP",
        }
        df.rename(columns=rename_map, inplace=True)

        col_order = [
            "Exerc\u00edcio",
            "N\u00ba EST",
            "N\u00ba EMP",
            "Empenho Atual",
            "Empenho RP",
            "N\u00ba PED",
            "Valor EMP",
            "Valor Est EMP (A LIQ/Em LIQ sem AQS)",
            "Valor Est EMP (Em LIQ com AQS)",
            "Valor EMP - (A LIQ/Em LIQ sem AQS) - (Em LIQ com AQS)",
            "UO",
            "Nome da Unidade Or\u00e7ament\u00e1ria",
            "UG",
            "Nome da Unidade Gestora",
            "Dota\u00e7\u00e3o Or\u00e7ament\u00e1ria",
            "Hist\u00f3rico",
            "Credor",
            "Nome do Credor",
            "CPF/CNPJ do Credor",
            "Data Cria\u00e7\u00e3o",
            "Data Emiss\u00e3o",
            "Situa\u00e7\u00e3o",
            "RP",
        ]
        col_order = [c for c in col_order if c in df.columns]
        if col_order:
            df = df[col_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="EST_EMP", header=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["EST_EMP"]
            cell_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            header_fmt = workbook.add_format({"font_name": "Helvetica", "font_size": 8})
            worksheet.set_default_row(12, cell_fmt)
            if len(df.columns) > 0:
                worksheet.set_column(0, len(df.columns) - 1, None, cell_fmt)
                worksheet.write_row(0, 0, df.columns, header_fmt)
                worksheet.set_row(0, None, header_fmt)
        output.seek(0)
        filename = f"est_emp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_bytes(output, filename)
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/relatorios/plan20-seduc/download", methods=["GET"])
@login_required
@require_feature("relatorios/plan20-seduc")
def api_relatorio_plan20_download():
    def _to_float(val):
        try:
            if val in (None, ""):
                return 0.0
            if isinstance(val, str):
                cleaned = val.replace(".", "").replace(",", ".")
                return float(cleaned)
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        publico_transversal_chave,
                        programa,
                        funcao,
                        unidade_orcamentaria,
                        acao_paoe,
                        subfuncao,
                        objetivo_especifico,
                        esfera,
                        responsavel_acao,
                        produto_acao,
                        unid_medida_produto,
                        regiao_produto,
                        meta_produto,
                        saldo_meta_produto,
                        publico_transversal,
                        subacao_entrega,
                        responsavel,
                        prazo,
                        unid_gestora,
                        unidade_setorial_planejamento,
                        produto_subacao,
                        unidade_medida,
                        regiao_subacao,
                        codigo,
                        municipios_entrega,
                        meta_subacao,
                        detalhamento_produto,
                        etapa,
                        responsavel_etapa,
                        prazo_etapa,
                        regiao_etapa,
                        natureza,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        subelemento,
                        fonte,
                        idu,
                        descricao_item_despesa,
                        unid_medida_item,
                        quantidade,
                        valor_unitario,
                        valor_total
                    FROM plan20_seduc
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404
        db.session.close()

        headers = [
            ("Exercício", "exercicio"),
            ("Chave de Planejamento", "chave_planejamento"),
            ("Região", "regiao"),
            ("Subfunção + UG", "subfuncao_ug"),
            ("ADJ", "adj"),
            ("Macropolitica", "macropolitica"),
            ("Pilar", "pilar"),
            ("Eixo", "eixo"),
            ("Politica_Decreto", "politica_decreto"),
            ("Público Transversal (chave)", "publico_transversal_chave"),
            ("Programa", "programa"),
            ("Função", "funcao"),
            ("Unidade Orçamentária", "unidade_orcamentaria"),
            ("Ação (P/A/OE)", "acao_paoe"),
            ("Subfunção", "subfuncao"),
            ("Objetivo Específico", "objetivo_especifico"),
            ("Esfera", "esfera"),
            ("Responsável pela Ação", "responsavel_acao"),
            ("Produto(s) da Ação", "produto_acao"),
            ("Unidade de Medida do Produto", "unid_medida_produto"),
            ("Região do Produto", "regiao_produto"),
            ("Meta do Produto", "meta_produto"),
            ("Saldo Meta do Produto", "saldo_meta_produto"),
            ("P\u00fablico Transversal", "publico_transversal"),
            ("SubA\u00e7\u00e3o/entrega", "subacao_entrega"),
            ("Respons\u00e1vel", "responsavel"),
            ("Prazo", "prazo"),
            ("Unid. Gestora", "unid_gestora"),
            ("Unidade Setorial de Planejamento", "unidade_setorial_planejamento"),
            ("Produto da SubA\u00e7\u00e3o", "produto_subacao"),
            ("Unidade de Medida", "unidade_medida"),
            ("Regi\u00e3o da SubA\u00e7\u00e3o", "regiao_subacao"),
            ("C\u00f3digo", "codigo"),
            ("Munic\u00edpio(s) da entrega", "municipios_entrega"),
            ("Meta da SubA\u00e7\u00e3o", "meta_subacao"),
            ("Detalhamento do produto", "detalhamento_produto"),
            ("Etapa", "etapa"),
            ("Respons\u00e1vel da Etapa", "responsavel_etapa"),
            ("Prazo da Etapa", "prazo_etapa"),
            ("Regi\u00e3o da Etapa", "regiao_etapa"),
            ("Natureza", "natureza"),
            ("Cat.Econ", "cat_econ"),
            ("Grupo", "grupo"),
            ("Modalidade", "modalidade"),
            ("Elemento", "elemento"),
            ("Subelemento", "subelemento"),
            ("Fonte", "fonte"),
            ("IDU", "idu"),
            ("Descri\u00e7\u00e3o do Item de Despesa", "descricao_item_despesa"),
            ("Unid. Medida", "unid_medida_item"),
            ("Quantidade", "quantidade"),
            ("Valor Unit\u00e1rio", "valor_unitario"),
            ("Valor Total", "valor_total"),
        ]

        data = []
        for r in rows:
            row_dict = {}
            for label, key in headers:
                val = r.get(key)
                if key in {"quantidade", "valor_unitario", "valor_total"}:
                    val = _to_float(val)
                row_dict[label] = val
            data.append(row_dict)

        df = None
        try:
            import pandas as pd
            from io import BytesIO
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            df = pd.DataFrame(data, columns=[h[0] for h in headers])
            output = BytesIO()
            df.to_excel(output, index=False)
            output.seek(0)

            wb = load_workbook(output)
            ws = wb.active
            font = Font(name="Helvetica", size=8)
            idx_map = {label: i + 1 for i, (label, _) in enumerate(headers)}
            numeric_cols = {
                idx_map.get("Quantidade"),
                idx_map.get("Valor Unit\u00e1rio"),
                idx_map.get("Valor Total"),
            }
            numeric_cols = {c for c in numeric_cols if c}
            number_format = "#,##0.00"
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    if cell.col_idx in numeric_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = number_format
                    if cell.col_idx == idx_map.get("Exerc\u00edcio") and isinstance(cell.value, (int, float, str)):
                        try:
                            cell.value = int(str(cell.value).split(".")[0])
                        except Exception:
                            pass
                        cell.number_format = "0"

            styled = BytesIO()
            wb.save(styled)
            styled.seek(0)

            filename = f"plan20_seduc_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return _send_excel_bytes(styled, filename)
        except Exception as exc:
            return jsonify({"error": f"Falha ao preparar planilha: {exc}"}), 500
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/relatorios/plan21-nger/download", methods=["GET"])
@login_required
@require_feature("relatorios/plan21-nger")
def api_relatorio_plan21_nger_download():
    def _to_float(val):
        dec = _parse_decimal(val)
        return float(dec) if dec is not None else 0.0

    try:
        rows = (
            db.session.execute(
                text(
                    """
                    SELECT
                        exercicio,
                        chave_planejamento,
                        regiao,
                        subfuncao_ug,
                        adj,
                        macropolitica,
                        pilar,
                        eixo,
                        politica_decreto,
                        publico_transversal_chave,
                        programa,
                        funcao,
                        unidade_orcamentaria,
                        acao_paoe,
                        subfuncao,
                        objetivo_especifico,
                        esfera,
                        responsavel_acao,
                        produto_acao,
                        unid_medida_produto,
                        regiao_produto,
                        meta_produto,
                        saldo_meta_produto,
                        meta_credito,
                        meta_anulada,
                        meta_atual,
                        publico_transversal,
                        subacao_entrega,
                        responsavel,
                        prazo,
                        unid_gestora,
                        unidade_setorial_planejamento,
                        produto_subacao,
                        unidade_medida,
                        regiao_subacao,
                        codigo,
                        municipios_entrega,
                        meta_subacao,
                        detalhamento_produto,
                        etapa,
                        responsavel_etapa,
                        prazo_etapa,
                        regiao_etapa,
                        natureza,
                        cat_econ,
                        grupo,
                        modalidade,
                        elemento,
                        subelemento,
                        fonte,
                        idu,
                        descricao_item_despesa,
                        unid_medida_item,
                        quantidade,
                        valor_unitario,
                        valor_total,
                        suplementacao,
                        anulacao,
                        valor_atual
                    FROM plan21_nger
                    WHERE ativo = 1
                    """
                )
            )
            .mappings()
            .all()
        )
        if not rows:
            return jsonify({"error": "Nenhum dado para exportar."}), 404
        db.session.close()

        headers = [
            ("Exerc\u00edcio", "exercicio"),
            ("Chave de Planejamento", "chave_planejamento"),
            ("Regi\u00e3o", "regiao"),
            ("Subfun\u00e7\u00e3o + UG", "subfuncao_ug"),
            ("ADJ", "adj"),
            ("Macropolitica", "macropolitica"),
            ("Pilar", "pilar"),
            ("Eixo", "eixo"),
            ("Politica_Decreto", "politica_decreto"),
            ("Público Transversal (chave)", "publico_transversal_chave"),
            ("Programa", "programa"),
            ("Função", "funcao"),
            ("Unidade Orçamentária", "unidade_orcamentaria"),
            ("Ação (P/A/OE)", "acao_paoe"),
            ("Subfunção", "subfuncao"),
            ("Objetivo Específico", "objetivo_especifico"),
            ("Esfera", "esfera"),
            ("Responsável pela Ação", "responsavel_acao"),
            ("Produto(s) da Ação", "produto_acao"),
            ("Unidade de Medida do Produto", "unid_medida_produto"),
            ("Região do Produto", "regiao_produto"),
            ("Meta do Produto", "meta_produto"),
            ("Meta Crédito", "meta_credito"),
            ("Meta Anulada", "meta_anulada"),
            ("Meta Atual", "meta_atual"),
            ("Saldo Meta do Produto", "saldo_meta_produto"),
            ("Público Transversal", "publico_transversal"),
            ("Subação/entrega", "subacao_entrega"),
            ("Responsável", "responsavel"),
            ("Prazo", "prazo"),
            ("Unid. Gestora", "unid_gestora"),
            ("Unidade Setorial de Planejamento", "unidade_setorial_planejamento"),
            ("Produto da Subação", "produto_subacao"),
            ("Unidade de Medida", "unidade_medida"),
            ("Região da Subação", "regiao_subacao"),
            ("Código", "codigo"),
            ("Município(s) da entrega", "municipios_entrega"),
            ("Meta da Subação", "meta_subacao"),
            ("Detalhamento do produto", "detalhamento_produto"),
            ("Etapa", "etapa"),
            ("Responsável da Etapa", "responsavel_etapa"),
            ("Prazo da Etapa", "prazo_etapa"),
            ("Região da Etapa", "regiao_etapa"),
            ("Natureza", "natureza"),
            ("Cat.Econ", "cat_econ"),
            ("Grupo", "grupo"),
            ("Modalidade", "modalidade"),
            ("Elemento", "elemento"),
            ("Subelemento", "subelemento"),
            ("Fonte", "fonte"),
            ("IDU", "idu"),
            ("Descrição do Item de Despesa", "descricao_item_despesa"),
            ("Unid. Medida", "unid_medida_item"),
            ("Quantidade", "quantidade"),
            ("Valor Unitário", "valor_unitario"),
            ("Valor Total", "valor_total"),
            ("Suplementação", "suplementacao"),
            ("Anulação", "anulacao"),
            ("Valor Atual", "valor_atual"),
        ]

        data = []
        numeric_keys = {
            "meta_produto",
            "saldo_meta_produto",
            "meta_credito",
            "meta_anulada",
            "meta_atual",
            "meta_subacao",
            "quantidade",
            "valor_unitario",
            "valor_total",
            "suplementacao",
            "anulacao",
            "valor_atual",
        }
        for r in rows:
            row_dict = {}
            for label, key in headers:
                val = r.get(key)
                if key in numeric_keys:
                    val = _to_float(val)
                row_dict[label] = val
            data.append(row_dict)

        df = None
        try:
            import pandas as pd
            from io import BytesIO
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            df = pd.DataFrame(data, columns=[h[0] for h in headers])
            output = BytesIO()
            df.to_excel(output, index=False)
            output.seek(0)

            wb = load_workbook(output)
            ws = wb.active
            font = Font(name="Helvetica", size=8)
            idx_map = {label: i + 1 for i, (label, _) in enumerate(headers)}
            numeric_cols = {
                idx_map.get("Meta do Produto"),
                idx_map.get("Saldo Meta do Produto"),
                idx_map.get("Meta Crédito"),
                idx_map.get("Meta Anulada"),
                idx_map.get("Meta Atual"),
                idx_map.get("Meta da Subação"),
                idx_map.get("Quantidade"),
                idx_map.get("Valor Unitário"),
                idx_map.get("Valor Total"),
                idx_map.get("Suplementação"),
                idx_map.get("Anulação"),
                idx_map.get("Valor Atual"),
            }
            numeric_cols = {c for c in numeric_cols if c}
            number_format = "#,##0.00"
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    if cell.col_idx in numeric_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = number_format
                    if cell.col_idx == idx_map.get("Exercício") and isinstance(
                        cell.value, (int, float, str)
                    ):
                        try:
                            cell.value = int(str(cell.value).split(".")[0])
                        except Exception:
                            pass
                        cell.number_format = "0"

            styled = BytesIO()
            wb.save(styled)
            styled.seek(0)

            filename = f"plan21_nger_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return _send_excel_bytes(styled, filename)
        except Exception as exc:
            return jsonify({"error": f"Falha ao preparar planilha: {exc}"}), 500
    except Exception as exc:
        return jsonify({"error": f"Falha ao exportar: {exc}"}), 500


@home_bp.route("/api/usuarios", methods=["POST"])
@login_required
@require_feature("usuarios/cadastrar")
def api_criar_usuario():
    data = request.get_json() or {}
    email = (data.get("email") or "").lower().strip()
    nome = (data.get("nome") or "").strip()
    perfil_id_raw = data.get("perfil_id")
    senha = (data.get("senha") or "").strip()
    ativo = bool(data.get("ativo", True))

    if not email or not nome or not senha:
        return jsonify({"error": "Campos obrigatorios ausentes."}), 400

    perfil_row = None
    if perfil_id_raw in (None, ""):
        return jsonify({"error": "Perfil obrigatorio."}), 400
    try:
        perfil_row = db.session.get(Perfil, int(perfil_id_raw))
    except (TypeError, ValueError):
        return jsonify({"error": "Perfil invalido."}), 400
    if not perfil_row:
        return jsonify({"error": "Perfil nao encontrado."}), 400

    caller_nivel = getattr(g, "user_nivel", None)
    # Apenas nivel 1 pode criar usuario com perfil nivel 1
    if caller_nivel != 1 and _is_nivel1(perfil_id=perfil_row.id):
        return jsonify({"error": "Apenas admin pode criar perfil admin."}), 403

    existing = Usuario.query.filter_by(email=email).first()
    if existing:
        return jsonify({"error": "Usuario ja existe."}), 400

    usuario = Usuario(
        email=email,
        nome=nome,
        perfil_id=perfil_row.id,
        perfil=(perfil_row.nome or "").strip(),
        ativo=ativo,
    )
    usuario.set_password(senha)
    db.session.add(usuario)
    db.session.commit()

    return jsonify({"ok": True, "message": "Usuario criado."}), 201


@home_bp.route("/api/usuarios/<email>", methods=["GET", "PUT", "DELETE", "POST"])
@login_required
def api_usuario(email):
    def _safe_unquote(value: str) -> str:
        v1 = unquote(value or "")
        v2 = unquote(v1)
        return v2

    def _normalize_email(value: str) -> str:
        return re.sub(r"[\s\u00a0\u200b\u200c\u200d\ufeff]+", "", value or "").lower()

    email_raw = _safe_unquote(email or "")
    email_norm = email_raw.strip().lower()
    email_compact = _normalize_email(email_raw)
    usuario = Usuario.query.filter(func.lower(func.trim(Usuario.email)) == email_norm).first()
    if not usuario and email_compact:
        expr = func.lower(Usuario.email)
        for ch in (" ", "\t", "\n", "\r", "\u00a0", "\u200b", "\u200c", "\u200d", "\ufeff"):
            expr = func.replace(expr, ch, "")
        usuario = Usuario.query.filter(expr == email_compact).first()
    if not usuario and email_compact:
        for cand in Usuario.query.all():
            if _normalize_email(cand.email or "") == email_compact:
                usuario = cand
                break
    if not usuario:
        return jsonify({"error": "Usuario nao encontrado."}), 404

    caller_nivel = getattr(g, "user_nivel", None)
    target_is_nivel1 = _is_nivel1(perfil_id=getattr(usuario, "perfil_id", None))
    if caller_nivel != 1 and target_is_nivel1:
        return jsonify({"error": "Apenas admin pode alterar usuario admin."}), 403

    if request.method == "GET":
        if not (has_permission("usuarios/editar") or has_permission("usuarios/senha") or getattr(g, "user_nivel", None) == 1):
            return jsonify({"error": "Sem permissao."}), 403
        return jsonify(
            {
                "email": usuario.email,
                "nome": usuario.nome,
                "perfil": usuario.perfil,
                "perfil_id": getattr(usuario, "perfil_id", None),
                "ativo": usuario.ativo,
            }
        )

    if request.method == "DELETE":
        if not (has_permission("usuarios/editar") or getattr(g, "user_nivel", None) == 1):
            return jsonify({"error": "Sem permissao."}), 403
        usuario.ativo = False
        db.session.commit()
        return jsonify({"ok": True, "message": "Usuario desativado."})

    if request.method not in ("PUT", "POST"):
        return jsonify({"error": "Metodo nao permitido."}), 405

    if not (has_permission("usuarios/editar") or getattr(g, "user_nivel", None) == 1):
        return jsonify({"error": "Sem permissao."}), 403

    data = request.get_json() or {}
    usuario.nome = (data.get("nome") or usuario.nome).strip()
    novo_perfil_id_raw = data.get("perfil_id")
    novo_perfil_row = None
    if novo_perfil_id_raw in (None, ""):
        return jsonify({"error": "Perfil obrigatorio."}), 400
    try:
        novo_perfil_row = db.session.get(Perfil, int(novo_perfil_id_raw))
    except (TypeError, ValueError):
        return jsonify({"error": "Perfil invalido."}), 400
    if not novo_perfil_row:
        return jsonify({"error": "Perfil nao encontrado."}), 400
    if caller_nivel != 1 and _is_nivel1(perfil_id=novo_perfil_row.id):
        return jsonify({"error": "Apenas admin pode definir perfil admin."}), 403
    usuario.perfil_id = novo_perfil_row.id
    usuario.perfil = (novo_perfil_row.nome or "").strip()
    usuario.ativo = bool(data.get("ativo", usuario.ativo))
    nova_senha = (data.get("senha") or "").strip()
    if nova_senha:
        usuario.set_password(nova_senha)
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao atualizar: {exc}"}), 500
    return jsonify({"ok": True, "message": "Usuario atualizado."})


@home_bp.route("/api/usuarios/id/<int:user_id>", methods=["GET", "PUT", "DELETE", "POST"])
@login_required
def api_usuario_por_id(user_id: int):
    usuario = Usuario.query.filter(Usuario.id == user_id).first()
    if not usuario:
        return jsonify({"error": "Usuario nao encontrado."}), 404

    caller_nivel = getattr(g, "user_nivel", None)
    target_is_nivel1 = _is_nivel1(perfil_id=getattr(usuario, "perfil_id", None))
    if caller_nivel != 1 and target_is_nivel1:
        return jsonify({"error": "Apenas admin pode alterar usuario admin."}), 403

    if request.method == "GET":
        if not (has_permission("usuarios/editar") or has_permission("usuarios/senha") or getattr(g, "user_nivel", None) == 1):
            return jsonify({"error": "Sem permissao."}), 403
        return jsonify(
            {
                "id": usuario.id,
                "email": usuario.email,
                "nome": usuario.nome,
                "perfil": usuario.perfil,
                "perfil_id": getattr(usuario, "perfil_id", None),
                "ativo": usuario.ativo,
            }
        )

    if request.method == "DELETE":
        if not (has_permission("usuarios/editar") or getattr(g, "user_nivel", None) == 1):
            return jsonify({"error": "Sem permissao."}), 403
        usuario.ativo = False
        db.session.commit()
        return jsonify({"ok": True, "message": "Usuario desativado."})

    if request.method not in ("PUT", "POST"):
        return jsonify({"error": "Metodo nao permitido."}), 405

    if not (has_permission("usuarios/editar") or getattr(g, "user_nivel", None) == 1):
        return jsonify({"error": "Sem permissao."}), 403

    data = request.get_json() or {}
    novo_email = (data.get("email") or "").strip().lower()
    if novo_email and novo_email != (usuario.email or "").lower():
        existing = Usuario.query.filter(func.lower(Usuario.email) == novo_email).first()
        if existing:
            return jsonify({"error": "Email ja em uso."}), 400
        usuario.email = novo_email

    usuario.nome = (data.get("nome") or usuario.nome).strip()
    novo_perfil_id_raw = data.get("perfil_id")
    novo_perfil_row = None
    if novo_perfil_id_raw in (None, ""):
        return jsonify({"error": "Perfil obrigatorio."}), 400
    try:
        novo_perfil_row = db.session.get(Perfil, int(novo_perfil_id_raw))
    except (TypeError, ValueError):
        return jsonify({"error": "Perfil invalido."}), 400
    if not novo_perfil_row:
        return jsonify({"error": "Perfil nao encontrado."}), 400
    if caller_nivel != 1 and _is_nivel1(perfil_id=novo_perfil_row.id):
        return jsonify({"error": "Apenas admin pode definir perfil admin."}), 403
    usuario.perfil_id = novo_perfil_row.id
    usuario.perfil = (novo_perfil_row.nome or "").strip()
    usuario.ativo = bool(data.get("ativo", usuario.ativo))
    nova_senha = (data.get("senha") or "").strip()
    if nova_senha:
        usuario.set_password(nova_senha)
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Falha ao atualizar: {exc}"}), 500
    return jsonify({"ok": True, "message": "Usuario atualizado."})


@home_bp.route("/api/usuarios/<email>/senha", methods=["POST"])
@login_required
def api_usuario_senha(email):
    def _safe_unquote(value: str) -> str:
        v1 = unquote(value or "")
        v2 = unquote(v1)
        return v2

    def _normalize_email(value: str) -> str:
        return re.sub(r"[\s\u00a0\u200b\u200c\u200d\ufeff]+", "", value or "").lower()

    email_raw = _safe_unquote(email or "")
    email_norm = email_raw.strip().lower()
    email_compact = _normalize_email(email_raw)
    usuario = Usuario.query.filter(func.lower(func.trim(Usuario.email)) == email_norm).first()
    if not usuario and email_compact:
        expr = func.lower(Usuario.email)
        for ch in (" ", "\t", "\n", "\r", "\u00a0", "\u200b", "\u200c", "\u200d", "\ufeff"):
            expr = func.replace(expr, ch, "")
        usuario = Usuario.query.filter(expr == email_compact).first()
    if not usuario and email_compact:
        for cand in Usuario.query.all():
            if _normalize_email(cand.email or "") == email_compact:
                usuario = cand
                break
    if not usuario:
        return jsonify({"error": "Usuario nao encontrado."}), 404
    caller_nivel = getattr(g, "user_nivel", None)
    target_is_nivel1 = _is_nivel1(perfil_id=getattr(usuario, "perfil_id", None))
    if caller_nivel != 1 and target_is_nivel1:
        return jsonify({"error": "Apenas admin pode alterar usuario admin."}), 403
    if not (has_permission("usuarios/senha") or getattr(g, "user_nivel", None) == 1):
        return jsonify({"error": "Sem permissao."}), 403

    data = request.get_json() or {}
    senha_atual = (data.get("senha_atual") or "").strip()
    senha_nova = (data.get("senha_nova") or "").strip()
    senha_confirmar = (data.get("senha_confirmar") or "").strip()

    if not senha_atual or not senha_nova or not senha_confirmar:
        return jsonify({"error": "Preencha todos os campos de senha."}), 400
    if senha_nova != senha_confirmar:
        return jsonify({"error": "Confirmacao diferente da nova senha."}), 400
    if not usuario.check_password(senha_atual):
        return jsonify({"error": "Senha atual incorreta."}), 400

    usuario.set_password(senha_nova)
    db.session.commit()
    return jsonify({"ok": True, "message": "Senha atualizada."})


@home_bp.route("/api/usuarios/id/<int:user_id>/senha", methods=["POST"])
@login_required
def api_usuario_senha_por_id(user_id: int):
    usuario = Usuario.query.filter(Usuario.id == user_id).first()
    if not usuario:
        return jsonify({"error": "Usuario nao encontrado."}), 404
    caller_nivel = getattr(g, "user_nivel", None)
    target_is_nivel1 = _is_nivel1(perfil_id=getattr(usuario, "perfil_id", None))
    if caller_nivel != 1 and target_is_nivel1:
        return jsonify({"error": "Apenas admin pode alterar usuario admin."}), 403
    if not (has_permission("usuarios/senha") or getattr(g, "user_nivel", None) == 1):
        return jsonify({"error": "Sem permissao."}), 403

    data = request.get_json() or {}
    senha_atual = (data.get("senha_atual") or "").strip()
    senha_nova = (data.get("senha_nova") or "").strip()
    senha_confirmar = (data.get("senha_confirmar") or "").strip()

    if not senha_atual or not senha_nova or not senha_confirmar:
        return jsonify({"error": "Preencha todos os campos de senha."}), 400
    if senha_nova != senha_confirmar:
        return jsonify({"error": "Confirmacao diferente da nova senha."}), 400
    if not usuario.check_password(senha_atual):
        return jsonify({"error": "Senha atual incorreta."}), 400

    usuario.set_password(senha_nova)
    db.session.commit()
    return jsonify({"ok": True, "message": "Senha atualizada."})


@home_bp.route("/api/perfis", methods=["GET", "POST"])
@login_required
@role_required("admin")
def api_perfis():
    if request.method == "GET":
        perfis = Perfil.query.order_by(Perfil.nivel, Perfil.nome).all()
        return jsonify(
            [
                {"id": p.id, "nome": p.nome, "nivel": p.nivel, "ativo": p.ativo}
                for p in perfis
            ]
        )
    data = request.get_json() or {}
    nome = (data.get("nome") or "").strip()
    nivel = data.get("nivel")
    ativo = bool(data.get("ativo", True))
    if not nome or nivel is None:
        return jsonify({"error": "Campos obrigatorios ausentes."}), 400
    existing = Perfil.query.filter_by(nome=nome).first()
    if existing:
        if existing.ativo:
            return jsonify({"error": "Perfil ja existe."}), 400
        # Reaproveita perfil inativo existente
        existing.nivel = int(nivel)
        existing.ativo = ativo
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify({"error": "Perfil ja existe."}), 400
        return jsonify({"ok": True, "message": "Perfil ativado.", "id": existing.id}), 200
    try:
        nivel_int = int(nivel)
    except (TypeError, ValueError):
        return jsonify({"error": "Nivel invalido."}), 400
    if nivel_int < 1 or nivel_int > 5:
        return jsonify({"error": "Nivel deve estar entre 1 e 5."}), 400
    perfil = Perfil(id=_next_pk(Perfil), nome=nome, nivel=nivel_int, ativo=ativo)
    db.session.add(perfil)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "Perfil ja existe."}), 400
    return jsonify({"ok": True, "message": "Perfil criado.", "id": perfil.id}), 201


@home_bp.route("/api/perfis/<int:perfil_id>", methods=["PUT", "DELETE"])
@login_required
@role_required("admin")
def api_perfil(perfil_id):
    perfil = db.session.get(Perfil, perfil_id)
    if not perfil:
        return jsonify({"error": "Perfil nao encontrado."}), 404

    if request.method == "DELETE":
        try:
            PerfilPermissao.query.filter_by(perfil_id=perfil_id).delete()
            db.session.delete(perfil)
            db.session.commit()
            return jsonify({"ok": True, "message": "Perfil excluido."})
        except IntegrityError:
            db.session.rollback()
            return jsonify({"error": "NÃƒÂ£o foi possÃƒÂ­vel excluir este perfil."}), 400

    data = request.get_json() or {}
    nome = (data.get("nome") or perfil.nome).strip()
    nivel = data.get("nivel", perfil.nivel)
    ativo = bool(data.get("ativo", perfil.ativo))
    existing = Perfil.query.filter(Perfil.nome == nome, Perfil.id != perfil_id).first()
    if existing:
        if existing.ativo:
            return jsonify({"error": "Perfil ja existe."}), 400
        # remover perfil inativo duplicado para liberar nome
        db.session.delete(existing)
    try:
        nivel_int = int(nivel)
    except (TypeError, ValueError):
        return jsonify({"error": "Nivel invalido."}), 400
    if nivel_int < 1 or nivel_int > 5:
        return jsonify({"error": "Nivel deve estar entre 1 e 5."}), 400

    perfil.nome = nome
    perfil.nivel = nivel_int
    perfil.ativo = ativo
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "Perfil ja existe."}), 400
    return jsonify({"ok": True, "message": "Perfil atualizado."})
    if chave_field == "chave_planejamento":
        ped_rows = [r for r in ped_rows if r]
        ped_rows = (
            PedRegistro.query.with_entities(PedRegistro.valor_ped)
            .filter(
                PedRegistro.ativo == True,  # noqa: E712
                PedRegistro.exercicio == exercicio,
                PedRegistro.programa_governo == programa_key,
                PedRegistro.paoe == acao_paoe_key,
                PedRegistro.fonte == fonte,
                PedRegistro.iduso == iduso,
                PedRegistro.uo == uo_norm,
                PedRegistro.subfuncao_ug.like(f"%.{ug_norm}"),
                PedRegistro.regiao == regiao,
                PedRegistro.chave_planejamento == chave_planejamento,
            )
            .all()
        )
    else:
        ped_rows = (
            PedRegistro.query.with_entities(PedRegistro.valor_ped)
            .filter(
                PedRegistro.ativo == True,  # noqa: E712
                PedRegistro.exercicio == exercicio,
                PedRegistro.programa_governo == programa_key,
                PedRegistro.paoe == acao_paoe_key,
                PedRegistro.fonte == fonte,
                PedRegistro.iduso == iduso,
                PedRegistro.uo == uo_norm,
                PedRegistro.subfuncao_ug.like(f"%.{ug_norm}"),
                PedRegistro.regiao == regiao,
                PedRegistro.chave == chave_planejamento,
            )
            .all()
        )
    valor_ped = sum((_dec_or_zero(r.valor_ped) for r in ped_rows), Decimal("0"))
    ped_count = len(ped_rows)









